"""
=============================================================================
COSME OFFLINE DASHBOARD GENERATOR
Generates a standalone HTML dashboard with all charts embedded.
No server required — just open the HTML file in any browser.

Usage:  python generate_offline_dashboard.py
Output: COSME_Offline_Dashboard.html (in same folder)

Requirements: pip install pandas numpy plotly openpyxl
=============================================================================
"""

import pandas as pd
import numpy as np
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import os, re, sys, json
from collections import Counter
from datetime import datetime

# ============================================================================
# CONFIGURATION (same as cosme_dashboard.py)
# ============================================================================
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

FORESTRY_EXCEL = os.path.join(SCRIPT_DIR, "Forest Functionality Basline_midline results.xlsx")
FORESTRY_SHEET = "Results"
WOMEN_EXCEL = os.path.join(SCRIPT_DIR, "Women Survey Basline_midline results.xlsx")
WOMEN_SHEET = "Results Women"
MEN_EXCEL = os.path.join(SCRIPT_DIR, "Men Survey Basline_midline results.xlsx")
MEN_SHEET = "Results Men"
GJJ_KAP_WOMEN_EXCEL = os.path.join(SCRIPT_DIR, "GJJ KAP Women Basline_endline results.xlsx")
GJJ_KAP_WOMEN_SHEET = "Results KAP Women Endline"
GJJ_KAP_MEN_EXCEL = os.path.join(SCRIPT_DIR, "GJJ KAP Men Basline_endline results.xlsx")
GJJ_KAP_MEN_SHEET = "Results KAP Men Endline"
FOREST_TRAINING_EXCEL = os.path.join(SCRIPT_DIR, "Forest Training Pre_post results.xlsx")
FOREST_TRAINING_SHEET = "Results"
MANGROVE_TRAINING_EXCEL = os.path.join(SCRIPT_DIR, "Mangrove Training Pre_post results.xlsx")
MANGROVE_TRAINING_SHEET = "Results"
SEAWEED_CSV = os.path.join(SCRIPT_DIR, "Seaweed Data Collection_06.12.2025.csv")
PROJECT_OUTPUTS_EXCEL = os.path.join(SCRIPT_DIR, "Project Outputs.xlsx")
VSLA_EXCEL = os.path.join(SCRIPT_DIR, "VSLA Functionality_(Q1-Q4) 2025.xlsx")
VSLA_SHEET = "Results (Across Qs)"

FOREST_TRAINING_DOMAINS = {
    'PFM Concepts': [1, 2, 3, 4, 11],
    'Gender & Inclusivity': [6, 7, 8],
    'Forest Management Practices': [9, 10, 12],
    'Biodiversity Conservation': [13, 14, 15],
    'Agroforestry': [16, 17, 18],
    'Nursery Management': [19],
    'Climate Change': [20, 21],
}

# Theme colors
COLORS = {
    "baseline": "#5B8DB8", "midline": "#2E7D32",
    "baseline_light": "#A3C4DC", "midline_light": "#81C784",
    "accent": "#FF8F00", "danger": "#E53935",
    "good": "#43A047", "medium": "#FB8C00",
    "poor": "#E53935", "low": "#43A047",
    "high": "#E53935", "decrease": "#43A047",
    "increase": "#E53935", "no_change": "#78909C",
    "header_gradient": "linear-gradient(135deg, #1B5E20 0%, #2E7D32 50%, #388E3C 100%)",
    "card_border": "#2E7D32", "card_value": "#1B5E20",
    "narrative_bg": "#E8F5E9", "narrative_border": "#2E7D32",
    "narrative_text": "#1B5E20",
    "radar_bl_fill": "rgba(91,141,184,0.3)",
    "radar_ml_fill": "rgba(46,125,50,0.3)",
}

# ============================================================================
# UTILITY helpers (same as cosme_dashboard.py)
# ============================================================================
def _val(df, row, col):
    try:
        v = df.iloc[row, col]
        if v is None or (isinstance(v, float) and np.isnan(v)):
            return 0
        return v
    except (IndexError, KeyError):
        return 0

def _clean_label(s):
    if isinstance(s, str):
        s = re.sub(r'[\t]+', '', s).strip().rstrip('-')
        s = s.replace('_', ' ').strip()
        if s and s[0].islower():
            s = s.title()
        return s
    return str(s)

def pct(value):
    if isinstance(value, (int, float)):
        return f"{value * 100:.1f}%"
    return str(value)

def pp_change(bl, ml):
    if isinstance(bl, (int, float)) and isinstance(ml, (int, float)):
        return (ml - bl) * 100
    return 0

# ============================================================================
# DATA LOADERS (copied from cosme_dashboard.py, without @st.cache_data)
# ============================================================================

def load_forestry_data(filepath):
    raw = pd.read_excel(filepath, sheet_name=FORESTRY_SHEET, header=None)
    data = {}
    data['functionality_threshold'] = pd.DataFrame({
        'Timepoint': ['Baseline', 'Midline'],
        'Functional_60_pct': [_val(raw, 7, 2), _val(raw, 8, 2)],
        'Functional_70_pct': [_val(raw, 7, 3), _val(raw, 8, 3)],
    })
    data['functionality_scores'] = pd.DataFrame({
        'Timepoint': ['Baseline', 'Midline'],
        'Respondents': [_val(raw, 13, 2), _val(raw, 14, 2)],
        'Average': [_val(raw, 13, 3), _val(raw, 14, 3)],
        'Max': [_val(raw, 13, 4), _val(raw, 14, 4)],
        'Min': [_val(raw, 13, 5), _val(raw, 14, 5)],
    })
    data['functionality_domain'] = pd.DataFrame({
        'Timepoint': ['Baseline', 'Midline'],
        'Management': [_val(raw, 19, 2), _val(raw, 20, 2)],
        'Gender': [_val(raw, 19, 3), _val(raw, 20, 3)],
        'Effectiveness': [_val(raw, 19, 4), _val(raw, 20, 4)],
        'Overall': [_val(raw, 19, 5), _val(raw, 20, 5)],
    })
    data['num_groups'] = pd.DataFrame({
        'Timepoint': ['Baseline', 'Midline'],
        'Groups_Assessed': [_val(raw, 27, 2), _val(raw, 27, 3)],
    })
    data['avg_membership'] = pd.DataFrame({
        'Category': ['Male', 'Female', 'All'],
        'Baseline': [_val(raw, 32, 2), _val(raw, 33, 2), _val(raw, 34, 2)],
        'Midline': [_val(raw, 32, 3), _val(raw, 33, 3), _val(raw, 34, 3)],
    })
    data['years_dist'] = pd.DataFrame({
        'Category': ['0-3 years', '4-5 years', '6-10 years', '11+ years'],
        'Baseline': [_val(raw, 57, 2), _val(raw, 58, 2), _val(raw, 59, 2), _val(raw, 60, 2)],
        'Midline': [_val(raw, 57, 3), _val(raw, 58, 3), _val(raw, 59, 3), _val(raw, 60, 3)],
    })
    data['tenure'] = pd.DataFrame({
        'Category': ['Community', 'Government', 'Govt Devolved', 'Individual', 'Other'],
        'Baseline': [_val(raw, 65, 2), _val(raw, 66, 2), _val(raw, 67, 2), _val(raw, 68, 2), _val(raw, 69, 2)],
        'Midline': [_val(raw, 65, 3), _val(raw, 66, 3), _val(raw, 67, 3), _val(raw, 68, 3), _val(raw, 69, 3)],
    })
    data['goals_defined'] = pd.DataFrame({
        'Category': ['No', 'Partially', 'Yes'],
        'Baseline': [_val(raw, 91, 2), _val(raw, 92, 2), _val(raw, 93, 2)],
        'Midline': [_val(raw, 91, 3), _val(raw, 92, 3), _val(raw, 93, 3)],
    })
    data['women_leadership'] = pd.DataFrame({
        'Category': ['No Leadership, No Voice', 'No Leadership, Occasional Voice',
                      'Some Leadership', 'Significant Leadership'],
        'Baseline': [_val(raw, 163, 12), _val(raw, 164, 12), _val(raw, 165, 12), _val(raw, 166, 12)],
        'Midline': [_val(raw, 163, 13), _val(raw, 164, 13), _val(raw, 165, 13), _val(raw, 166, 13)],
    })
    data['training_coverage'] = pd.DataFrame({
        'Category': ['None', 'A Few Members', 'Many Members', 'Most Members'],
        'Baseline': [_val(raw, 186, 4), _val(raw, 187, 4), _val(raw, 188, 4), _val(raw, 189, 4)],
        'Midline': [_val(raw, 186, 5), _val(raw, 187, 5), _val(raw, 188, 5), _val(raw, 189, 5)],
    })
    cond_labels = ['Area', 'Volume/Biomass', 'Regeneration', 'Biodiversity', 'Ecosystem Services']
    data['forest_condition'] = pd.DataFrame({
        'Characteristic': cond_labels,
        'Baseline_Good': [_val(raw, r, 2) for r in range(271, 276)],
        'Baseline_Medium': [_val(raw, r, 3) for r in range(271, 276)],
        'Baseline_Poor': [_val(raw, r, 4) for r in range(271, 276)],
        'Midline_Good': [_val(raw, r, 5) for r in range(271, 276)],
        'Midline_Medium': [_val(raw, r, 6) for r in range(271, 276)],
        'Midline_Poor': [_val(raw, r, 7) for r in range(271, 276)],
    })
    threat_labels = ['Fire', 'Fuelwood', 'Charcoal', 'Other Wood', 'Poaching', 'Encroachment', 'Land Grabbing', 'Poles']
    data['threats'] = pd.DataFrame({
        'Threat': threat_labels,
        'Baseline_Low': [_val(raw, r, 2) for r in range(290, 298)],
        'Baseline_Medium': [_val(raw, r, 3) for r in range(290, 298)],
        'Baseline_High': [_val(raw, r, 4) for r in range(290, 298)],
        'Midline_Low': [_val(raw, r, 5) for r in range(290, 298)],
        'Midline_Medium': [_val(raw, r, 6) for r in range(290, 298)],
        'Midline_High': [_val(raw, r, 7) for r in range(290, 298)],
    })
    data['income_gen'] = pd.DataFrame({
        'Category': ['Yes', 'No'],
        'Baseline': [_val(raw, 337, 2), _val(raw, 338, 2)],
        'Midline': [_val(raw, 337, 3), _val(raw, 338, 3)],
    })
    data['income_sources'] = pd.DataFrame({
        'Source': ['Timber', 'Fuelwood', 'Wildlife', 'NWFPs', 'PES', 'Poles', 'Other'],
        'Baseline': [_val(raw, r, 2) for r in range(342, 349)],
        'Midline': [_val(raw, r, 3) for r in range(342, 349)],
    })
    data['agroforestry'] = pd.DataFrame({
        'Category': ['Yes', 'No'],
        'Baseline': [_val(raw, 362, 2), _val(raw, 363, 2)],
        'Midline': [_val(raw, 362, 3), _val(raw, 363, 3)],
    })
    data['af_types'] = pd.DataFrame({
        'Practice': ['Intercropping', 'Silvopasture', 'Alley Cropping', 'Forest Farming', 'Beekeeping'],
        'Baseline': [_val(raw, r, 2) for r in range(367, 372)],
        'Midline': [_val(raw, r, 3) for r in range(367, 372)],
    })
    return data


def load_women_data(filepath):
    raw = pd.read_excel(filepath, sheet_name=WOMEN_SHEET, header=None)
    w = {}
    w['location'] = pd.DataFrame({'Category': ['Marine', 'Terrestrial'], 'Baseline': [_val(raw, 11, 2), _val(raw, 12, 2)], 'Midline': [_val(raw, 11, 3), _val(raw, 12, 3)]})
    w['hh_type'] = pd.DataFrame({'Category': ['Female-Headed', 'Male-Headed'], 'Baseline': [_val(raw, 11, 11), _val(raw, 12, 11)], 'Midline': [_val(raw, 11, 12), _val(raw, 12, 12)]})
    w['marital'] = pd.DataFrame({'Category': ['Monogamously Married', 'Widowed', 'Polygamously Married', 'Separated', 'Single', 'Divorced'], 'Baseline': [_val(raw, r, 2) for r in range(16, 22)], 'Midline': [_val(raw, r, 3) for r in range(16, 22)]})
    w['education'] = pd.DataFrame({'Category': ['Primary', 'Pre-primary/None/Other', 'Secondary/Vocational', 'College+'], 'Baseline': [_val(raw, r, 11) for r in range(16, 20)], 'Midline': [_val(raw, r, 12) for r in range(16, 20)]})
    shock_labels = ['Drought', 'Heat/Cold Waves', 'Flooding', 'Climate Variability', 'Typhoons/Hurricanes', 'Wildfires', 'Severe Thunderstorms', 'Land Degradation', 'Tornadoes', 'Landslides']
    shock_rows = [64, 66, 67, 68, 69, 70, 71, 72, 73, 74]
    w['shocks'] = pd.DataFrame({'Shock': shock_labels, 'Baseline': [_val(raw, r, 2) for r in shock_rows], 'Midline': [_val(raw, r, 3) for r in shock_rows]})
    coping_labels = ['Skipped Meals', 'Took a Loan', 'Sold Small Livestock', 'Engaged in Wage Labour', 'Used Savings', 'Replaced Food with Less Preferred', 'Additional Livelihood Activities', 'Sold Large Livestock', 'Sold Household Assets', 'No Strategies', 'Informal Transfers', 'Formal Transfers (Govt/NGO)', 'Reduced Caloric Intake']
    w['coping'] = pd.DataFrame({'Strategy': coping_labels, 'Baseline': [_val(raw, r, 2) for r in range(78, 91)], 'Midline': [_val(raw, r, 3) for r in range(78, 91)]})
    w['water'] = pd.DataFrame({'Category': ['Yes', 'No'], 'Baseline': [_val(raw, 44, 2), _val(raw, 45, 2)], 'Midline': [_val(raw, 44, 3), _val(raw, 45, 3)]})
    w['personal_saving'] = pd.DataFrame({'Response': ['Yes', 'No'], 'Baseline': [_val(raw, 195, 11), _val(raw, 196, 11)], 'Midline': [_val(raw, 195, 12), _val(raw, 196, 12)]})
    w['saving_mechanism'] = pd.DataFrame({'Mechanism': ['At Home', 'Bank', 'Informal Lender', 'Informal Savings Group', 'MPESA', 'VSLA'], 'Baseline': [_val(raw, r, 2) for r in range(209, 215)], 'Midline': [_val(raw, r, 3) for r in range(209, 215)]})
    role_labels = ['Fetch Firewood', 'Income Provider', 'Clean/Sweep', 'Look After Livestock', 'Care for Ill/Elderly', 'Maintain Garden', 'Cook Food', 'Care for Children', 'Fetch Water', 'Wash Clothes']
    w['roles_should_joint'] = pd.DataFrame({'Role': role_labels, 'Baseline': [_val(raw, r, 2) for r in range(238, 248)], 'Midline': [_val(raw, r, 3) for r in range(238, 248)]})
    w['time_summary'] = pd.DataFrame({'Category': ['Unpaid Care Work', 'Productive Work', 'Community Conservation', 'Personal Development', 'Personal Care', 'Leisure', 'Other'], 'Baseline': [_val(raw, 273, 2), _val(raw, 283, 2), _val(raw, 289, 2), _val(raw, 291, 2), _val(raw, 296, 2), _val(raw, 298, 2), _val(raw, 300, 2)], 'Midline': [_val(raw, 273, 3), _val(raw, 283, 3), _val(raw, 289, 3), _val(raw, 291, 3), _val(raw, 296, 3), _val(raw, 298, 3), _val(raw, 300, 3)]})
    decision_labels = ['Routine HH Purchases', 'Women Work on Farm', 'Use of HH Income', 'Women Go to Market', 'Mangrove/Seaweed/Forest Work', 'Using HH Savings', 'Invest Borrowed/Saved Money', 'Children Education', 'Large HH Purchases', 'Small Business', 'Taking Out Loans']
    w['decision_should_joint'] = pd.DataFrame({'Decision': decision_labels, 'Baseline': [_val(raw, r, 2) for r in range(308, 319)], 'Midline': [_val(raw, r, 3) for r in range(308, 319)]})
    w['cc_heard'] = pd.DataFrame({'Response': ['Yes', 'No'], 'Baseline': [_val(raw, 356, 2), _val(raw, 357, 2)], 'Midline': [_val(raw, 356, 3), _val(raw, 357, 3)]})
    w['nbs_heard'] = pd.DataFrame({'Response': ['Yes', 'No'], 'Baseline': [_val(raw, 383, 2), _val(raw, 384, 2)], 'Midline': [_val(raw, 383, 3), _val(raw, 384, 3)]})
    w['cc_env_effects'] = pd.DataFrame({'Effect': ['Hotter Temperatures', 'Extreme Weather Events', 'Increased Drought', 'Warming Oceans', 'Sea Level Rise', 'Loss of Species', 'Migration of Species', 'Animal Migration Changes', 'Spread of Disease', 'Ocean Acidification', 'Increased Invasive Species', "I Don't Know"], 'Baseline': [_val(raw, r, 2) for r in range(368, 380)], 'Midline': [_val(raw, r, 3) for r in range(368, 380)]})
    w['nbs_examples'] = pd.DataFrame({'Example': ['Mangrove Restoration', 'Coral Reef Restoration', 'Reforestation', 'Sustainable Seaweed Farming', "I Don't Know", 'Rainwater Harvesting', 'Sustainable Agriculture', 'Solar Energy', 'Rainwater Harvesting 2'], 'Baseline': [_val(raw, r, 2) for r in range(388, 397)], 'Midline': [_val(raw, r, 3) for r in range(388, 397)]})
    ls_labels = ['I have many positive qualities', 'Respected in community', 'Life has meaning', 'Think about future', 'Specific goals', 'Know what to do for goals', 'Confident to achieve goals', 'Ability to lead', 'Bring people together', 'People seek my advice', 'Prioritize tasks', 'Convince others to join cause']
    ls_domains = ['Self Esteem']*3 + ['Aspirations']*4 + ['Leadership']*5
    ls_rows = list(range(476, 488))
    w['lifeskills_agree'] = pd.DataFrame({'Domain': ls_domains, 'Statement': ls_labels, 'Baseline': [_val(raw, r, 2) for r in ls_rows], 'Midline': [_val(raw, r, 3) for r in ls_rows]})
    sn_labels = ['Income → Husband Controls', 'Men Better Business Ideas', "Men Earn, Women Look After Home", 'Inappropriate Dress → Her Fault', 'Cook & Clean → Good Marriage', 'Embarrassing for Men to Do Chores', 'Planting Crops (Family Food)', 'Restoring Ecosystems (Mangrove/Forest)', 'Only Men Drive Boats', 'Ok for Women to Express Emotions']
    sn_rows = list(range(506, 516))
    w['socialnorms_agree'] = pd.DataFrame({'Norm': sn_labels, 'Baseline': [_val(raw, r, 2) for r in sn_rows], 'Midline': [_val(raw, r, 3) for r in sn_rows]})
    return w


def load_men_data(filepath):
    raw = pd.read_excel(filepath, sheet_name=MEN_SHEET, header=None)
    m = {}
    m['location'] = pd.DataFrame({'Category': ['Marine', 'Terrestrial'], 'Baseline': [_val(raw, 11, 2), _val(raw, 12, 2)], 'Midline': [_val(raw, 11, 3), _val(raw, 12, 3)]})
    m['education'] = pd.DataFrame({'Category': ['College Level or Higher', 'Pre-primary/None/Other', 'Primary', 'Secondary/Vocational'], 'Baseline': [_val(raw, 11, 11), _val(raw, 12, 11), _val(raw, 13, 11), _val(raw, 14, 11)], 'Midline': [_val(raw, 11, 12), _val(raw, 12, 12), _val(raw, 13, 12), _val(raw, 14, 12)]})
    m['cc_heard'] = pd.DataFrame({'Response': ['Yes', 'No'], 'Baseline': [_val(raw, 35, 2), _val(raw, 36, 2)], 'Midline': [_val(raw, 35, 3), _val(raw, 36, 3)]})
    m['nbs_heard'] = pd.DataFrame({'Response': ['Yes', 'No'], 'Baseline': [_val(raw, 62, 2), _val(raw, 63, 2)], 'Midline': [_val(raw, 62, 3), _val(raw, 63, 3)]})
    m['cc_env_effects'] = pd.DataFrame({'Effect': ['Hotter Temperatures', 'More Extreme Weather Events', 'Increased Drought', 'Warming Oceans/Water Bodies', 'Sea Level Rise', 'Loss of Species', 'Migration of Species', 'Animal Migration Pattern Changes', 'Spread of Disease/Algal Blooms', 'Ocean Acidification (Coral Reefs)', 'Increased Species Invasions', "I Don't Know"], 'Baseline': [_val(raw, r, 2) for r in range(47, 59)], 'Midline': [_val(raw, r, 3) for r in range(47, 59)]})
    role_labels = ['Cook Food', 'Clean/Sweep House', 'Look After Livestock', 'Maintain Garden', 'Care for Ill/Elderly/Disabled', 'Fetch Firewood', 'Fetch Water', 'Wash Clothes', 'Main Income Provider', 'Care for Children']
    m['roles_should_joint'] = pd.DataFrame({'Role': role_labels, 'Baseline': [_val(raw, r, 2) for r in range(151, 161)], 'Midline': [_val(raw, r, 3) for r in range(151, 161)]})
    decision_labels = ['Start Small Business', 'Large HH Purchases', 'Using HH Savings', 'Taking Out Loans', 'Children Education', 'Routine HH Purchases', 'Women Go to Mangrove/Seaweed/Forest', 'Use of HH Income', 'Women Go to Market', 'Women Go to Farm', 'Invest Borrowed/Saved Money']
    m['decision_should_joint'] = pd.DataFrame({'Decision': decision_labels, 'Baseline': [_val(raw, r, 2) for r in range(217, 228)], 'Midline': [_val(raw, r, 3) for r in range(217, 228)]})
    m['time_summary'] = pd.DataFrame({'Category': ['Unpaid Care Work', 'Productive Work', 'Community Conservation', 'Personal Development', 'Personal Care', 'Leisure', 'Other'], 'Baseline': [_val(raw, 183, 2), _val(raw, 192, 2), _val(raw, 198, 2), _val(raw, 200, 2), _val(raw, 205, 2), _val(raw, 207, 2), _val(raw, 209, 2)], 'Midline': [_val(raw, 183, 3), _val(raw, 192, 3), _val(raw, 198, 3), _val(raw, 200, 3), _val(raw, 205, 3), _val(raw, 207, 3), _val(raw, 209, 3)]})
    sn_labels = ['Income -> Husband Controls', 'Men Better Business Ideas', "Men Earn, Women Look After Home", 'Inappropriate Dress -> Her Fault', 'Cook & Clean -> Good Marriage', 'Embarrassing for Men to Do Chores', 'Planting Crops (Family Food)', 'Restoring Ecosystems (Mangrove/Forest)', 'Only Men Drive Boats', 'Ok for Women to Express Emotions', 'Supportive of Diverse People', 'Women Gain Rights -> Men Lose', 'Stronger Women -> Stronger Families']
    sn_rows = list(range(250, 263))
    m['socialnorms_agree'] = pd.DataFrame({'Norm': sn_labels, 'Baseline': [_val(raw, r, 2) for r in sn_rows], 'Midline': [_val(raw, r, 3) for r in sn_rows]})
    # Support types
    support_labels = ['Encouraged Participation', 'Sought Community Support', 'Supported with HH Chores', 'Supported with Restoration Work', 'Supported with Materials Purchase', 'None', 'Other']
    m['mangrove_support_type'] = pd.DataFrame({'Type': support_labels, 'Baseline': [_val(raw, r, 2) for r in range(91, 98)], 'Midline': [_val(raw, r, 3) for r in range(91, 98)]})
    return m


def load_gjj_kap_women_data(filepath):
    raw = pd.read_excel(filepath, sheet_name=GJJ_KAP_WOMEN_SHEET, header=None)
    g = {}
    self_statements = [_clean_label(_val(raw, 9, 1)), _clean_label(_val(raw, 10, 1)), _clean_label(_val(raw, 11, 1))]
    g['self_strongly_agree'] = pd.DataFrame({'Statement': self_statements, 'Baseline': [_val(raw, r, 2) for r in range(16, 19)], 'Endline': [_val(raw, r, 3) for r in range(16, 19)]})
    g['self_agreement'] = pd.DataFrame({'Statement': self_statements, 'Agreement_BL': [_val(raw, r, 2) for r in range(30, 33)], 'Agreement_EL': [_val(raw, r, 3) for r in range(30, 33)], 'Disagreement_BL': [_val(raw, r, 4) for r in range(30, 33)], 'Disagreement_EL': [_val(raw, r, 5) for r in range(30, 33)]})
    rel_statements = [_clean_label(_val(raw, r, 1)) for r in range(47, 54)]
    g['rel_af_rn'] = pd.DataFrame({'Statement': rel_statements, 'AF_Baseline': [_val(raw, r, 2) for r in range(69, 76)], 'AF_Endline': [_val(raw, r, 3) for r in range(69, 76)], 'RN_Baseline': [_val(raw, r, 4) for r in range(69, 76)], 'RN_Endline': [_val(raw, r, 5) for r in range(69, 76)]})
    g['shared_chores_yn'] = pd.DataFrame({'Response': [_clean_label(_val(raw, r, 1)) for r in range(99, 101)], 'Baseline': [_val(raw, r, 2) for r in range(99, 101)], 'Endline': [_val(raw, r, 3) for r in range(99, 101)]})
    g['equal_say'] = pd.DataFrame({'Response': [_clean_label(_val(raw, r, 1)) for r in range(197, 199)], 'Baseline': [_val(raw, r, 2) for r in range(197, 199)], 'Endline': [_val(raw, r, 3) for r in range(197, 199)]})
    g['decision_conversations'] = pd.DataFrame({'Response': [_clean_label(_val(raw, r, 1)) for r in range(167, 169)], 'Baseline': [_val(raw, r, 2) for r in range(167, 169)], 'Endline': [_val(raw, r, 3) for r in range(167, 169)]})
    g['decision_types'] = pd.DataFrame({'Decision': [_clean_label(_val(raw, r, 1)) for r in range(177, 188)], 'Baseline': [_val(raw, r, 2) for r in range(177, 188)], 'Endline': [_val(raw, r, 3) for r in range(177, 188)]})
    g['support_leader'] = pd.DataFrame({'Category': [_clean_label(_val(raw, r, 1)) for r in range(213, 219)], 'Baseline': [_val(raw, r, 2) for r in range(213, 219)], 'Endline': [_val(raw, r, 3) for r in range(213, 219)]})
    g['support_business'] = pd.DataFrame({'Category': [_clean_label(_val(raw, r, 1)) for r in range(222, 228)], 'Baseline': [_val(raw, r, 2) for r in range(222, 228)], 'Endline': [_val(raw, r, 3) for r in range(222, 228)]})
    return g


def load_gjj_kap_men_data(filepath):
    raw = pd.read_excel(filepath, sheet_name=GJJ_KAP_MEN_SHEET, header=None)
    g = {}
    def _men_label(r, c=1):
        v = raw.iloc[r, c]
        if pd.isna(v): return 'Other'
        return _clean_label(v)
    self_statements = [_men_label(r) for r in range(9, 12)]
    g['self_agreement'] = pd.DataFrame({'Statement': self_statements, 'Agreement_BL': [_val(raw, r, 2) for r in range(23, 26)], 'Agreement_EL': [_val(raw, r, 3) for r in range(23, 26)], 'Disagreement_BL': [_val(raw, r, 4) for r in range(23, 26)], 'Disagreement_EL': [_val(raw, r, 5) for r in range(23, 26)]})
    rel_statements = [_men_label(r) for r in range(41, 48)]
    g['rel_af_rn'] = pd.DataFrame({'Statement': rel_statements, 'AF_Baseline': [_val(raw, r, 2) for r in range(63, 70)], 'AF_Endline': [_val(raw, r, 3) for r in range(63, 70)], 'RN_Baseline': [_val(raw, r, 4) for r in range(63, 70)], 'RN_Endline': [_val(raw, r, 5) for r in range(63, 70)]})
    g['shared_chores_yn'] = pd.DataFrame({'Response': [_men_label(r) for r in range(83, 85)], 'Baseline': [_val(raw, r, 2) for r in range(83, 85)], 'Endline': [_val(raw, r, 3) for r in range(83, 85)]})
    g['decision_conversations'] = pd.DataFrame({'Response': [_men_label(r) for r in range(121, 123)], 'Baseline': [_val(raw, r, 2) for r in range(121, 123)], 'Endline': [_val(raw, r, 3) for r in range(121, 123)]})
    g['support_leader'] = pd.DataFrame({'Category': [_men_label(r) for r in range(133, 139)], 'Baseline': [_val(raw, r, 2) for r in range(133, 139)], 'Endline': [_val(raw, r, 3) for r in range(133, 139)]})
    g['support_business'] = pd.DataFrame({'Category': [_men_label(r) for r in range(143, 148)], 'Baseline': [_val(raw, r, 2) for r in range(143, 148)], 'Endline': [_val(raw, r, 3) for r in range(143, 148)]})
    return g


def load_forest_training_data(filepath):
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[FOREST_TRAINING_SHEET]
    def _cell(r, c):
        v = ws.cell(r, c).value
        if v is None: return 0.0
        try: return float(v)
        except: return 0.0
    def _cell_str(r, c):
        v = ws.cell(r, c).value
        return str(v).strip() if v is not None else ''
    threshold_labels = []
    for c in range(3, 7):
        lbl = _cell_str(7, c)
        if not lbl: lbl = f'>={[50,60,70,80][c-3]}'
        threshold_labels.append(lbl)
    threshold_rows = []
    for tp_row, tp_name in [(8, 'Baseline'), (9, 'Endline')]:
        for i, lbl in enumerate(threshold_labels):
            proportion = _cell(tp_row, 3 + i)
            threshold_rows.append({'Timepoint': tp_name, 'Threshold': lbl, 'Proportion': proportion})
    thresholds_df = pd.DataFrame(threshold_rows)
    score_rows = []
    for tp_row, tp_name in [(14, 'Baseline'), (15, 'Endline')]:
        score_rows.append({'Timepoint': tp_name, 'Respondents': int(_cell(tp_row, 3)), 'AverageScore': round(_cell(tp_row, 4), 2), 'MaxScore': round(_cell(tp_row, 5), 1), 'MinScore': round(_cell(tp_row, 6), 1)})
    scores_df = pd.DataFrame(score_rows)
    question_rows = []
    r = 20
    while r <= ws.max_row:
        q_text = _cell_str(r, 2)
        if not q_text or not q_text[0].isdigit(): r += 1; continue
        parts = q_text.split('.', 1)
        try: q_num = int(parts[0].strip())
        except ValueError: r += 1; continue
        q_label = parts[1].strip() if len(parts) > 1 else q_text
        bl_pct = _cell(r, 12); el_pct = _cell(r, 13)
        question_rows.append({'QuestionNumber': q_num, 'QuestionText': q_label, 'Baseline': bl_pct, 'Endline': el_pct})
        r += 1
    questions_df = pd.DataFrame(question_rows)
    if not questions_df.empty: questions_df = questions_df.sort_values('QuestionNumber').reset_index(drop=True)
    wb.close()
    return {'thresholds': thresholds_df, 'scores': scores_df, 'questions': questions_df}


def load_mangrove_training_data(filepath):
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[MANGROVE_TRAINING_SHEET]
    def _cell(r, c):
        v = ws.cell(r, c).value
        if v is None: return 0.0
        try: return float(v)
        except: return 0.0
    def _cell_str(r, c):
        v = ws.cell(r, c).value
        return str(v).strip() if v is not None else ''
    county_rows = []
    for row_idx in [22, 23, 24]:
        county = _cell_str(row_idx, 2)
        if not county: continue
        pre_val = _cell(row_idx, 3); post_val = _cell(row_idx, 4)
        county_rows.append({'County': county, 'Timepoint': 'Pre-Test', 'Value': pre_val})
        county_rows.append({'County': county, 'Timepoint': 'Post-Test', 'Value': post_val})
    adequate_county_df = pd.DataFrame(county_rows)
    score_rows = []
    for row_idx in [16, 17, 18]:
        county = _cell_str(row_idx, 2)
        if not county: continue
        score_rows.append({'County': county, 'Respondents': int(_cell(row_idx, 3)), 'AvgScore': round(_cell(row_idx, 4), 2), 'MaxScore': round(_cell(row_idx, 5), 2), 'MinScore': round(_cell(row_idx, 6), 2)})
    scores_df = pd.DataFrame(score_rows)
    sex_rows = []
    for row_idx in [28, 29, 30]:
        sex = _cell_str(row_idx, 2)
        if not sex: continue
        pre_val = _cell(row_idx, 3); post_val = _cell(row_idx, 4)
        sex_rows.append({'Sex': sex, 'Timepoint': 'Pre-Test', 'Value': pre_val})
        sex_rows.append({'Sex': sex, 'Timepoint': 'Post-Test', 'Value': post_val})
    adequate_sex_df = pd.DataFrame(sex_rows)
    wb.close()
    return {'scores': scores_df, 'adequate_county': adequate_county_df, 'adequate_sex': adequate_sex_df}


def load_seaweed_data(filepath):
    for enc in ['utf-16', 'utf-16-le', 'latin1', 'utf-8']:
        try:
            df = pd.read_csv(filepath, encoding=enc); break
        except: continue
    else:
        df = pd.read_csv(filepath, encoding='latin1')
    df = df[[c for c in df.columns if not c.startswith('Unnamed')]]
    df = df.dropna(subset=['Group', 'Member'], how='all').copy()
    num_cols = ['Ropes_Ocean', 'Ropes_Home', 'Ropes_Total', 'Target_Ropes', 'Ropes Required', 'Gap', 'Dried_KG', 'Wet_KG', 'Total_KG', 'x', 'y']
    for c in num_cols:
        if c in df.columns: df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0)
    flag_cols = ['flag_transport', 'flag_market', 'flag_disease', 'flag_equipment', 'flag_storage', 'flag_labour', 'flag_sand_tide']
    for c in flag_cols:
        if c in df.columns: df[c] = df[c].astype(str).str.strip().str.lower().isin(['true', '1', 'yes'])
    df['Ropes_Achievement_pct'] = df.apply(lambda r: round(r['Ropes_Total'] / r['Target_Ropes'] * 100, 1) if r['Target_Ropes'] > 0 else 0.0, axis=1)
    df['Production_per_rope_kg'] = df.apply(lambda r: round(r['Total_KG'] / r['Ropes_Total'], 2) if r['Ropes_Total'] > 0 else 0.0, axis=1)
    return df


def prepare_seaweed_aggregates(df):
    flag_cols = ['flag_transport', 'flag_market', 'flag_disease', 'flag_equipment', 'flag_storage', 'flag_labour', 'flag_sand_tide']
    flag_labels = {'flag_transport': 'Transport', 'flag_market': 'Market Access', 'flag_disease': 'Disease', 'flag_equipment': 'Equipment', 'flag_storage': 'Storage', 'flag_labour': 'Labour', 'flag_sand_tide': 'Sand / Tide'}
    grp = df.groupby('Group', dropna=False).agg(Members=('Member', 'count'), Total_KG=('Total_KG', 'sum'), Ropes_Total=('Ropes_Total', 'sum'), Target_Ropes=('Target_Ropes', 'sum'), Gap=('Gap', 'sum')).reset_index()
    ch_rows = []
    valid_df = df.dropna(subset=flag_cols[:1])
    for fc in flag_cols:
        if fc in df.columns:
            lbl = flag_labels.get(fc, fc); count = int(df[fc].sum()); pct_v = round(count / max(len(valid_df), 1) * 100, 1)
            ch_rows.append({'Challenge': lbl, 'Count': count, 'Pct': pct_v})
    challenge_df = pd.DataFrame(ch_rows)
    total_kg = float(df['Total_KG'].sum()); n_farmers = int(df['Member'].nunique()); ropes_ocean = float(df['Ropes_Ocean'].sum()); total_ropes = float(df['Ropes_Total'].sum())
    overall = {
        'total_kg': total_kg, 'n_farmers': n_farmers, 'n_groups': int(df['Group'].nunique()),
        'ropes_ocean': ropes_ocean, 'ropes_total': total_ropes,
        'avg_prod_per_rope': round(total_kg / max(total_ropes, 1), 2),
        'dried_kg': float(df['Dried_KG'].sum()), 'wet_kg': float(df['Wet_KG'].sum()),
        'dried_wet_ratio': round(float(df['Dried_KG'].sum()) / max(float(df['Wet_KG'].sum()), 1), 2),
        'avg_production_per_farmer': round(total_kg / max(n_farmers, 1), 1),
        'pct_meeting_target': round(len(df[df['Ropes_Total'] >= df['Target_Ropes']]) / max(len(df), 1) * 100, 1),
        'gap_total': float(df['Gap'].sum()),
        'multi_challenge_pct': round(len(df[df[flag_cols].sum(axis=1) >= 2]) / max(len(valid_df), 1) * 100, 1) if all(fc in df.columns for fc in flag_cols) else 0.0,
    }
    return {'group_summary': grp, 'challenge_counts': challenge_df, 'overall': overall}


# ============================================================================
# CHART BUILDERS
# ============================================================================

def _layout_defaults(title, height=400):
    return dict(
        title=title, height=height,
        font=dict(size=12, color='#333', family='Inter, Segoe UI, sans-serif'),
        title_font=dict(size=15, color='#222'),
        plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
        margin=dict(l=20, r=20, t=50, b=20),
        legend=dict(orientation='h', yanchor='bottom', y=1.02, font=dict(size=11)),
    )


def make_comparison_bar(df, cat_col, title, multiply=True, height=400, orientation='v'):
    plot_df = df.copy()
    bl_col = 'Baseline' if 'Baseline' in df.columns else [c for c in df.columns if 'BL' in c or 'Baseline' in c][0] if any('BL' in c or 'Baseline' in c for c in df.columns) else df.columns[1]
    ml_col = 'Midline' if 'Midline' in df.columns else 'Endline' if 'Endline' in df.columns else df.columns[2]
    if multiply:
        plot_df[bl_col] = plot_df[bl_col].apply(lambda x: x*100 if isinstance(x, (int, float)) else 0)
        plot_df[ml_col] = plot_df[ml_col].apply(lambda x: x*100 if isinstance(x, (int, float)) else 0)
    fig = go.Figure()
    bl_label = 'Baseline' if 'Baseline' in df.columns else 'Pre'
    ml_label = 'Midline' if 'Midline' in df.columns else 'Endline' if 'Endline' in df.columns else 'Post'
    if orientation == 'h':
        fig.add_trace(go.Bar(y=plot_df[cat_col], x=plot_df[bl_col], name=bl_label, orientation='h', marker_color=COLORS['baseline'], text=plot_df[bl_col].apply(lambda x: f"{x:.1f}%"), textposition='auto'))
        fig.add_trace(go.Bar(y=plot_df[cat_col], x=plot_df[ml_col], name=ml_label, orientation='h', marker_color=COLORS['midline'], text=plot_df[ml_col].apply(lambda x: f"{x:.1f}%"), textposition='auto'))
    else:
        fig.add_trace(go.Bar(x=plot_df[cat_col], y=plot_df[bl_col], name=bl_label, marker_color=COLORS['baseline'], text=plot_df[bl_col].apply(lambda x: f"{x:.1f}%"), textposition='auto'))
        fig.add_trace(go.Bar(x=plot_df[cat_col], y=plot_df[ml_col], name=ml_label, marker_color=COLORS['midline'], text=plot_df[ml_col].apply(lambda x: f"{x:.1f}%"), textposition='auto'))
    fig.update_layout(**_layout_defaults(title, height), barmode='group')
    return fig


def make_delta_bar(df, cat_col, title, multiply=True, height=400):
    plot_df = df.copy()
    bl_col = 'Baseline' if 'Baseline' in df.columns else df.columns[1]
    ml_col = 'Midline' if 'Midline' in df.columns else 'Endline' if 'Endline' in df.columns else df.columns[2]
    factor = 100 if multiply else 1
    plot_df['Change'] = (plot_df[ml_col] - plot_df[bl_col]) * factor
    plot_df = plot_df.sort_values('Change')
    colors = [COLORS['good'] if v >= 0 else COLORS['danger'] for v in plot_df['Change']]
    fig = go.Figure()
    fig.add_trace(go.Bar(y=plot_df[cat_col], x=plot_df['Change'], orientation='h', marker_color=colors, text=plot_df['Change'].apply(lambda x: f"{x:+.1f}pp"), textposition='auto'))
    fig.update_layout(**_layout_defaults(title, height), xaxis_title="Change (pp)")
    return fig


def fig_to_html(fig, div_id=None):
    """Convert a Plotly figure to an embedded HTML div string."""
    return fig.to_html(full_html=False, include_plotlyjs=False, div_id=div_id)


# ============================================================================
# KPI CARD BUILDER
# ============================================================================

def kpi_card(title, value, delta_text="", delta_positive=True):
    cls = "delta-positive" if delta_positive else "delta-negative"
    return f'''<div class="kpi-card">
        <h3>{title}</h3>
        <div class="value">{value}</div>
        <div class="{cls}">{delta_text}</div>
    </div>'''


def kpi_row(cards_html, cols=4):
    grid = f'style="display:grid;grid-template-columns:repeat({cols},1fr);gap:1rem;margin-bottom:1.5rem;"'
    return f'<div {grid}>{"".join(cards_html)}</div>'


# ============================================================================
# SECTION BUILDERS
# ============================================================================

def build_forestry_section(data):
    html = []
    ft = data['functionality_threshold']
    fs = data['functionality_scores']
    fd = data['functionality_domain']
    bl_60 = ft.loc[ft['Timepoint']=='Baseline','Functional_60_pct'].values[0]
    ml_60 = ft.loc[ft['Timepoint']=='Midline','Functional_60_pct'].values[0]
    bl_avg = fs.loc[fs['Timepoint']=='Baseline','Average'].values[0]
    ml_avg = fs.loc[fs['Timepoint']=='Midline','Average'].values[0]
    bl_grp = data['num_groups'].loc[data['num_groups']['Timepoint']=='Baseline','Groups_Assessed'].values[0]
    ml_grp = data['num_groups'].loc[data['num_groups']['Timepoint']=='Midline','Groups_Assessed'].values[0]
    af_bl = data['agroforestry'].loc[data['agroforestry']['Category']=='Yes','Baseline'].values[0]
    af_ml = data['agroforestry'].loc[data['agroforestry']['Category']=='Yes','Midline'].values[0]

    html.append('<div class="section-narrative"><strong>Forestry Conservation Groups:</strong> Key indicators comparing baseline vs midline for community forestry groups — functionality, governance, forest condition, income generation & agroforestry.</div>')

    html.append(kpi_row([
        kpi_card("Groups Assessed", f"{int(ml_grp)}", f"{int(ml_grp-bl_grp):+d} from BL", ml_grp >= bl_grp),
        kpi_card("Functional ≥60%", f"{ml_60*100:.1f}%", f"{(ml_60-bl_60)*100:+.1f}pp", ml_60 >= bl_60),
        kpi_card("Overall Score", f"{ml_avg:.1f}/100", f"{ml_avg-bl_avg:+.1f} pts", ml_avg >= bl_avg),
        kpi_card("Agroforestry Yes", f"{af_ml*100:.0f}%", f"{(af_ml-af_bl)*100:+.1f}pp", af_ml >= af_bl),
    ]))

    # Domain radar
    domain_df = pd.DataFrame({'Domain': ['Management','Gender','Effectiveness','Overall'], 'Baseline': fd[['Management','Gender','Effectiveness','Overall']].iloc[0].values, 'Midline': fd[['Management','Gender','Effectiveness','Overall']].iloc[1].values})
    fig_r = go.Figure()
    fig_r.add_trace(go.Scatterpolar(r=list(domain_df['Baseline'])+[domain_df['Baseline'].iloc[0]], theta=list(domain_df['Domain'])+[domain_df['Domain'].iloc[0]], fill='toself', name='Baseline', fillcolor=COLORS['radar_bl_fill'], line=dict(color=COLORS['baseline'])))
    fig_r.add_trace(go.Scatterpolar(r=list(domain_df['Midline'])+[domain_df['Midline'].iloc[0]], theta=list(domain_df['Domain'])+[domain_df['Domain'].iloc[0]], fill='toself', name='Midline', fillcolor=COLORS['radar_ml_fill'], line=dict(color=COLORS['midline'])))
    fig_r.update_layout(**_layout_defaults("Forestry Domain Performance", 420), polar=dict(radialaxis=dict(visible=True, range=[0,100])))
    html.append('<div class="chart-row">')
    html.append(f'<div class="chart-half">{fig_to_html(fig_r)}</div>')

    # Membership comparison
    fig_mem = make_comparison_bar(data['avg_membership'], 'Category', 'Average Membership by Gender', multiply=False, height=380)
    html.append(f'<div class="chart-half">{fig_to_html(fig_mem)}</div>')
    html.append('</div>')

    # Goals defined
    fig_goals = make_comparison_bar(data['goals_defined'], 'Category', 'Conservation Goals Defined', height=350)
    html.append('<div class="chart-row">')
    html.append(f'<div class="chart-half">{fig_to_html(fig_goals)}</div>')
    # Women leadership
    fig_wl = make_comparison_bar(data['women_leadership'], 'Category', 'Women in Leadership', height=350, orientation='h')
    html.append(f'<div class="chart-half">{fig_to_html(fig_wl)}</div>')
    html.append('</div>')

    # Training coverage
    fig_tc = make_comparison_bar(data['training_coverage'], 'Category', 'Training Coverage', height=350)
    html.append('<div class="chart-row">')
    html.append(f'<div class="chart-half">{fig_to_html(fig_tc)}</div>')

    # Forest condition stacked
    fc_data = data['forest_condition']
    fig_fc = make_subplots(rows=1, cols=2, subplot_titles=['Baseline', 'Midline'], horizontal_spacing=0.1)
    for i, prefix in enumerate(['Baseline', 'Midline']):
        fig_fc.add_trace(go.Bar(y=fc_data['Characteristic'], x=fc_data[f'{prefix}_Good']*100, name='Good', marker_color=COLORS['good'], orientation='h', showlegend=i==0), row=1, col=i+1)
        fig_fc.add_trace(go.Bar(y=fc_data['Characteristic'], x=fc_data[f'{prefix}_Medium']*100, name='Medium', marker_color=COLORS['medium'], orientation='h', showlegend=i==0), row=1, col=i+1)
        fig_fc.add_trace(go.Bar(y=fc_data['Characteristic'], x=fc_data[f'{prefix}_Poor']*100, name='Poor', marker_color=COLORS['poor'], orientation='h', showlegend=i==0), row=1, col=i+1)
    fig_fc.update_layout(**_layout_defaults("Forest Condition Assessment", 400), barmode='stack')
    html.append(f'<div class="chart-half">{fig_to_html(fig_fc)}</div>')
    html.append('</div>')

    # Threats
    fig_threats = make_comparison_bar(pd.DataFrame({'Threat': data['threats']['Threat'], 'Baseline': data['threats']['Baseline_High'], 'Midline': data['threats']['Midline_High']}), 'Threat', 'High-Level Threats (Baseline vs Midline)', height=400, orientation='h')
    html.append('<div class="chart-row">')
    html.append(f'<div class="chart-half">{fig_to_html(fig_threats)}</div>')

    # Income sources
    fig_inc = make_comparison_bar(data['income_sources'], 'Source', 'Forest Income Sources', height=400, orientation='h')
    html.append(f'<div class="chart-half">{fig_to_html(fig_inc)}</div>')
    html.append('</div>')

    # Agroforestry types
    fig_af = make_comparison_bar(data['af_types'], 'Practice', 'Agroforestry Practices', height=350, orientation='h')
    html.append(fig_to_html(fig_af))

    return '\n'.join(html)


def build_women_section(w):
    html = []
    cc_bl = w['cc_heard'].loc[w['cc_heard']['Response']=='Yes','Baseline'].values[0]
    cc_ml = w['cc_heard'].loc[w['cc_heard']['Response']=='Yes','Midline'].values[0]
    nbs_bl = w['nbs_heard'].loc[w['nbs_heard']['Response']=='Yes','Baseline'].values[0]
    nbs_ml = w['nbs_heard'].loc[w['nbs_heard']['Response']=='Yes','Midline'].values[0]
    sav_bl = w['personal_saving'].loc[w['personal_saving']['Response']=='Yes','Baseline'].values[0]
    sav_ml = w['personal_saving'].loc[w['personal_saving']['Response']=='Yes','Midline'].values[0]
    ls_bl = w['lifeskills_agree']['Baseline'].mean()
    ls_ml = w['lifeskills_agree']['Midline'].mean()

    html.append('<div class="section-narrative"><strong>Women\'s Survey:</strong> Household-level analysis of women\'s empowerment & climate resilience — CC/NbS awareness, savings, life skills, social norms, time use, and decision-making.</div>')

    html.append(kpi_row([
        kpi_card("CC Awareness", f"{cc_ml*100:.0f}%", f"{(cc_ml-cc_bl)*100:+.1f}pp", cc_ml >= cc_bl),
        kpi_card("NbS Awareness", f"{nbs_ml*100:.0f}%", f"{(nbs_ml-nbs_bl)*100:+.1f}pp", nbs_ml >= nbs_bl),
        kpi_card("Women Saving", f"{sav_ml*100:.0f}%", f"{(sav_ml-sav_bl)*100:+.1f}pp", sav_ml >= sav_bl),
        kpi_card("Life Skills (Avg)", f"{ls_ml*100:.0f}%", f"{(ls_ml-ls_bl)*100:+.1f}pp", ls_ml >= ls_bl),
    ]))

    # Demographics
    html.append('<div class="chart-row">')
    fig1 = make_comparison_bar(w['location'], 'Category', 'Women by Location Type', height=320)
    html.append(f'<div class="chart-half">{fig_to_html(fig1)}</div>')
    fig2 = make_comparison_bar(w['hh_type'], 'Category', 'Household Type', height=320)
    html.append(f'<div class="chart-half">{fig_to_html(fig2)}</div>')
    html.append('</div>')

    html.append('<div class="chart-row">')
    fig3 = make_comparison_bar(w['marital'], 'Category', 'Marital Status', height=380, orientation='h')
    html.append(f'<div class="chart-half">{fig_to_html(fig3)}</div>')
    fig4 = make_comparison_bar(w['education'], 'Category', 'Education Level', height=380, orientation='h')
    html.append(f'<div class="chart-half">{fig_to_html(fig4)}</div>')
    html.append('</div>')

    # Shocks & Coping
    html.append('<h3 class="section-title">Shocks & Coping</h3>')
    html.append('<div class="chart-row">')
    fig5 = make_comparison_bar(w['shocks'], 'Shock', 'Shocks Experienced', height=450, orientation='h')
    html.append(f'<div class="chart-half">{fig_to_html(fig5)}</div>')
    fig6 = make_comparison_bar(w['coping'], 'Strategy', 'Coping Strategies', height=450, orientation='h')
    html.append(f'<div class="chart-half">{fig_to_html(fig6)}</div>')
    html.append('</div>')

    # Savings
    html.append('<h3 class="section-title">Savings & Financial Inclusion</h3>')
    html.append('<div class="chart-row">')
    fig7 = make_comparison_bar(w['personal_saving'], 'Response', 'Personal Saving', height=320)
    html.append(f'<div class="chart-half">{fig_to_html(fig7)}</div>')
    fig8 = make_comparison_bar(w['saving_mechanism'], 'Mechanism', 'Saving Mechanisms', height=380, orientation='h')
    html.append(f'<div class="chart-half">{fig_to_html(fig8)}</div>')
    html.append('</div>')

    # Roles & Time
    html.append('<h3 class="section-title">Roles, Time Use & Decisions</h3>')
    html.append('<div class="chart-row">')
    fig9 = make_comparison_bar(w['roles_should_joint'], 'Role', 'Roles Should Be Joint (%)', height=450, orientation='h')
    html.append(f'<div class="chart-half">{fig_to_html(fig9)}</div>')
    fig10 = make_comparison_bar(w['time_summary'], 'Category', 'Time Use (hrs/day)', multiply=False, height=400)
    html.append(f'<div class="chart-half">{fig_to_html(fig10)}</div>')
    html.append('</div>')

    # Climate & NbS
    html.append('<h3 class="section-title">Climate Change & NbS</h3>')
    html.append('<div class="chart-row">')
    fig11 = make_comparison_bar(w['cc_env_effects'], 'Effect', 'CC Environmental Effects', height=480, orientation='h')
    html.append(f'<div class="chart-half">{fig_to_html(fig11)}</div>')
    fig12 = make_comparison_bar(w['nbs_examples'], 'Example', 'NbS Examples Cited', height=400, orientation='h')
    html.append(f'<div class="chart-half">{fig_to_html(fig12)}</div>')
    html.append('</div>')

    # Life Skills Radar
    ls_a = w['lifeskills_agree']
    ls_domain = ls_a.groupby('Domain')[['Baseline','Midline']].mean().reset_index()
    domains_l = list(ls_domain['Domain']) + [ls_domain['Domain'].iloc[0]]
    bl_v = list(ls_domain['Baseline']*100) + [ls_domain['Baseline'].iloc[0]*100]
    ml_v = list(ls_domain['Midline']*100) + [ls_domain['Midline'].iloc[0]*100]
    fig_ls = go.Figure()
    fig_ls.add_trace(go.Scatterpolar(r=bl_v, theta=domains_l, fill='toself', name='Baseline', fillcolor=COLORS['radar_bl_fill'], line=dict(color=COLORS['baseline'])))
    fig_ls.add_trace(go.Scatterpolar(r=ml_v, theta=domains_l, fill='toself', name='Midline', fillcolor=COLORS['radar_ml_fill'], line=dict(color=COLORS['midline'])))
    fig_ls.update_layout(**_layout_defaults("Women's Life Skills Domains", 420), polar=dict(radialaxis=dict(visible=True, range=[0,100])))

    html.append('<h3 class="section-title">Life Skills & Social Norms</h3>')
    html.append('<div class="chart-row">')
    html.append(f'<div class="chart-half">{fig_to_html(fig_ls)}</div>')

    # Social norms change
    sn = w['socialnorms_agree'].copy()
    sn['Change'] = (sn['Midline'] - sn['Baseline'])*100
    sn_sorted = sn.sort_values('Change')
    colors_sn = [COLORS['good'] if v <= 0 else COLORS['danger'] for v in sn_sorted['Change']]
    fig_sn = go.Figure()
    fig_sn.add_trace(go.Bar(y=sn_sorted['Norm'], x=sn_sorted['Change'], orientation='h', marker_color=colors_sn, text=sn_sorted['Change'].apply(lambda x: f"{x:+.1f}pp"), textposition='auto'))
    fig_sn.update_layout(**_layout_defaults("Social Norms Change (↓ = positive)", 450), xaxis_title="Change (pp)")
    html.append(f'<div class="chart-half">{fig_to_html(fig_sn)}</div>')
    html.append('</div>')

    # Decision making
    fig_dec = make_comparison_bar(w['decision_should_joint'], 'Decision', 'Decisions Should Be Joint', height=450, orientation='h')
    html.append(fig_to_html(fig_dec))

    return '\n'.join(html)


def build_men_section(m):
    html = []
    m_cc_bl = m['cc_heard'].loc[m['cc_heard']['Response']=='Yes','Baseline'].values[0]
    m_cc_ml = m['cc_heard'].loc[m['cc_heard']['Response']=='Yes','Midline'].values[0]
    m_nbs_bl = m['nbs_heard'].loc[m['nbs_heard']['Response']=='Yes','Baseline'].values[0]
    m_nbs_ml = m['nbs_heard'].loc[m['nbs_heard']['Response']=='Yes','Midline'].values[0]
    m_rj_bl = m['roles_should_joint']['Baseline'].mean()
    m_rj_ml = m['roles_should_joint']['Midline'].mean()
    m_dj_bl = m['decision_should_joint']['Baseline'].mean()
    m_dj_ml = m['decision_should_joint']['Midline'].mean()

    html.append('<div class="section-narrative"><strong>Men\'s Survey:</strong> Men\'s perspectives on climate awareness, gender roles, household decision-making, support for women in NbS, and social norms.</div>')

    html.append(kpi_row([
        kpi_card("CC Awareness (Men)", f"{m_cc_ml*100:.0f}%", f"{(m_cc_ml-m_cc_bl)*100:+.1f}pp", m_cc_ml >= m_cc_bl),
        kpi_card("NbS Awareness (Men)", f"{m_nbs_ml*100:.0f}%", f"{(m_nbs_ml-m_nbs_bl)*100:+.1f}pp", m_nbs_ml >= m_nbs_bl),
        kpi_card("Joint Roles (Should)", f"{m_rj_ml*100:.0f}%", f"{(m_rj_ml-m_rj_bl)*100:+.1f}pp", m_rj_ml >= m_rj_bl),
        kpi_card("Joint Decisions (Should)", f"{m_dj_ml*100:.0f}%", f"{(m_dj_ml-m_dj_bl)*100:+.1f}pp", m_dj_ml >= m_dj_bl),
    ]))

    html.append('<div class="chart-row">')
    fig1 = make_comparison_bar(m['location'], 'Category', 'Men by Location', height=320)
    html.append(f'<div class="chart-half">{fig_to_html(fig1)}</div>')
    fig2 = make_comparison_bar(m['education'], 'Category', 'Education Level', height=320)
    html.append(f'<div class="chart-half">{fig_to_html(fig2)}</div>')
    html.append('</div>')

    html.append('<div class="chart-row">')
    fig3 = make_comparison_bar(m['cc_env_effects'], 'Effect', 'CC Environmental Effects (Men)', height=480, orientation='h')
    html.append(f'<div class="chart-half">{fig_to_html(fig3)}</div>')
    fig4 = make_comparison_bar(m['mangrove_support_type'], 'Type', 'Support Type for Mangrove Restoration', height=400, orientation='h')
    html.append(f'<div class="chart-half">{fig_to_html(fig4)}</div>')
    html.append('</div>')

    html.append('<div class="chart-row">')
    fig5 = make_comparison_bar(m['roles_should_joint'], 'Role', 'Roles Should Be Joint (Men)', height=450, orientation='h')
    html.append(f'<div class="chart-half">{fig_to_html(fig5)}</div>')
    fig6 = make_comparison_bar(m['decision_should_joint'], 'Decision', 'Decisions Should Be Joint (Men)', height=450, orientation='h')
    html.append(f'<div class="chart-half">{fig_to_html(fig6)}</div>')
    html.append('</div>')

    html.append('<div class="chart-row">')
    fig7 = make_comparison_bar(m['time_summary'], 'Category', 'Men Time Use (hrs/day)', multiply=False, height=400)
    html.append(f'<div class="chart-half">{fig_to_html(fig7)}</div>')
    # Social norms change
    sn = m['socialnorms_agree'].copy()
    sn['Change'] = (sn['Midline'] - sn['Baseline'])*100
    sn_sorted = sn.sort_values('Change')
    colors_sn = [COLORS['good'] if v <= 0 else COLORS['danger'] for v in sn_sorted['Change']]
    fig_sn = go.Figure()
    fig_sn.add_trace(go.Bar(y=sn_sorted['Norm'], x=sn_sorted['Change'], orientation='h', marker_color=colors_sn, text=sn_sorted['Change'].apply(lambda x: f"{x:+.1f}pp"), textposition='auto'))
    fig_sn.update_layout(**_layout_defaults("Social Norms Change — Men (↓ = positive)", 500), xaxis_title="Change (pp)")
    html.append(f'<div class="chart-half">{fig_to_html(fig_sn)}</div>')
    html.append('</div>')

    return '\n'.join(html)


def build_gjj_women_section(g):
    html = []
    sa = g['self_strongly_agree']
    sa_bl = float(sa['Baseline'].mean()); sa_el = float(sa['Endline'].mean())
    af_rn = g['rel_af_rn']
    af_bl = float(af_rn['AF_Baseline'].mean()); af_el = float(af_rn['AF_Endline'].mean())
    ch_yes = g['shared_chores_yn'][g['shared_chores_yn']['Response'].str.strip().str.lower() == 'yes']
    ch_bl = float(ch_yes['Baseline'].values[0]) if len(ch_yes) else 0.0
    ch_el = float(ch_yes['Endline'].values[0]) if len(ch_yes) else 0.0
    eq_yes = g['equal_say'][g['equal_say']['Response'].str.strip().str.lower() == 'yes']
    eq_bl = float(eq_yes['Baseline'].values[0]) if len(eq_yes) else 0.0
    eq_el = float(eq_yes['Endline'].values[0]) if len(eq_yes) else 0.0

    html.append('<div class="section-narrative"><strong>GJJ KAP Women:</strong> Gender Justice Journey — women\'s self-esteem, relational wellbeing, shared responsibility & power, autonomy & leadership. Baseline vs Endline.</div>')

    html.append(kpi_row([
        kpi_card("Self-Esteem (SA)", f"{sa_el*100:.0f}%", f"{(sa_el-sa_bl)*100:+.1f}pp", sa_el >= sa_bl),
        kpi_card("Relational (AF)", f"{af_el*100:.0f}%", f"{(af_el-af_bl)*100:+.1f}pp", af_el >= af_bl),
        kpi_card("Husband Chores", f"{ch_el*100:.0f}%", f"{(ch_el-ch_bl)*100:+.1f}pp", ch_el >= ch_bl),
        kpi_card("Equal Say", f"{eq_el*100:.0f}%", f"{(eq_el-eq_bl)*100:+.1f}pp", eq_el >= eq_bl),
    ]))

    html.append('<div class="chart-row">')
    fig1 = make_comparison_bar(g['self_strongly_agree'], 'Statement', 'Self-Esteem (Strongly Agree)', height=350)
    html.append(f'<div class="chart-half">{fig_to_html(fig1)}</div>')
    fig2 = make_comparison_bar(g['shared_chores_yn'], 'Response', 'Husband Supports Chores', height=320)
    html.append(f'<div class="chart-half">{fig_to_html(fig2)}</div>')
    html.append('</div>')

    html.append('<div class="chart-row">')
    fig3 = make_comparison_bar(g['decision_types'], 'Decision', 'Decision Types', height=450, orientation='h')
    html.append(f'<div class="chart-half">{fig_to_html(fig3)}</div>')
    fig4 = make_comparison_bar(g['support_leader'], 'Category', 'Support for Women Leaders', height=380, orientation='h')
    html.append(f'<div class="chart-half">{fig_to_html(fig4)}</div>')
    html.append('</div>')

    fig5 = make_comparison_bar(g['support_business'], 'Category', 'Support for Women\'s Business', height=380, orientation='h')
    html.append(fig_to_html(fig5))

    return '\n'.join(html)


def build_gjj_men_section(g):
    html = []
    agr = g['self_agreement']
    agr_bl = float(agr['Agreement_BL'].mean()); agr_el = float(agr['Agreement_EL'].mean())
    af_rn = g['rel_af_rn']
    af_bl_m = float(af_rn['AF_Baseline'].mean()); af_el_m = float(af_rn['AF_Endline'].mean())
    ch = g['shared_chores_yn']
    ch_yes = ch[ch['Response'].str.strip().str.lower() == 'yes']
    ch_bl_m = float(ch_yes['Baseline'].values[0]) if len(ch_yes) else 0.0
    ch_el_m = float(ch_yes['Endline'].values[0]) if len(ch_yes) else 0.0
    dc = g['decision_conversations']
    dc_yes = dc[dc['Response'].str.strip().str.lower() == 'yes']
    dc_bl_m = float(dc_yes['Baseline'].values[0]) if len(dc_yes) else 0.0
    dc_el_m = float(dc_yes['Endline'].values[0]) if len(dc_yes) else 0.0

    html.append('<div class="section-narrative"><strong>GJJ KAP Men:</strong> Gender Justice Journey — men\'s self-esteem, relational wellbeing, shared responsibility & power, leadership & business support. Baseline vs Endline.</div>')

    html.append(kpi_row([
        kpi_card("Self-Esteem (Men)", f"{agr_el*100:.0f}%", f"{(agr_el-agr_bl)*100:+.1f}pp", agr_el >= agr_bl),
        kpi_card("Relational AF (Men)", f"{af_el_m*100:.0f}%", f"{(af_el_m-af_bl_m)*100:+.1f}pp", af_el_m >= af_bl_m),
        kpi_card("Supports Chores (Men)", f"{ch_el_m*100:.0f}%", f"{(ch_el_m-ch_bl_m)*100:+.1f}pp", ch_el_m >= ch_bl_m),
        kpi_card("Decision Convos (Men)", f"{dc_el_m*100:.0f}%", f"{(dc_el_m-dc_bl_m)*100:+.1f}pp", dc_el_m >= dc_bl_m),
    ]))

    html.append('<div class="chart-row">')
    fig1 = make_comparison_bar(g['shared_chores_yn'], 'Response', 'Men Supporting Chores', height=320)
    html.append(f'<div class="chart-half">{fig_to_html(fig1)}</div>')
    fig2 = make_comparison_bar(g['decision_conversations'], 'Response', 'Decision Conversations', height=320)
    html.append(f'<div class="chart-half">{fig_to_html(fig2)}</div>')
    html.append('</div>')

    html.append('<div class="chart-row">')
    fig3 = make_comparison_bar(g['support_leader'], 'Category', 'Support Wife as Leader', height=380, orientation='h')
    html.append(f'<div class="chart-half">{fig_to_html(fig3)}</div>')
    fig4 = make_comparison_bar(g['support_business'], 'Category', 'Support Wife Business', height=380, orientation='h')
    html.append(f'<div class="chart-half">{fig_to_html(fig4)}</div>')
    html.append('</div>')

    return '\n'.join(html)


def build_forest_training_section(t_data):
    html = []
    t_scores = t_data['scores']; t_thresh = t_data['thresholds']; t_questions = t_data['questions']
    t_bl_avg = float(t_scores[t_scores['Timepoint']=='Baseline']['AverageScore'].values[0])
    t_el_avg = float(t_scores[t_scores['Timepoint']=='Endline']['AverageScore'].values[0])
    t70 = t_thresh[t_thresh['Threshold'].str.contains('70')]
    t70_bl = float(t70[t70['Timepoint']=='Baseline']['Proportion'].values[0])
    t70_el = float(t70[t70['Timepoint']=='Endline']['Proportion'].values[0])
    t_bl_n = int(t_scores[t_scores['Timepoint']=='Baseline']['Respondents'].values[0])
    t_el_n = int(t_scores[t_scores['Timepoint']=='Endline']['Respondents'].values[0])

    html.append('<div class="section-narrative"><strong>Forest Training:</strong> Pre-Training vs Post-Training knowledge assessment results. Question-by-question analysis with domain-level grouping.</div>')

    html.append(kpi_row([
        kpi_card("≥70% Pass Rate", f"{t70_el*100:.1f}%", f"{(t70_el-t70_bl)*100:+.1f}pp", t70_el >= t70_bl),
        kpi_card("Average Score", f"{t_el_avg:.1f}%", f"{t_el_avg-t_bl_avg:+.1f}pp", t_el_avg >= t_bl_avg),
        kpi_card("Pre-Training N", f"{t_bl_n:,}", "baseline"),
        kpi_card("Post-Training N", f"{t_el_n:,}", f"{t_el_n - t_bl_n:+,} vs Pre", t_el_n >= t_bl_n),
    ]))

    if not t_questions.empty:
        # Question-by-question
        fig_q = go.Figure()
        fig_q.add_trace(go.Bar(y=t_questions['QuestionText'], x=t_questions['Baseline']*100, name='Pre-Training', orientation='h', marker_color=COLORS['baseline']))
        fig_q.add_trace(go.Bar(y=t_questions['QuestionText'], x=t_questions['Endline']*100, name='Post-Training', orientation='h', marker_color=COLORS['midline']))
        fig_q.update_layout(**_layout_defaults("Knowledge by Question (% Correct)", 600), barmode='group', xaxis_title='% Correct')
        html.append(fig_to_html(fig_q))

        # Domain grouping
        domain_rows = []
        for domain, q_nums in FOREST_TRAINING_DOMAINS.items():
            d_q = t_questions[t_questions['QuestionNumber'].isin(q_nums)]
            if not d_q.empty:
                domain_rows.append({'Domain': domain, 'Baseline': d_q['Baseline'].mean()*100, 'Endline': d_q['Endline'].mean()*100})
        if domain_rows:
            domain_df = pd.DataFrame(domain_rows)
            fig_d = go.Figure()
            fig_d.add_trace(go.Bar(x=domain_df['Domain'], y=domain_df['Baseline'], name='Pre-Training', marker_color=COLORS['baseline'], text=domain_df['Baseline'].apply(lambda x: f"{x:.1f}%"), textposition='auto'))
            fig_d.add_trace(go.Bar(x=domain_df['Domain'], y=domain_df['Endline'], name='Post-Training', marker_color=COLORS['midline'], text=domain_df['Endline'].apply(lambda x: f"{x:.1f}%"), textposition='auto'))
            fig_d.update_layout(**_layout_defaults("Domain-Level Knowledge Gains", 400), barmode='group', yaxis_title='% Correct')
            html.append(fig_to_html(fig_d))

    return '\n'.join(html)


def build_mangrove_training_section(mg_data):
    html = []
    mg_county = mg_data['adequate_county']
    mg_scores = mg_data['scores']
    mg_all = mg_county[mg_county['County'] == 'All']
    mg_pre_val = float(mg_all[mg_all['Timepoint'] == 'Pre-Test']['Value'].values[0]) * 100
    mg_post_val = float(mg_all[mg_all['Timepoint'] == 'Post-Test']['Value'].values[0]) * 100
    mg_change = mg_post_val - mg_pre_val
    mg_all_scores = mg_scores[mg_scores['County'] == 'All']
    mg_avg = float(mg_all_scores['AvgScore'].values[0]) if len(mg_all_scores) else 0
    mg_n = int(mg_all_scores['Respondents'].values[0]) if len(mg_all_scores) else 0

    html.append('<div class="section-narrative"><strong>Mangrove Training:</strong> Pre-Training vs Post-Training knowledge assessment for mangrove restoration. County-level and sex-disaggregated results.</div>')

    html.append(kpi_row([
        kpi_card("≥60% Pass (Post)", f"{mg_post_val:.1f}%", f"{mg_change:+.1f}pp", mg_change > 0),
        kpi_card("≥60% Pass (Pre)", f"{mg_pre_val:.1f}%", "Baseline"),
        kpi_card("Avg Score", f"{mg_avg:.1f}%", "Overall"),
        kpi_card("Respondents", f"{mg_n:,}", "All counties"),
    ]))

    # County comparison
    fig_county = go.Figure()
    for tp in ['Pre-Test', 'Post-Test']:
        d = mg_county[mg_county['Timepoint'] == tp]
        color = COLORS['baseline'] if tp == 'Pre-Test' else COLORS['midline']
        fig_county.add_trace(go.Bar(x=d['County'], y=d['Value']*100, name=tp, marker_color=color, text=d['Value'].apply(lambda x: f"{x*100:.1f}%"), textposition='auto'))
    fig_county.update_layout(**_layout_defaults("Adequate Knowledge (≥60%) by County", 400), barmode='group', yaxis_title='%')
    html.append(fig_to_html(fig_county))

    # Sex disaggregation
    if 'adequate_sex' in mg_data and not mg_data['adequate_sex'].empty:
        fig_sex = go.Figure()
        for tp in ['Pre-Test', 'Post-Test']:
            d = mg_data['adequate_sex'][mg_data['adequate_sex']['Timepoint'] == tp]
            color = COLORS['baseline'] if tp == 'Pre-Test' else COLORS['midline']
            fig_sex.add_trace(go.Bar(x=d['Sex'], y=d['Value']*100, name=tp, marker_color=color, text=d['Value'].apply(lambda x: f"{x*100:.1f}%"), textposition='auto'))
        fig_sex.update_layout(**_layout_defaults("Adequate Knowledge by Sex", 380), barmode='group', yaxis_title='%')
        html.append(fig_to_html(fig_sex))

    return '\n'.join(html)


def build_seaweed_section(sw_df):
    html = []
    sw_agg = prepare_seaweed_aggregates(sw_df)
    sw_ov = sw_agg['overall']; sw_grp = sw_agg['group_summary']; sw_ch = sw_agg['challenge_counts']
    sw_avg_ach = sw_df['Ropes_Achievement_pct'].mean()

    html.append('<div class="section-narrative"><strong>Seaweed Production & Challenges (2025):</strong> Group performance, production metrics, rope achievement, and constraint analysis across all seaweed farming groups.</div>')

    html.append(kpi_row([
        kpi_card("Total Production", f"{sw_ov['total_kg']:,.0f} kg", f"{sw_ov['n_farmers']:,} farmers"),
        kpi_card("Ropes in Ocean", f"{sw_ov['ropes_ocean']:,.0f}", f"of {sw_ov['ropes_total']:,.0f} total"),
        kpi_card("Avg Prod/Rope", f"{sw_ov['avg_prod_per_rope']:.2f} kg", f"{sw_ov['avg_production_per_farmer']:.1f} kg/farmer"),
        kpi_card("Target Achievement", f"{sw_avg_ach:.1f}%", "On track" if sw_avg_ach >= 70 else "Below target", sw_avg_ach >= 70),
    ]))

    html.append(kpi_row([
        kpi_card("Dried/Wet Ratio", f"{sw_ov['dried_wet_ratio']:.2f}", f"{sw_ov['dried_kg']:,.0f} vs {sw_ov['wet_kg']:,.0f}"),
        kpi_card("Meet Target", f"{sw_ov['pct_meeting_target']:.0f}%", "of farmers", sw_ov['pct_meeting_target'] >= 50),
        kpi_card("Multi-Challenge", f"{sw_ov['multi_challenge_pct']:.1f}%", "face 2+ challenges", sw_ov['multi_challenge_pct'] <= 30),
    ], cols=3))

    html.append('<div class="chart-row">')
    # Production by group
    grp_sorted = sw_grp.sort_values('Total_KG', ascending=False)
    fig_prod = go.Figure(go.Bar(x=grp_sorted['Group'], y=grp_sorted['Total_KG'], marker_color=COLORS['baseline'], text=grp_sorted['Total_KG'].apply(lambda v: f"{v:,.0f}"), textposition='auto'))
    fig_prod.update_layout(**_layout_defaults('Production by Group (kg)', 350), xaxis_tickangle=-25, yaxis_title='Total KG')
    html.append(f'<div class="chart-half">{fig_to_html(fig_prod)}</div>')

    # Challenges
    if not sw_ch.empty:
        fig_ch = go.Figure(go.Bar(x=sw_ch.sort_values('Pct', ascending=True)['Pct'], y=sw_ch.sort_values('Pct', ascending=True)['Challenge'], orientation='h', marker_color='#FF9800', text=sw_ch.sort_values('Pct', ascending=True)['Pct'].apply(lambda v: f"{v:.0f}%"), textposition='auto'))
        fig_ch.update_layout(**_layout_defaults('Challenges (% farmers)', 350), xaxis_title='%')
        html.append(f'<div class="chart-half">{fig_to_html(fig_ch)}</div>')
    html.append('</div>')

    return '\n'.join(html)


def build_comparative_section(f_data, w_data, m_data):
    html = []
    html.append('<h3 class="section-title">Men vs Women: Key Comparisons</h3>')

    w_cc_ml = w_data['cc_heard'].loc[w_data['cc_heard']['Response']=='Yes','Midline'].values[0]
    m_cc_ml = m_data['cc_heard'].loc[m_data['cc_heard']['Response']=='Yes','Midline'].values[0]
    w_nbs_ml = w_data['nbs_heard'].loc[w_data['nbs_heard']['Response']=='Yes','Midline'].values[0]
    m_nbs_ml = m_data['nbs_heard'].loc[m_data['nbs_heard']['Response']=='Yes','Midline'].values[0]

    html.append('<div class="chart-row">')
    # CC/NbS comparison
    cmp_df = pd.DataFrame({'Indicator': ['CC Awareness', 'NbS Awareness'], 'Women': [w_cc_ml*100, w_nbs_ml*100], 'Men': [m_cc_ml*100, m_nbs_ml*100]})
    fig_cmp = go.Figure()
    fig_cmp.add_trace(go.Bar(x=cmp_df['Indicator'], y=cmp_df['Women'], name='Women', marker_color=COLORS['midline'], text=cmp_df['Women'].apply(lambda x: f"{x:.1f}%"), textposition='auto'))
    fig_cmp.add_trace(go.Bar(x=cmp_df['Indicator'], y=cmp_df['Men'], name='Men', marker_color=COLORS['baseline'], text=cmp_df['Men'].apply(lambda x: f"{x:.1f}%"), textposition='auto'))
    fig_cmp.update_layout(**_layout_defaults("Climate & NbS Awareness (Midline)", 380), barmode='group', yaxis_title='%')
    html.append(f'<div class="chart-half">{fig_to_html(fig_cmp)}</div>')

    # Social norms comparison
    w_sn = w_data['socialnorms_agree'].copy()
    w_harmful = w_sn[~w_sn['Norm'].str.contains('Planting Crops|Restoring Ecosystems|Express Emotions', regex=True)]
    w_sn_bl = w_harmful['Baseline'].mean() * 100; w_sn_ml = w_harmful['Midline'].mean() * 100
    m_sn = m_data['socialnorms_agree'].copy()
    m_harmful = m_sn[~m_sn['Norm'].str.contains('Planting Crops|Restoring Ecosystems|Express Emotions', regex=True)]
    m_sn_bl = m_harmful['Baseline'].mean() * 100; m_sn_ml = m_harmful['Midline'].mean() * 100
    norms_df = pd.DataFrame({'Timepoint': ['Baseline', 'Midline'], 'Women': [w_sn_bl, w_sn_ml], 'Men': [m_sn_bl, m_sn_ml]})
    fig_norms = go.Figure()
    fig_norms.add_trace(go.Bar(x=norms_df['Timepoint'], y=norms_df['Women'], name='Women', marker_color=COLORS['midline'], text=norms_df['Women'].apply(lambda x: f"{x:.1f}%"), textposition='auto'))
    fig_norms.add_trace(go.Bar(x=norms_df['Timepoint'], y=norms_df['Men'], name='Men', marker_color=COLORS['baseline'], text=norms_df['Men'].apply(lambda x: f"{x:.1f}%"), textposition='auto'))
    fig_norms.update_layout(**_layout_defaults("Harmful Social Norms (Avg Agree %)", 380), barmode='group', yaxis_title='Agree/SA (%)')
    html.append(f'<div class="chart-half">{fig_to_html(fig_norms)}</div>')
    html.append('</div>')

    return '\n'.join(html)


# ============================================================================
# HTML TEMPLATE
# ============================================================================

def build_full_html(sections):
    """Build the complete standalone HTML file with all sections as tabs."""
    tab_buttons = ''
    tab_contents = ''
    for i, (tab_id, tab_label, content) in enumerate(sections):
        active = ' active' if i == 0 else ''
        tab_buttons += f'<button class="tab-btn{active}" onclick="showTab(\'{tab_id}\')">{tab_label}</button>\n'
        display = 'block' if i == 0 else 'none'
        tab_contents += f'<div class="tab-content" id="{tab_id}" style="display:{display}">\n{content}\n</div>\n'

    return f'''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>COSME M&E Dashboard — Offline</title>
<script src="https://cdn.plot.ly/plotly-2.27.0.min.js"></script>
<style>
/* ---------- RESET & BASE ---------- */
*, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}
html {{ scroll-behavior: smooth; }}
body {{
    font-family: 'Inter', 'Segoe UI', -apple-system, BlinkMacSystemFont, sans-serif;
    background: #f5f5f5; color: #333; line-height: 1.6;
}}
.container {{ max-width: 1280px; margin: 0 auto; padding: 1rem 2rem 3rem; }}

/* ---------- HEADER ---------- */
.main-header {{
    background: {COLORS['header_gradient']};
    color: white; padding: 2rem 2.5rem; border-radius: 14px;
    margin-bottom: 1.5rem; text-align: center;
    box-shadow: 0 4px 20px rgba(0,0,0,0.18);
    position: relative; overflow: hidden;
}}
.main-header::before {{
    content: ''; position: absolute; top: -50%; left: -50%;
    width: 200%; height: 200%;
    background: radial-gradient(circle at 20% 50%, rgba(255,255,255,0.08) 0%, transparent 50%),
                radial-gradient(circle at 80% 20%, rgba(255,255,255,0.05) 0%, transparent 40%);
    pointer-events: none;
}}
.main-header h1 {{ font-size: 2.2rem; font-weight: 800; letter-spacing: -0.02em;
    text-shadow: 0 2px 4px rgba(0,0,0,0.2); position: relative; }}
.main-header p {{ margin: 0.5rem 0 0; opacity: 0.95; font-size: 1rem; font-weight: 500; position: relative; }}

/* ---------- TAB NAVIGATION ---------- */
.tab-nav {{
    display: flex; flex-wrap: wrap; gap: 0.3rem; background: white;
    padding: 0.5rem; border-radius: 12px; box-shadow: 0 2px 10px rgba(0,0,0,0.08);
    margin-bottom: 1.5rem; border: 1px solid rgba(0,0,0,0.06);
}}
.tab-btn {{
    padding: 0.6rem 1.1rem; border: none; background: transparent;
    border-radius: 8px; font-size: 0.85rem; font-weight: 600; color: #555;
    cursor: pointer; transition: all 0.2s ease; white-space: nowrap;
}}
.tab-btn:hover {{ background: {COLORS['narrative_bg']}; color: {COLORS['card_value']}; }}
.tab-btn.active {{
    background: {COLORS['card_border']}; color: white;
    box-shadow: 0 2px 8px rgba(46,125,50,0.3);
}}

/* ---------- KPI CARDS ---------- */
.kpi-card {{
    background: white; border-radius: 12px; padding: 1.3rem 1rem;
    box-shadow: 0 2px 14px rgba(0,0,0,0.08); text-align: center;
    border-left: 5px solid {COLORS['card_border']};
    transition: transform 0.2s ease, box-shadow 0.2s ease;
}}
.kpi-card:hover {{ transform: translateY(-3px); box-shadow: 0 6px 24px rgba(0,0,0,0.14); }}
.kpi-card h3 {{ font-size: 0.78rem; color: #555; margin: 0 0 0.4rem;
    text-transform: uppercase; letter-spacing: 0.05em; font-weight: 700; }}
.kpi-card .value {{ font-size: 2rem; font-weight: 800; color: {COLORS['card_value']}; line-height: 1.15; }}
.delta-positive {{ color: {COLORS['good']}; font-weight: 700; font-size: 0.85rem; margin-top: 0.3rem; }}
.delta-negative {{ color: {COLORS['danger']}; font-weight: 700; font-size: 0.85rem; margin-top: 0.3rem; }}

/* ---------- SECTION NARRATIVE ---------- */
.section-narrative {{
    background: {COLORS['narrative_bg']};
    border-left: 5px solid {COLORS['narrative_border']};
    padding: 1rem 1.3rem; border-radius: 0 10px 10px 0;
    margin-bottom: 1.2rem; font-size: 0.95rem;
    color: {COLORS['narrative_text']}; line-height: 1.65;
    box-shadow: 0 1px 6px rgba(0,0,0,0.05);
}}
.section-narrative strong {{ font-size: 1rem; }}

/* ---------- CHART LAYOUT ---------- */
.chart-row {{ display: flex; gap: 1rem; margin-bottom: 1rem; flex-wrap: wrap; }}
.chart-half {{ flex: 1; min-width: 400px; background: white; border-radius: 10px;
    padding: 0.5rem; box-shadow: 0 1px 6px rgba(0,0,0,0.04);
    border: 1px solid rgba(0,0,0,0.06); }}

/* ---------- SECTION TITLES ---------- */
.section-title {{
    font-size: 1.2rem; font-weight: 700; color: {COLORS['card_value']};
    padding: 0.6rem 0; margin: 1.2rem 0 0.6rem;
    border-bottom: 3px solid {COLORS['card_border']};
}}

/* ---------- FOOTER ---------- */
.dashboard-footer {{
    text-align: center; color: #777; font-size: 0.85rem;
    padding: 2rem 0 1rem; border-top: 2px solid #e0e0e0; margin-top: 2rem;
}}
.dashboard-footer strong {{ color: #444; }}

/* ---------- PRINT ---------- */
@media print {{
    .tab-nav {{ display: none; }}
    .tab-content {{ display: block !important; page-break-before: always; }}
    .main-header {{ print-color-adjust: exact; -webkit-print-color-adjust: exact; }}
}}

/* ---------- RESPONSIVE ---------- */
@media (max-width: 768px) {{
    .container {{ padding: 0.5rem 1rem; }}
    .chart-half {{ min-width: 100%; }}
    .tab-btn {{ font-size: 0.75rem; padding: 0.4rem 0.7rem; }}
}}
</style>
</head>
<body>
<div class="container">

<div class="main-header">
    <h1>COSME Baseline–Midline Dashboard</h1>
    <p>Integrated M&amp;E Analysis | Offline Version | Generated {datetime.now().strftime('%B %d, %Y at %H:%M')}</p>
</div>

<div class="tab-nav">
{tab_buttons}
</div>

{tab_contents}

<div class="dashboard-footer">
    <strong>COSME Baseline–Midline Dashboard — Offline</strong><br>
    Community Forest Conservation Groups, Women's Survey, Men's Survey, GJJ KAP Women &amp; Men, Forest Training, Mangrove Training, Seaweed Production<br>
    <span>Generated with Python + Plotly | {datetime.now().strftime('%B %Y')}</span>
</div>

</div>

<script>
function showTab(tabId) {{
    document.querySelectorAll('.tab-content').forEach(el => el.style.display = 'none');
    document.querySelectorAll('.tab-btn').forEach(el => el.classList.remove('active'));
    document.getElementById(tabId).style.display = 'block';
    event.target.classList.add('active');
    window.scrollTo({{top: 0, behavior: 'smooth'}});
    // Trigger Plotly resize for the newly visible tab
    setTimeout(() => {{
        const plots = document.getElementById(tabId).querySelectorAll('.js-plotly-plot');
        plots.forEach(p => Plotly.Plots.resize(p));
    }}, 100);
}}
</script>
</body>
</html>'''


# ============================================================================
# MAIN — GENERATE OFFLINE DASHBOARD
# ============================================================================

def main():
    print("=" * 70)
    print("  COSME OFFLINE DASHBOARD GENERATOR")
    print("=" * 70)
    print()

    sections = []

    # 1. FORESTRY
    print("[1/8] Loading Forestry data...")
    try:
        f_data = load_forestry_data(FORESTRY_EXCEL)
        sections.append(('tab-forestry', 'Forestry Groups', build_forestry_section(f_data)))
        print("      ✓ Forestry section built")
    except Exception as e:
        print(f"      ✗ Forestry failed: {e}")
        f_data = None

    # 2. WOMEN
    print("[2/8] Loading Women Survey data...")
    try:
        w_data = load_women_data(WOMEN_EXCEL)
        sections.append(('tab-women', "Women's Survey", build_women_section(w_data)))
        print("      ✓ Women section built")
    except Exception as e:
        print(f"      ✗ Women failed: {e}")
        w_data = None

    # 3. MEN
    print("[3/8] Loading Men Survey data...")
    try:
        m_data = load_men_data(MEN_EXCEL)
        sections.append(('tab-men', "Men's Survey", build_men_section(m_data)))
        print("      ✓ Men section built")
    except Exception as e:
        print(f"      ✗ Men failed: {e}")
        m_data = None

    # 4. GJJ WOMEN
    print("[4/8] Loading GJJ KAP Women data...")
    try:
        gjj_data = load_gjj_kap_women_data(GJJ_KAP_WOMEN_EXCEL)
        sections.append(('tab-gjj-women', 'GJJ KAP Women', build_gjj_women_section(gjj_data)))
        print("      ✓ GJJ Women section built")
    except Exception as e:
        print(f"      ✗ GJJ Women failed: {e}")

    # 5. GJJ MEN
    print("[5/8] Loading GJJ KAP Men data...")
    try:
        gjj_men_data = load_gjj_kap_men_data(GJJ_KAP_MEN_EXCEL)
        sections.append(('tab-gjj-men', 'GJJ KAP Men', build_gjj_men_section(gjj_men_data)))
        print("      ✓ GJJ Men section built")
    except Exception as e:
        print(f"      ✗ GJJ Men failed: {e}")

    # 6. FOREST TRAINING
    print("[6/8] Loading Forest Training data...")
    try:
        ft_data = load_forest_training_data(FOREST_TRAINING_EXCEL)
        sections.append(('tab-ft', 'Forest Training', build_forest_training_section(ft_data)))
        print("      ✓ Forest Training section built")
    except Exception as e:
        print(f"      ✗ Forest Training failed: {e}")

    # 7. MANGROVE TRAINING
    print("[7/8] Loading Mangrove Training data...")
    try:
        mg_data = load_mangrove_training_data(MANGROVE_TRAINING_EXCEL)
        sections.append(('tab-mg', 'Mangrove Training', build_mangrove_training_section(mg_data)))
        print("      ✓ Mangrove Training section built")
    except Exception as e:
        print(f"      ✗ Mangrove Training failed: {e}")

    # 8. SEAWEED
    print("[8/8] Loading Seaweed data...")
    try:
        sw_data = load_seaweed_data(SEAWEED_CSV)
        sections.append(('tab-sw', 'Seaweed Production', build_seaweed_section(sw_data)))
        print("      ✓ Seaweed section built")
    except Exception as e:
        print(f"      ✗ Seaweed failed: {e}")

    # COMPARATIVE
    if f_data and w_data and m_data:
        print("[+]   Building comparative section...")
        sections.append(('tab-compare', 'Comparisons', build_comparative_section(f_data, w_data, m_data)))
        print("      ✓ Comparative section built")

    # ASSEMBLE HTML
    print()
    print("Assembling full HTML dashboard...")
    html_content = build_full_html(sections)

    output_path = os.path.join(SCRIPT_DIR, "COSME_Offline_Dashboard.html")
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html_content)

    file_size_mb = os.path.getsize(output_path) / (1024 * 1024)
    print(f"✓ Dashboard saved to: {output_path}")
    print(f"  File size: {file_size_mb:.1f} MB")
    print(f"  Sections: {len(sections)}")
    print()
    print("Open the HTML file in any browser — no server needed!")
    print("=" * 70)


if __name__ == "__main__":
    main()
