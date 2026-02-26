import openpyxl
wb = openpyxl.load_workbook(r"C:\Users\Bmuiruri.PLANKE-KILIFI\Desktop\Benard\My Project\My work\GJJ KAP Women Basline_endline results.xlsx", data_only=True)
print("Sheets:", wb.sheetnames)
for sname in wb.sheetnames:
    ws = wb[sname]
    print()
    print("=" * 80)
    print("Sheet:", sname, " Rows:", ws.max_row, " Cols:", ws.max_column)
    print("=" * 80)
    for r in range(1, min(ws.max_row+1, 300)):
        vals = []
        for c in range(1, min(ws.max_column+1, 10)):
            v = ws.cell(r, c).value
            vals.append(str(v)[:80] if v is not None else "")
        line = " | ".join(vals)
        if any(v for v in vals):
            print("  R" + str(r).rjust(3) + ": " + line)
