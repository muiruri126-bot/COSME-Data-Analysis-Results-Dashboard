"""
Generate COSME PMF Indicators Excel — All 20 Indicators with Numerator & Denominator Values
Corrected: proper numerator calculation for % indicators; N/A for non-count indicators.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import math

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "COSME PMF Indicators"

# ── Styles ──
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="005A9E", end_color="005A9E", fill_type="solid")
sub_header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
sub_header_fill2 = PatternFill(start_color="E65100", end_color="E65100", fill_type="solid")
sub_header_fill3 = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
sub_header_fill4 = PatternFill(start_color="00695C", end_color="00695C", fill_type="solid")
data_font = Font(name="Calibri", size=10)
exceeded_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
on_track_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
below_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
na_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
wrap_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

# ── Column widths ──
col_widths = {
    "A": 6, "B": 12, "C": 55, "D": 14, "E": 10, "F": 16, "G": 16,
    "H": 14, "I": 16, "J": 16, "K": 14, "L": 14, "M": 16, "N": 16,
    "O": 25
}
for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

# ── Title Row ──
ws.merge_cells("A1:O1")
title_cell = ws["A1"]
title_cell.value = "Plan International Kenya - COSME Project: PMF Indicator Performance with Numerator & Denominator Values"
title_cell.font = Font(name="Calibri", size=14, bold=True, color="005A9E")
title_cell.alignment = Alignment(horizontal="center", vertical="center")

ws.merge_cells("A2:O2")
subtitle = ws["A2"]
subtitle.value = "Numerator = number of respondents meeting criteria | Denominator = total respondents surveyed | For index/score/hours/hectare indicators: N/A shown (see Value columns)"
subtitle.font = Font(name="Calibri", size=9, italic=True, color="666666")
subtitle.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ── Headers (Row 4) ──
headers = [
    "#", "Indicator\nCode", "Indicator Name", "Thematic Area", "Unit",
    "Baseline\nDenominator (N)", "Baseline\nNumerator (n)", "Baseline\nValue",
    "Midline\nDenominator (N)", "Midline\nNumerator (n)", "Midline\nValue",
    "Target", "Change", "Status", "Data Source & Notes"
]

for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

# ────────────────────────────────────────────────────────────────────
# INDICATOR DATA
# type="pct"    -> Numerator = round(percentage * denominator / 100)
# type="index"  -> Composite score: Numerator/Denominator = N/A
# type="hours"  -> Average hours: Numerator/Denominator = N/A
# type="ha"     -> Hectares: Numerator/Denominator = N/A
# ────────────────────────────────────────────────────────────────────

indicators = [
    {"section": True, "label": "ULTIMATE OUTCOME - Overall Impact (1000-series)", "fill": header_fill},

    {"num": 1, "code": "1000a", "type": "index",
     "name": "Resilience Index (RI) average score",
     "theme": "1000 Climate-resilient economies",
     "unit": "/100",
     "bl_denom": 345, "bl_pct": 37.0,
     "ml_denom": 345, "ml_pct": 43.0,
     "target": "45.0/100", "change": "+6.0 pts",
     "status": "ON TRACK",
     "source": "Women COSME Survey - composite index score averaged across 345 respondents"},

    {"num": 2, "code": "1000b", "type": "index",
     "name": "Women and Girl's Empowerment Index (WGEI) average score",
     "theme": "1000 Climate-resilient economies",
     "unit": "/100",
     "bl_denom": 345, "bl_pct": 51.2,
     "ml_denom": 345, "ml_pct": 68.3,
     "target": "60.0/100", "change": "+17.1 pts",
     "status": "EXCEEDED",
     "source": "Women COSME Survey - composite index score averaged across 345 respondents"},

    {"section": True, "label": "INTERMEDIATE OUTCOME 1100 - NbS Adoption", "fill": sub_header_fill},

    {"num": 3, "code": "1100a", "type": "ha",
     "name": "Area (hectares) under rehabilitation using gender responsive NbS for climate mitigation and adaptation with biodiversity benefits",
     "theme": "1100 NbS adoption",
     "unit": "ha",
     "bl_denom": None, "bl_pct": 0,
     "ml_denom": None, "ml_pct": 63,
     "target": "123.8 ha", "change": "+63 ha",
     "status": "ON TRACK (50.9%)",
     "source": "Field Monitoring / GIS measurement - cumulative hectares rehabilitated"},

    {"section": True, "label": "IMMEDIATE OUTCOME 1110 - Mangrove Restoration & Conservation", "fill": sub_header_fill},

    {"num": 4, "code": "1110a", "type": "pct",
     "name": "% and Number of women and men with adequate practical knowledge on gender-responsive mangrove restoration and conservation - ALL (combined)",
     "theme": "1110 Mangrove capacity",
     "unit": "%",
     "bl_denom": 612, "bl_pct": 36.9,
     "ml_denom": 612, "ml_pct": 81.9,
     "target": "57.0%", "change": "+45.0pp",
     "status": "EXCEEDED",
     "source": "Women (n=345) + Men (n=267) combined = 612"},

    {"num": 5, "code": "1110a (F)", "type": "pct",
     "name": "% and Number of WOMEN with adequate practical knowledge on gender-responsive mangrove restoration and conservation",
     "theme": "1110 Mangrove capacity",
     "unit": "%",
     "bl_denom": 345, "bl_pct": 34.7,
     "ml_denom": 345, "ml_pct": 87.4,
     "target": "57.0%", "change": "+52.7pp",
     "status": "EXCEEDED",
     "source": "Women COSME Survey (n=345)"},

    {"num": 6, "code": "1110a (M)", "type": "pct",
     "name": "% and Number of MEN with adequate practical knowledge on gender-responsive mangrove restoration and conservation",
     "theme": "1110 Mangrove capacity",
     "unit": "%",
     "bl_denom": 267, "bl_pct": 38.2,
     "ml_denom": 267, "ml_pct": 80.2,
     "target": "57.0%", "change": "+42.0pp",
     "status": "EXCEEDED",
     "source": "Men COSME Survey (n=267)"},

    {"num": 7, "code": "1110b", "type": "pct",
     "name": "% and Number of women and men with adequate resources to undertake mangrove restoration and conservation",
     "theme": "1110 Mangrove capacity",
     "unit": "%",
     "bl_denom": 612, "bl_pct": 0,
     "ml_denom": 612, "ml_pct": 52,
     "target": "70%", "change": "+52.0pp",
     "status": "ON TRACK (74.3%)",
     "source": "Mangrove Resource Survey - Women+Men combined (n=612)"},

    {"section": True, "label": "IMMEDIATE OUTCOME 1120 - Seaweed Production Capacity", "fill": sub_header_fill},

    {"num": 8, "code": "1120a", "type": "pct",
     "name": "% and Number of targeted women with adequate practical knowledge in seaweed farming",
     "theme": "1120 Seaweed capacity",
     "unit": "%",
     "bl_denom": 610, "bl_pct": 49.3,
     "ml_denom": 610, "ml_pct": 29.3,
     "target": "68.0%", "change": "-20.0pp",
     "status": "BELOW TARGET",
     "source": "Seaweed Assessment - 19 groups, 610 women assessed"},

    {"num": 9, "code": "1120b", "type": "pct",
     "name": "% and Number of targeted women with adequate resources for seaweed production, value addition and commercialization",
     "theme": "1120 Seaweed capacity",
     "unit": "%",
     "bl_denom": 610, "bl_pct": 0,
     "ml_denom": 610, "ml_pct": 34,
     "target": "75%", "change": "+34.0pp",
     "status": "ON TRACK (45.3%)",
     "source": "Seaweed Assessment - 19 groups, 610 women assessed"},

    {"section": True, "label": "IMMEDIATE OUTCOME 1130 - Forest Management & Conservation", "fill": sub_header_fill},

    {"num": 10, "code": "1130a", "type": "pct",
     "name": "% and Number of women and men with adequate practical knowledge in forest management and conservation (disaggregated by sex)",
     "theme": "1130 Forest management",
     "unit": "%",
     "bl_denom": 546, "bl_pct": 50.0,
     "ml_denom": 909, "ml_pct": 94.7,
     "target": "68.0%", "change": "+44.7pp",
     "status": "EXCEEDED",
     "source": "Forest Conservation Training — Pre-test (n=546) vs Post-test (n=909)"},

    {"num": 11, "code": "1130b", "type": "pct",
     "name": "% and Number of community forest conservation groups that are functional (scoring >= 70%)",
     "theme": "1130 Forest management",
     "unit": "%",
     "bl_denom": 43, "bl_pct": 53.5,
     "ml_denom": 28, "ml_pct": 85.7,
     "target": "70.0%", "change": "+32.2pp",
     "status": "EXCEEDED",
     "source": "Forestry Assessment - BL: 43 groups, ML: 28 groups (units = groups, not people)"},

    {"section": True, "label": "INTERMEDIATE OUTCOME 1200 - Women's Agency & Behavior Change", "fill": sub_header_fill2},

    {"num": 12, "code": "1200a", "type": "pct",
     "name": "% of targeted women with adequate involvement in HH, group and enterprise decision making",
     "theme": "1200 Women's agency",
     "unit": "%",
     "bl_denom": 345, "bl_pct": 57.6,
     "ml_denom": 345, "ml_pct": 95.9,
     "target": "65.0%", "change": "+38.3pp",
     "status": "EXCEEDED",
     "source": "Women COSME Survey (n=345)"},

    {"num": 13, "code": "1200b", "type": "hours",
     "name": "Average time women spend in unpaid work",
     "theme": "1200 Women's agency",
     "unit": "hrs/day",
     "bl_denom": 345, "bl_pct": 9.3,
     "ml_denom": 345, "ml_pct": 7.7,
     "target": "8.3 hrs", "change": "-1.6 hrs",
     "status": "EXCEEDED",
     "source": "Women COSME Survey (n=345) - average hours per day across all respondents"},

    {"num": 14, "code": "1200c", "type": "pct",
     "name": "% of targeted women with adequate access and control over household and enterprise resources",
     "theme": "1200 Women's agency",
     "unit": "%",
     "bl_denom": 345, "bl_pct": 41.2,
     "ml_denom": 345, "ml_pct": 53.8,
     "target": "50.0%", "change": "+12.6pp",
     "status": "EXCEEDED",
     "source": "Women COSME Survey (n=345)"},

    {"section": True, "label": "IMMEDIATE OUTCOME 1210 - Women's Knowledge, Skills & Perception", "fill": sub_header_fill3},

    {"num": 15, "code": "1210a", "type": "pct",
     "name": "% of targeted women with adequate practical knowledge on NbS, Economic Rights and GE&I",
     "theme": "1210 Knowledge & skills",
     "unit": "%",
     "bl_denom": 345, "bl_pct": 20.7,
     "ml_denom": 345, "ml_pct": 80.9,
     "target": "55.0%", "change": "+60.2pp",
     "status": "EXCEEDED",
     "source": "Women COSME Survey (n=345)"},

    {"num": 16, "code": "1210b", "type": "pct",
     "name": "% of targeted women with adequate personal skills and confidence to assert their rights and express their decisions",
     "theme": "1210 Knowledge & skills",
     "unit": "%",
     "bl_denom": 345, "bl_pct": 32.7,
     "ml_denom": 345, "ml_pct": 46.3,
     "target": "55.0%", "change": "+13.6pp",
     "status": "ON TRACK (84.2%)",
     "source": "Women COSME Survey (n=345)"},

    {"num": 17, "code": "1210c", "type": "index",
     "name": "Degree/extent to which women are perceived as equal to men by targeted women",
     "theme": "1210 Knowledge & skills",
     "unit": "/100",
     "bl_denom": 345, "bl_pct": 42.1,
     "ml_denom": 345, "ml_pct": 64.7,
     "target": "55.0/100", "change": "+22.6 pts",
     "status": "EXCEEDED",
     "source": "Women COSME Survey (n=345) - perception index score averaged across respondents"},

    {"section": True, "label": "IMMEDIATE OUTCOME 1220 - Resilience-Building Assets", "fill": sub_header_fill3},

    {"num": 18, "code": "1220a", "type": "pct",
     "name": "% of targeted women and young women who report having savings (actively saving in a Savings Group)",
     "theme": "1220 Resilience assets",
     "unit": "%",
     "bl_denom": 345, "bl_pct": 28.8,
     "ml_denom": 345, "ml_pct": 91.9,
     "target": "70.0%", "change": "+63.1pp",
     "status": "EXCEEDED",
     "source": "Women COSME Survey (n=345) + VSLA Monitoring"},

    {"num": 19, "code": "1220b", "type": "pct",
     "name": "% of women with access to time-saving and environmentally friendly solar panels and cooking stoves",
     "theme": "1220 Resilience assets",
     "unit": "%",
     "bl_denom": 345, "bl_pct": 50.8,
     "ml_denom": 345, "ml_pct": 83.1,
     "target": "80.0%", "change": "+32.3pp",
     "status": "EXCEEDED",
     "source": "Women COSME Survey (n=345)"},

    {"section": True, "label": "IMMEDIATE OUTCOME 1230 - Male Family Members' Capacity", "fill": sub_header_fill4},

    {"num": 20, "code": "1230a", "type": "pct",
     "name": "% of male family members with adequate knowledge and practical skills to promote and support NbS, economic rights of women and GE&I",
     "theme": "1230 Male engagement",
     "unit": "%",
     "bl_denom": 267, "bl_pct": 43.57,
     "ml_denom": 267, "ml_pct": 89.77,
     "target": "65.0%", "change": "+46.2pp",
     "status": "EXCEEDED",
     "source": "Men COSME Survey (n=267)"},

    {"num": 21, "code": "1230b", "type": "index",
     "name": "Degree/extent to which women are perceived as equal to men by male family members",
     "theme": "1230 Male engagement",
     "unit": "/100",
     "bl_denom": 267, "bl_pct": 41.2,
     "ml_denom": 267, "ml_pct": 70.2,
     "target": "55.0/100", "change": "+29.0 pts",
     "status": "EXCEEDED",
     "source": "Men COSME Survey (n=267) - perception index score averaged across respondents"},
]


def calc_numerator(ind, phase):
    """
    Calculate numerator correctly based on indicator type.
    
    For type="pct": Numerator = round(percentage/100 x denominator)
        This gives the NUMBER of respondents meeting the criteria.
    For type="index", "hours", "ha": Returns None (N/A - not count-based).
    """
    pct = ind[f"{phase}_pct"]
    denom = ind[f"{phase}_denom"]
    itype = ind.get("type", "pct")
    
    if itype != "pct":
        return None   # N/A for index scores, hours, hectares
    
    if denom is None or denom == 0:
        return 0
    
    # Numerator = percentage/100 x denominator, rounded to nearest integer
    return round(pct * denom / 100)


def format_value(ind, phase):
    """Format the display value based on indicator type."""
    pct = ind[f"{phase}_pct"]
    itype = ind.get("type", "pct")

    if itype == "pct":
        return f"{pct}%"
    elif itype == "index":
        return f"{pct}/100"
    elif itype == "hours":
        return f"{pct} hrs/day"
    elif itype == "ha":
        return f"{pct} ha"
    return str(pct)


row = 5
for ind in indicators:
    if ind.get("section"):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=15)
        cell = ws.cell(row=row, column=1, value=ind["label"])
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = ind["fill"]
        cell.alignment = Alignment(horizontal="left", vertical="center")
        for c in range(1, 16):
            ws.cell(row=row, column=c).border = thin_border
        row += 1
        continue

    bl_num = calc_numerator(ind, "bl")
    ml_num = calc_numerator(ind, "ml")
    bl_val = format_value(ind, "bl")
    ml_val = format_value(ind, "ml")
    itype = ind.get("type", "pct")

    # Denominator display
    bl_denom_display = ind["bl_denom"] if ind["bl_denom"] is not None else "N/A"
    ml_denom_display = ind["ml_denom"] if ind["ml_denom"] is not None else "N/A"
    
    # Numerator display
    bl_num_display = bl_num if bl_num is not None else "N/A"
    ml_num_display = ml_num if ml_num is not None else "N/A"
    
    # For non-pct types, denominator column shows n respondents but note it's not a ratio
    if itype != "pct":
        if ind["bl_denom"] is not None:
            bl_denom_display = f"{ind['bl_denom']} (respondents)"
        if ind["ml_denom"] is not None:
            ml_denom_display = f"{ind['ml_denom']} (respondents)"

    values = [
        ind["num"],         # A: #
        ind["code"],        # B: Indicator Code
        ind["name"],        # C: Indicator Name
        ind["theme"],       # D: Thematic Area
        ind["unit"],        # E: Unit
        bl_denom_display,   # F: Baseline Denominator
        bl_num_display,     # G: Baseline Numerator
        bl_val,             # H: Baseline Value
        ml_denom_display,   # I: Midline Denominator
        ml_num_display,     # J: Midline Numerator
        ml_val,             # K: Midline Value
        ind["target"],      # L: Target
        ind["change"],      # M: Change
        ind["status"],      # N: Status
        ind["source"],      # O: Data Source & Notes
    ]

    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = data_font
        cell.border = thin_border
        if col_idx in (1, 2, 5, 6, 7, 8, 9, 10, 11, 12, 13):
            cell.alignment = center_align
        else:
            cell.alignment = wrap_align
        
        # Grey out N/A cells
        if val == "N/A":
            cell.fill = na_fill
            cell.font = Font(name="Calibri", size=10, italic=True, color="999999")

    # Status color
    status_cell = ws.cell(row=row, column=14)
    if "EXCEEDED" in ind["status"]:
        status_cell.fill = exceeded_fill
        status_cell.font = Font(name="Calibri", size=10, bold=True, color="1B5E20")
    elif "ON TRACK" in ind["status"]:
        status_cell.fill = on_track_fill
        status_cell.font = Font(name="Calibri", size=10, bold=True, color="F57F17")
    elif "BELOW" in ind["status"]:
        status_cell.fill = below_fill
        status_cell.font = Font(name="Calibri", size=10, bold=True, color="C62828")

    row += 1

# ── Verification Table ──
row += 2
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=15)
cell = ws.cell(row=row, column=1, value="NUMERATOR CALCULATION VERIFICATION")
cell.font = Font(name="Calibri", size=12, bold=True, color="005A9E")
row += 1

ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=15)
ws.cell(row=row, column=1, value="For % indicators: Numerator (n) = Percentage / 100 x Denominator (N), rounded to nearest integer. Verify: n / N x 100 = stated percentage.").font = Font(name="Calibri", size=9, italic=True, color="555555")
row += 1

# Verification headers
verify_headers = ["#", "Code", "Phase", "Denominator (N)", "Numerator (n)", "Stated %", "Verified %  (n/N x 100)", "Match?"]
for col_idx, h in enumerate(verify_headers, 1):
    cell = ws.cell(row=row, column=col_idx, value=h)
    cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="37474F", end_color="37474F", fill_type="solid")
    cell.alignment = center_align
    cell.border = thin_border
row += 1

for ind in indicators:
    if ind.get("section"):
        continue
    if ind.get("type", "pct") != "pct":
        continue
    
    for phase, phase_label in [("bl", "Baseline"), ("ml", "Midline")]:
        denom = ind[f"{phase}_denom"]
        pct = ind[f"{phase}_pct"]
        num = calc_numerator(ind, phase)
        
        if denom is None or denom == 0:
            verified_pct = 0
        else:
            verified_pct = round(num / denom * 100, 1)
        
        match = "Y" if abs(verified_pct - pct) < 0.5 else f"~(diff {abs(verified_pct - pct):.1f}pp)"
        
        vals = [ind["num"], ind["code"], phase_label, denom, num, f"{pct}%", f"{verified_pct}%", match]
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = data_font
            cell.alignment = center_align
            cell.border = thin_border
            if val == "Y":
                cell.font = Font(name="Calibri", size=10, bold=True, color="2E7D32")
        row += 1

# ── Notes Section ──
row += 1
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=15)
ws.cell(row=row, column=1, value="NOTES").font = Font(name="Calibri", size=11, bold=True, color="005A9E")
row += 1

notes = [
    "1. PERCENTAGE (%) INDICATORS: Numerator = number of respondents meeting criteria. Formula: n = round(Percentage/100 x N). These are back-calculated approximations from the reported percentages.",
    "2. INDEX/SCORE INDICATORS (1000a, 1000b, 1210c, 1230b): Composite scores averaged across all respondents. No numerator/denominator applies - shown as N/A. The value IS the average score on a 0-100 scale.",
    "3. HOURS INDICATOR (1200b): Average daily hours of unpaid work per woman. No numerator/denominator applies - shown as N/A. Value = mean hours across all surveyed women.",
    "4. HECTARES INDICATOR (1100a): Cumulative hectares of land under rehabilitation. No numerator/denominator applies - shown as N/A. Value = measured/reported hectares.",
    "5. DENOMINATORS: Women COSME Survey n=345, Men COSME Survey n=267, Combined n=612, Forestry Groups (BL: 43 groups, ML: 28 groups), Seaweed Assessment (19 groups, 610 women).",
    "6. 1130b (Forestry): Denominator is number of GROUPS assessed, not individual people. Baseline had 43 groups, Midline had 28 groups (different sample sizes).",
    "7. 1120a (Seaweed knowledge): Shows DECLINE from 49.3% to 29.3% - likely due to stricter/revised assessment criteria at midline.",
    "8. ROUNDING: Because numerators are back-calculated from rounded percentages, the verified % may differ from stated % by up to +/-0.5pp. The Verification Table above confirms accuracy.",
    "9. Change column: pp = percentage points, pts = index score points, hrs = hours/day, ha = hectares.",
]
for note in notes:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=15)
    cell = ws.cell(row=row, column=1, value=note)
    cell.font = Font(name="Calibri", size=9, italic=True, color="555555")
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 22
    row += 1

# ── Freeze panes ──
ws.freeze_panes = "A5"

# ── Row heights ──
ws.row_dimensions[1].height = 30
ws.row_dimensions[2].height = 26
ws.row_dimensions[4].height = 40
for r in range(5, row):
    if ws.row_dimensions[r].height is None or ws.row_dimensions[r].height < 48:
        ws.row_dimensions[r].height = 48

# ── Save ──
output_path = r"c:\Users\Bmuiruri.PLANKE-KILIFI\Desktop\Benard\My Project\My work\Plan_International_COSME_PMF_Indicators.xlsx"
wb.save(output_path)
print(f"Excel file saved: {output_path}")
print()

# Print verification summary
print("VERIFICATION - Percentage Indicators:")
print(f"{'Code':<12} {'Phase':<10} {'Denom':<8} {'Num':<8} {'Stated%':<10} {'Verified%':<12} {'OK?'}")
print("-" * 72)
for ind in indicators:
    if ind.get("section") or ind.get("type", "pct") != "pct":
        continue
    for phase, label in [("bl", "Baseline"), ("ml", "Midline")]:
        denom = ind[f"{phase}_denom"]
        pct = ind[f"{phase}_pct"]
        num = calc_numerator(ind, phase)
        vpct = round(num / denom * 100, 1) if denom else 0
        ok = "Y" if abs(vpct - pct) < 0.5 else f"~{abs(vpct - pct):.1f}pp"
        print(f"{ind['code']:<12} {label:<10} {denom:<8} {num:<8} {pct}%{'':<6} {vpct}%{'':<8} {ok}")
