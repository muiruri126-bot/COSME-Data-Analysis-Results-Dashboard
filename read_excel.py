import openpyxl
wb = openpyxl.load_workbook('Forest Functionality Basline_midline results.xlsx', read_only=True, data_only=True)
ws = wb['Results']

# Print ALL rows
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False)):
    vals = [cell.value for cell in row]
    # Only print non-empty rows
    if any(v is not None for v in vals):
        print(f'Row {i+1}: {vals}')
wb.close()
