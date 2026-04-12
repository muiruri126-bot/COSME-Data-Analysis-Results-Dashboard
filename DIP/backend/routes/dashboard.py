"""Dashboard + Reports API: KPIs, charts, exports."""
from datetime import date, timedelta
from io import BytesIO
from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from sqlalchemy import func, case
from models import (
    db, Task, Activity, Output, ImmediateOutcome, IntermediateOutcome,
    User, BudgetHolder, Notification, AuditLog, TASK_STATUSES
)
from auth_utils import get_current_user, require_roles, log_audit

dashboard_bp = Blueprint('dashboard', __name__, url_prefix='/api/v1')


def compute_kpis(task_query):
    """Compute KPIs from a task query."""
    tasks = task_query.filter_by(is_deleted=False).all()
    total = len(tasks)
    if total == 0:
        return {
            'total_tasks': 0, 'complete': 0, 'in_progress': 0,
            'pending': 0, 'delayed': 0, 'cancelled': 0,
            'percent_complete': 0, 'on_time_rate': 0,
            'avg_days_overdue': 0, 'tasks_by_status': {},
        }

    complete = [t for t in tasks if t.status == 'Complete']
    in_progress = [t for t in tasks if t.status == 'In progress']
    pending = [t for t in tasks if t.status == 'Pending']
    delayed = [t for t in tasks if t.status == 'Delayed']
    cancelled = [t for t in tasks if t.status == 'Cancelled']

    # On-time: complete tasks where actual end <= planned end
    on_time = 0
    for t in complete:
        # Find matching planned task
        planned = Task.query.filter_by(
            activity_id=t.activity_id, plan_actual='Planned', is_deleted=False
        ).first()
        if planned and planned.end_date and t.end_date <= planned.end_date:
            on_time += 1
        elif not planned:
            on_time += 1  # No planned reference = count as on-time

    on_time_rate = round(on_time / len(complete) * 100, 1) if complete else 0

    # Average days overdue
    today = date.today()
    overdue_tasks = [t for t in tasks if t.status in ('In progress', 'Delayed')
                     and t.end_date and t.end_date < today]
    avg_overdue = 0
    if overdue_tasks:
        total_overdue_days = sum((today - t.end_date).days for t in overdue_tasks)
        avg_overdue = round(total_overdue_days / len(overdue_tasks), 1)

    return {
        'total_tasks': total,
        'complete': len(complete),
        'in_progress': len(in_progress),
        'pending': len(pending),
        'delayed': len(delayed),
        'cancelled': len(cancelled),
        'percent_complete': round(len(complete) / total * 100, 1),
        'on_time_rate': on_time_rate,
        'avg_days_overdue': avg_overdue,
        'tasks_by_status': {
            'Pending': len(pending),
            'In progress': len(in_progress),
            'Complete': len(complete),
            'Delayed': len(delayed),
            'Cancelled': len(cancelled),
        },
    }


# ── Executive / M&E Dashboard ──────────────────────────────────

@dashboard_bp.get('/dashboard/executive')
@jwt_required()
def executive_dashboard():
    query = Task.query.filter_by(is_deleted=False)

    # Optional budget holder filter
    bh_id = request.args.get('budget_holder_id')
    if bh_id:
        act_ids = [a.id for a in Activity.query.filter_by(budget_holder_id=bh_id, is_deleted=False).all()]
        query = query.filter(Task.activity_id.in_(act_ids))

    kpis = compute_kpis(query)

    # Top delayed tasks
    today = date.today()
    delayed_tasks = Task.query.filter(
        Task.is_deleted == False,
        Task.status.in_(['In progress', 'Delayed']),
        Task.end_date < today,
    ).order_by(Task.end_date).limit(10).all()

    top_delayed = [{
        'id': t.id,
        'name': t.name,
        'activity_code': t.activity.code if t.activity else None,
        'end_date': t.end_date.strftime('%d/%m/%Y'),
        'days_overdue': (today - t.end_date).days,
        'responsible': t.responsible.full_name if t.responsible else None,
    } for t in delayed_tasks]

    # Workload by responsible
    workload_query = db.session.query(
        User.full_name,
        Task.status,
        func.count(Task.id),
    ).join(User, Task.responsible_id == User.id)\
     .filter(Task.is_deleted == False)\
     .group_by(User.full_name, Task.status).all()

    workload = {}
    for name, status, count in workload_query:
        if name not in workload:
            workload[name] = {'name': name, 'total': 0}
        workload[name][status] = count
        workload[name]['total'] += count

    # Delayed ageing buckets
    ageing = {'1-7': 0, '8-14': 0, '15-30': 0, '30+': 0}
    for t in delayed_tasks:
        days = (today - t.end_date).days
        if days <= 7:
            ageing['1-7'] += 1
        elif days <= 14:
            ageing['8-14'] += 1
        elif days <= 30:
            ageing['15-30'] += 1
        else:
            ageing['30+'] += 1

    return jsonify({
        'kpis': kpis,
        'top_delayed': top_delayed,
        'workload': list(workload.values()),
        'delayed_ageing': ageing,
    }), 200


# ── Budget Holder Dashboard ────────────────────────────────────

@dashboard_bp.get('/dashboard/budget-holder/<string:bh_id>')
@jwt_required()
def budget_holder_dashboard(bh_id):
    bh = BudgetHolder.query.get_or_404(bh_id)
    activities = Activity.query.filter_by(budget_holder_id=bh_id, is_deleted=False).all()
    act_ids = [a.id for a in activities]

    query = Task.query.filter(Task.activity_id.in_(act_ids), Task.is_deleted == False)
    kpis = compute_kpis(query)

    today = date.today()

    # Due soon (14 days)
    due_soon = Task.query.filter(
        Task.activity_id.in_(act_ids),
        Task.is_deleted == False,
        Task.status.notin_(['Complete', 'Cancelled']),
        Task.end_date >= today,
        Task.end_date <= today + timedelta(days=14),
    ).order_by(Task.end_date).all()

    # Overdue
    overdue = Task.query.filter(
        Task.activity_id.in_(act_ids),
        Task.is_deleted == False,
        Task.status.notin_(['Complete', 'Cancelled']),
        Task.end_date < today,
    ).order_by(Task.end_date).all()

    # Activities summary
    activity_summary = []
    for act in activities:
        act_tasks = Task.query.filter_by(activity_id=act.id, is_deleted=False).all()
        total = len(act_tasks)
        complete = sum(1 for t in act_tasks if t.status == 'Complete')
        activity_summary.append({
            'code': act.code,
            'description': act.description,
            'total_tasks': total,
            'percent_complete': round(complete / total * 100, 1) if total > 0 else 0,
        })

    return jsonify({
        'budget_holder': bh.to_dict(),
        'kpis': kpis,
        'due_soon': [t.to_dict() for t in due_soon],
        'overdue': [t.to_dict() for t in overdue],
        'activities': activity_summary,
    }), 200


# ── Implementer Dashboard ──────────────────────────────────────

@dashboard_bp.get('/dashboard/my-tasks')
@jwt_required()
def implementer_dashboard():
    user_id = get_jwt_identity()
    today = date.today()

    query = Task.query.filter_by(responsible_id=user_id, is_deleted=False)
    kpis = compute_kpis(query)

    due_soon = query.filter(
        Task.status.notin_(['Complete', 'Cancelled']),
        Task.end_date >= today,
        Task.end_date <= today + timedelta(days=14),
    ).order_by(Task.end_date).all()

    overdue = query.filter(
        Task.status.notin_(['Complete', 'Cancelled']),
        Task.end_date < today,
    ).order_by(Task.end_date).all()

    completed = query.filter_by(status='Complete').order_by(Task.updated_at.desc()).limit(10).all()

    return jsonify({
        'kpis': kpis,
        'due_soon': [t.to_dict() for t in due_soon],
        'overdue': [t.to_dict() for t in overdue],
        'recent_completed': [t.to_dict() for t in completed],
    }), 200


# ── Reports ─────────────────────────────────────────────────────

@dashboard_bp.get('/reports/activity-progress')
@jwt_required()
def activity_progress_report():
    activities = Activity.query.filter_by(is_deleted=False).order_by(Activity.sort_order).all()
    report = []
    for act in activities:
        tasks = Task.query.filter_by(activity_id=act.id, is_deleted=False).all()
        total = len(tasks)
        complete = sum(1 for t in tasks if t.status == 'Complete')
        in_progress = sum(1 for t in tasks if t.status == 'In progress')
        report.append({
            'code': act.code,
            'description': act.description,
            'budget_holder': act.budget_holder.name if act.budget_holder else None,
            'total_tasks': total,
            'complete': complete,
            'in_progress': in_progress,
            'percent_complete': round(complete / total * 100, 1) if total > 0 else 0,
        })
    return jsonify({'data': report}), 200


@dashboard_bp.get('/reports/output-completion')
@jwt_required()
def output_completion_report():
    outputs = Output.query.filter_by(is_deleted=False).order_by(Output.sort_order).all()
    report = []
    for out in outputs:
        activities = Activity.query.filter_by(output_id=out.id, is_deleted=False).all()
        act_ids = [a.id for a in activities]
        tasks = Task.query.filter(Task.activity_id.in_(act_ids), Task.is_deleted == False).all()
        total = len(tasks)
        complete = sum(1 for t in tasks if t.status == 'Complete')
        report.append({
            'code': out.code,
            'description': out.description,
            'activity_count': len(activities),
            'total_tasks': total,
            'complete': complete,
            'percent_complete': round(complete / total * 100, 1) if total > 0 else 0,
        })
    return jsonify({'data': report}), 200


@dashboard_bp.get('/reports/variance')
@jwt_required()
def variance_report():
    """Planned vs actual variance report."""
    actual_tasks = Task.query.filter_by(plan_actual='Actual', is_deleted=False).all()
    report = []
    for t in actual_tasks:
        planned = Task.query.filter_by(
            activity_id=t.activity_id, plan_actual='Planned', is_deleted=False
        ).first()
        if planned:
            start_var = (t.start_date - planned.start_date).days if t.start_date and planned.start_date else None
            end_var = (t.end_date - planned.end_date).days if t.end_date and planned.end_date else None
            report.append({
                'task_name': t.name,
                'activity_code': t.activity.code if t.activity else None,
                'planned_start': planned.start_date.strftime('%d/%m/%Y') if planned.start_date else None,
                'planned_end': planned.end_date.strftime('%d/%m/%Y') if planned.end_date else None,
                'actual_start': t.start_date.strftime('%d/%m/%Y') if t.start_date else None,
                'actual_end': t.end_date.strftime('%d/%m/%Y') if t.end_date else None,
                'start_variance_days': start_var,
                'end_variance_days': end_var,
                'responsible': t.responsible.full_name if t.responsible else None,
            })

    return jsonify({'data': report}), 200


@dashboard_bp.get('/reports/workload')
@jwt_required()
def workload_report():
    """Responsible person workload report."""
    users = User.query.filter_by(is_active=True).all()
    report = []
    for u in users:
        tasks = Task.query.filter_by(responsible_id=u.id, is_deleted=False).all()
        if tasks:
            status_counts = {}
            for s in TASK_STATUSES:
                status_counts[s] = sum(1 for t in tasks if t.status == s)
            report.append({
                'user': u.full_name,
                'email': u.email,
                'total_tasks': len(tasks),
                **status_counts,
            })

    return jsonify({'data': sorted(report, key=lambda x: -x['total_tasks'])}), 200


@dashboard_bp.get('/reports/dip-export')
@jwt_required()
def dip_export():
    """Export full DIP to Excel."""
    fmt = request.args.get('format', 'excel')

    if fmt == 'excel':
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'DIP Framework'

            # Headers
            headers = ['Level', 'Code', 'Description', 'Parent', 'Status',
                       'Task Count', 'Complete', '% Complete']
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                cell.font = Font(bold=True, color='FFFFFF')

            row = 2
            int_outcomes = IntermediateOutcome.query.filter_by(is_deleted=False).order_by(IntermediateOutcome.sort_order).all()

            for io in int_outcomes:
                ws.cell(row=row, column=1, value='Intermediate Outcome')
                ws.cell(row=row, column=2, value=io.code)
                ws.cell(row=row, column=3, value=io.description)
                row += 1

                for imo in io.immediate_outcomes.filter_by(is_deleted=False).order_by(ImmediateOutcome.sort_order).all():
                    ws.cell(row=row, column=1, value='Immediate Outcome')
                    ws.cell(row=row, column=2, value=imo.code)
                    ws.cell(row=row, column=3, value=imo.description)
                    ws.cell(row=row, column=4, value=io.code)
                    row += 1

                    for out in imo.outputs.filter_by(is_deleted=False).order_by(Output.sort_order).all():
                        ws.cell(row=row, column=1, value='Output')
                        ws.cell(row=row, column=2, value=out.code)
                        ws.cell(row=row, column=3, value=out.description)
                        ws.cell(row=row, column=4, value=imo.code)
                        row += 1

                        for act in out.activities.filter_by(is_deleted=False).order_by(Activity.sort_order).all():
                            tasks = Task.query.filter_by(activity_id=act.id, is_deleted=False).all()
                            total = len(tasks)
                            complete = sum(1 for t in tasks if t.status == 'Complete')
                            pct = round(complete / total * 100, 1) if total > 0 else 0

                            ws.cell(row=row, column=1, value='Activity')
                            ws.cell(row=row, column=2, value=act.code)
                            ws.cell(row=row, column=3, value=act.description)
                            ws.cell(row=row, column=4, value=out.code)
                            ws.cell(row=row, column=5, value=act.status)
                            ws.cell(row=row, column=6, value=total)
                            ws.cell(row=row, column=7, value=complete)
                            ws.cell(row=row, column=8, value=pct)
                            row += 1

            # Tasks sheet
            ws2 = wb.create_sheet('Tasks')
            task_headers = ['Activity Code', 'Task Name', 'Responsible', 'Plan/Actual',
                           'Start Date', 'End Date', 'Status', 'Created By']
            for col, h in enumerate(task_headers, 1):
                cell = ws2.cell(row=1, column=col, value=h)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                cell.font = Font(bold=True, color='FFFFFF')

            row = 2
            all_tasks = Task.query.filter_by(is_deleted=False).join(Activity).order_by(Activity.sort_order, Task.start_date).all()
            for t in all_tasks:
                ws2.cell(row=row, column=1, value=t.activity.code if t.activity else '')
                ws2.cell(row=row, column=2, value=t.name)
                ws2.cell(row=row, column=3, value=t.responsible.full_name if t.responsible else '')
                ws2.cell(row=row, column=4, value=t.plan_actual)
                ws2.cell(row=row, column=5, value=t.start_date.strftime('%d/%m/%Y') if t.start_date else '')
                ws2.cell(row=row, column=6, value=t.end_date.strftime('%d/%m/%Y') if t.end_date else '')
                ws2.cell(row=row, column=7, value=t.status)
                ws2.cell(row=row, column=8, value=t.creator.full_name if t.creator else '')
                row += 1

            # Auto-fit columns
            for ws_obj in [ws, ws2]:
                for col in ws_obj.columns:
                    max_len = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    ws_obj.column_dimensions[col_letter].width = min(max_len + 2, 60)

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return send_file(
                buffer,
                as_attachment=True,
                download_name='COSME_DIP_Export.xlsx',
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )

        except ImportError:
            return jsonify({'error': 'openpyxl not installed'}), 500

    return jsonify({'error': 'Unsupported format. Use ?format=excel'}), 400


# ── Notifications ───────────────────────────────────────────────

@dashboard_bp.get('/notifications')
@jwt_required()
def list_notifications():
    user_id = get_jwt_identity()
    notifs = Notification.query.filter_by(user_id=user_id)\
        .order_by(Notification.created_at.desc()).limit(50).all()
    unread = Notification.query.filter_by(user_id=user_id, is_read=False).count()
    return jsonify({
        'data': [n.to_dict() for n in notifs],
        'unread_count': unread,
    }), 200


@dashboard_bp.put('/notifications/<string:notif_id>/read')
@jwt_required()
def mark_notification_read(notif_id):
    user_id = get_jwt_identity()
    n = Notification.query.filter_by(id=notif_id, user_id=user_id).first_or_404()
    n.is_read = True
    db.session.commit()
    return jsonify({'message': 'Marked as read'}), 200


@dashboard_bp.put('/notifications/read-all')
@jwt_required()
def mark_all_read():
    user_id = get_jwt_identity()
    Notification.query.filter_by(user_id=user_id, is_read=False).update({'is_read': True})
    db.session.commit()
    return jsonify({'message': 'All marked as read'}), 200


# ── Admin: Audit Logs ──────────────────────────────────────────

@dashboard_bp.get('/admin/audit-logs')
@require_roles('Admin', 'ME_Specialist')
def list_audit_logs():
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)

    query = AuditLog.query.order_by(AuditLog.created_at.desc())

    entity_type = request.args.get('entity_type')
    if entity_type:
        query = query.filter_by(entity_type=entity_type)

    user_id = request.args.get('user_id')
    if user_id:
        query = query.filter_by(user_id=user_id)

    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    return jsonify({
        'data': [log.to_dict() for log in pagination.items],
        'total': pagination.total,
        'page': pagination.page,
        'pages': pagination.pages,
    }), 200


# ── Admin: User Management ─────────────────────────────────────

@dashboard_bp.get('/admin/users')
@require_roles('Admin')
def admin_list_users():
    users = User.query.order_by(User.full_name).all()
    return jsonify({'data': [u.to_dict() for u in users]}), 200


@dashboard_bp.post('/admin/users')
@require_roles('Admin')
def admin_create_user():
    from models import Role
    data = request.get_json()
    required = ('email', 'full_name', 'password')
    if not data or not all(data.get(f) for f in required):
        return jsonify({'error': 'email, full_name, and password required'}), 400

    if User.query.filter_by(email=data['email'].strip().lower()).first():
        return jsonify({'error': 'Email already exists'}), 409

    user = User(
        email=data['email'].strip().lower(),
        full_name=data['full_name'].strip(),
        phone=data.get('phone'),
        budget_holder_id=data.get('budget_holder_id'),
    )
    user.set_password(data['password'])

    # Assign roles
    for role_name in data.get('roles', []):
        role = Role.query.filter_by(name=role_name).first()
        if role:
            user.roles.append(role)

    db.session.add(user)
    db.session.flush()
    log_audit(get_jwt_identity(), 'user', user.id, 'create')
    db.session.commit()
    return jsonify(user.to_dict()), 201


@dashboard_bp.put('/admin/users/<string:uid>')
@require_roles('Admin')
def admin_update_user(uid):
    from models import Role
    user = User.query.get_or_404(uid)
    data = request.get_json()
    changes = {}

    for field in ('full_name', 'phone', 'budget_holder_id', 'is_active'):
        if field in data and getattr(user, field) != data[field]:
            changes[field] = {'old': str(getattr(user, field)), 'new': str(data[field])}
            setattr(user, field, data[field])

    if 'password' in data and data['password']:
        user.set_password(data['password'])
        changes['password'] = {'old': '***', 'new': '***'}

    if 'roles' in data:
        user.roles = []
        for role_name in data['roles']:
            role = Role.query.filter_by(name=role_name).first()
            if role:
                user.roles.append(role)
        changes['roles'] = {'new': data['roles']}

    if changes:
        log_audit(get_jwt_identity(), 'user', user.id, 'update', changes)

    db.session.commit()
    return jsonify(user.to_dict()), 200


@dashboard_bp.delete('/admin/users/<string:uid>')
@require_roles('Admin')
def admin_deactivate_user(uid):
    user = User.query.get_or_404(uid)
    user.is_active = False
    log_audit(get_jwt_identity(), 'user', user.id, 'update', {'is_active': {'old': 'True', 'new': 'False'}})
    db.session.commit()
    return jsonify({'message': 'User deactivated'}), 200
