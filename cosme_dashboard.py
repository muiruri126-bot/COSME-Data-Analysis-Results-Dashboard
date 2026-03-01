"""
=============================================================================
COSME INTEGRATED DASHBOARD
Baseline vs Midline Assessment
  1. Community Forest Conservation Groups (Group-level)
  2. Women's Survey (Household-level)
  3. Men's Survey (Household-level)
=============================================================================
A Streamlit + Plotly interactive dashboard for M&E analysis of the COSME
project, combining Forestry Conservation Groups, Women Survey, and Men Survey
datasets.

Excel files (place in the same folder as this script):
  • Forest Functionality Basline_midline results.xlsx (sheet "Results")
  • Women Survey Basline_midline results.xlsx (sheet "Results Women")
  • Men Survey Basline_midline results.xlsx (sheet "Results Men")

Run with: streamlit run cosme_dashboard.py
Requirements: pip install streamlit pandas numpy plotly openpyxl
=============================================================================
"""

import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import os, re
from collections import Counter

# ============================================================================
# CONFIGURATION
# ============================================================================
FORESTRY_EXCEL = "Forest Functionality Basline_midline results.xlsx"
FORESTRY_SHEET = "Results"

WOMEN_EXCEL = "Women Survey Basline_midline results.xlsx"
WOMEN_SHEET = "Results Women"

MEN_EXCEL = "Men Survey Basline_midline results.xlsx"
MEN_SHEET = "Results Men"

GJJ_KAP_WOMEN_EXCEL = "GJJ KAP Women Basline_endline results.xlsx"
GJJ_KAP_WOMEN_SHEET = "Results KAP Women Endline"

GJJ_KAP_MEN_EXCEL = "GJJ KAP Men Basline_endline results.xlsx"
GJJ_KAP_MEN_SHEET = "Results KAP Men Endline"

FOREST_TRAINING_EXCEL = "Forest Training Pre_post results.xlsx"
FOREST_TRAINING_SHEET = "Results"

MANGROVE_TRAINING_EXCEL = "Mangrove Training Pre_post results.xlsx"
MANGROVE_TRAINING_SHEET = "Results"

SEAWEED_CSV = "Seaweed Data Collection_06.12.2025.csv"

PROJECT_OUTPUTS_EXCEL = "Project Outputs.xlsx"

VSLA_EXCEL = "VSLA Functionality_(Q1-Q4) 2025.xlsx"
VSLA_SHEET = "Results (Across Qs)"

# Domain-level question grouping for Forest Training
FOREST_TRAINING_DOMAINS = {
    'PFM Concepts': [1, 2, 3, 4, 11],
    'Gender & Inclusivity': [6, 7, 8],
    'Forest Management Practices': [9, 10, 12],
    'Biodiversity Conservation': [13, 14, 15],
    'Agroforestry': [16, 17, 18],
    'Nursery Management': [19],
    'Climate Change': [20, 21],
}

# ============================================================================
# THEMES
# ============================================================================
THEMES = {
    "Forest Green": {
        "baseline": "#5B8DB8", "midline": "#2E7D32",
        "baseline_light": "#A3C4DC", "midline_light": "#81C784",
        "accent": "#FF8F00", "danger": "#E53935",
        "good": "#43A047", "medium": "#FB8C00",
        "poor": "#E53935", "low": "#43A047",
        "high": "#E53935", "decrease": "#43A047",
        "increase": "#E53935", "no_change": "#78909C",
        "header_gradient": "linear-gradient(135deg, #1B5E20 0%, #2E7D32 50%, #388E3C 100%)",
        "card_border": "#2E7D32", "card_value": "#1B5E20",
        "narrative_bg": "#E8F5E9", "narrative_border": "#2E7D32",
        "narrative_text": "#1B5E20",
        "radar_bl_fill": "rgba(91,141,184,0.3)",
        "radar_ml_fill": "rgba(46,125,50,0.3)",
    },
    "Ocean Blue": {
        "baseline": "#1565C0", "midline": "#00838F",
        "baseline_light": "#90CAF9", "midline_light": "#80DEEA",
        "accent": "#FF6F00", "danger": "#D32F2F",
        "good": "#00897B", "medium": "#F9A825",
        "poor": "#D32F2F", "low": "#00897B",
        "high": "#D32F2F", "decrease": "#00897B",
        "increase": "#D32F2F", "no_change": "#78909C",
        "header_gradient": "linear-gradient(135deg, #0D47A1 0%, #1565C0 50%, #1976D2 100%)",
        "card_border": "#1565C0", "card_value": "#0D47A1",
        "narrative_bg": "#E3F2FD", "narrative_border": "#1565C0",
        "narrative_text": "#0D47A1",
        "radar_bl_fill": "rgba(21,101,192,0.3)",
        "radar_ml_fill": "rgba(0,131,143,0.3)",
    },
    "Sunset": {
        "baseline": "#E65100", "midline": "#AD1457",
        "baseline_light": "#FFB74D", "midline_light": "#F48FB1",
        "accent": "#FDD835", "danger": "#B71C1C",
        "good": "#388E3C", "medium": "#F9A825",
        "poor": "#B71C1C", "low": "#388E3C",
        "high": "#B71C1C", "decrease": "#388E3C",
        "increase": "#B71C1C", "no_change": "#78909C",
        "header_gradient": "linear-gradient(135deg, #BF360C 0%, #E65100 50%, #F4511E 100%)",
        "card_border": "#E65100", "card_value": "#BF360C",
        "narrative_bg": "#FBE9E7", "narrative_border": "#E65100",
        "narrative_text": "#BF360C",
        "radar_bl_fill": "rgba(230,81,0,0.3)",
        "radar_ml_fill": "rgba(173,20,87,0.3)",
    },
    "Earth Tones": {
        "baseline": "#6D4C41", "midline": "#33691E",
        "baseline_light": "#BCAAA4", "midline_light": "#A5D6A7",
        "accent": "#FF8F00", "danger": "#C62828",
        "good": "#558B2F", "medium": "#F9A825",
        "poor": "#C62828", "low": "#558B2F",
        "high": "#C62828", "decrease": "#558B2F",
        "increase": "#C62828", "no_change": "#8D6E63",
        "header_gradient": "linear-gradient(135deg, #3E2723 0%, #5D4037 50%, #6D4C41 100%)",
        "card_border": "#6D4C41", "card_value": "#3E2723",
        "narrative_bg": "#EFEBE9", "narrative_border": "#6D4C41",
        "narrative_text": "#3E2723",
        "radar_bl_fill": "rgba(109,76,65,0.3)",
        "radar_ml_fill": "rgba(51,105,30,0.3)",
    },
    "Professional": {
        "baseline": "#37474F", "midline": "#1565C0",
        "baseline_light": "#B0BEC5", "midline_light": "#90CAF9",
        "accent": "#FF6F00", "danger": "#C62828",
        "good": "#2E7D32", "medium": "#EF6C00",
        "poor": "#C62828", "low": "#2E7D32",
        "high": "#C62828", "decrease": "#2E7D32",
        "increase": "#C62828", "no_change": "#78909C",
        "header_gradient": "linear-gradient(135deg, #263238 0%, #37474F 50%, #455A64 100%)",
        "card_border": "#37474F", "card_value": "#263238",
        "narrative_bg": "#ECEFF1", "narrative_border": "#37474F",
        "narrative_text": "#263238",
        "radar_bl_fill": "rgba(55,71,79,0.3)",
        "radar_ml_fill": "rgba(21,101,192,0.3)",
    },
}

COLORS = THEMES["Forest Green"]

# ============================================================================
# UTILITY helpers
# ============================================================================

def _val(df, row, col):
    """Safely extract a value from the raw DataFrame (0-based row/col)."""
    try:
        v = df.iloc[row, col]
        if v is None or (isinstance(v, float) and np.isnan(v)):
            return 0
        return v
    except (IndexError, KeyError):
        return 0


def _clean_label(s):
    """Strip trailing tabs/whitespace and normalise asset/indicator labels."""
    if isinstance(s, str):
        s = re.sub(r'[\t]+', '', s).strip().rstrip('-')
        # Title-case and replace underscores
        s = s.replace('_', ' ').strip()
        if s and s[0].islower():
            s = s.title()
        return s
    return str(s)


def pct(value):
    if isinstance(value, (int, float)):
        return f"{value * 100:.1f}%"
    return str(value)


def pp_change(bl, ml):
    if isinstance(bl, (int, float)) and isinstance(ml, (int, float)):
        return (ml - bl) * 100
    return 0


# ============================================================================
# FORESTRY DATA LOADER (unchanged from forestry_dashboard.py)
# ============================================================================

@st.cache_data
def load_forestry_data(filepath):
    try:
        raw = pd.read_excel(filepath, sheet_name=FORESTRY_SHEET, header=None)
    except FileNotFoundError:
        st.error(f"Forestry Excel not found: {filepath}")
        st.stop()

    data = {}

    data['functionality_threshold'] = pd.DataFrame({
        'Timepoint': ['Baseline', 'Midline'],
        'Functional_60_pct': [_val(raw, 7, 2), _val(raw, 8, 2)],
        'Functional_70_pct': [_val(raw, 7, 3), _val(raw, 8, 3)],
    })
    data['functionality_scores'] = pd.DataFrame({
        'Timepoint': ['Baseline', 'Midline'],
        'Respondents': [_val(raw, 13, 2), _val(raw, 14, 2)],
        'Average': [_val(raw, 13, 3), _val(raw, 14, 3)],
        'Max': [_val(raw, 13, 4), _val(raw, 14, 4)],
        'Min': [_val(raw, 13, 5), _val(raw, 14, 5)],
    })
    data['functionality_domain'] = pd.DataFrame({
        'Timepoint': ['Baseline', 'Midline'],
        'Management': [_val(raw, 19, 2), _val(raw, 20, 2)],
        'Gender': [_val(raw, 19, 3), _val(raw, 20, 3)],
        'Effectiveness': [_val(raw, 19, 4), _val(raw, 20, 4)],
        'Overall': [_val(raw, 19, 5), _val(raw, 20, 5)],
    })
    data['num_groups'] = pd.DataFrame({
        'Timepoint': ['Baseline', 'Midline'],
        'Groups_Assessed': [_val(raw, 27, 2), _val(raw, 27, 3)],
    })
    data['avg_membership'] = pd.DataFrame({
        'Category': ['Male', 'Female', 'All'],
        'Baseline': [_val(raw, 32, 2), _val(raw, 33, 2), _val(raw, 34, 2)],
        'Midline': [_val(raw, 32, 3), _val(raw, 33, 3), _val(raw, 34, 3)],
    })
    data['max_membership'] = pd.DataFrame({
        'Category': ['Male', 'Female', 'All'],
        'Baseline': [_val(raw, 39, 2), _val(raw, 40, 2), _val(raw, 41, 2)],
        'Midline': [_val(raw, 39, 3), _val(raw, 40, 3), _val(raw, 41, 3)],
    })
    data['min_membership'] = pd.DataFrame({
        'Category': ['Male', 'Female', 'All'],
        'Baseline': [_val(raw, 46, 2), _val(raw, 47, 2), _val(raw, 48, 2)],
        'Midline': [_val(raw, 46, 3), _val(raw, 47, 3), _val(raw, 48, 3)],
    })
    data['years_stats'] = pd.DataFrame({
        'Stat': ['Average', 'Maximum', 'Minimum'],
        'Baseline': [_val(raw, 53, 2), _val(raw, 54, 2), _val(raw, 55, 2)],
        'Midline': [_val(raw, 53, 3), _val(raw, 54, 3), _val(raw, 55, 3)],
    })
    data['years_dist'] = pd.DataFrame({
        'Category': ['0-3 years', '4-5 years', '6-10 years', '11+ years'],
        'Baseline': [_val(raw, 57, 2), _val(raw, 58, 2), _val(raw, 59, 2), _val(raw, 60, 2)],
        'Midline': [_val(raw, 57, 3), _val(raw, 58, 3), _val(raw, 59, 3), _val(raw, 60, 3)],
    })
    data['tenure'] = pd.DataFrame({
        'Category': ['Community', 'Government', 'Govt Devolved', 'Individual', 'Other'],
        'Baseline': [_val(raw, 65, 2), _val(raw, 66, 2), _val(raw, 67, 2), _val(raw, 68, 2), _val(raw, 69, 2)],
        'Midline': [_val(raw, 65, 3), _val(raw, 66, 3), _val(raw, 67, 3), _val(raw, 68, 3), _val(raw, 69, 3)],
    })
    data['forest_size_stats'] = pd.DataFrame({
        'Stat': ['Average', 'Maximum', 'Minimum'],
        'Baseline': [_val(raw, 74, 2), _val(raw, 75, 2), _val(raw, 76, 2)],
        'Midline': [_val(raw, 74, 3), _val(raw, 75, 3), _val(raw, 76, 3)],
    })
    data['forest_size_dist'] = pd.DataFrame({
        'Category': ['0 ha', '1-5 ha', '6-10 ha', '11-100 ha', '101-999 ha', '1000+ ha'],
        'Baseline': [_val(raw, 78, 2), _val(raw, 79, 2), _val(raw, 80, 2), _val(raw, 81, 2), _val(raw, 82, 2), _val(raw, 83, 2)],
        'Midline': [_val(raw, 78, 3), _val(raw, 79, 3), _val(raw, 80, 3), _val(raw, 81, 3), _val(raw, 82, 3), _val(raw, 83, 3)],
    })
    data['goals_defined'] = pd.DataFrame({
        'Category': ['No', 'Partially', 'Yes'],
        'Baseline': [_val(raw, 91, 2), _val(raw, 92, 2), _val(raw, 93, 2)],
        'Midline': [_val(raw, 91, 3), _val(raw, 92, 3), _val(raw, 93, 3)],
    })
    data['goals_stated'] = pd.DataFrame({
        'Goal': ['Sustain Forest Resources', 'Local Livelihoods/Poverty',
                 'Protect Local Rights', 'Public Rights/State Control', 'Other'],
        'Baseline': [_val(raw, 98, 9), _val(raw, 99, 9), _val(raw, 100, 9), _val(raw, 101, 9), _val(raw, 102, 9)],
        'Midline': [_val(raw, 98, 10), _val(raw, 99, 10), _val(raw, 100, 10), _val(raw, 101, 10), _val(raw, 102, 10)],
    })
    data['rights'] = pd.DataFrame({
        'Right': ['Access', 'Withdrawal (Subsistence)', 'Withdrawal (Commercial)',
                  'Management', 'Exclusion', 'Alienation', 'Compensation', "Don't Know"],
        'Baseline': [_val(raw, r, 9) for r in range(107, 115)],
        'Midline': [_val(raw, r, 10) for r in range(107, 115)],
    })
    resp_labels = ['Register Group', 'EIA', 'Management Plan', 'Forestry Inventory',
                   'Approval NWFP', 'Approval Grazing', 'Approval Fuelwood', 'Approval Timber',
                   'Approval Transport/Sale', 'Pay Tax', 'Certification',
                   'Monitor Restrictions', 'Other', "Don't Know"]
    data['responsibilities'] = pd.DataFrame({
        'Responsibility': resp_labels,
        'Baseline': [_val(raw, r, 8) for r in range(119, 133)],
        'Midline': [_val(raw, r, 9) for r in range(119, 133)],
    })
    data['board_roles'] = pd.DataFrame({
        'Category': ['Well Defined & Understood', 'Well Defined, Not All Understand',
                      'Not Defined at All', 'Not Well Defined, Not Understood'],
        'Baseline': [_val(raw, 140, 11), _val(raw, 141, 11), _val(raw, 142, 11), _val(raw, 143, 11)],
        'Midline': [_val(raw, 140, 12), _val(raw, 141, 12), _val(raw, 142, 12), _val(raw, 143, 12)],
    })
    data['guidelines'] = pd.DataFrame({
        'Category': ['Adopted & Known Most', 'Approved & Known Some',
                      'Drafted & Known Few', 'No Guidelines'],
        'Baseline': [_val(raw, 148, 5), _val(raw, 149, 5), _val(raw, 150, 5), _val(raw, 151, 5)],
        'Midline': [_val(raw, 148, 6), _val(raw, 149, 6), _val(raw, 150, 6), _val(raw, 151, 6)],
    })
    data['meetings'] = pd.DataFrame({
        'Category': ['Occasional / Low Attendance', 'Often / Moderate Attendance',
                      'Regular / Full Quorum'],
        'Baseline': [_val(raw, 156, 11), _val(raw, 157, 11), _val(raw, 158, 11)],
        'Midline': [_val(raw, 156, 12), _val(raw, 157, 12), _val(raw, 158, 12)],
    })
    data['women_leadership'] = pd.DataFrame({
        'Category': ['No Leadership, No Voice', 'No Leadership, Occasional Voice',
                      'Some Leadership', 'Significant Leadership'],
        'Baseline': [_val(raw, 163, 12), _val(raw, 164, 12), _val(raw, 165, 12), _val(raw, 166, 12)],
        'Midline': [_val(raw, 163, 13), _val(raw, 164, 13), _val(raw, 165, 13), _val(raw, 166, 13)],
    })
    mgmt_labels = ['Transparent Elections', 'Inclusive Decisions', 'Clear Procedures',
                   'Grievance Mechanism', 'Benefit Sharing Rules', 'Decisions Communicated']
    data['mgmt_practices'] = pd.DataFrame({
        'Practice': mgmt_labels,
        'Agree_Baseline': [_val(raw, r, 8) for r in range(173, 179)],
        'Agree_Midline': [_val(raw, r, 9) for r in range(173, 179)],
        'StronglyAgree_Baseline': [_val(raw, r, 10) for r in range(173, 179)],
        'StronglyAgree_Midline': [_val(raw, r, 11) for r in range(173, 179)],
    })
    data['training_coverage'] = pd.DataFrame({
        'Category': ['None', 'A Few Members', 'Many Members', 'Most Members'],
        'Baseline': [_val(raw, 186, 4), _val(raw, 187, 4), _val(raw, 188, 4), _val(raw, 189, 4)],
        'Midline': [_val(raw, 186, 5), _val(raw, 187, 5), _val(raw, 188, 5), _val(raw, 189, 5)],
    })
    topic_labels = ['Climate Change & Forests', 'People & Forests', 'CFM Concepts',
                    'Policy & Regulatory', 'Group Establishment', 'Participation',
                    'Leadership', 'Forest Assessment', 'Socioeconomic Assessment',
                    'Visioning & Objectives', 'Group Management', 'Group Registration',
                    'Environmental Impact Assessment', 'Management Plan',
                    'Forest Livelihoods', 'Forest Enterprises',
                    'Gender & Governance', 'Gender & Environment',
                    'Disaster Risk Management', 'Other']
    data['training_topics'] = pd.DataFrame({
        'Topic': topic_labels,
        'Baseline': [_val(raw, r, 6) for r in range(194, 214)],
        'Midline': [_val(raw, r, 7) for r in range(194, 214)],
    })
    data['assets_received'] = pd.DataFrame({
        'Category': ['No', 'Yes'],
        'Baseline': [_val(raw, 218, 2), _val(raw, 219, 2)],
        'Midline': [_val(raw, 218, 3), _val(raw, 219, 3)],
    })
    data['asset_types'] = pd.DataFrame({
        'Asset': ['Seedlings/Nursery', 'Agricultural Inputs', 'Energy (Solar/Stoves)',
                  'Water & Irrigation', 'Tools & Equipment', 'Beekeeping Inputs', 'Infrastructure'],
        'Baseline': [_val(raw, 224, 2), _val(raw, 225, 2), _val(raw, 226, 2), _val(raw, 227, 2),
                     _val(raw, 228, 2), _val(raw, 229, 2), _val(raw, 230, 2)],
        'Midline': [_val(raw, 224, 3), _val(raw, 225, 3), _val(raw, 226, 3), _val(raw, 227, 3),
                     _val(raw, 228, 3), _val(raw, 229, 3), _val(raw, 230, 3)],
    })
    data['ge_discussion'] = pd.DataFrame({
        'Category': ['No', 'Yes'],
        'Baseline': [_val(raw, 238, 2), _val(raw, 239, 2)],
        'Midline': [_val(raw, 238, 3), _val(raw, 239, 3)],
    })
    data['ge_topics'] = pd.DataFrame({
        'Topic': ['Equitable Roles/Responsibilities', 'Resource Sharing',
                  'Women Leadership & Voice', 'Other'],
        'Baseline': [_val(raw, 244, 6), _val(raw, 245, 6), _val(raw, 246, 6), _val(raw, 247, 6)],
        'Midline': [_val(raw, 244, 7), _val(raw, 245, 7), _val(raw, 246, 7), _val(raw, 247, 7)],
    })
    data['ge_actions'] = pd.DataFrame({
        'Category': ['No', 'Partially', 'Yes'],
        'Baseline': [_val(raw, 252, 2), _val(raw, 253, 2), _val(raw, 254, 2)],
        'Midline': [_val(raw, 252, 3), _val(raw, 253, 3), _val(raw, 254, 3)],
    })
    data['ge_completion'] = pd.DataFrame({
        'Category': ['None', 'A Few', 'Many', 'All'],
        'Baseline': [_val(raw, 259, 2), _val(raw, 260, 2), _val(raw, 261, 2), _val(raw, 262, 2)],
        'Midline': [_val(raw, 259, 3), _val(raw, 260, 3), _val(raw, 261, 3), _val(raw, 262, 3)],
    })
    cond_labels = ['Area', 'Volume/Biomass', 'Regeneration', 'Biodiversity', 'Ecosystem Services']
    data['forest_condition'] = pd.DataFrame({
        'Characteristic': cond_labels,
        'Baseline_Good': [_val(raw, r, 2) for r in range(271, 276)],
        'Baseline_Medium': [_val(raw, r, 3) for r in range(271, 276)],
        'Baseline_Poor': [_val(raw, r, 4) for r in range(271, 276)],
        'Midline_Good': [_val(raw, r, 5) for r in range(271, 276)],
        'Midline_Medium': [_val(raw, r, 6) for r in range(271, 276)],
        'Midline_Poor': [_val(raw, r, 7) for r in range(271, 276)],
    })
    data['forest_change'] = pd.DataFrame({
        'Characteristic': cond_labels,
        'Baseline_Decrease': [_val(raw, r, 2) for r in range(280, 285)],
        'Baseline_Increase': [_val(raw, r, 3) for r in range(280, 285)],
        'Baseline_NoChange': [_val(raw, r, 4) for r in range(280, 285)],
        'Midline_Decrease': [_val(raw, r, 5) for r in range(280, 285)],
        'Midline_Increase': [_val(raw, r, 6) for r in range(280, 285)],
        'Midline_NoChange': [_val(raw, r, 7) for r in range(280, 285)],
    })
    threat_labels = ['Fire', 'Fuelwood', 'Charcoal', 'Other Wood', 'Poaching',
                     'Encroachment', 'Land Grabbing', 'Poles']
    data['threats'] = pd.DataFrame({
        'Threat': threat_labels,
        'Baseline_Low': [_val(raw, r, 2) for r in range(290, 298)],
        'Baseline_Medium': [_val(raw, r, 3) for r in range(290, 298)],
        'Baseline_High': [_val(raw, r, 4) for r in range(290, 298)],
        'Midline_Low': [_val(raw, r, 5) for r in range(290, 298)],
        'Midline_Medium': [_val(raw, r, 6) for r in range(290, 298)],
        'Midline_High': [_val(raw, r, 7) for r in range(290, 298)],
    })
    data['threat_changes'] = pd.DataFrame({
        'Threat': threat_labels,
        'Baseline_Decrease': [_val(raw, r, 2) for r in range(302, 310)],
        'Baseline_Increase': [_val(raw, r, 3) for r in range(302, 310)],
        'Baseline_NoChange': [_val(raw, r, 4) for r in range(302, 310)],
        'Midline_Decrease': [_val(raw, r, 5) for r in range(302, 310)],
        'Midline_Increase': [_val(raw, r, 6) for r in range(302, 310)],
        'Midline_NoChange': [_val(raw, r, 7) for r in range(302, 310)],
    })
    harvest_labels = ['Timber', 'Woodfuel', 'Poles', 'Fodder', 'Leaf Mulch',
                      'Wildlife', 'Food', 'NWFPs']
    data['harvest_amount'] = pd.DataFrame({
        'Product': harvest_labels,
        'Baseline_None': [_val(raw, r, 2) for r in range(314, 322)],
        'Baseline_Medium': [_val(raw, r, 3) for r in range(314, 322)],
        'Baseline_Substantial': [_val(raw, r, 4) for r in range(314, 322)],
        'Midline_None': [_val(raw, r, 5) for r in range(314, 322)],
        'Midline_Medium': [_val(raw, r, 6) for r in range(314, 322)],
        'Midline_Substantial': [_val(raw, r, 7) for r in range(314, 322)],
    })
    data['harvest_changes'] = pd.DataFrame({
        'Product': harvest_labels,
        'Baseline_Decrease': [_val(raw, r, 2) for r in range(326, 334)],
        'Baseline_NoChange': [_val(raw, r, 3) for r in range(326, 334)],
        'Baseline_Increase': [_val(raw, r, 4) for r in range(326, 334)],
        'Baseline_DontKnow': [_val(raw, r, 5) for r in range(326, 334)],
        'Midline_Decrease': [_val(raw, r, 6) for r in range(326, 334)],
        'Midline_NoChange': [_val(raw, r, 7) for r in range(326, 334)],
        'Midline_Increase': [_val(raw, r, 8) for r in range(326, 334)],
        'Midline_DontKnow': [_val(raw, r, 9) for r in range(326, 334)],
    })
    data['income_gen'] = pd.DataFrame({
        'Category': ['Yes', 'No'],
        'Baseline': [_val(raw, 337, 2), _val(raw, 338, 2)],
        'Midline': [_val(raw, 337, 3), _val(raw, 338, 3)],
    })
    data['income_sources'] = pd.DataFrame({
        'Source': ['Timber', 'Fuelwood', 'Wildlife', 'NWFPs', 'PES', 'Poles', 'Other'],
        'Baseline': [_val(raw, r, 2) for r in range(342, 349)],
        'Midline': [_val(raw, r, 3) for r in range(342, 349)],
    })
    data['income_use'] = pd.DataFrame({
        'Use': ['HH Needs', 'Vulnerable Support', 'Reinvest Forest',
                'Community Initiatives', 'Individual Enterprises',
                'Community Enterprises', 'Other'],
        'Baseline': [_val(raw, r, 2) for r in range(352, 359)],
        'Midline': [_val(raw, r, 3) for r in range(352, 359)],
    })
    data['agroforestry'] = pd.DataFrame({
        'Category': ['Yes', 'No'],
        'Baseline': [_val(raw, 362, 2), _val(raw, 363, 2)],
        'Midline': [_val(raw, 362, 3), _val(raw, 363, 3)],
    })
    data['af_types'] = pd.DataFrame({
        'Practice': ['Intercropping', 'Silvopasture', 'Alley Cropping', 'Forest Farming', 'Beekeeping'],
        'Baseline': [_val(raw, r, 2) for r in range(367, 372)],
        'Midline': [_val(raw, r, 3) for r in range(367, 372)],
    })
    data['af_objectives'] = pd.DataFrame({
        'Objective': ['Biodiversity', 'Soil Fertility', 'Income', 'Water Retention', 'Reduce Deforestation', 'Food Security'],
        'Baseline': [_val(raw, r, 2) for r in range(375, 381)],
        'Midline': [_val(raw, r, 3) for r in range(375, 381)],
    })
    data['af_training'] = pd.DataFrame({
        'Category': ['Yes', 'No'],
        'Baseline': [_val(raw, 384, 2), _val(raw, 385, 2)],
        'Midline': [_val(raw, 384, 3), _val(raw, 385, 3)],
    })
    data['af_support'] = pd.DataFrame({
        'Support': ['On-site Workshops', 'Materials/Seedlings', 'Market Access',
                    'Financial Assistance', 'Technical Support', 'Online Resources'],
        'Baseline': [_val(raw, r, 2) for r in range(389, 395)],
        'Midline': [_val(raw, r, 3) for r in range(389, 395)],
    })
    data['af_challenges'] = pd.DataFrame({
        'Challenge': ['Lack of Knowledge', 'Insufficient Funding', 'Market Access',
                      'Land/Property Rights', 'Environmental Constraints',
                      'Policy/Regulatory', 'Cultural Resistance'],
        'Baseline': [_val(raw, r, 2) for r in range(398, 405)],
        'Midline': [_val(raw, r, 3) for r in range(398, 405)],
    })
    data['af_reinvest'] = pd.DataFrame({
        'Category': ['Education & Training', 'Not Applicable', 'Land Purchase', 'Conservation Projects'],
        'Baseline': [_val(raw, r, 2) for r in range(408, 412)],
        'Midline': [_val(raw, r, 3) for r in range(408, 412)],
    })
    data['af_potential'] = pd.DataFrame({
        'Category': ['Moderate', 'Significant', 'Unsure'],
        'Baseline': [_val(raw, 415, 2), _val(raw, 416, 2), _val(raw, 417, 2)],
        'Midline': [_val(raw, 415, 3), _val(raw, 416, 3), _val(raw, 417, 3)],
    })
    return data


# ============================================================================
# WOMEN SURVEY DATA LOADER
# ============================================================================

@st.cache_data
def load_women_data(filepath):
    """
    Parse the Women Survey Excel file.
    Sheet 'Results Women' — 516 rows × 21 columns, non-standard layout.
    Each section is parsed from specific row/col positions.
    """
    try:
        raw = pd.read_excel(filepath, sheet_name=WOMEN_SHEET, header=None)
    except FileNotFoundError:
        st.error(f"Women Survey Excel not found: {filepath}")
        st.stop()

    w = {}

    # ---- A. HOUSEHOLD CHARACTERISTICS ----
    # Location (rows 11-12, cols 1-3)
    w['location'] = pd.DataFrame({
        'Category': ['Marine', 'Terrestrial'],
        'Baseline': [_val(raw, 11, 2), _val(raw, 12, 2)],
        'Midline': [_val(raw, 11, 3), _val(raw, 12, 3)],
    })
    # HH type (rows 11-12, cols 10-12)
    w['hh_type'] = pd.DataFrame({
        'Category': ['Female-Headed', 'Male-Headed'],
        'Baseline': [_val(raw, 11, 11), _val(raw, 12, 11)],
        'Midline': [_val(raw, 11, 12), _val(raw, 12, 12)],
    })
    # Marital status (rows 16-22, cols 1-3)
    w['marital'] = pd.DataFrame({
        'Category': ['Monogamously Married', 'Widowed', 'Polygamously Married',
                     'Separated', 'Single', 'Divorced'],
        'Baseline': [_val(raw, r, 2) for r in range(16, 22)],
        'Midline': [_val(raw, r, 3) for r in range(16, 22)],
    })
    # Education (rows 16-19, cols 10-12)
    w['education'] = pd.DataFrame({
        'Category': ['Primary', 'Pre-primary/None/Other', 'Secondary/Vocational', 'College+'],
        'Baseline': [_val(raw, r, 11) for r in range(16, 20)],
        'Midline': [_val(raw, r, 12) for r in range(16, 20)],
    })
    # Main economic activity (rows 26-39, cols 1-3)
    main_econ_labels = [_clean_label(_val(raw, r, 1)) for r in range(26, 40)
                        if _val(raw, r, 1) != 0]
    main_econ_bl = [_val(raw, r, 2) for r in range(26, 40) if _val(raw, r, 1) != 0]
    main_econ_ml = [_val(raw, r, 3) for r in range(26, 40) if _val(raw, r, 1) != 0]
    w['main_econ'] = pd.DataFrame({
        'Activity': main_econ_labels, 'Baseline': main_econ_bl, 'Midline': main_econ_ml})

    # Secondary economic activity (rows 26-39, cols 10-12)
    sec_econ_labels = [_clean_label(_val(raw, r, 10)) for r in range(26, 40)
                       if _val(raw, r, 10) != 0]
    sec_econ_bl = [_val(raw, r, 11) for r in range(26, 40) if _val(raw, r, 10) != 0]
    sec_econ_ml = [_val(raw, r, 12) for r in range(26, 40) if _val(raw, r, 10) != 0]
    w['sec_econ'] = pd.DataFrame({
        'Activity': sec_econ_labels, 'Baseline': sec_econ_bl, 'Midline': sec_econ_ml})

    # Drinking water (rows 44-45, cols 1-3)
    w['water'] = pd.DataFrame({
        'Category': ['Yes', 'No'],
        'Baseline': [_val(raw, 44, 2), _val(raw, 45, 2)],
        'Midline': [_val(raw, 44, 3), _val(raw, 45, 3)],
    })
    # Toilet (rows 44-45, cols 10-12)
    w['toilet'] = pd.DataFrame({
        'Category': ['Yes', 'No'],
        'Baseline': [_val(raw, 44, 11), _val(raw, 45, 11)],
        'Midline': [_val(raw, 44, 12), _val(raw, 45, 12)],
    })
    # Electricity (rows 49-50, cols 1-3)
    w['electricity'] = pd.DataFrame({
        'Category': ['Yes', 'No'],
        'Baseline': [_val(raw, 49, 2), _val(raw, 50, 2)],
        'Midline': [_val(raw, 49, 3), _val(raw, 50, 3)],
    })
    # Wall materials (rows 54-56, cols 1-3)
    w['walls'] = pd.DataFrame({
        'Category': ['Natural/Mud', 'Finished (Cement/Brick)', 'Uncovered/Iron'],
        'Baseline': [_val(raw, 54, 2), _val(raw, 55, 2), _val(raw, 56, 2)],
        'Midline': [_val(raw, 54, 3), _val(raw, 55, 3), _val(raw, 56, 3)],
    })
    # Floor materials (rows 54-55, cols 10-12)
    w['floors'] = pd.DataFrame({
        'Category': ['Natural (Earth/Sand)', 'Other (Wood/Tiles/Cement)'],
        'Baseline': [_val(raw, 54, 11), _val(raw, 55, 11)],
        'Midline': [_val(raw, 54, 12), _val(raw, 55, 12)],
    })

    # ---- B. SHOCKS AND STRESSES ----
    # Shocks experienced (rows 64-74, cols 1-3) - row 65 has no label (continuation of drought)
    shock_labels = ['Drought', 'Heat/Cold Waves', 'Flooding', 'Climate Variability',
                    'Typhoons/Hurricanes', 'Wildfires', 'Severe Thunderstorms',
                    'Land Degradation', 'Tornadoes', 'Landslides']
    shock_rows = [64, 66, 67, 68, 69, 70, 71, 72, 73, 74]
    w['shocks'] = pd.DataFrame({
        'Shock': shock_labels,
        'Baseline': [_val(raw, r, 2) for r in shock_rows],
        'Midline': [_val(raw, r, 3) for r in shock_rows],
    })
    # Shock impact extent (rows 64-69, cols 10-12)
    w['shock_impact'] = pd.DataFrame({
        'Extent': ['Very Large', 'Large', 'Moderate', 'Small', 'Very Small', 'Not at All'],
        'Baseline': [_val(raw, r, 11) for r in range(64, 70)],
        'Midline': [_val(raw, r, 12) for r in range(64, 70)],
    })
    # Coping strategies (rows 78-90, cols 1-3)
    coping_labels = ['Skipped Meals', 'Took a Loan', 'Sold Small Livestock',
                     'Engaged in Wage Labour', 'Used Savings',
                     'Replaced Food with Less Preferred', 'Additional Livelihood Activities',
                     'Sold Large Livestock', 'Sold Household Assets', 'No Strategies',
                     'Informal Transfers', 'Formal Transfers (Govt/NGO)', 'Reduced Caloric Intake']
    w['coping'] = pd.DataFrame({
        'Strategy': coping_labels,
        'Baseline': [_val(raw, r, 2) for r in range(78, 91)],
        'Midline': [_val(raw, r, 3) for r in range(78, 91)],
    })

    # ---- C. ACCESS TO PREPAREDNESS INFO ----
    # Knowledge of plans (rows 98-101, cols 1-3)
    w['prep_knowledge'] = pd.DataFrame({
        'Response': ['Yes', 'No', "I Don't Know", 'Not Applicable'],
        'Baseline': [_val(raw, r, 2) for r in range(98, 102)],
        'Midline': [_val(raw, r, 3) for r in range(98, 102)],
    })
    # Participation in plan development (rows 98-101, cols 10-12)
    w['prep_participation'] = pd.DataFrame({
        'Response': ['Yes, I did', 'Yes, HH member', 'No', 'Not Applicable'],
        'Baseline': [_val(raw, 98, 11), _val(raw, 99, 11), _val(raw, 100, 11), _val(raw, 101, 11)],
        'Midline': [_val(raw, 98, 12), _val(raw, 99, 12), _val(raw, 100, 12), _val(raw, 101, 12)],
    })
    # Awareness of plan actions (rows 104-110, cols 1-3)
    w['prep_awareness'] = pd.DataFrame({
        'Extent': ['Not at All', 'Large Extent', 'Moderate Extent',
                   'Small Extent', 'Very Large Extent', 'Very Small Extent', 'Not Applicable'],
        'Baseline': [_val(raw, r, 2) for r in range(104, 111)],
        'Midline': [_val(raw, r, 3) for r in range(104, 111)],
    })
    # Know what to do in disaster (rows 105-107, cols 10-12)
    w['prep_know_action'] = pd.DataFrame({
        'Response': ['Yes', 'Partially', 'No'],
        'Baseline': [_val(raw, 105, 11), _val(raw, 106, 11), _val(raw, 107, 11)],
        'Midline': [_val(raw, 105, 12), _val(raw, 106, 12), _val(raw, 107, 12)],
    })
    # Weather forecast access (rows 114-116, cols 1-3)
    w['weather_forecast'] = pd.DataFrame({
        'Response': ['No', 'Yes, I do', 'Yes, HH member'],
        'Baseline': [_val(raw, 114, 2), _val(raw, 115, 2), _val(raw, 116, 2)],
        'Midline': [_val(raw, 114, 3), _val(raw, 115, 3), _val(raw, 116, 3)],
    })
    # Tidal forecast (rows 114-116, cols 10-12)
    w['tidal_forecast'] = pd.DataFrame({
        'Response': ['No', 'Yes, I do', 'Yes, HH member'],
        'Baseline': [_val(raw, 114, 11), _val(raw, 115, 11), _val(raw, 116, 11)],
        'Midline': [_val(raw, 114, 12), _val(raw, 115, 12), _val(raw, 116, 12)],
    })
    # Early warning (rows 120-122, cols 1-3)
    w['early_warning'] = pd.DataFrame({
        'Response': ['No', 'Yes, I do', 'Yes, HH member'],
        'Baseline': [_val(raw, 120, 2), _val(raw, 121, 2), _val(raw, 122, 2)],
        'Midline': [_val(raw, 120, 3), _val(raw, 121, 3), _val(raw, 122, 3)],
    })

    # ---- D. ACCESS & CONTROL OVER RESOURCES ----
    # Assets in HH (rows 129-148, cols 1-3)
    asset_names_raw = [_val(raw, r, 1) for r in range(129, 149)]
    asset_rows_valid = [(r, _clean_label(asset_names_raw[r-129]))
                        for r in range(129, 149) if asset_names_raw[r-129] != 0]
    w['hh_assets'] = pd.DataFrame({
        'Asset': [a[1] for a in asset_rows_valid],
        'Baseline': [_val(raw, a[0], 2) for a in asset_rows_valid],
        'Midline': [_val(raw, a[0], 3) for a in asset_rows_valid],
    })
    # Asset ownership — jointly (rows 130-149, cols 10-12)
    own_names_raw = [_val(raw, r, 10) for r in range(130, 149)]
    own_rows_valid = [(r, _clean_label(own_names_raw[r-130]))
                      for r in range(130, 149) if own_names_raw[r-130] != 0]
    w['asset_ownership'] = pd.DataFrame({
        'Asset': [a[1] for a in own_rows_valid],
        'Joint_BL': [_val(raw, a[0], 11) for a in own_rows_valid],
        'Joint_ML': [_val(raw, a[0], 12) for a in own_rows_valid],
        'Sole_BL': [_val(raw, a[0], 13) for a in own_rows_valid],
        'Sole_ML': [_val(raw, a[0], 14) for a in own_rows_valid],
        'All_BL': [_val(raw, a[0], 15) for a in own_rows_valid],
        'All_ML': [_val(raw, a[0], 16) for a in own_rows_valid],
    })
    # Use/Sell assets (rows 154-173, cols 1-5)
    usesell_names = [_val(raw, r, 1) for r in range(154, 174)]
    usesell_valid = [(r, _clean_label(usesell_names[r-154]))
                     for r in range(154, 174) if usesell_names[r-154] != 0]
    w['asset_use_sell'] = pd.DataFrame({
        'Asset': [a[1] for a in usesell_valid],
        'Use_BL': [_val(raw, a[0], 2) for a in usesell_valid],
        'Use_ML': [_val(raw, a[0], 3) for a in usesell_valid],
        'Sell_BL': [_val(raw, a[0], 4) for a in usesell_valid],
        'Sell_ML': [_val(raw, a[0], 5) for a in usesell_valid],
    })
    # Inputs (rows 154-161, cols 10-14)
    input_names = [_val(raw, r, 10) for r in range(154, 162)]
    input_valid = [(r, _clean_label(input_names[r-154]))
                   for r in range(154, 162) if input_names[r-154] != 0]
    w['inputs'] = pd.DataFrame({
        'Input': [a[1] for a in input_valid],
        'Access_BL': [_val(raw, a[0], 11) for a in input_valid],
        'Access_ML': [_val(raw, a[0], 12) for a in input_valid],
        'PersonalUse_BL': [_val(raw, a[0], 13) for a in input_valid],
        'PersonalUse_ML': [_val(raw, a[0], 14) for a in input_valid],
    })
    # Land size (rows 177-182, cols 1-3)
    w['land_size'] = pd.DataFrame({
        'Category': ['0 ha', '1 ha', '2 ha', '3-5 ha', '6-10 ha', '11+ ha'],
        'Baseline': [_val(raw, r, 2) for r in range(177, 183)],
        'Midline': [_val(raw, r, 3) for r in range(177, 183)],
    })
    # Land status (rows 177-179, cols 10-12)
    w['land_status'] = pd.DataFrame({
        'Category': ['Owned', 'Leased', 'Both'],
        'Baseline': [_val(raw, 177, 11), _val(raw, 178, 11), _val(raw, 179, 11)],
        'Midline': [_val(raw, 177, 12), _val(raw, 178, 12), _val(raw, 179, 12)],
    })
    # Land use (rows 186-187, cols 1-3)
    w['land_use'] = pd.DataFrame({
        'Category': ['Yes', 'No'],
        'Baseline': [_val(raw, 186, 2), _val(raw, 187, 2)],
        'Midline': [_val(raw, 186, 3), _val(raw, 187, 3)],
    })

    # ---- E. SAVINGS AND LOANS ----
    # Family saving (rows 195-197, cols 1-3)
    w['family_saving'] = pd.DataFrame({
        'Response': ['Yes', 'No', "I Don't Know"],
        'Baseline': [_val(raw, 195, 2), _val(raw, 196, 2), _val(raw, 197, 2)],
        'Midline': [_val(raw, 195, 3), _val(raw, 196, 3), _val(raw, 197, 3)],
    })
    # Personal saving (rows 195-196, cols 10-12)
    w['personal_saving'] = pd.DataFrame({
        'Response': ['Yes', 'No'],
        'Baseline': [_val(raw, 195, 11), _val(raw, 196, 11)],
        'Midline': [_val(raw, 195, 12), _val(raw, 196, 12)],
    })
    # Saving frequency (rows 201-205, cols 1-3)
    w['saving_freq'] = pd.DataFrame({
        'Frequency': ['Daily', 'Weekly', 'Monthly', 'Every 3+ Months', 'Varies'],
        'Baseline': [_val(raw, r, 2) for r in range(201, 206)],
        'Midline': [_val(raw, r, 3) for r in range(201, 206)],
    })
    # Saving amount (rows 201-204, cols 10-12)
    w['saving_amount'] = pd.DataFrame({
        'Amount': ['0-100 KES', '101-500 KES', '501-1000 KES', 'Above 1000 KES'],
        'Baseline': [_val(raw, r, 11) for r in range(201, 205)],
        'Midline': [_val(raw, r, 12) for r in range(201, 205)],
    })
    # Saving mechanism (rows 209-214, cols 1-3)
    w['saving_mechanism'] = pd.DataFrame({
        'Mechanism': ['At Home', 'Bank', 'Informal Lender', 'Informal Savings Group', 'MPESA', 'VSLA'],
        'Baseline': [_val(raw, r, 2) for r in range(209, 215)],
        'Midline': [_val(raw, r, 3) for r in range(209, 215)],
    })
    # Saving balance (rows 209-212, cols 10-12)
    w['saving_balance'] = pd.DataFrame({
        'Balance': ['0-100 KES', '101-500 KES', '501-1000 KES', 'Above 1000 KES'],
        'Baseline': [_val(raw, r, 11) for r in range(209, 213)],
        'Midline': [_val(raw, r, 12) for r in range(209, 213)],
    })
    # Intended use (rows 218-224, cols 1-3)
    w['saving_use'] = pd.DataFrame({
        'Use': ['Pay Debts', 'Purchase Food', 'Education Children',
                'Small HH Assets', 'Large HH Assets', 'Agriculture/Fishing', 'Productive Business'],
        'Baseline': [_val(raw, r, 2) for r in range(218, 225)],
        'Midline': [_val(raw, r, 3) for r in range(218, 225)],
    })
    # Family loan (rows 218-220, cols 10-12)
    w['family_loan'] = pd.DataFrame({
        'Response': ['Yes', 'No', "I Don't Know"],
        'Baseline': [_val(raw, 218, 11), _val(raw, 219, 11), _val(raw, 220, 11)],
        'Midline': [_val(raw, 218, 12), _val(raw, 219, 12), _val(raw, 220, 12)],
    })
    # Average loan size (family) — row 222, col 10-12
    w['family_loan_size'] = pd.DataFrame({
        'Metric': ['Average Loan Size (KES)'],
        'Baseline': [_val(raw, 222, 11)],
        'Midline': [_val(raw, 222, 12)],
    })
    # Personal loan (rows 228-229, cols 1-3)
    w['personal_loan'] = pd.DataFrame({
        'Response': ['Yes', 'No'],
        'Baseline': [_val(raw, 228, 2), _val(raw, 229, 2)],
        'Midline': [_val(raw, 228, 3), _val(raw, 229, 3)],
    })
    # Personal loan size — row 231
    w['personal_loan_size'] = pd.DataFrame({
        'Metric': ['Average Loan Size (KES)'],
        'Baseline': [_val(raw, 231, 2)],
        'Midline': [_val(raw, 231, 3)],
    })
    # Loan source (rows 228-232, cols 10-12)
    w['loan_source'] = pd.DataFrame({
        'Source': ['Bank', 'Informal Lender', 'Informal Credit/Savings Group', 'MPESA', 'VSLA'],
        'Baseline': [_val(raw, r, 11) for r in range(228, 233)],
        'Midline': [_val(raw, r, 12) for r in range(228, 233)],
    })

    # ---- F. ROLES, RESPONSIBILITIES & TIME USE ----
    # Who SHOULD do (norms) — rows 238-247, cols 1-3 (joint %)
    role_labels = ['Fetch Firewood', 'Income Provider', 'Clean/Sweep', 'Look After Livestock',
                   'Care for Ill/Elderly', 'Maintain Garden', 'Cook Food',
                   'Care for Children', 'Fetch Water', 'Wash Clothes']
    w['roles_should_joint'] = pd.DataFrame({
        'Role': role_labels,
        'Baseline': [_val(raw, r, 2) for r in range(238, 248)],
        'Midline': [_val(raw, r, 3) for r in range(238, 248)],
    })
    # Who SHOULD do — women only (rows 239-248, cols 10-12)
    w['roles_should_women'] = pd.DataFrame({
        'Role': role_labels,
        'Baseline': [_val(raw, r, 11) for r in range(239, 249)],
        'Midline': [_val(raw, r, 12) for r in range(239, 249)],
    })
    # Who DOES — Joint (rows 252-261, cols 1-3)
    w['roles_does_joint'] = pd.DataFrame({
        'Role': role_labels,
        'Baseline': [_val(raw, r, 2) for r in range(252, 262)],
        'Midline': [_val(raw, r, 3) for r in range(252, 262)],
    })
    # Who DOES — Self/Spouse only (rows 253-262, cols 10-12)
    w['roles_does_self'] = pd.DataFrame({
        'Role': role_labels,
        'Baseline': [_val(raw, r, 11) for r in range(253, 263)],
        'Midline': [_val(raw, r, 12) for r in range(253, 263)],
    })

    # Time use (rows 267-300, cols 0-3 for hours, cols 9-12 for participation %)
    # Unpaid care work
    unpaid_labels = ['Child/Elderly Care', 'Cooking/Dishes', 'Cleaning/Washing', 'Fetching Water', 'Fetching Firewood']
    w['time_unpaid'] = pd.DataFrame({
        'Activity': unpaid_labels,
        'Hours_BL': [_val(raw, r, 2) for r in range(267, 272)],
        'Hours_ML': [_val(raw, r, 3) for r in range(267, 272)],
        'Pct_BL': [_val(raw, r, 11) for r in range(267, 272)],
        'Pct_ML': [_val(raw, r, 12) for r in range(267, 272)],
    })
    # Total unpaid care
    w['time_unpaid_total'] = pd.DataFrame({
        'Category': ['Unpaid Care Work'],
        'Baseline': [_val(raw, 273, 2)], 'Midline': [_val(raw, 273, 3)]})

    # Productive work
    prod_labels = ['Farm Work (Others)', 'Farm Work (Own)', 'Fishing', 'Seaweed Farming',
                   'Tending Livestock', 'Selling at Market', 'Formal Employment']
    w['time_productive'] = pd.DataFrame({
        'Activity': prod_labels,
        'Hours_BL': [_val(raw, r, 2) for r in range(275, 282)],
        'Hours_ML': [_val(raw, r, 3) for r in range(275, 282)],
    })
    w['time_productive_total'] = pd.DataFrame({
        'Category': ['Productive Work'],
        'Baseline': [_val(raw, 283, 2)], 'Midline': [_val(raw, 283, 3)]})

    # Community conservation
    conservation_labels = ['Mangrove Restoration', 'Seaweed Farming', 'Forest Mgmt & Conservation']
    w['time_conservation'] = pd.DataFrame({
        'Activity': conservation_labels,
        'Hours_BL': [_val(raw, r, 2) for r in range(285, 288)],
        'Hours_ML': [_val(raw, r, 3) for r in range(285, 288)],
    })
    w['time_conservation_total'] = pd.DataFrame({
        'Category': ['Community Conservation'],
        'Baseline': [_val(raw, 289, 2)], 'Midline': [_val(raw, 289, 3)]})

    # Time summary
    w['time_summary'] = pd.DataFrame({
        'Category': ['Unpaid Care Work', 'Productive Work', 'Community Conservation',
                     'Personal Development', 'Personal Care', 'Leisure', 'Other'],
        'Baseline': [_val(raw, 273, 2), _val(raw, 283, 2), _val(raw, 289, 2),
                     _val(raw, 291, 2), _val(raw, 296, 2), _val(raw, 298, 2), _val(raw, 300, 2)],
        'Midline': [_val(raw, 273, 3), _val(raw, 283, 3), _val(raw, 289, 3),
                    _val(raw, 291, 3), _val(raw, 296, 3), _val(raw, 298, 3), _val(raw, 300, 3)],
    })

    # ---- G. DECISION MAKING ----
    decision_labels = ['Routine HH Purchases', 'Women Work on Farm', 'Use of HH Income',
                       'Women Go to Market', 'Mangrove/Seaweed/Forest Work',
                       'Using HH Savings', 'Invest Borrowed/Saved Money',
                       'Children Education', 'Large HH Purchases',
                       'Small Business', 'Taking Out Loans']
    # Should be Joint (rows 308-318, cols 1-3)
    w['decision_should_joint'] = pd.DataFrame({
        'Decision': decision_labels,
        'Baseline': [_val(raw, r, 2) for r in range(308, 319)],
        'Midline': [_val(raw, r, 3) for r in range(308, 319)],
    })
    # Should be — Men/Women only (rows 309-319, cols 10-12)
    w['decision_should_women'] = pd.DataFrame({
        'Decision': decision_labels,
        'Baseline': [_val(raw, r, 11) for r in range(309, 320)],
        'Midline': [_val(raw, r, 12) for r in range(309, 320)],
    })
    # Does — Joint (rows 323-333, cols 1-3)
    w['decision_does_joint'] = pd.DataFrame({
        'Decision': decision_labels,
        'Baseline': [_val(raw, r, 2) for r in range(323, 334)],
        'Midline': [_val(raw, r, 3) for r in range(323, 334)],
    })
    # Does — Self/Spouse (rows 324-334, cols 10-12)
    w['decision_does_self'] = pd.DataFrame({
        'Decision': decision_labels,
        'Baseline': [_val(raw, r, 11) for r in range(324, 335)],
        'Midline': [_val(raw, r, 12) for r in range(324, 335)],
    })
    # Can influence to a large extent (rows 338-348, cols 1-3)
    w['decision_influence'] = pd.DataFrame({
        'Decision': decision_labels,
        'Baseline': [_val(raw, r, 2) for r in range(338, 349)],
        'Midline': [_val(raw, r, 3) for r in range(338, 349)],
    })

    # ---- H. CLIMATE CHANGE & NBS ----
    # Heard of CC (rows 356-357, cols 1-3)
    w['cc_heard'] = pd.DataFrame({
        'Response': ['Yes', 'No'],
        'Baseline': [_val(raw, 356, 2), _val(raw, 357, 2)],
        'Midline': [_val(raw, 356, 3), _val(raw, 357, 3)],
    })
    # Adequately define CC (rows 361-364, cols 1-3)
    w['cc_define'] = pd.DataFrame({
        'Response': ['Completely Acceptable', 'Somewhat Acceptable', 'Not Acceptable', "Doesn't Know"],
        'Baseline': [_val(raw, r, 2) for r in range(361, 365)],
        'Midline': [_val(raw, r, 3) for r in range(361, 365)],
    })
    # CC effects on environment (rows 368-379, cols 1-3)
    env_effect_labels = ['Hotter Temperatures', 'Extreme Weather Events', 'Increased Drought',
                         'Warming Oceans', 'Sea Level Rise', 'Loss of Species',
                         'Migration of Species', 'Animal Migration Changes',
                         'Spread of Disease', 'Ocean Acidification',
                         'Increased Invasive Species', "I Don't Know"]
    w['cc_env_effects'] = pd.DataFrame({
        'Effect': env_effect_labels,
        'Baseline': [_val(raw, r, 2) for r in range(368, 380)],
        'Midline': [_val(raw, r, 3) for r in range(368, 380)],
    })
    # CC effects on livelihoods (rows 356-367, cols 10-12)
    livelihood_labels = ['Increased Food Insecurity', 'Increased Water Insecurity',
                         'Personal Risk from Extremes', 'Forced Migration',
                         'Changes in Livelihood', 'Reduced Livelihood Activities',
                         'Extreme Weather on Infrastructure', 'Reduced Ag Productivity',
                         'Crop Loss', 'Loss of Livestock', 'Reduced Livestock Productivity',
                         'Loss of Savings/Assets', "I Don't Know"]
    w['cc_livelihood_effects'] = pd.DataFrame({
        'Effect': livelihood_labels,
        'Baseline': [_val(raw, r, 11) for r in range(356, 369)],
        'Midline': [_val(raw, r, 12) for r in range(356, 369)],
    })
    # Heard of NbS (rows 383-384, cols 1-3)
    w['nbs_heard'] = pd.DataFrame({
        'Response': ['Yes', 'No'],
        'Baseline': [_val(raw, 383, 2), _val(raw, 384, 2)],
        'Midline': [_val(raw, 383, 3), _val(raw, 384, 3)],
    })
    # Define NbS (rows 372-375, cols 10-12)
    w['nbs_define'] = pd.DataFrame({
        'Response': ['Completely Acceptable', 'Somewhat Acceptable', 'Not Acceptable', "Doesn't Know"],
        'Baseline': [_val(raw, r, 11) for r in range(372, 376)],
        'Midline': [_val(raw, r, 12) for r in range(372, 376)],
    })
    # NbS examples (rows 388-396, cols 1-3)
    nbs_example_labels = ['Mangrove Restoration', 'Coral Reef Restoration', 'Reforestation',
                          'Sustainable Seaweed Farming', "I Don't Know",
                          'Rainwater Harvesting', 'Sustainable Agriculture', 'Solar Energy',
                          'Rainwater Harvesting 2']
    w['nbs_examples'] = pd.DataFrame({
        'Example': nbs_example_labels,
        'Baseline': [_val(raw, r, 2) for r in range(388, 397)],
        'Midline': [_val(raw, r, 3) for r in range(388, 397)],
    })
    # NbS benefits (rows 379-390, cols 10-12)
    nbs_benefit_labels = ['Carbon Capture', 'Reduce Storm Surges', 'Defense Against Salination',
                          'Improve Air Quality', 'Water Conservation', 'Soil Health',
                          'Support Pollinators', 'Local Biodiversity', 'Fish Diversity',
                          'Seawater Quality', 'Improved Livelihoods', "I Don't Know"]
    w['nbs_benefits'] = pd.DataFrame({
        'Benefit': nbs_benefit_labels,
        'Baseline': [_val(raw, r, 11) for r in range(379, 391)],
        'Midline': [_val(raw, r, 12) for r in range(379, 391)],
    })

    # ---- H1. MANGROVE RESTORATION ----
    w['mangrove_heard'] = pd.DataFrame({
        'Response': ['Heard & Importance', 'Heard Only', 'Not Heard'],
        'Baseline': [_val(raw, 403, 2), _val(raw, 404, 2), _val(raw, 405, 2)],
        'Midline': [_val(raw, 403, 3), _val(raw, 404, 3), _val(raw, 405, 3)],
    })
    w['mangrove_ever'] = pd.DataFrame({
        'Response': ['Yes', 'No'],
        'Baseline': [_val(raw, 403, 11), _val(raw, 404, 11)],
        'Midline': [_val(raw, 403, 12), _val(raw, 404, 12)],
    })
    w['mangrove_current'] = pd.DataFrame({
        'Response': ['Yes', 'No'],
        'Baseline': [_val(raw, 410, 2), _val(raw, 411, 2)],
        'Midline': [_val(raw, 410, 3), _val(raw, 411, 3)],
    })
    w['mangrove_modality'] = pd.DataFrame({
        'Modality': ['Individually', 'As a Group', 'As a Family'],
        'Baseline': [_val(raw, 408, 11), _val(raw, 409, 11), _val(raw, 410, 11)],
        'Midline': [_val(raw, 408, 12), _val(raw, 409, 12), _val(raw, 410, 12)],
    })
    w['mangrove_training'] = pd.DataFrame({
        'Response': ['Yes', 'Partially', 'No'],
        'Baseline': [_val(raw, 415, 2), _val(raw, 416, 2), _val(raw, 417, 2)],
        'Midline': [_val(raw, 415, 3), _val(raw, 416, 3), _val(raw, 417, 3)],
    })
    w['mangrove_assets'] = pd.DataFrame({
        'Response': ['Yes', 'Partially', 'No'],
        'Baseline': [_val(raw, 414, 11), _val(raw, 415, 11), _val(raw, 416, 11)],
        'Midline': [_val(raw, 414, 12), _val(raw, 415, 12), _val(raw, 416, 12)],
    })
    w['mangrove_interest'] = pd.DataFrame({
        'Response': ['Yes, Definitely', 'Yes, Maybe', 'No, Not Really'],
        'Baseline': [_val(raw, 421, 2), _val(raw, 422, 2), _val(raw, 423, 2)],
        'Midline': [_val(raw, 421, 3), _val(raw, 422, 3), _val(raw, 423, 3)],
    })

    # ---- H2. SEAWEED FARMING ----
    w['seaweed_heard'] = pd.DataFrame({
        'Response': ['Heard & Importance', 'Heard Only', 'Not Heard'],
        'Baseline': [_val(raw, 430, 2), _val(raw, 431, 2), _val(raw, 432, 2)],
        'Midline': [_val(raw, 430, 3), _val(raw, 431, 3), _val(raw, 432, 3)],
    })
    w['seaweed_ever'] = pd.DataFrame({
        'Response': ['Yes', "Don't Know", 'No'],
        'Baseline': [_val(raw, 430, 11), _val(raw, 431, 11), _val(raw, 432, 11)],
        'Midline': [_val(raw, 430, 12), _val(raw, 431, 12), _val(raw, 432, 12)],
    })
    w['seaweed_current'] = pd.DataFrame({
        'Response': ['Yes', 'No'],
        'Baseline': [_val(raw, 436, 2), _val(raw, 437, 2)],
        'Midline': [_val(raw, 436, 3), _val(raw, 437, 3)],
    })
    w['seaweed_modality'] = pd.DataFrame({
        'Modality': ['Individually', 'As a Group', 'As a Family'],
        'Baseline': [_val(raw, 436, 11), _val(raw, 437, 11), _val(raw, 438, 11)],
        'Midline': [_val(raw, 436, 12), _val(raw, 437, 12), _val(raw, 438, 12)],
    })
    w['seaweed_training'] = pd.DataFrame({
        'Response': ['Yes', 'Partially', 'No'],
        'Baseline': [_val(raw, 441, 2), _val(raw, 442, 2), _val(raw, 443, 2)],
        'Midline': [_val(raw, 441, 3), _val(raw, 442, 3), _val(raw, 443, 3)],
    })
    w['seaweed_assets'] = pd.DataFrame({
        'Response': ['Yes', 'No', 'Partially'],
        'Baseline': [_val(raw, 442, 11), _val(raw, 443, 11), _val(raw, 444, 11)],
        'Midline': [_val(raw, 442, 12), _val(raw, 443, 12), _val(raw, 444, 12)],
    })
    w['seaweed_interest'] = pd.DataFrame({
        'Response': ['Yes, Definitely', 'Yes, Maybe', 'No, Not Really'],
        'Baseline': [_val(raw, 447, 2), _val(raw, 448, 2), _val(raw, 449, 2)],
        'Midline': [_val(raw, 447, 3), _val(raw, 448, 3), _val(raw, 449, 3)],
    })

    # ---- H3. FOREST MANAGEMENT ----
    w['forest_heard'] = pd.DataFrame({
        'Response': ['Heard & Importance', 'Not Heard'],
        'Baseline': [_val(raw, 456, 2), _val(raw, 457, 2)],
        'Midline': [_val(raw, 456, 3), _val(raw, 457, 3)],
    })
    w['forest_ever'] = pd.DataFrame({
        'Response': ['Yes', 'No'],
        'Baseline': [_val(raw, 456, 11), _val(raw, 457, 11)],
        'Midline': [_val(raw, 456, 12), _val(raw, 457, 12)],
    })
    w['forest_current'] = pd.DataFrame({
        'Response': ['Yes', 'No'],
        'Baseline': [_val(raw, 462, 2), _val(raw, 463, 2)],
        'Midline': [_val(raw, 462, 3), _val(raw, 463, 3)],
    })
    w['forest_modality'] = pd.DataFrame({
        'Modality': ['Individually', 'As a Group', 'As a Family'],
        'Baseline': [_val(raw, 462, 11), _val(raw, 463, 11), _val(raw, 464, 11)],
        'Midline': [_val(raw, 462, 12), _val(raw, 463, 12), _val(raw, 464, 12)],
    })
    w['forest_training'] = pd.DataFrame({
        'Response': ['Yes', 'Partially', 'No'],
        'Baseline': [_val(raw, 467, 2), _val(raw, 468, 2), _val(raw, 469, 2)],
        'Midline': [_val(raw, 467, 3), _val(raw, 468, 3), _val(raw, 469, 3)],
    })
    w['forest_assets'] = pd.DataFrame({
        'Response': ['Yes', 'Partially', 'No'],
        'Baseline': [_val(raw, 468, 11), _val(raw, 469, 11), _val(raw, 470, 11)],
        'Midline': [_val(raw, 468, 12), _val(raw, 469, 12), _val(raw, 470, 12)],
    })

    # ---- I. LIFE SKILLS ----
    # Agree or Strongly Agree (rows 476-499, cols 1-3)
    # Strongly Agree (rows 476-499, cols 10-12)
    ls_labels = [
        'I have many positive qualities',
        'Respected in community',
        'Life has meaning',
        'Think about future',
        'Specific goals',
        'Know what to do for goals',
        'Confident to achieve goals',
        'Ability to lead',
        'Bring people together',
        'People seek my advice',
        'Prioritize tasks',
        'Convince others to join cause',
    ]
    ls_domains = ['Self Esteem']*3 + ['Aspirations']*4 + ['Leadership']*5
    ls_rows = list(range(476, 488))
    w['lifeskills_agree'] = pd.DataFrame({
        'Domain': ls_domains, 'Statement': ls_labels,
        'Baseline': [_val(raw, r, 2) for r in ls_rows],
        'Midline': [_val(raw, r, 3) for r in ls_rows],
    })
    w['lifeskills_strong'] = pd.DataFrame({
        'Domain': ls_domains, 'Statement': ls_labels,
        'Baseline': [_val(raw, r, 11) for r in ls_rows],
        'Midline': [_val(raw, r, 12) for r in ls_rows],
    })
    # Communication & Conflict Resolution
    comm_labels = ['Express Opinion (Family)', 'Express Opinion (Community)',
                   'Express What I Think', 'Listen Carefully', 'Easy to Understand Ideas']
    cr_labels = ['Seek Advice/Opinions', 'Understand Other POV',
                 'Not Offended by Disagreement', 'Think Long-Term Before Deciding']
    comm_rows = list(range(491, 496))
    cr_rows = list(range(496, 500))
    w['communication_agree'] = pd.DataFrame({
        'Statement': comm_labels,
        'Baseline': [_val(raw, r, 2) for r in comm_rows],
        'Midline': [_val(raw, r, 3) for r in comm_rows],
    })
    w['communication_strong'] = pd.DataFrame({
        'Statement': comm_labels,
        'Baseline': [_val(raw, r, 11) for r in comm_rows],
        'Midline': [_val(raw, r, 12) for r in comm_rows],
    })
    w['conflict_agree'] = pd.DataFrame({
        'Statement': cr_labels,
        'Baseline': [_val(raw, r, 2) for r in cr_rows],
        'Midline': [_val(raw, r, 3) for r in cr_rows],
    })
    w['conflict_strong'] = pd.DataFrame({
        'Statement': cr_labels,
        'Baseline': [_val(raw, r, 11) for r in cr_rows],
        'Midline': [_val(raw, r, 12) for r in cr_rows],
    })

    # ---- J. SOCIAL NORMS ----
    sn_labels = ['Income → Husband Controls', 'Men Better Business Ideas',
                 "Men Earn, Women Look After Home", 'Inappropriate Dress → Her Fault',
                 'Cook & Clean → Good Marriage', 'Embarrassing for Men to Do Chores',
                 'Planting Crops (Family Food)', 'Restoring Ecosystems (Mangrove/Forest)',
                 'Only Men Drive Boats', 'Ok for Women to Express Emotions']
    sn_rows = list(range(506, 516))
    w['socialnorms_agree'] = pd.DataFrame({
        'Norm': sn_labels,
        'Baseline': [_val(raw, r, 2) for r in sn_rows],
        'Midline': [_val(raw, r, 3) for r in sn_rows],
    })
    w['socialnorms_strong'] = pd.DataFrame({
        'Norm': sn_labels,
        'Baseline': [_val(raw, r, 11) for r in sn_rows],
        'Midline': [_val(raw, r, 12) for r in sn_rows],
    })

    return w


# ============================================================================
# MEN SURVEY DATA LOADER
# ============================================================================

@st.cache_data
def load_men_data(filepath):
    """
    Parse the Men Survey Excel file.
    Sheet 'Results Men' — 263 rows × 19 columns, non-standard dual-table layout.
    Left table cols 1-3 (label / BL / ML), right table cols 11-13 (label / BL / ML).
    All proportions are 0-1 scale; charts multiply ×100 for display.
    """
    try:
        raw = pd.read_excel(filepath, sheet_name=MEN_SHEET, header=None)
    except FileNotFoundError:
        st.error(f"Men Survey Excel not found: {filepath}")
        st.stop()

    m = {}

    # ---- A. HOUSEHOLD CHARACTERISTICS ----
    # Location type (Excel R12-13, _val rows 11-12, BL=col2, ML=col3)
    m['location'] = pd.DataFrame({
        'Category': ['Marine', 'Terrestrial'],
        'Baseline': [_val(raw, 11, 2), _val(raw, 12, 2)],
        'Midline': [_val(raw, 11, 3), _val(raw, 12, 3)],
    })

    # Education level (Excel R12-15 right, _val rows 11-14, cols 11-12)
    m['education'] = pd.DataFrame({
        'Category': ['College Level or Higher', 'Pre-primary/None/Other',
                     'Primary', 'Secondary/Vocational'],
        'Baseline': [_val(raw, 11, 11), _val(raw, 12, 11),
                     _val(raw, 13, 11), _val(raw, 14, 11)],
        'Midline': [_val(raw, 11, 12), _val(raw, 12, 12),
                    _val(raw, 13, 12), _val(raw, 14, 12)],
    })

    # Main HH economic activity (Excel R17-28, _val rows 16-27, labels=col1, BL=col2, ML=col3)
    econ_labels = [_clean_label(_val(raw, r, 1)) for r in range(16, 28)
                   if _val(raw, r, 1) != 0]
    econ_bl = [_val(raw, r, 2) for r in range(16, 28) if _val(raw, r, 1) != 0]
    econ_ml = [_val(raw, r, 3) for r in range(16, 28) if _val(raw, r, 1) != 0]
    m['main_econ'] = pd.DataFrame({
        'Activity': econ_labels, 'Baseline': econ_bl, 'Midline': econ_ml})

    # ---- B. CLIMATE CHANGE & NBS KNOWLEDGE ----
    # CC heard (Excel R36-37, _val rows 35-36, col 2-3)
    m['cc_heard'] = pd.DataFrame({
        'Response': ['Yes', 'No'],
        'Baseline': [_val(raw, 35, 2), _val(raw, 36, 2)],
        'Midline': [_val(raw, 35, 3), _val(raw, 36, 3)],
    })

    # CC define adequacy (Excel R41-44, _val rows 40-43, col 2-3)
    m['cc_define'] = pd.DataFrame({
        'Response': ['Completely Acceptable', 'Somewhat Acceptable',
                     'Not Acceptable', "Doesn't Know/Answer"],
        'Baseline': [_val(raw, r, 2) for r in range(40, 44)],
        'Midline': [_val(raw, r, 3) for r in range(40, 44)],
    })

    # CC environmental effects (Excel R48-59, _val rows 47-58, col 2-3)
    cc_env_labels = ['Hotter Temperatures', 'More Extreme Weather Events',
                     'Increased Drought', 'Warming Oceans/Water Bodies',
                     'Sea Level Rise', 'Loss of Species', 'Migration of Species',
                     'Animal Migration Pattern Changes', 'Spread of Disease/Algal Blooms',
                     'Ocean Acidification (Coral Reefs)',
                     'Increased Species Invasions', "I Don't Know"]
    m['cc_env_effects'] = pd.DataFrame({
        'Effect': cc_env_labels,
        'Baseline': [_val(raw, r, 2) for r in range(47, 59)],
        'Midline': [_val(raw, r, 3) for r in range(47, 59)],
    })

    # CC livelihood effects (Excel R36-48 right, _val rows 35-47, cols 11-12)
    cc_live_labels = ['Increased Food Insecurity', 'Increased Water Insecurity',
                      'Personal Risk from Extremes', 'Forced Migration (Disaster)',
                      'Changes in Livelihood Activities', 'Reduced Livelihood Activities',
                      'Extreme Weather on Infrastructure', 'Reduced Agricultural Productivity',
                      'Crop Loss', 'Loss of Livestock', 'Reduced Livestock Productivity',
                      'Loss of Savings/Assets', "I Don't Know"]
    m['cc_livelihood_effects'] = pd.DataFrame({
        'Effect': cc_live_labels,
        'Baseline': [_val(raw, r, 11) for r in range(35, 48)],
        'Midline': [_val(raw, r, 12) for r in range(35, 48)],
    })

    # NbS heard (Excel R63-64, _val rows 62-63, col 2-3)
    m['nbs_heard'] = pd.DataFrame({
        'Response': ['Yes', 'No'],
        'Baseline': [_val(raw, 62, 2), _val(raw, 63, 2)],
        'Midline': [_val(raw, 62, 3), _val(raw, 63, 3)],
    })

    # NbS define (Excel R52-55 right, _val rows 51-54, cols 11-12)
    m['nbs_define'] = pd.DataFrame({
        'Response': ['Completely Acceptable', 'Somewhat Acceptable',
                     'Not Acceptable', "Doesn't Know/Answer"],
        'Baseline': [_val(raw, r, 11) for r in range(51, 55)],
        'Midline': [_val(raw, r, 12) for r in range(51, 55)],
    })

    # NbS examples cited (Excel R68-75, _val rows 67-74, col 2-3)
    nbs_ex_labels = ['Mangrove Restoration', 'Coral Reef Restoration',
                     'Reforestation/Prevention of Deforestation',
                     'Sustainable Seaweed Farming', "I Don't Know",
                     'Rainwater Harvesting', 'Sustainable Agriculture',
                     'Solar Energy']
    m['nbs_examples'] = pd.DataFrame({
        'Example': nbs_ex_labels,
        'Baseline': [_val(raw, r, 2) for r in range(67, 75)],
        'Midline': [_val(raw, r, 3) for r in range(67, 75)],
    })

    # NbS benefits (Excel R59-70 right, _val rows 58-69, cols 11-12)
    nbs_ben_labels = ['Carbon Capture', 'Reduce Storm Surges/Flooding',
                      'Defense Against Salination', 'Improve Air Quality',
                      'Water Conservation', 'Soil Health & Fertility',
                      'Support Pollinators/Biodiversity',
                      'Local Biodiversity/Ecosystem', 'Fish Diversity & Quantity',
                      'Seawater Quality & Clarity',
                      'Improved/Diversified Livelihoods', "I Don't Know"]
    m['nbs_benefits'] = pd.DataFrame({
        'Benefit': nbs_ben_labels,
        'Baseline': [_val(raw, r, 11) for r in range(58, 70)],
        'Midline': [_val(raw, r, 12) for r in range(58, 70)],
    })

    # ---- B1. MANGROVE RESTORATION ----
    # Heard of mangrove (Excel R81-83, _val rows 80-82, col 2-3)
    m['mangrove_heard'] = pd.DataFrame({
        'Response': ['Heard & Importance', 'Heard Only', 'Not Heard'],
        'Baseline': [_val(raw, 80, 2), _val(raw, 81, 2), _val(raw, 82, 2)],
        'Midline': [_val(raw, 80, 3), _val(raw, 81, 3), _val(raw, 82, 3)],
    })
    # Female HH ever participated (Excel R81-82 right, _val rows 80-81, cols 11-12)
    m['mangrove_ever'] = pd.DataFrame({
        'Response': ['Yes', 'No'],
        'Baseline': [_val(raw, 80, 11), _val(raw, 81, 11)],
        'Midline': [_val(raw, 80, 12), _val(raw, 81, 12)],
    })
    # Female HH currently involved (Excel R87-88, _val rows 86-87, col 2-3)
    m['mangrove_current'] = pd.DataFrame({
        'Response': ['Yes', 'No'],
        'Baseline': [_val(raw, 86, 2), _val(raw, 87, 2)],
        'Midline': [_val(raw, 86, 3), _val(raw, 87, 3)],
    })
    # Men supporting (Excel R86-88 right, _val rows 85-87, cols 11-12)
    m['mangrove_support'] = pd.DataFrame({
        'Response': ['Yes', 'No', 'Somehow'],
        'Baseline': [_val(raw, 85, 11), _val(raw, 86, 11), _val(raw, 87, 11)],
        'Midline': [_val(raw, 85, 12), _val(raw, 86, 12), _val(raw, 87, 12)],
    })
    # Type of support (Excel R92-98, _val rows 91-97, col 2-3)
    support_labels = ['Encouraged Participation', 'Sought Community Support',
                      'Supported with HH Chores', 'Supported with Restoration Work',
                      'Supported with Materials Purchase', 'None', 'Other']
    m['mangrove_support_type'] = pd.DataFrame({
        'Type': support_labels,
        'Baseline': [_val(raw, r, 2) for r in range(91, 98)],
        'Midline': [_val(raw, r, 3) for r in range(91, 98)],
    })

    # ---- B2. SEAWEED FARMING ----
    m['seaweed_heard'] = pd.DataFrame({
        'Response': ['Heard & Importance', 'Heard Only', 'Not Heard'],
        'Baseline': [_val(raw, 103, 2), _val(raw, 104, 2), _val(raw, 105, 2)],
        'Midline': [_val(raw, 103, 3), _val(raw, 104, 3), _val(raw, 105, 3)],
    })
    m['seaweed_ever'] = pd.DataFrame({
        'Response': ['Yes', 'No'],
        'Baseline': [_val(raw, 103, 11), _val(raw, 104, 11)],
        'Midline': [_val(raw, 103, 12), _val(raw, 104, 12)],
    })
    m['seaweed_current'] = pd.DataFrame({
        'Response': ['Yes', 'No'],
        'Baseline': [_val(raw, 109, 2), _val(raw, 110, 2)],
        'Midline': [_val(raw, 109, 3), _val(raw, 110, 3)],
    })
    m['seaweed_support'] = pd.DataFrame({
        'Response': ['Yes', 'No', 'Somehow'],
        'Baseline': [_val(raw, 108, 11), _val(raw, 109, 11), _val(raw, 110, 11)],
        'Midline': [_val(raw, 108, 12), _val(raw, 109, 12), _val(raw, 110, 12)],
    })
    m['seaweed_support_type'] = pd.DataFrame({
        'Type': support_labels,
        'Baseline': [_val(raw, r, 2) for r in range(114, 121)],
        'Midline': [_val(raw, r, 3) for r in range(114, 121)],
    })

    # ---- B3. FOREST MANAGEMENT ----
    m['forest_heard'] = pd.DataFrame({
        'Response': ['Yes', 'Not Heard'],
        'Baseline': [_val(raw, 126, 2), _val(raw, 127, 2)],
        'Midline': [_val(raw, 126, 3), _val(raw, 127, 3)],
    })
    m['forest_ever'] = pd.DataFrame({
        'Response': ['Yes', 'No'],
        'Baseline': [_val(raw, 126, 11), _val(raw, 127, 11)],
        'Midline': [_val(raw, 126, 12), _val(raw, 127, 12)],
    })
    m['forest_current'] = pd.DataFrame({
        'Response': ['Yes', 'No'],
        'Baseline': [_val(raw, 132, 2), _val(raw, 133, 2)],
        'Midline': [_val(raw, 132, 3), _val(raw, 133, 3)],
    })
    m['forest_support'] = pd.DataFrame({
        'Response': ['Yes', 'No', 'Somehow'],
        'Baseline': [_val(raw, 132, 11), _val(raw, 133, 11), _val(raw, 134, 11)],
        'Midline': [_val(raw, 132, 12), _val(raw, 133, 12), _val(raw, 134, 12)],
    })
    m['forest_support_type'] = pd.DataFrame({
        'Type': support_labels,
        'Baseline': [_val(raw, r, 2) for r in range(137, 144)],
        'Midline': [_val(raw, r, 3) for r in range(137, 144)],
    })

    # ---- C. ROLES, RESPONSIBILITIES & TIME POVERTY ----
    role_labels = ['Cook Food', 'Clean/Sweep House', 'Look After Livestock',
                   'Maintain Garden', 'Care for Ill/Elderly/Disabled',
                   'Fetch Firewood', 'Fetch Water', 'Wash Clothes',
                   'Main Income Provider', 'Care for Children']

    # Roles SHOULD — Joint % (Excel R152-161, _val rows 151-160, col 2-3)
    m['roles_should_joint'] = pd.DataFrame({
        'Role': role_labels,
        'Baseline': [_val(raw, r, 2) for r in range(151, 161)],
        'Midline': [_val(raw, r, 3) for r in range(151, 161)],
    })
    # Roles SHOULD — Women only (Excel R152-161 right, _val rows 151-160, cols 11-12)
    m['roles_should_women'] = pd.DataFrame({
        'Role': role_labels,
        'Baseline': [_val(raw, r, 11) for r in range(151, 161)],
        'Midline': [_val(raw, r, 12) for r in range(151, 161)],
    })
    # Roles SHOULD — Men only (Excel R152-161 right, _val rows 151-160, cols 13-14)
    m['roles_should_men'] = pd.DataFrame({
        'Role': role_labels,
        'Baseline': [_val(raw, r, 13) for r in range(151, 161)],
        'Midline': [_val(raw, r, 14) for r in range(151, 161)],
    })

    # Roles DOES — Joint % (Excel R166-175, _val rows 165-174, col 2-3)
    m['roles_does_joint'] = pd.DataFrame({
        'Role': role_labels,
        'Baseline': [_val(raw, r, 2) for r in range(165, 175)],
        'Midline': [_val(raw, r, 3) for r in range(165, 175)],
    })
    # Roles DOES — Men(Myself) (Excel R166-175 right, _val rows 165-174, cols 11-12)
    m['roles_does_men'] = pd.DataFrame({
        'Role': role_labels,
        'Baseline': [_val(raw, r, 11) for r in range(165, 175)],
        'Midline': [_val(raw, r, 12) for r in range(165, 175)],
    })
    # Roles DOES — Women(MySpouse) (Excel R166-175 right, _val rows 165-174, cols 13-14)
    m['roles_does_women'] = pd.DataFrame({
        'Role': role_labels,
        'Baseline': [_val(raw, r, 13) for r in range(165, 175)],
        'Midline': [_val(raw, r, 14) for r in range(165, 175)],
    })

    # Time use (hours per day)
    # Unpaid care activities (Excel R178-182, _val rows 177-181, col 2-3)
    unpaid_labels = ['Child/Elderly/Disabled Care', 'Cooking/Dishes',
                     'Cleaning/Washing', 'Fetching Water', 'Fetching Firewood']
    m['time_unpaid'] = pd.DataFrame({
        'Activity': unpaid_labels,
        'Baseline': [_val(raw, r, 2) for r in range(177, 182)],
        'Midline': [_val(raw, r, 3) for r in range(177, 182)],
    })
    m['time_unpaid_total'] = pd.DataFrame({
        'Category': ['Unpaid Care Work'],
        'Baseline': [_val(raw, 183, 2)], 'Midline': [_val(raw, 183, 3)]})

    # Productive work (Excel R186-191, _val rows 185-190, col 2-3)
    prod_labels = ['Farm Work (Others)', 'Farm Work (Own)', 'Fishing',
                   'Seaweed Farming', 'Selling at Market', 'Formal Employment']
    m['time_productive'] = pd.DataFrame({
        'Activity': prod_labels,
        'Baseline': [_val(raw, r, 2) for r in range(185, 191)],
        'Midline': [_val(raw, r, 3) for r in range(185, 191)],
    })
    m['time_productive_total'] = pd.DataFrame({
        'Category': ['Productive Work'],
        'Baseline': [_val(raw, 192, 2)], 'Midline': [_val(raw, 192, 3)]})

    # Community conservation (Excel R195-197, _val rows 194-196, col 2-3)
    cons_labels = ['Mangrove Restoration', 'Seaweed Farming', 'Forest Mgmt & Conservation']
    m['time_conservation'] = pd.DataFrame({
        'Activity': cons_labels,
        'Baseline': [_val(raw, r, 2) for r in range(194, 197)],
        'Midline': [_val(raw, r, 3) for r in range(194, 197)],
    })
    m['time_conservation_total'] = pd.DataFrame({
        'Category': ['Community Conservation Work'],
        'Baseline': [_val(raw, 198, 2)], 'Midline': [_val(raw, 198, 3)]})

    # Time summary (all categories)
    m['time_summary'] = pd.DataFrame({
        'Category': ['Unpaid Care Work', 'Productive Work', 'Community Conservation',
                     'Personal Development', 'Personal Care', 'Leisure', 'Other'],
        'Baseline': [_val(raw, 183, 2), _val(raw, 192, 2), _val(raw, 198, 2),
                     _val(raw, 200, 2), _val(raw, 205, 2), _val(raw, 207, 2), _val(raw, 209, 2)],
        'Midline': [_val(raw, 183, 3), _val(raw, 192, 3), _val(raw, 198, 3),
                    _val(raw, 200, 3), _val(raw, 205, 3), _val(raw, 207, 3), _val(raw, 209, 3)],
    })

    # ---- D. DECISION-MAKING ----
    decision_labels = ['Start Small Business', 'Large HH Purchases',
                       'Using HH Savings', 'Taking Out Loans',
                       'Children Education', 'Routine HH Purchases',
                       'Women Go to Mangrove/Seaweed/Forest',
                       'Use of HH Income', 'Women Go to Market',
                       'Women Go to Farm', 'Invest Borrowed/Saved Money']

    # Decision SHOULD — Joint (Excel R218-228, _val rows 217-227, col 2-3)
    m['decision_should_joint'] = pd.DataFrame({
        'Decision': decision_labels,
        'Baseline': [_val(raw, r, 2) for r in range(217, 228)],
        'Midline': [_val(raw, r, 3) for r in range(217, 228)],
    })
    # Decision SHOULD — Women only (Excel R219-229 right, _val rows 218-228, cols 11-12)
    m['decision_should_women'] = pd.DataFrame({
        'Decision': decision_labels,
        'Baseline': [_val(raw, r, 11) for r in range(218, 229)],
        'Midline': [_val(raw, r, 12) for r in range(218, 229)],
    })
    # Decision SHOULD — Men only (Excel R219-229 right, _val rows 218-228, cols 13-14)
    m['decision_should_men'] = pd.DataFrame({
        'Decision': decision_labels,
        'Baseline': [_val(raw, r, 13) for r in range(218, 229)],
        'Midline': [_val(raw, r, 14) for r in range(218, 229)],
    })

    # Decision DOES — Joint (Excel R234-244, _val rows 233-243, col 2-3)
    m['decision_does_joint'] = pd.DataFrame({
        'Decision': decision_labels,
        'Baseline': [_val(raw, r, 2) for r in range(233, 244)],
        'Midline': [_val(raw, r, 3) for r in range(233, 244)],
    })
    # Decision DOES — Men(Myself) (Excel R234-244 right, _val rows 233-243, cols 11-12)
    m['decision_does_men'] = pd.DataFrame({
        'Decision': decision_labels,
        'Baseline': [_val(raw, r, 11) for r in range(233, 244)],
        'Midline': [_val(raw, r, 12) for r in range(233, 244)],
    })
    # Decision DOES — Women(MySpouse) (Excel R234-244 right, _val rows 233-243, cols 13-14)
    m['decision_does_women'] = pd.DataFrame({
        'Decision': decision_labels,
        'Baseline': [_val(raw, r, 13) for r in range(233, 244)],
        'Midline': [_val(raw, r, 14) for r in range(233, 244)],
    })

    # ---- E. SOCIAL NORMS ----
    sn_labels = [
        'Income -> Husband Controls', 'Men Better Business Ideas',
        "Men Earn, Women Look After Home", 'Inappropriate Dress -> Her Fault',
        'Cook & Clean -> Good Marriage', 'Embarrassing for Men to Do Chores',
        'Planting Crops (Family Food)', 'Restoring Ecosystems (Mangrove/Forest)',
        'Only Men Drive Boats', 'Ok for Women to Express Emotions',
        'Supportive of Diverse People', 'Women Gain Rights -> Men Lose',
        'Stronger Women -> Stronger Families',
    ]
    sn_rows = list(range(250, 263))

    # Agree or Strongly Agree (Excel R251-263, _val rows 250-262, col 2-3)
    m['socialnorms_agree'] = pd.DataFrame({
        'Norm': sn_labels,
        'Baseline': [_val(raw, r, 2) for r in sn_rows],
        'Midline': [_val(raw, r, 3) for r in sn_rows],
    })
    # Strongly Agree only (Excel R251-263 right, _val rows 250-262, cols 11-12)
    m['socialnorms_strong'] = pd.DataFrame({
        'Norm': sn_labels,
        'Baseline': [_val(raw, r, 11) for r in sn_rows],
        'Midline': [_val(raw, r, 12) for r in sn_rows],
    })

    return m


# ============================================================================
# GJJ KAP WOMEN (BASELINE / ENDLINE) DATA LOADER
# ============================================================================

@st.cache_data
def load_gjj_kap_women_data(filepath):
    """
    Parse the GJJ KAP Women Survey Excel file.
    Sheet 'Results KAP Women Endline' — 228 rows x 9 columns.
    All data lives in Column B (labels) + Columns C-H (values).
    Proportions are 0-1 scale; charts multiply x100 for display.
    Row references below are 1-based Excel rows; _val() uses 0-based.
    """
    try:
        raw = pd.read_excel(filepath, sheet_name=GJJ_KAP_WOMEN_SHEET, header=None)
    except FileNotFoundError:
        st.error(f"GJJ KAP Women Excel not found: {filepath}")
        st.stop()

    g = {}

    # ---- A. SELF — Self-Esteem, Self-Compassion, Confidence ----
    # 3 statements: R10-R12 (0-based 9-11), Baseline Likert: StronglyAgree(c3) Agree(c4) Disagree(c5) StronglyDisagree(c6)
    self_statements = [
        _clean_label(_val(raw, 9, 1)),   # "I have many strengths and qualities"
        _clean_label(_val(raw, 10, 1)),  # "I believe my feelings and opinions are important"
        _clean_label(_val(raw, 11, 1)),  # "I see myself as an equal..."
    ]

    g['self_baseline_likert'] = pd.DataFrame({
        'Statement': self_statements,
        'Strongly agree': [_val(raw, r, 2) for r in range(9, 12)],
        'Agree': [_val(raw, r, 3) for r in range(9, 12)],
        'Disagree': [_val(raw, r, 4) for r in range(9, 12)],
        'Strongly disagree': [_val(raw, r, 5) for r in range(9, 12)],
    })

    # Endline Likert: R24-R26 (0-based 23-25)
    g['self_endline_likert'] = pd.DataFrame({
        'Statement': self_statements,
        'Strongly agree': [_val(raw, r, 2) for r in range(23, 26)],
        'Agree': [_val(raw, r, 3) for r in range(23, 26)],
        'Disagree': [_val(raw, r, 4) for r in range(23, 26)],
        'Strongly disagree': [_val(raw, r, 5) for r in range(23, 26)],
    })

    # Strongly Agree comparison: R17-R19 (0-based 16-18), BL=c3, EL=c4
    g['self_strongly_agree'] = pd.DataFrame({
        'Statement': self_statements,
        'Baseline': [_val(raw, r, 2) for r in range(16, 19)],
        'Endline': [_val(raw, r, 3) for r in range(16, 19)],
    })

    # Agreement vs Disagreement: R31-R33 (0-based 30-32)
    g['self_agreement'] = pd.DataFrame({
        'Statement': self_statements,
        'Agreement_BL': [_val(raw, r, 2) for r in range(30, 33)],
        'Agreement_EL': [_val(raw, r, 3) for r in range(30, 33)],
        'Disagreement_BL': [_val(raw, r, 4) for r in range(30, 33)],
        'Disagreement_EL': [_val(raw, r, 5) for r in range(30, 33)],
    })

    # Self-compassion frequency: R38-R41 (0-based 37-40), BL=c3, EL=c4
    compassion_cats = [
        _clean_label(_val(raw, r, 1)) for r in range(37, 41)
    ]
    g['self_compassion'] = pd.DataFrame({
        'Category': compassion_cats,
        'Baseline': [_val(raw, r, 2) for r in range(37, 41)],
        'Endline': [_val(raw, r, 3) for r in range(37, 41)],
    })

    # ---- B. RELATIONAL WELLBEING ----
    rel_statements = [
        _clean_label(_val(raw, r, 1)) for r in range(47, 54)
    ]

    # Baseline frequency: R48-R54 (0-based 47-53), 6 cols: Always(c3) Frequently(c4) Sometimes(c5) Rarely(c6) Never(c7) NA(c8)
    g['rel_baseline_freq'] = pd.DataFrame({
        'Statement': rel_statements,
        'Always': [_val(raw, r, 2) for r in range(47, 54)],
        'Frequently': [_val(raw, r, 3) for r in range(47, 54)],
        'Sometimes': [_val(raw, r, 4) for r in range(47, 54)],
        'Rarely': [_val(raw, r, 5) for r in range(47, 54)],
        'Never': [_val(raw, r, 6) for r in range(47, 54)],
        'NA': [_val(raw, r, 7) for r in range(47, 54)],
    })

    # Endline frequency: R59-R65 (0-based 58-64)
    g['rel_endline_freq'] = pd.DataFrame({
        'Statement': rel_statements,
        'Always': [_val(raw, r, 2) for r in range(58, 65)],
        'Frequently': [_val(raw, r, 3) for r in range(58, 65)],
        'Sometimes': [_val(raw, r, 4) for r in range(58, 65)],
        'Rarely': [_val(raw, r, 5) for r in range(58, 65)],
        'Never': [_val(raw, r, 6) for r in range(58, 65)],
        'NA': [_val(raw, r, 7) for r in range(58, 65)],
    })

    # Always/Frequently vs Rarely/Never comparison: R70-R76 (0-based 69-75)
    # Cols: AF_BL(c3) AF_EL(c4) RN_BL(c5) RN_EL(c6) DIF_AF(c8) DIF_RN(c9)
    g['rel_af_rn'] = pd.DataFrame({
        'Statement': rel_statements,
        'AF_Baseline': [_val(raw, r, 2) for r in range(69, 76)],
        'AF_Endline': [_val(raw, r, 3) for r in range(69, 76)],
        'RN_Baseline': [_val(raw, r, 4) for r in range(69, 76)],
        'RN_Endline': [_val(raw, r, 5) for r in range(69, 76)],
        'DIF_AF': [_val(raw, r, 7) for r in range(69, 76)],
        'DIF_RN': [_val(raw, r, 8) for r in range(69, 76)],
    })

    # Relational agreement statements — Baseline: R81-R82 (0-based 80-81)
    rel_agree_stmts = [
        _clean_label(_val(raw, 80, 1)),
        _clean_label(_val(raw, 81, 1)),
    ]
    g['rel_agree_baseline'] = pd.DataFrame({
        'Statement': rel_agree_stmts,
        'Strongly agree': [_val(raw, r, 2) for r in range(80, 82)],
        'Agree': [_val(raw, r, 3) for r in range(80, 82)],
        'Disagree': [_val(raw, r, 4) for r in range(80, 82)],
        'Strongly disagree': [_val(raw, r, 5) for r in range(80, 82)],
    })

    # Endline: R87-R88 (0-based 86-87)
    g['rel_agree_endline'] = pd.DataFrame({
        'Statement': rel_agree_stmts,
        'Strongly agree': [_val(raw, r, 2) for r in range(86, 88)],
        'Agree': [_val(raw, r, 3) for r in range(86, 88)],
        'Disagree': [_val(raw, r, 4) for r in range(86, 88)],
        'Strongly disagree': [_val(raw, r, 5) for r in range(86, 88)],
    })

    # Agreement vs Disagreement: R93-R94 (0-based 92-93)
    rel_agree_summary_stmts = [
        _clean_label(_val(raw, 92, 1)),
        _clean_label(_val(raw, 93, 1)),
    ]
    g['rel_agreement_summary'] = pd.DataFrame({
        'Statement': rel_agree_summary_stmts,
        'Agreement_BL': [_val(raw, r, 2) for r in range(92, 94)],
        'Agreement_EL': [_val(raw, r, 3) for r in range(92, 94)],
        'Disagreement_BL': [_val(raw, r, 4) for r in range(92, 94)],
        'Disagreement_EL': [_val(raw, r, 5) for r in range(92, 94)],
    })

    # ---- C. GENDER TRANSFORMATION: SHARED RESPONSIBILITY ----
    # Husband supports household chores: R100-R101 (0-based 99-100), BL=c3, EL=c4
    g['shared_chores_yn'] = pd.DataFrame({
        'Response': [_clean_label(_val(raw, r, 1)) for r in range(99, 101)],
        'Baseline': [_val(raw, r, 2) for r in range(99, 101)],
        'Endline': [_val(raw, r, 3) for r in range(99, 101)],
    })

    # Time since husband supporting: R105-R111 (0-based 104-110)
    g['chore_duration'] = pd.DataFrame({
        'Category': [_clean_label(_val(raw, r, 1)) for r in range(104, 111)],
        'Baseline': [_val(raw, r, 2) for r in range(104, 111)],
        'Endline': [_val(raw, r, 3) for r in range(104, 111)],
    })

    # Frequency of husband chores: R115-R119 (0-based 114-118)
    g['chore_frequency'] = pd.DataFrame({
        'Category': [_clean_label(_val(raw, r, 1)) for r in range(114, 119)],
        'Baseline': [_val(raw, r, 2) for r in range(114, 119)],
        'Endline': [_val(raw, r, 3) for r in range(114, 119)],
    })

    # Discussed sharing unpaid chores: R123-R124 (0-based 122-123)
    g['chore_discussed'] = pd.DataFrame({
        'Response': [_clean_label(_val(raw, r, 1)) for r in range(122, 124)],
        'Baseline': [_val(raw, r, 2) for r in range(122, 124)],
        'Endline': [_val(raw, r, 3) for r in range(122, 124)],
    })

    # Person who started conversation: R128-R130 (0-based 127-129)
    g['chore_initiator'] = pd.DataFrame({
        'Person': [_clean_label(_val(raw, r, 1)) for r in range(127, 130)],
        'Baseline': [_val(raw, r, 2) for r in range(127, 130)],
        'Endline': [_val(raw, r, 3) for r in range(127, 130)],
    })

    # Husband completed significant chore: R134-R135 (0-based 133-134)
    g['chore_completed'] = pd.DataFrame({
        'Response': [_clean_label(_val(raw, r, 1)) for r in range(133, 135)],
        'Baseline': [_val(raw, r, 2) for r in range(133, 135)],
        'Endline': [_val(raw, r, 3) for r in range(133, 135)],
    })

    # Type of chore: R139-R143 (0-based 138-142)
    g['chore_type'] = pd.DataFrame({
        'Chore': [_clean_label(_val(raw, r, 1)) for r in range(138, 143)],
        'Baseline': [_val(raw, r, 2) for r in range(138, 143)],
        'Endline': [_val(raw, r, 3) for r in range(138, 143)],
    })

    # Hours saved: R147-R152 (0-based 146-151)
    g['hours_saved'] = pd.DataFrame({
        'Hours': [_clean_label(_val(raw, r, 1)) for r in range(146, 152)],
        'Baseline': [_val(raw, r, 2) for r in range(146, 152)],
        'Endline': [_val(raw, r, 3) for r in range(146, 152)],
    })

    # Frequency husband supports women's self-time: R156-R161 (0-based 155-160)
    g['support_self_time'] = pd.DataFrame({
        'Category': [_clean_label(_val(raw, r, 1)) for r in range(155, 161)],
        'Baseline': [_val(raw, r, 2) for r in range(155, 161)],
        'Endline': [_val(raw, r, 3) for r in range(155, 161)],
    })

    # ---- D. GENDER TRANSFORMATION: SHARED POWER ----
    # Conversations to change decisions: R168-R169 (0-based 167-168)
    g['decision_conversations'] = pd.DataFrame({
        'Response': [_clean_label(_val(raw, r, 1)) for r in range(167, 169)],
        'Baseline': [_val(raw, r, 2) for r in range(167, 169)],
        'Endline': [_val(raw, r, 3) for r in range(167, 169)],
    })

    # Made important decisions: R173-R174 (0-based 172-173)
    g['decisions_made'] = pd.DataFrame({
        'Response': [_clean_label(_val(raw, r, 1)) for r in range(172, 174)],
        'Baseline': [_val(raw, r, 2) for r in range(172, 174)],
        'Endline': [_val(raw, r, 3) for r in range(172, 174)],
    })

    # Type of decision: R178-R188 (0-based 177-187)
    g['decision_types'] = pd.DataFrame({
        'Decision': [_clean_label(_val(raw, r, 1)) for r in range(177, 188)],
        'Baseline': [_val(raw, r, 2) for r in range(177, 188)],
        'Endline': [_val(raw, r, 3) for r in range(177, 188)],
    })

    # Person making decision: R192-R194 (0-based 191-193)
    g['decision_maker'] = pd.DataFrame({
        'Person': [_clean_label(_val(raw, r, 1)) for r in range(191, 194)],
        'Baseline': [_val(raw, r, 2) for r in range(191, 194)],
        'Endline': [_val(raw, r, 3) for r in range(191, 194)],
    })

    # Equal say in joint decisions: R198-R199 (0-based 197-198)
    g['equal_say'] = pd.DataFrame({
        'Response': [_clean_label(_val(raw, r, 1)) for r in range(197, 199)],
        'Baseline': [_val(raw, r, 2) for r in range(197, 199)],
        'Endline': [_val(raw, r, 3) for r in range(197, 199)],
    })

    # ---- E. GENDER TRANSFORMATION: AUTONOMY & LEADERSHIP ----
    # Hide money to save: R205-R210 (0-based 204-209)
    g['hide_money'] = pd.DataFrame({
        'Category': [_clean_label(_val(raw, r, 1)) for r in range(204, 210)],
        'Baseline': [_val(raw, r, 2) for r in range(204, 210)],
        'Endline': [_val(raw, r, 3) for r in range(204, 210)],
    })

    # Husband support for women becoming community leader: R214-R219 (0-based 213-218)
    g['support_leader'] = pd.DataFrame({
        'Category': [_clean_label(_val(raw, r, 1)) for r in range(213, 219)],
        'Baseline': [_val(raw, r, 2) for r in range(213, 219)],
        'Endline': [_val(raw, r, 3) for r in range(213, 219)],
    })

    # Husband support if started/grew own business: R223-R228 (0-based 222-227)
    g['support_business'] = pd.DataFrame({
        'Category': [_clean_label(_val(raw, r, 1)) for r in range(222, 228)],
        'Baseline': [_val(raw, r, 2) for r in range(222, 228)],
        'Endline': [_val(raw, r, 3) for r in range(222, 228)],
    })

    return g


# ============================================================================
# GJJ KAP MEN DATA LOADER
# ============================================================================

@st.cache_data
def load_gjj_kap_men_data(filepath):
    """
    Parse the GJJ KAP Men Survey Excel file.
    Sheet 'Results KAP Men Endline' — 148 rows x 7 columns.
    All data in Col B (labels) + Cols C-G (values).
    Proportions are 0-1 scale; charts multiply x100 for display.
    Row references use 0-based indexing for _val() / iloc.
    """
    try:
        raw = pd.read_excel(filepath, sheet_name=GJJ_KAP_MEN_SHEET, header=None)
    except FileNotFoundError:
        st.error(f"GJJ KAP Men Excel not found: {filepath}")
        st.stop()

    g = {}

    def _men_label(r, c=1):
        """Get label, replacing NaN with 'Other'."""
        v = raw.iloc[r, c]
        if pd.isna(v):
            return 'Other'
        return _clean_label(v)

    # ---- A. SELF — Self-Esteem, Self-Compassion, Confidence ----
    # 3 statements: R9-R11, BL Likert: SA(c2) A(c3) D(c4) SD(c5)
    self_statements = [_men_label(r) for r in range(9, 12)]

    g['self_baseline_likert'] = pd.DataFrame({
        'Statement': self_statements,
        'Strongly agree': [_val(raw, r, 2) for r in range(9, 12)],
        'Agree': [_val(raw, r, 3) for r in range(9, 12)],
        'Disagree': [_val(raw, r, 4) for r in range(9, 12)],
        'Strongly disagree': [_val(raw, r, 5) for r in range(9, 12)],
    })

    # EL Likert: R16-R18
    g['self_endline_likert'] = pd.DataFrame({
        'Statement': self_statements,
        'Strongly agree': [_val(raw, r, 2) for r in range(16, 19)],
        'Agree': [_val(raw, r, 3) for r in range(16, 19)],
        'Disagree': [_val(raw, r, 4) for r in range(16, 19)],
        'Strongly disagree': [_val(raw, r, 5) for r in range(16, 19)],
    })

    # Agreement vs Disagreement: R23-R25
    g['self_agreement'] = pd.DataFrame({
        'Statement': self_statements,
        'Agreement_BL': [_val(raw, r, 2) for r in range(23, 26)],
        'Agreement_EL': [_val(raw, r, 3) for r in range(23, 26)],
        'Disagreement_BL': [_val(raw, r, 4) for r in range(23, 26)],
        'Disagreement_EL': [_val(raw, r, 5) for r in range(23, 26)],
    })

    # Self-compassion frequency: R30-R33, BL=c2, EL=c3
    compassion_cats = [_men_label(r) for r in range(30, 34)]
    g['self_compassion'] = pd.DataFrame({
        'Category': compassion_cats,
        'Baseline': [_val(raw, r, 2) for r in range(30, 34)],
        'Endline': [_val(raw, r, 3) for r in range(30, 34)],
    })

    # ---- B. RELATIONAL WELLBEING ----
    rel_statements = [_men_label(r) for r in range(41, 48)]

    # BL frequency: R41-R47, Always(c2) Freq(c3) Sometimes(c4) Rarely(c5) Never(c6)
    g['rel_baseline_freq'] = pd.DataFrame({
        'Statement': rel_statements,
        'Always': [_val(raw, r, 2) for r in range(41, 48)],
        'Frequently': [_val(raw, r, 3) for r in range(41, 48)],
        'Sometimes': [_val(raw, r, 4) for r in range(41, 48)],
        'Rarely': [_val(raw, r, 5) for r in range(41, 48)],
        'Never': [_val(raw, r, 6) for r in range(41, 48)],
    })

    # EL frequency: R52-R58
    g['rel_endline_freq'] = pd.DataFrame({
        'Statement': rel_statements,
        'Always': [_val(raw, r, 2) for r in range(52, 59)],
        'Frequently': [_val(raw, r, 3) for r in range(52, 59)],
        'Sometimes': [_val(raw, r, 4) for r in range(52, 59)],
        'Rarely': [_val(raw, r, 5) for r in range(52, 59)],
        'Never': [_val(raw, r, 6) for r in range(52, 59)],
    })

    # Always/Frequently vs Rarely/Never comparison: R63-R69
    g['rel_af_rn'] = pd.DataFrame({
        'Statement': rel_statements,
        'AF_Baseline': [_val(raw, r, 2) for r in range(63, 70)],
        'AF_Endline': [_val(raw, r, 3) for r in range(63, 70)],
        'RN_Baseline': [_val(raw, r, 4) for r in range(63, 70)],
        'RN_Endline': [_val(raw, r, 5) for r in range(63, 70)],
    })

    # ---- GJJ Journey Support (Baseline only) ----
    # R73-R76: Strongly Agree / Agree / Disagree / Strongly Disagree, col 2 only
    gjj_cats = [_men_label(r) for r in range(73, 77)]
    g['gjj_journey_support'] = pd.DataFrame({
        'Category': gjj_cats,
        'Baseline': [_val(raw, r, 2) for r in range(73, 77)],
    })

    # ---- C. SHARED RESPONSIBILITY ----
    # Supports household chores Y/N: R83-R84
    g['shared_chores_yn'] = pd.DataFrame({
        'Response': [_men_label(r) for r in range(83, 85)],
        'Baseline': [_val(raw, r, 2) for r in range(83, 85)],
        'Endline': [_val(raw, r, 3) for r in range(83, 85)],
    })

    # Time since supporting: R89-R95
    g['chore_duration'] = pd.DataFrame({
        'Category': [_men_label(r) for r in range(89, 96)],
        'Baseline': [_val(raw, r, 2) for r in range(89, 96)],
        'Endline': [_val(raw, r, 3) for r in range(89, 96)],
    })

    # Frequency of chore support: R99-R103
    g['chore_frequency'] = pd.DataFrame({
        'Category': [_men_label(r) for r in range(99, 104)],
        'Baseline': [_val(raw, r, 2) for r in range(99, 104)],
        'Endline': [_val(raw, r, 3) for r in range(99, 104)],
    })

    # Discussed sharing unpaid chores: R107-R108
    g['chore_discussed'] = pd.DataFrame({
        'Response': [_men_label(r) for r in range(107, 109)],
        'Baseline': [_val(raw, r, 2) for r in range(107, 109)],
        'Endline': [_val(raw, r, 3) for r in range(107, 109)],
    })

    # Encourages wife to make time for self: R113-R114
    g['support_self_time'] = pd.DataFrame({
        'Response': [_men_label(r) for r in range(113, 115)],
        'Baseline': [_val(raw, r, 2) for r in range(113, 115)],
        'Endline': [_val(raw, r, 3) for r in range(113, 115)],
    })

    # ---- D. SHARED POWER ----
    # Conversations to change decisions: R121-R122
    g['decision_conversations'] = pd.DataFrame({
        'Response': [_men_label(r) for r in range(121, 123)],
        'Baseline': [_val(raw, r, 2) for r in range(121, 123)],
        'Endline': [_val(raw, r, 3) for r in range(121, 123)],
    })

    # Made important decisions: R127-R128
    g['decisions_made'] = pd.DataFrame({
        'Response': [_men_label(r) for r in range(127, 129)],
        'Baseline': [_val(raw, r, 2) for r in range(127, 129)],
        'Endline': [_val(raw, r, 3) for r in range(127, 129)],
    })

    # ---- E. LEADERSHIP & BUSINESS SUPPORT ----
    # Support wife becoming community leader: R133-R138 (R135 label is NaN → 'Other')
    g['support_leader'] = pd.DataFrame({
        'Category': [_men_label(r) for r in range(133, 139)],
        'Baseline': [_val(raw, r, 2) for r in range(133, 139)],
        'Endline': [_val(raw, r, 3) for r in range(133, 139)],
    })

    # Support wife starting/growing business: R143-R147 (R147 label is NaN → 'Other')
    g['support_business'] = pd.DataFrame({
        'Category': [_men_label(r) for r in range(143, 148)],
        'Baseline': [_val(raw, r, 2) for r in range(143, 148)],
        'Endline': [_val(raw, r, 3) for r in range(143, 148)],
    })

    return g


# ============================================================================
# FOREST TRAINING DATA LOADER
# ============================================================================

def load_forest_training_data(filepath):
    """Load Forest Training Pre/Post Knowledge Assessment data.

    Returns a dict with keys:
        'thresholds'   – DataFrame with columns [Timepoint, Threshold, Proportion]
        'scores'       – DataFrame with columns [Timepoint, Respondents, AverageScore, MaxScore, MinScore]
        'questions'    – DataFrame with columns [QuestionNumber, QuestionText, Baseline, Endline]
    """
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[FOREST_TRAINING_SHEET]

    def _cell(r, c):
        v = ws.cell(r, c).value
        if v is None:
            return 0.0
        try:
            return float(v)
        except (ValueError, TypeError):
            return 0.0

    def _cell_str(r, c):
        v = ws.cell(r, c).value
        return str(v).strip() if v is not None else ''

    # ---- A. Pass Thresholds (rows 7-9, cols B-F) ----
    threshold_labels = []
    for c in range(3, 7):  # cols C..F  (>=50, >=60, >=70, >=80)
        lbl = _cell_str(7, c)  # row 7 has header labels
        if not lbl:
            lbl = f'>={[50,60,70,80][c-3]}'
        threshold_labels.append(lbl)

    threshold_rows = []
    for tp_row, tp_name in [(8, 'Baseline'), (9, 'Endline')]:
        for i, lbl in enumerate(threshold_labels):
            proportion = _cell(tp_row, 3 + i)
            threshold_rows.append({'Timepoint': tp_name, 'Threshold': lbl, 'Proportion': proportion})
    thresholds_df = pd.DataFrame(threshold_rows)

    # ---- B. Standardized Scores (rows 13-15, cols B-F) ----
    score_rows = []
    for tp_row, tp_name in [(14, 'Baseline'), (15, 'Endline')]:
        score_rows.append({
            'Timepoint': tp_name,
            'Respondents': int(_cell(tp_row, 3)),
            'AverageScore': round(_cell(tp_row, 4), 2),
            'MaxScore': round(_cell(tp_row, 5), 1),
            'MinScore': round(_cell(tp_row, 6), 1),
        })
    scores_df = pd.DataFrame(score_rows)

    # ---- C. Question-by-Question (rows 20-39+, cols B, L, M) ----
    question_rows = []
    r = 20
    while r <= ws.max_row:
        q_text = _cell_str(r, 2)  # col B
        if not q_text or not q_text[0].isdigit():
            r += 1
            continue
        # Extract question number from text like "1. Can you define..."
        parts = q_text.split('.', 1)
        try:
            q_num = int(parts[0].strip())
        except ValueError:
            r += 1
            continue
        q_label = parts[1].strip() if len(parts) > 1 else q_text
        bl_pct = _cell(r, 12)  # col L
        el_pct = _cell(r, 13)  # col M
        question_rows.append({
            'QuestionNumber': q_num,
            'QuestionText': q_label,
            'Baseline': bl_pct,
            'Endline': el_pct,
        })
        r += 1

    questions_df = pd.DataFrame(question_rows)
    if not questions_df.empty:
        questions_df = questions_df.sort_values('QuestionNumber').reset_index(drop=True)

    wb.close()

    return {
        'thresholds': thresholds_df,
        'scores': scores_df,
        'questions': questions_df,
    }


# ============================================================================
# MANGROVE TRAINING DATA LOADER
# ============================================================================

def load_mangrove_training_data(filepath):
    """Load Mangrove Training Pre/Post Knowledge Assessment data.

    Returns dict with keys:
        'thresholds'      – Pass thresholds by county (long-form)
        'scores'          – Standardized scores by county & sex
        'adequate_county'  – Adequate knowledge (≥60) by county (Pre/Post)
        'adequate_sex'     – Adequate knowledge (≥60) by sex (Pre/Post)
    """
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[MANGROVE_TRAINING_SHEET]

    def _cell(r, c):
        v = ws.cell(r, c).value
        if v is None:
            return 0.0
        try:
            return float(v)
        except (ValueError, TypeError):
            return 0.0

    def _cell_str(r, c):
        v = ws.cell(r, c).value
        return str(v).strip() if v is not None else ''

    # ---- A. Pass Thresholds (rows 8-11, cols B-D) ----
    # Header row 8 has threshold labels: >=50, >=55, >=60
    threshold_labels = []
    for c in range(3, 6):  # cols C, D, E
        lbl = _cell_str(8, c)
        if not lbl:
            lbl = f'>={[50,55,60][c-3]}'
        threshold_labels.append(lbl)

    threshold_rows = []
    for row_idx in [9, 10, 11]:  # Kilifi, Kwale, All
        county = _cell_str(row_idx, 2)  # col B
        if not county or 'follow' in county.lower():
            continue
        for i, thr_lbl in enumerate(threshold_labels):
            val = _cell(row_idx, 3 + i)
            threshold_rows.append({
                'County': county, 'Threshold': thr_lbl, 'Value': val,
            })
    thresholds_df = pd.DataFrame(threshold_rows)

    # ---- Also capture M/F values from rows 10-11 cols H-I ----
    sex_threshold_rows = []
    for row_idx in [10, 11]:
        sex_label = _cell_str(row_idx, 8)  # col H (M or F)
        sex_val = _cell(row_idx, 9)  # col I
        if sex_label and sex_label in ('M', 'F'):
            full_sex = 'Male' if sex_label == 'M' else 'Female'
            sex_threshold_rows.append({
                'Sex': full_sex,
                'Value_60': sex_val,  # appears to be >=60 pass rate
            })
    sex_threshold_df = pd.DataFrame(sex_threshold_rows) if sex_threshold_rows else pd.DataFrame()

    # ---- B. Standardized Scores (rows 15-18, cols B-F) ----
    score_rows = []
    for row_idx in [16, 17, 18]:  # Kilifi, Kwale, All
        county = _cell_str(row_idx, 2)
        if not county:
            continue
        score_rows.append({
            'County': county,
            'Respondents': int(_cell(row_idx, 3)),
            'AvgScore': round(_cell(row_idx, 4), 2),
            'MaxScore': round(_cell(row_idx, 5), 2),
            'MinScore': round(_cell(row_idx, 6), 2),
        })
    scores_df = pd.DataFrame(score_rows)

    # ---- C. Adequate Knowledge by County (rows 22-24, Pre/Post) ----
    county_rows = []
    for row_idx in [22, 23, 24]:
        county = _cell_str(row_idx, 2)
        if not county:
            continue
        pre_val = _cell(row_idx, 3)
        post_val = _cell(row_idx, 4)
        county_rows.append({'County': county, 'Timepoint': 'Pre-Test', 'Value': pre_val})
        county_rows.append({'County': county, 'Timepoint': 'Post-Test', 'Value': post_val})
    adequate_county_df = pd.DataFrame(county_rows)

    # ---- D. Adequate Knowledge by Sex (rows 28-30, Pre/Post) ----
    sex_rows = []
    for row_idx in [28, 29, 30]:
        sex = _cell_str(row_idx, 2)
        if not sex:
            continue
        pre_val = _cell(row_idx, 3)
        post_val = _cell(row_idx, 4)
        sex_rows.append({'Sex': sex, 'Timepoint': 'Pre-Test', 'Value': pre_val})
        sex_rows.append({'Sex': sex, 'Timepoint': 'Post-Test', 'Value': post_val})
    adequate_sex_df = pd.DataFrame(sex_rows)

    wb.close()

    return {
        'thresholds': thresholds_df,
        'sex_threshold': sex_threshold_df,
        'scores': scores_df,
        'adequate_county': adequate_county_df,
        'adequate_sex': adequate_sex_df,
    }


# ============================================================================
# SEAWEED DATA LOADER & AGGREGATOR
# ============================================================================

def load_seaweed_data(filepath):
    """Load Seaweed Production & Challenges CSV and return cleaned DataFrame.

    The CSV is UTF-16 encoded. Numeric columns stored as strings are converted.
    Challenge flag columns are coerced to boolean.
    Derived columns: Ropes_Achievement_pct, Production_per_rope_kg.
    """
    # Try utf-16 first, fall back to other encodings
    for enc in ['utf-16', 'utf-16-le', 'latin1', 'utf-8']:
        try:
            df = pd.read_csv(filepath, encoding=enc)
            break
        except Exception:
            continue
    else:
        df = pd.read_csv(filepath, encoding='latin1')

    # Drop unnamed trailing columns
    df = df[[c for c in df.columns if not c.startswith('Unnamed')]]

    # Drop rows where key identifiers are missing
    df = df.dropna(subset=['Group', 'Member'], how='all').copy()

    # ---- Numeric coercion ----
    num_cols = ['Ropes_Ocean', 'Ropes_Home', 'Ropes_Total', 'Target_Ropes',
                'Ropes Required', 'Gap', 'Dried_KG', 'Wet_KG', 'Total_KG',
                'x', 'y']
    for c in num_cols:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0)

    # ---- Boolean coercion for flag columns ----
    flag_cols = ['flag_transport', 'flag_market', 'flag_disease',
                 'flag_equipment', 'flag_storage', 'flag_labour', 'flag_sand_tide']
    for c in flag_cols:
        if c in df.columns:
            df[c] = df[c].astype(str).str.strip().str.lower().isin(['true', '1', 'yes'])

    # ---- Derived metrics ----
    df['Ropes_Achievement_pct'] = df.apply(
        lambda r: round(r['Ropes_Total'] / r['Target_Ropes'] * 100, 1)
        if r['Target_Ropes'] > 0 else 0.0, axis=1)

    df['Production_per_rope_kg'] = df.apply(
        lambda r: round(r['Total_KG'] / r['Ropes_Total'], 2)
        if r['Ropes_Total'] > 0 else 0.0, axis=1)

    return df


def prepare_seaweed_aggregates(df):
    """Compute group-level aggregations and challenge counts.

    Returns dict with:
        'group_summary'   – DataFrame of per-group metrics
        'challenge_counts' – DataFrame of challenge flag totals/percentages
        'overall'          – dict of overall KPIs
    """
    flag_cols = ['flag_transport', 'flag_market', 'flag_disease',
                 'flag_equipment', 'flag_storage', 'flag_labour', 'flag_sand_tide']
    flag_labels = {
        'flag_transport': 'Transport', 'flag_market': 'Market Access',
        'flag_disease': 'Disease', 'flag_equipment': 'Equipment',
        'flag_storage': 'Storage', 'flag_labour': 'Labour',
        'flag_sand_tide': 'Sand / Tide',
    }

    # ---- Group-level summary ----
    grp = df.groupby('Group', dropna=False).agg(
        Members=('Member', 'count'),
        Ropes_Ocean=('Ropes_Ocean', 'sum'),
        Ropes_Home=('Ropes_Home', 'sum'),
        Ropes_Total=('Ropes_Total', 'sum'),
        Target_Ropes=('Target_Ropes', 'sum'),
        Ropes_Required=('Ropes Required', 'sum'),
        Gap=('Gap', 'sum'),
        Dried_KG=('Dried_KG', 'sum'),
        Wet_KG=('Wet_KG', 'sum'),
        Total_KG=('Total_KG', 'sum'),
        Avg_Ropes_per_Member=('Ropes_Total', 'mean'),
        Avg_Achievement_pct=('Ropes_Achievement_pct', 'mean'),
        Avg_Prod_per_Rope=('Production_per_rope_kg', 'mean'),
    ).reset_index()
    grp['Casual_Worker_pct'] = df.groupby('Group')['Casual_Workers'].apply(
        lambda s: round((s.str.lower() == 'yes').sum() / max(len(s), 1) * 100, 1)
    ).values

    # Per-group challenge proportions
    for fc in flag_cols:
        if fc in df.columns:
            lbl = flag_labels.get(fc, fc)
            grp[f'pct_{lbl}'] = df.groupby('Group')[fc].apply(
                lambda s: round(s.sum() / max(len(s), 1) * 100, 1)
            ).values

    # ---- Overall challenge counts ----
    ch_rows = []
    valid_df = df.dropna(subset=flag_cols[:1])  # rows with non-null flags
    for fc in flag_cols:
        if fc in df.columns:
            lbl = flag_labels.get(fc, fc)
            count = int(df[fc].sum())
            pct = round(count / max(len(valid_df), 1) * 100, 1)
            ch_rows.append({'Challenge': lbl, 'Count': count, 'Pct': pct})
    challenge_df = pd.DataFrame(ch_rows)

    # ---- Overall KPIs ----
    total_kg = float(df['Total_KG'].sum())
    n_farmers = int(df['Member'].nunique())
    ropes_ocean = float(df['Ropes_Ocean'].sum())
    total_ropes = float(df['Ropes_Total'].sum())
    avg_prod = round(total_kg / max(total_ropes, 1), 2)
    casual_pct = round(
        (df['Casual_Workers'].str.lower() == 'yes').sum() / max(len(df), 1) * 100, 1)

    overall = {
        'total_kg': total_kg,
        'n_farmers': n_farmers,
        'n_groups': int(df['Group'].nunique()),
        'ropes_ocean': ropes_ocean,
        'ropes_total': total_ropes,
        'target_ropes_total': float(df['Target_Ropes'].sum()),
        'avg_prod_per_rope': avg_prod,
        'casual_pct': casual_pct,
        'dried_kg': float(df['Dried_KG'].sum()),
        'wet_kg': float(df['Wet_KG'].sum()),
        'dried_wet_ratio': round(float(df['Dried_KG'].sum()) / max(float(df['Wet_KG'].sum()), 1), 2),
        'avg_achievement_pct': round(float(df['Ropes_Achievement_pct'].mean()), 1),
        'pct_meeting_target': round(
            len(df[df['Ropes_Total'] >= df['Target_Ropes']]) / max(len(df), 1) * 100, 1),
        'avg_ropes_per_farmer': round(total_ropes / max(n_farmers, 1), 1),
        'avg_production_per_farmer': round(total_kg / max(n_farmers, 1), 1),
        'gap_total': float(df['Gap'].sum()),
        'multi_challenge_pct': round(
            len(df[df[flag_cols].sum(axis=1) >= 2]) / max(len(valid_df), 1) * 100, 1)
            if all(fc in df.columns for fc in flag_cols) else 0.0,
    }

    return {
        'group_summary': grp,
        'challenge_counts': challenge_df,
        'overall': overall,
    }


# ============================================================================
# CHART HELPERS
# ============================================================================

def _export_data_csv(data_dict, prefix='data'):
    """Combine all DataFrames in a dict into a single CSV for download."""
    import io
    buf = io.StringIO()
    first = True
    for key, df in data_dict.items():
        if not isinstance(df, pd.DataFrame) or df.empty:
            continue
        if not first:
            buf.write('\n')
        buf.write(f'--- {key} ---\n')
        df.to_csv(buf, index=False)
        first = False
    return buf.getvalue()


def make_comparison_bar(df, cat_col, title, y_label="Percentage (%)",
                        multiply=True, height=450, orientation='v',
                        color_baseline=None, color_midline=None):
    cb = color_baseline or COLORS["baseline"]
    cm = color_midline or COLORS["midline"]
    plot_df = df.copy()
    if multiply:
        plot_df['Baseline'] = plot_df['Baseline'].apply(lambda x: x*100 if isinstance(x,(int,float)) else 0)
        plot_df['Midline'] = plot_df['Midline'].apply(lambda x: x*100 if isinstance(x,(int,float)) else 0)
    if orientation == 'h':
        # Sort so longest bars appear at top (Plotly renders bottom-to-top)
        plot_df['_sort'] = plot_df[['Baseline', 'Midline']].max(axis=1)
        plot_df = plot_df.sort_values('_sort', ascending=True).drop(columns='_sort')
        fig = go.Figure()
        fig.add_trace(go.Bar(y=plot_df[cat_col], x=plot_df['Baseline'], name='Baseline',
                             orientation='h', marker_color=cb,
                             text=plot_df['Baseline'].apply(lambda x: f"{x:.1f}%"), textposition='auto'))
        fig.add_trace(go.Bar(y=plot_df[cat_col], x=plot_df['Midline'], name='Midline',
                             orientation='h', marker_color=cm,
                             text=plot_df['Midline'].apply(lambda x: f"{x:.1f}%"), textposition='auto'))
        fig.update_layout(title=title, barmode='group', height=height,
                          xaxis_title=y_label, legend=dict(orientation='h', yanchor='bottom', y=1.02),
                          yaxis=dict(categoryorder='trace'))
    else:
        fig = go.Figure()
        fig.add_trace(go.Bar(x=plot_df[cat_col], y=plot_df['Baseline'], name='Baseline',
                             marker_color=cb, text=plot_df['Baseline'].apply(lambda x: f"{x:.1f}%"),
                             textposition='auto'))
        fig.add_trace(go.Bar(x=plot_df[cat_col], y=plot_df['Midline'], name='Midline',
                             marker_color=cm, text=plot_df['Midline'].apply(lambda x: f"{x:.1f}%"),
                             textposition='auto'))
        fig.update_layout(title=title, barmode='group', height=height, yaxis_title=y_label,
                          legend=dict(orientation='h', yanchor='bottom', y=1.02))
    fig.update_layout(
        font=dict(size=13, color='#333'),
        title_font=dict(size=16, color='#222'),
        plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
        margin=dict(l=20, r=20, t=60, b=20),
    )
    return fig


def make_stacked_bar(df, cat_col, columns, colors_list, title,
                     y_label="Percentage (%)", height=400, multiply=True, orientation='v'):
    plot_df = df.copy()
    if multiply:
        for c in columns:
            plot_df[c] = plot_df[c].apply(lambda x: x*100 if isinstance(x,(int,float)) else 0)
    fig = go.Figure()
    for col, color in zip(columns, colors_list):
        label = col.replace('Baseline_','').replace('Midline_','')
        if orientation == 'h':
            fig.add_trace(go.Bar(y=plot_df[cat_col], x=plot_df[col], name=label, orientation='h',
                                 marker_color=color,
                                 text=plot_df[col].apply(lambda x: f"{x:.1f}%" if x > 3 else ""),
                                 textposition='inside'))
        else:
            fig.add_trace(go.Bar(x=plot_df[cat_col], y=plot_df[col], name=label, marker_color=color,
                                 text=plot_df[col].apply(lambda x: f"{x:.1f}%" if x > 3 else ""),
                                 textposition='inside'))
    fig.update_layout(title=title, barmode='stack', height=height,
                      yaxis_title=y_label if orientation=='v' else '',
                      xaxis_title=y_label if orientation=='h' else '',
                      legend=dict(orientation='h', yanchor='bottom', y=1.02),
                      font=dict(size=13, color='#333'),
                      title_font=dict(size=16, color='#222'),
                      plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                      margin=dict(l=20, r=20, t=60, b=20))
    return fig


def make_delta_bar(df, cat_col, title, multiply=True, height=400):
    plot_df = df.copy()
    factor = 100 if multiply else 1
    plot_df['Change'] = (plot_df['Midline'] - plot_df['Baseline']) * factor
    plot_df = plot_df.sort_values('Change')
    colors = [COLORS['good'] if v >= 0 else COLORS['danger'] for v in plot_df['Change']]
    fig = go.Figure()
    fig.add_trace(go.Bar(y=plot_df[cat_col], x=plot_df['Change'], orientation='h',
                         marker_color=colors,
                         text=plot_df['Change'].apply(lambda x: f"{x:+.1f}pp"), textposition='auto'))
    fig.update_layout(title=title, height=height, xaxis_title="Change (pp)",
                      yaxis=dict(categoryorder='trace'),
                      font=dict(size=13, color='#333'),
                      title_font=dict(size=16, color='#222'),
                      plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                      margin=dict(l=20, r=20, t=60, b=20))
    return fig


def make_two_col_bar(df1, df2, col1_name, col2_name, cat_col, title, height=450):
    """Side-by-side comparison of two related dataframes (e.g., norms vs experience)."""
    # Sort so longest bars appear at top (Plotly renders bottom-to-top)
    sort_vals = (df1[['Baseline','Midline']].max(axis=1)*100 + df2[['Baseline','Midline']].max(axis=1)*100) / 2
    sort_order = sort_vals.sort_values(ascending=True).index
    df1 = df1.loc[sort_order]
    df2 = df2.loc[sort_order]
    fig = go.Figure()
    fig.add_trace(go.Bar(y=df1[cat_col], x=df1['Baseline']*100, name=f'{col1_name} (BL)',
                         orientation='h', marker_color=COLORS['baseline_light'],
                         text=df1['Baseline'].apply(lambda x: f"{x*100:.0f}%"), textposition='auto'))
    fig.add_trace(go.Bar(y=df1[cat_col], x=df1['Midline']*100, name=f'{col1_name} (ML)',
                         orientation='h', marker_color=COLORS['baseline'],
                         text=df1['Midline'].apply(lambda x: f"{x*100:.0f}%"), textposition='auto'))
    fig.add_trace(go.Bar(y=df2[cat_col], x=df2['Baseline']*100, name=f'{col2_name} (BL)',
                         orientation='h', marker_color=COLORS['midline_light'],
                         text=df2['Baseline'].apply(lambda x: f"{x*100:.0f}%"), textposition='auto'))
    fig.add_trace(go.Bar(y=df2[cat_col], x=df2['Midline']*100, name=f'{col2_name} (ML)',
                         orientation='h', marker_color=COLORS['midline'],
                         text=df2['Midline'].apply(lambda x: f"{x*100:.0f}%"), textposition='auto'))
    fig.update_layout(title=title, barmode='group', height=height, xaxis_title='%',
                      legend=dict(orientation='h', yanchor='bottom', y=1.02),
                      yaxis=dict(categoryorder='trace'),
                      font=dict(size=13, color='#333'),
                      title_font=dict(size=16, color='#222'),
                      plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                      margin=dict(l=20, r=20, t=60, b=20))
    return fig


# ============================================================================
# WOMEN SURVEY TAB RENDERERS
# ============================================================================

def _section_header(icon, title, badge_text=None):
    """Render a styled section header with optional badge."""
    badge = f'<span class="badge">{badge_text}</span>' if badge_text else ''
    heading = f'{icon} {title}' if icon else title
    anchor_id = re.sub(r'[^a-z0-9]+', '-', title.lower()).strip('-')
    st.markdown(f'<div class="section-header" id="{anchor_id}"><h2>{heading}</h2>{badge}</div>',
                unsafe_allow_html=True)


def _quick_nav_pills(items):
    """Render quick navigation pills for in-tab section jumping."""
    pills = ''.join([
        f'<a href="#{re.sub(r"[^a-z0-9]+", "-", item.lower()).strip("-")}" '
        f'class="quick-nav-pill" style="text-decoration:none;color:inherit;">{item}</a>'
        for item in items
    ])
    st.markdown(f'<div class="quick-nav">{pills}</div>', unsafe_allow_html=True)


def render_women_tab1(w):
    """Tab 1: Household Profile & Services."""
    st.markdown("""<div class="section-narrative">
    <strong>Household Profile:</strong> Demographics of surveyed women — location type, household
    headship, marital status, education levels, economic activities, and access to basic services
    (water, sanitation, electricity, housing materials).
    </div>""", unsafe_allow_html=True)

    _quick_nav_pills(['Demographics', 'Economic Activities', 'Basic Services'])

    c1, c2 = st.columns(2)
    with c1:
        st.plotly_chart(make_comparison_bar(w['location'], 'Category',
                        'Women by Location Type', height=300), use_container_width=True)
    with c2:
        st.plotly_chart(make_comparison_bar(w['hh_type'], 'Category',
                        'Household Type', height=300), use_container_width=True)

    c3, c4 = st.columns(2)
    with c3:
        st.plotly_chart(make_comparison_bar(w['marital'], 'Category',
                        'Marital Status', height=380, orientation='h'), use_container_width=True)
    with c4:
        st.plotly_chart(make_comparison_bar(w['education'], 'Category',
                        'Education Level', height=380, orientation='h'), use_container_width=True)

    st.markdown("---")
    _section_header('', 'Economic Activities', 'Section A')
    c5, c6 = st.columns(2)
    with c5:
        st.plotly_chart(make_comparison_bar(w['main_econ'], 'Activity',
                        'Main HH Economic Activity', height=500, orientation='h'), use_container_width=True)
    with c6:
        st.plotly_chart(make_comparison_bar(w['sec_econ'], 'Activity',
                        'Secondary HH Economic Activities', height=500, orientation='h'), use_container_width=True)

    st.markdown("---")
    _section_header('', 'Access to Basic Services', 'Section A')
    c7, c8, c9 = st.columns(3)
    with c7:
        st.plotly_chart(make_comparison_bar(w['water'], 'Category',
                        'Access to Safe Drinking Water', height=300), use_container_width=True)
    with c8:
        st.plotly_chart(make_comparison_bar(w['toilet'], 'Category',
                        'Access to Improved Toilet', height=300), use_container_width=True)
    with c9:
        st.plotly_chart(make_comparison_bar(w['electricity'], 'Category',
                        'Access to Electricity', height=300), use_container_width=True)

    c10, c11 = st.columns(2)
    with c10:
        st.plotly_chart(make_comparison_bar(w['walls'], 'Category',
                        'Wall Material', height=350), use_container_width=True)
    with c11:
        st.plotly_chart(make_comparison_bar(w['floors'], 'Category',
                        'Floor Material', height=350), use_container_width=True)


def render_women_tab2(w):
    """Tab 2: Shocks, Coping & Preparedness."""
    st.markdown("""<div class="section-narrative">
    <strong>Shocks & Preparedness:</strong> Shocks and stresses experienced by women's households,
    their perceived impact, coping strategies employed, and access to disaster preparedness information
    including weather/tidal forecasts and early warning systems.
    </div>""", unsafe_allow_html=True)

    _quick_nav_pills(['Shocks & Impact', 'Coping Strategies', 'Disaster Preparedness'])

    c1, c2 = st.columns(2)
    with c1:
        st.plotly_chart(make_comparison_bar(w['shocks'], 'Shock',
                        'Shocks Experienced (Past 12 Months)', height=450, orientation='h'),
                        use_container_width=True)
    with c2:
        st.plotly_chart(make_comparison_bar(w['shock_impact'], 'Extent',
                        'Perceived Impact on Wellbeing', height=400, orientation='h'),
                        use_container_width=True)

    _section_header('', 'Coping Strategies', 'Section B')
    st.plotly_chart(make_comparison_bar(w['coping'], 'Strategy',
                    'Coping Strategies Used', height=500, orientation='h'), use_container_width=True)

    st.markdown("---")
    _section_header('', 'Disaster Preparedness', 'Section C')
    c3, c4 = st.columns(2)
    with c3:
        st.plotly_chart(make_comparison_bar(w['prep_knowledge'], 'Response',
                        'Knowledge of Preparedness Plans', height=350), use_container_width=True)
    with c4:
        st.plotly_chart(make_comparison_bar(w['prep_participation'], 'Response',
                        'Participation in Plan Development', height=350), use_container_width=True)

    c5, c6 = st.columns(2)
    with c5:
        st.plotly_chart(make_comparison_bar(w['prep_know_action'], 'Response',
                        'Know What to Do in Disaster', height=300), use_container_width=True)
    with c6:
        st.plotly_chart(make_comparison_bar(w['early_warning'], 'Response',
                        'Access to Early Warning Info', height=300), use_container_width=True)

    c7, c8 = st.columns(2)
    with c7:
        st.plotly_chart(make_comparison_bar(w['weather_forecast'], 'Response',
                        'Access to Weather Forecasts', height=300), use_container_width=True)
    with c8:
        st.plotly_chart(make_comparison_bar(w['tidal_forecast'], 'Response',
                        'Access to Tidal Forecasts', height=300), use_container_width=True)

    # Awareness of plan actions
    _section_header('', 'Awareness of Preparedness Plan Actions', 'Section C')
    st.plotly_chart(make_comparison_bar(w['prep_awareness'], 'Extent',
                    'Awareness of Plan Actions (by Extent)', height=400, orientation='h'),
                    use_container_width=True)


def render_women_tab3(w):
    """Tab 3: Assets, Land, Savings & Loans."""
    st.markdown("""<div class="section-narrative">
    <strong>Assets & Financial Inclusion:</strong> Household assets, ownership patterns (joint/sole/all),
    ability to use or sell assets, access to productive inputs, land size and tenure, savings behaviour,
    and borrowing patterns.
    </div>""", unsafe_allow_html=True)

    _quick_nav_pills(['Household Assets', 'Land', 'Savings', 'Loans'])

    _section_header('', 'Household Assets', 'Section D')
    st.plotly_chart(make_comparison_bar(w['hh_assets'], 'Asset',
                    'Assets Present in Household', height=600, orientation='h'), use_container_width=True)

    # Asset ownership breakdown
    _section_header('', 'Asset Ownership (Joint vs Sole)', 'Section D')
    ao = w['asset_ownership']
    fig_ao = go.Figure()
    fig_ao.add_trace(go.Bar(y=ao['Asset'], x=ao['Joint_BL'].apply(lambda x: x*100 if isinstance(x,(int,float)) and x<=1 else x),
                            name='Joint (BL)', orientation='h', marker_color=COLORS['baseline_light']))
    fig_ao.add_trace(go.Bar(y=ao['Asset'], x=ao['Joint_ML'].apply(lambda x: x*100 if isinstance(x,(int,float)) and x<=1 else x),
                            name='Joint (ML)', orientation='h', marker_color=COLORS['baseline']))
    fig_ao.add_trace(go.Bar(y=ao['Asset'], x=ao['Sole_BL'].apply(lambda x: x*100 if isinstance(x,(int,float)) and x<=1 else x),
                            name='Sole (BL)', orientation='h', marker_color=COLORS['midline_light']))
    fig_ao.add_trace(go.Bar(y=ao['Asset'], x=ao['Sole_ML'].apply(lambda x: x*100 if isinstance(x,(int,float)) and x<=1 else x),
                            name='Sole (ML)', orientation='h', marker_color=COLORS['midline']))
    fig_ao.update_layout(title='Asset Ownership: Joint vs Sole', barmode='group', height=550,
                         xaxis_title='%', legend=dict(orientation='h', yanchor='bottom', y=1.02),
                         font=dict(size=13, color='#333'), title_font=dict(size=16, color='#222'),
                         plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                         margin=dict(l=20, r=20, t=60, b=20))
    st.plotly_chart(fig_ao, use_container_width=True)

    # Ability to use/sell assets
    aus = w['asset_use_sell']
    aus_c1, aus_c2 = st.columns(2)
    with aus_c1:
        fig_use = go.Figure()
        fig_use.add_trace(go.Bar(y=aus['Asset'], x=aus['Use_BL'].apply(lambda x: x*100 if isinstance(x,(int,float)) and x<=1 else x),
                                 name='BL', orientation='h', marker_color=COLORS['baseline']))
        fig_use.add_trace(go.Bar(y=aus['Asset'], x=aus['Use_ML'].apply(lambda x: x*100 if isinstance(x,(int,float)) and x<=1 else x),
                                 name='ML', orientation='h', marker_color=COLORS['midline']))
        fig_use.update_layout(title='Ability to Use Assets', barmode='group', height=500,
                              xaxis_title='%', legend=dict(orientation='h', yanchor='bottom', y=1.02),
                              font=dict(size=13, color='#333'), title_font=dict(size=16, color='#222'),
                              plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)')  
        st.plotly_chart(fig_use, use_container_width=True)
    with aus_c2:
        fig_sell = go.Figure()
        fig_sell.add_trace(go.Bar(y=aus['Asset'], x=aus['Sell_BL'].apply(lambda x: x*100 if isinstance(x,(int,float)) and x<=1 else x),
                                  name='BL', orientation='h', marker_color=COLORS['baseline']))
        fig_sell.add_trace(go.Bar(y=aus['Asset'], x=aus['Sell_ML'].apply(lambda x: x*100 if isinstance(x,(int,float)) and x<=1 else x),
                                  name='ML', orientation='h', marker_color=COLORS['midline']))
        fig_sell.update_layout(title='Ability to Sell Assets', barmode='group', height=500,
                               xaxis_title='%', legend=dict(orientation='h', yanchor='bottom', y=1.02),
                               font=dict(size=13, color='#333'), title_font=dict(size=16, color='#222'),
                               plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)') 
        st.plotly_chart(fig_sell, use_container_width=True)

    # Access to productive inputs
    _section_header('', 'Access to Productive Inputs', 'Section D')
    inp = w['inputs']
    inp_c1, inp_c2 = st.columns(2)
    with inp_c1:
        fig_inp_a = go.Figure()
        fig_inp_a.add_trace(go.Bar(y=inp['Input'], x=inp['Access_BL'].apply(lambda x: x*100 if isinstance(x,(int,float)) and x<=1 else x),
                                   name='BL', orientation='h', marker_color=COLORS['baseline']))
        fig_inp_a.add_trace(go.Bar(y=inp['Input'], x=inp['Access_ML'].apply(lambda x: x*100 if isinstance(x,(int,float)) and x<=1 else x),
                                   name='ML', orientation='h', marker_color=COLORS['midline']))
        fig_inp_a.update_layout(title='Access to Inputs', barmode='group', height=400,
                                xaxis_title='%', legend=dict(orientation='h', yanchor='bottom', y=1.02),
                                font=dict(size=13, color='#333'), title_font=dict(size=16, color='#222'),
                                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)')  
        st.plotly_chart(fig_inp_a, use_container_width=True)
    with inp_c2:
        fig_inp_p = go.Figure()
        fig_inp_p.add_trace(go.Bar(y=inp['Input'], x=inp['PersonalUse_BL'].apply(lambda x: x*100 if isinstance(x,(int,float)) and x<=1 else x),
                                   name='BL', orientation='h', marker_color=COLORS['baseline']))
        fig_inp_p.add_trace(go.Bar(y=inp['Input'], x=inp['PersonalUse_ML'].apply(lambda x: x*100 if isinstance(x,(int,float)) and x<=1 else x),
                                   name='ML', orientation='h', marker_color=COLORS['midline']))
        fig_inp_p.update_layout(title='Personal Use of Inputs', barmode='group', height=400,
                                xaxis_title='%', legend=dict(orientation='h', yanchor='bottom', y=1.02),
                                font=dict(size=13, color='#333'), title_font=dict(size=16, color='#222'),
                                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)')  
        st.plotly_chart(fig_inp_p, use_container_width=True)

    st.markdown("---")
    _section_header('', 'Land', 'Section D')
    c1, c2, c3 = st.columns(3)
    with c1:
        st.plotly_chart(make_comparison_bar(w['land_size'], 'Category',
                        'Agricultural Land Size', height=350), use_container_width=True)
    with c2:
        st.plotly_chart(make_comparison_bar(w['land_status'], 'Category',
                        'Land Tenure Status', height=350), use_container_width=True)
    with c3:
        st.plotly_chart(make_comparison_bar(w['land_use'], 'Category',
                        'Currently Using Land for Agriculture', height=300), use_container_width=True)

    st.markdown("---")
    _section_header('', 'Savings', 'Section E')
    c4, c5 = st.columns(2)
    with c4:
        st.plotly_chart(make_comparison_bar(w['personal_saving'], 'Response',
                        'Women Personally Saving', height=300), use_container_width=True)
    with c5:
        st.plotly_chart(make_comparison_bar(w['family_saving'], 'Response',
                        'Family Members Saving', height=300), use_container_width=True)

    c6, c7 = st.columns(2)
    with c6:
        st.plotly_chart(make_comparison_bar(w['saving_freq'], 'Frequency',
                        'Saving Frequency', height=350, orientation='h'), use_container_width=True)
    with c7:
        st.plotly_chart(make_comparison_bar(w['saving_amount'], 'Amount',
                        'Amount Saved per Period', height=350, orientation='h'), use_container_width=True)

    c8, c9 = st.columns(2)
    with c8:
        st.plotly_chart(make_comparison_bar(w['saving_mechanism'], 'Mechanism',
                        'Saving Mechanism', height=380, orientation='h'), use_container_width=True)
    with c9:
        st.plotly_chart(make_comparison_bar(w['saving_use'], 'Use',
                        'Intended Use of Savings', height=380, orientation='h'), use_container_width=True)

    # Savings balance distribution
    st.plotly_chart(make_comparison_bar(w['saving_balance'], 'Balance',
                    'Savings Balance Distribution', height=350, orientation='h'),
                    use_container_width=True)

    st.markdown('---')
    _section_header('', 'Loans', 'Section E')
    c10, c11 = st.columns(2)
    with c10:
        st.plotly_chart(make_comparison_bar(w['personal_loan'], 'Response',
                        'Women Personally Taking a Loan', height=300), use_container_width=True)
        pl = w['personal_loan_size']
        st.metric('Avg Personal Loan (BL \u2192 ML)',
                  f"KES {pl['Baseline'].values[0]:,.0f} \u2192 {pl['Midline'].values[0]:,.0f}")
    with c11:
        st.plotly_chart(make_comparison_bar(w['loan_source'], 'Source',
                        'Loan Sources', height=380, orientation='h'), use_container_width=True)

    # Family loan data
    fl_c1, fl_c2 = st.columns(2)
    with fl_c1:
        st.plotly_chart(make_comparison_bar(w['family_loan'], 'Response',
                        'Family Members Taking Loans', height=300), use_container_width=True)
    with fl_c2:
        fls = w['family_loan_size']
        st.metric('Avg Family Loan (BL \u2192 ML)',
                  f"KES {fls['Baseline'].values[0]:,.0f} \u2192 {fls['Midline'].values[0]:,.0f}")


def render_women_tab4(w):
    """Tab 4: Roles, Time Use & Decision-Making."""
    st.markdown("""<div class="section-narrative">
    <strong>Roles & Decisions:</strong> Gender norms around household roles (who SHOULD do tasks vs.
    who DOES them), women's time use across categories (unpaid care, productive work, community
    conservation, personal development), and decision-making power including influence on key
    household decisions.
    </div>""", unsafe_allow_html=True)

    _quick_nav_pills(['Roles & Norms', 'Time Use', 'Decision-Making'])

    _section_header('', 'Roles & Responsibilities: Norms vs. Experience', 'Section F')
    st.plotly_chart(make_two_col_bar(w['roles_should_joint'], w['roles_does_joint'],
                    'Should be Joint', 'Actually Joint', 'Role',
                    'Roles: Should be Joint vs. Actually Joint', height=500), use_container_width=True)

    # Women-only norms vs actual practice
    rw_c1, rw_c2 = st.columns(2)
    with rw_c1:
        st.plotly_chart(make_comparison_bar(w['roles_should_women'], 'Role',
                        'Roles: Should be Done by Women Only (Norms)', height=500, orientation='h'),
                        use_container_width=True)
    with rw_c2:
        st.plotly_chart(make_comparison_bar(w['roles_does_self'], 'Role',
                        'Roles: Actually Done by Self/Spouse', height=500, orientation='h'),
                        use_container_width=True)

    st.markdown('---')
    _section_header('', 'Time Use (Average Hours per Day)', 'Section G')
    ts = w['time_summary']
    fig_time = go.Figure()
    fig_time.add_trace(go.Bar(x=ts['Category'], y=ts['Baseline'], name='Baseline',
                              marker_color=COLORS['baseline'],
                              text=ts['Baseline'].apply(lambda x: f"{x:.1f}h"), textposition='auto'))
    fig_time.add_trace(go.Bar(x=ts['Category'], y=ts['Midline'], name='Midline',
                              marker_color=COLORS['midline'],
                              text=ts['Midline'].apply(lambda x: f"{x:.1f}h"), textposition='auto'))
    fig_time.update_layout(title="Average Hours per Day by Activity Category",
                           barmode='group', height=450, yaxis_title='Hours',
                           legend=dict(orientation='h', yanchor='bottom', y=1.02),
                           font=dict(size=13, color='#333'), title_font=dict(size=16, color='#222'),
                           plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)')
    st.plotly_chart(fig_time, use_container_width=True)

    c1, c2 = st.columns(2)
    with c1:
        tu = w['time_unpaid']
        fig_uc = go.Figure()
        fig_uc.add_trace(go.Bar(y=tu['Activity'], x=tu['Hours_BL'], name='Baseline',
                                orientation='h', marker_color=COLORS['baseline']))
        fig_uc.add_trace(go.Bar(y=tu['Activity'], x=tu['Hours_ML'], name='Midline',
                                orientation='h', marker_color=COLORS['midline']))
        fig_uc.update_layout(title='Unpaid Care Work (Hours/Day)', barmode='group', height=350,
                             xaxis_title='Hours', legend=dict(orientation='h', yanchor='bottom', y=1.02),
                             font=dict(size=13, color='#333'), title_font=dict(size=16, color='#222'),
                             plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)')
        st.plotly_chart(fig_uc, use_container_width=True)
    with c2:
        tp = w['time_productive']
        fig_pw = go.Figure()
        fig_pw.add_trace(go.Bar(y=tp['Activity'], x=tp['Hours_BL'], name='Baseline',
                                orientation='h', marker_color=COLORS['baseline']))
        fig_pw.add_trace(go.Bar(y=tp['Activity'], x=tp['Hours_ML'], name='Midline',
                                orientation='h', marker_color=COLORS['midline']))
        fig_pw.update_layout(title='Productive Work (Hours/Day)', barmode='group', height=350,
                             xaxis_title='Hours', legend=dict(orientation='h', yanchor='bottom', y=1.02),
                             font=dict(size=13, color='#333'), title_font=dict(size=16, color='#222'),
                             plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)')
        st.plotly_chart(fig_pw, use_container_width=True)

    # Community conservation time breakdown
    tc = w['time_conservation']
    fig_tc = go.Figure()
    fig_tc.add_trace(go.Bar(y=tc['Activity'], x=tc['Hours_BL'], name='Baseline',
                            orientation='h', marker_color=COLORS['baseline']))
    fig_tc.add_trace(go.Bar(y=tc['Activity'], x=tc['Hours_ML'], name='Midline',
                            orientation='h', marker_color=COLORS['midline']))
    fig_tc.update_layout(title='Community Conservation Work (Hours/Day)', barmode='group', height=300,
                         xaxis_title='Hours', legend=dict(orientation='h', yanchor='bottom', y=1.02),
                         font=dict(size=13, color='#333'), title_font=dict(size=16, color='#222'),
                         plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)')
    st.plotly_chart(fig_tc, use_container_width=True)

    # Total hours summary metrics
    tt_c1, tt_c2, tt_c3 = st.columns(3)
    tut = w['time_unpaid_total']
    tpt = w['time_productive_total']
    tct = w['time_conservation_total']
    with tt_c1:
        st.metric('Unpaid Care (BL / ML)',
                  f"{tut['Baseline'].values[0]:.1f}h / {tut['Midline'].values[0]:.1f}h")
    with tt_c2:
        st.metric('Productive (BL / ML)',
                  f"{tpt['Baseline'].values[0]:.1f}h / {tpt['Midline'].values[0]:.1f}h")
    with tt_c3:
        st.metric('Conservation (BL / ML)',
                  f"{tct['Baseline'].values[0]:.1f}h / {tct['Midline'].values[0]:.1f}h")

    st.markdown('---')
    _section_header('', 'Decision-Making', 'Section H')
    st.plotly_chart(make_two_col_bar(w['decision_should_joint'], w['decision_does_joint'],
                    'Should be Joint', 'Actually Joint', 'Decision',
                    'Decision-Making: Norms vs. Experience (Joint %)', height=550), use_container_width=True)

    # Women-only decision norms vs actual
    dw_c1, dw_c2 = st.columns(2)
    with dw_c1:
        st.plotly_chart(make_comparison_bar(w['decision_should_women'], 'Decision',
                        'Should be Decided by Women Only', height=550, orientation='h'),
                        use_container_width=True)
    with dw_c2:
        st.plotly_chart(make_comparison_bar(w['decision_does_self'], 'Decision',
                        'Actually Decided by Self/Spouse', height=550, orientation='h'),
                        use_container_width=True)

    _section_header('', "Influence on HH Decisions ('To a Large Extent')", 'Section H')
    st.plotly_chart(make_comparison_bar(w['decision_influence'], 'Decision',
                    'Women Who Can Influence Decisions to a Large Extent',
                    height=500, orientation='h'), use_container_width=True)


def render_women_tab5(w):
    """Tab 5: Climate Change & NbS."""
    st.markdown("""<div class="section-narrative">
    <strong>Climate & Nature-based Solutions:</strong> Women's awareness of climate change, ability to
    define it, perceived effects on livelihoods and environment, knowledge of Nature-based Solutions
    (NbS), and participation in mangrove restoration, seaweed farming, and forest management.
    </div>""", unsafe_allow_html=True)

    _quick_nav_pills(['Climate Awareness', 'NbS Knowledge', 'Mangrove', 'Seaweed', 'Forest'])

    c1, c2 = st.columns(2)
    with c1:
        st.plotly_chart(make_comparison_bar(w['cc_heard'], 'Response',
                        'Heard of Climate Change', height=300), use_container_width=True)
    with c2:
        st.plotly_chart(make_comparison_bar(w['cc_define'], 'Response',
                        'Ability to Define Climate Change', height=350), use_container_width=True)

    c3, c4 = st.columns(2)
    with c3:
        st.plotly_chart(make_comparison_bar(w['cc_env_effects'], 'Effect',
                        'Perceived Environmental Effects of CC', height=500, orientation='h'),
                        use_container_width=True)
    with c4:
        st.plotly_chart(make_comparison_bar(w['cc_livelihood_effects'], 'Effect',
                        'Perceived Livelihood Effects of CC', height=500, orientation='h'),
                        use_container_width=True)

    st.markdown("---")
    _section_header('', 'Nature-based Solutions', 'Section H')
    c5, c6 = st.columns(2)
    with c5:
        st.plotly_chart(make_comparison_bar(w['nbs_heard'], 'Response',
                        'Heard of NbS', height=300), use_container_width=True)
    with c6:
        st.plotly_chart(make_comparison_bar(w['nbs_define'], 'Response',
                        'Ability to Define NbS', height=350), use_container_width=True)

    c7, c8 = st.columns(2)
    with c7:
        st.plotly_chart(make_comparison_bar(w['nbs_examples'], 'Example',
                        'NbS Examples Women Can Cite', height=450, orientation='h'), use_container_width=True)
    with c8:
        st.plotly_chart(make_comparison_bar(w['nbs_benefits'], 'Benefit',
                        'Perceived Benefits of NbS', height=500, orientation='h'), use_container_width=True)

    st.markdown("---")
    # NbS Modules
    for module, prefix, icon in [('Mangrove Restoration', 'mangrove', ''),
                                  ('Seaweed Farming', 'seaweed', ''),
                                  ('Forest Management', 'forest', '')]:
        _section_header(icon, module, 'NbS Module')
        ca, cb = st.columns(2)
        with ca:
            st.plotly_chart(make_comparison_bar(w[f'{prefix}_heard'], 'Response',
                            f'Awareness of {module}', height=300), use_container_width=True)
            st.plotly_chart(make_comparison_bar(w[f'{prefix}_current'], 'Response',
                            f'Currently Involved', height=280), use_container_width=True)
        with cb:
            st.plotly_chart(make_comparison_bar(w[f'{prefix}_ever'], 'Response',
                            f'Ever Participated', height=300), use_container_width=True)
            st.plotly_chart(make_comparison_bar(w[f'{prefix}_modality'], 'Modality',
                            f'Participation Modality', height=300), use_container_width=True)

        cc, cd = st.columns(2)
        with cc:
            st.plotly_chart(make_comparison_bar(w[f'{prefix}_training'], 'Response',
                            f'Training Received', height=300), use_container_width=True)
        with cd:
            st.plotly_chart(make_comparison_bar(w[f'{prefix}_assets'], 'Response',
                            f'Assets/Inputs Received', height=300), use_container_width=True)

        if f'{prefix}_interest' in w:
            st.plotly_chart(make_comparison_bar(w[f'{prefix}_interest'], 'Response',
                            f'Interest in Participating', height=280), use_container_width=True)
        st.markdown("---")


def render_women_tab6(w):
    """Tab 6: Life Skills & Social Norms."""
    st.markdown("""<div class="section-narrative">
    <strong>Life Skills & Social Norms:</strong> Women's self-assessed life skills (self-esteem,
    aspirations, leadership, communication, conflict resolution) and prevailing social norms around
    gender roles, economic control, domestic work, ecosystem participation, and emotional expression.
    </div>""", unsafe_allow_html=True)

    _quick_nav_pills(['Life Skills Radar', 'Communication', 'Social Norms'])

    # Life skills radar
    _section_header('', 'Life Skills — Agree/Strongly Agree', 'Section I')
    ls_a = w['lifeskills_agree'].copy()
    ls_s = w['lifeskills_strong'].copy()

    # Aggregate by domain
    ls_domain = ls_a.groupby('Domain')[['Baseline','Midline']].mean().reset_index()
    ls_domain_s = ls_s.groupby('Domain')[['Baseline','Midline']].mean().reset_index()

    fig_radar = go.Figure()
    domains = list(ls_domain['Domain']) + [ls_domain['Domain'].iloc[0]]
    bl_vals = list(ls_domain['Baseline']*100) + [ls_domain['Baseline'].iloc[0]*100]
    ml_vals = list(ls_domain['Midline']*100) + [ls_domain['Midline'].iloc[0]*100]
    fig_radar.add_trace(go.Scatterpolar(r=bl_vals, theta=domains, fill='toself',
                                         name='Baseline', fillcolor=COLORS['radar_bl_fill'],
                                         line=dict(color=COLORS['baseline'])))
    fig_radar.add_trace(go.Scatterpolar(r=ml_vals, theta=domains, fill='toself',
                                         name='Midline', fillcolor=COLORS['radar_ml_fill'],
                                         line=dict(color=COLORS['midline'])))
    fig_radar.update_layout(title='Life Skills by Domain (Agree/SA %)',
                            polar=dict(radialaxis=dict(visible=True, range=[0, 100])),
                            height=450, legend=dict(orientation='h', yanchor='bottom', y=-0.15),
                            font=dict(size=13, color='#333'), title_font=dict(size=16, color='#222'))
    st.plotly_chart(fig_radar, use_container_width=True)

    # Detailed bars
    c1, c2 = st.columns(2)
    with c1:
        st.plotly_chart(make_comparison_bar(ls_a, 'Statement',
                        'Life Skills — Agree/SA %', height=550, orientation='h'), use_container_width=True)
    with c2:
        st.plotly_chart(make_comparison_bar(ls_s, 'Statement',
                        'Life Skills — Strongly Agree %', height=550, orientation='h'), use_container_width=True)

    st.markdown("---")
    _section_header('', 'Communication & Conflict Resolution', 'Section I')
    c3, c4 = st.columns(2)
    with c3:
        st.plotly_chart(make_comparison_bar(w['communication_agree'], 'Statement',
                        'Communication — Agree/SA', height=350, orientation='h'), use_container_width=True)
    with c4:
        st.plotly_chart(make_comparison_bar(w['conflict_agree'], 'Statement',
                        'Conflict Resolution — Agree/SA', height=350, orientation='h'), use_container_width=True)

    c3b, c4b = st.columns(2)
    with c3b:
        st.plotly_chart(make_comparison_bar(w['communication_strong'], 'Statement',
                        'Communication — Strongly Agree Only', height=350, orientation='h'),
                        use_container_width=True)
    with c4b:
        st.plotly_chart(make_comparison_bar(w['conflict_strong'], 'Statement',
                        'Conflict Resolution — Strongly Agree Only', height=350, orientation='h'),
                        use_container_width=True)

    st.markdown("---")
    _section_header('', 'Social Norms', 'Section J')
    c5, c6 = st.columns(2)
    with c5:
        st.plotly_chart(make_comparison_bar(w['socialnorms_agree'], 'Norm',
                        'Social Norms — Agree/SA %', height=500, orientation='h'), use_container_width=True)
    with c6:
        st.plotly_chart(make_comparison_bar(w['socialnorms_strong'], 'Norm',
                        'Social Norms — Strongly Agree %', height=500, orientation='h'), use_container_width=True)


# ============================================================================
# FORESTRY TAB RENDERERS (extracted from forestry_dashboard.py main())
# ============================================================================

def render_forestry_tabs(data, show_change):
    """Render all 6 forestry tabs exactly as the original forestry_dashboard.py."""
    tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
        " Overview", " Group Characteristics", " Governance & Gender",
        " Training & Assets", " Forest Condition & Threats", " Income & Agroforestry"
    ])

    # ---- TAB 1: OVERVIEW ----
    with tab1:
        st.markdown("""<div class="section-narrative">
        <strong>Overview:</strong> Key performance indicators for Community Forestry Conservation Groups,
        comparing baseline and midline. Functionality thresholds and domain scores (Management, Gender, Effectiveness).
        </div>""", unsafe_allow_html=True)

        ft = data['functionality_threshold']
        fs = data['functionality_scores']
        fd = data['functionality_domain']
        bl_60 = ft.loc[ft['Timepoint']=='Baseline','Functional_60_pct'].values[0]
        ml_60 = ft.loc[ft['Timepoint']=='Midline','Functional_60_pct'].values[0]
        bl_70 = ft.loc[ft['Timepoint']=='Baseline','Functional_70_pct'].values[0]
        ml_70 = ft.loc[ft['Timepoint']=='Midline','Functional_70_pct'].values[0]
        bl_avg = fs.loc[fs['Timepoint']=='Baseline','Average'].values[0]
        ml_avg = fs.loc[fs['Timepoint']=='Midline','Average'].values[0]
        bl_grp = data['num_groups'].loc[data['num_groups']['Timepoint']=='Baseline','Groups_Assessed'].values[0]
        ml_grp = data['num_groups'].loc[data['num_groups']['Timepoint']=='Midline','Groups_Assessed'].values[0]

        c1,c2,c3,c4 = st.columns(4)
        with c1: st.metric("Functional ≥60%", f"{ml_60*100:.1f}%", delta=f"{(ml_60-bl_60)*100:+.1f}pp")
        with c2: st.metric("Functional ≥70%", f"{ml_70*100:.1f}%", delta=f"{(ml_70-bl_70)*100:+.1f}pp")
        with c3: st.metric("Overall Score", f"{ml_avg:.1f}/100", delta=f"{ml_avg-bl_avg:+.1f} pts")
        with c4: st.metric("Groups Assessed", f"{int(ml_grp)}", delta=f"{int(ml_grp-bl_grp)}")
        st.markdown("---")

        col_l, col_r = st.columns([3,2])
        with col_l:
            domain_df = pd.DataFrame({
                'Domain': ['Management','Gender','Effectiveness','Overall'],
                'Baseline': fd[['Management','Gender','Effectiveness','Overall']].iloc[0].values,
                'Midline': fd[['Management','Gender','Effectiveness','Overall']].iloc[1].values,
            })
            fig = go.Figure()
            fig.add_trace(go.Bar(x=domain_df['Domain'], y=domain_df['Baseline'], name='Baseline',
                                 marker_color=COLORS['baseline'],
                                 text=domain_df['Baseline'].apply(lambda x: f"{x:.1f}"), textposition='auto'))
            fig.add_trace(go.Bar(x=domain_df['Domain'], y=domain_df['Midline'], name='Midline',
                                 marker_color=COLORS['midline'],
                                 text=domain_df['Midline'].apply(lambda x: f"{x:.1f}"), textposition='auto'))
            fig.update_layout(title="Functionality Scores by Domain (0-100)", barmode='group', height=450,
                              yaxis=dict(title='Score', range=[0,105]),
                              legend=dict(orientation='h', yanchor='bottom', y=1.02),
                              font=dict(size=13, color='#333'), title_font=dict(size=16, color='#222'),
                              plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)')
            st.plotly_chart(fig, use_container_width=True)
        with col_r:
            fig_r = go.Figure()
            fig_r.add_trace(go.Scatterpolar(
                r=list(domain_df['Baseline'])+[domain_df['Baseline'].iloc[0]],
                theta=list(domain_df['Domain'])+[domain_df['Domain'].iloc[0]],
                fill='toself', name='Baseline',
                fillcolor=COLORS['radar_bl_fill'], line=dict(color=COLORS['baseline'])))
            fig_r.add_trace(go.Scatterpolar(
                r=list(domain_df['Midline'])+[domain_df['Midline'].iloc[0]],
                theta=list(domain_df['Domain'])+[domain_df['Domain'].iloc[0]],
                fill='toself', name='Midline',
                fillcolor=COLORS['radar_ml_fill'], line=dict(color=COLORS['midline'])))
            fig_r.update_layout(title="Domain Performance Radar",
                                polar=dict(radialaxis=dict(visible=True, range=[0,100])),
                                height=450, legend=dict(orientation='h', yanchor='bottom', y=-0.15),
                                font=dict(size=13, color='#333'), title_font=dict(size=16, color='#222'))
            st.plotly_chart(fig_r, use_container_width=True)

        st.subheader("Score Statistics")
        col_s1, col_s2 = st.columns(2)
        with col_s1:
            s = fs[fs['Timepoint']=='Baseline'].iloc[0]
            st.markdown(f"**Baseline** — N={int(s['Respondents'])} | Avg={s['Average']:.1f} | Max={s['Max']:.1f} | Min={s['Min']:.1f}")
        with col_s2:
            s = fs[fs['Timepoint']=='Midline'].iloc[0]
            st.markdown(f"**Midline** — N={int(s['Respondents'])} | Avg={s['Average']:.1f} | Max={s['Max']:.1f} | Min={s['Min']:.1f}")

    # ---- TAB 2: GROUP CHARACTERISTICS ----
    with tab2:
        st.markdown("""<div class="section-narrative">
        <strong>Group Characteristics:</strong> Membership composition, years of operation, land tenure, and forest size.
        </div>""", unsafe_allow_html=True)

        # Summary statistics metric cards
        ys = data['years_stats']
        fss = data['forest_size_stats']
        sc1, sc2, sc3, sc4, sc5, sc6 = st.columns(6)
        with sc1: st.metric('Avg Years (BL)', f"{ys.loc[ys['Stat']=='Average','Baseline'].values[0]:.1f}")
        with sc2: st.metric('Avg Years (ML)', f"{ys.loc[ys['Stat']=='Average','Midline'].values[0]:.1f}")
        with sc3: st.metric('Max Years (ML)', f"{ys.loc[ys['Stat']=='Maximum','Midline'].values[0]:.0f}")
        with sc4: st.metric('Avg Forest ha (BL)', f"{fss.loc[fss['Stat']=='Average','Baseline'].values[0]:.1f}")
        with sc5: st.metric('Avg Forest ha (ML)', f"{fss.loc[fss['Stat']=='Average','Midline'].values[0]:.1f}")
        with sc6: st.metric('Max Forest ha (ML)', f"{fss.loc[fss['Stat']=='Maximum','Midline'].values[0]:.0f}")
        st.markdown('---')

        c1,c2,c3 = st.columns(3)
        with c1:
            st.plotly_chart(make_comparison_bar(data['avg_membership'],'Category','Avg Members',
                            y_label='Members', multiply=False, height=350), use_container_width=True)
        with c2:
            st.plotly_chart(make_comparison_bar(data['max_membership'],'Category','Max Members',
                            y_label='Members', multiply=False, height=350), use_container_width=True)
        with c3:
            st.plotly_chart(make_comparison_bar(data['min_membership'],'Category','Min Members',
                            y_label='Members', multiply=False, height=350), use_container_width=True)
        st.markdown("---")
        c4,c5,c6 = st.columns(3)
        with c4:
            fn = make_delta_bar if show_change else make_comparison_bar
            kw = {'title':'Change in Years Dist'} if show_change else {'title':'Years of Operation', 'height':400}
            st.plotly_chart(fn(data['years_dist'],'Category',**kw), use_container_width=True)
        with c5:
            fn = make_delta_bar if show_change else make_comparison_bar
            kw = {'title':'Change in Tenure'} if show_change else {'title':'Land Tenure Type', 'height':400}
            st.plotly_chart(fn(data['tenure'],'Category',**kw), use_container_width=True)
        with c6:
            fn = make_delta_bar if show_change else make_comparison_bar
            kw = {'title':'Change in Forest Size'} if show_change else {'title':'Forest Size (ha)', 'height':400}
            st.plotly_chart(fn(data['forest_size_dist'],'Category',**kw), use_container_width=True)

    # ---- TAB 3: GOVERNANCE & GENDER ----
    with tab3:
        st.markdown("""<div class="section-narrative">
        <strong>Governance & Gender:</strong> Board clarity, guidelines, meetings, women's leadership,
        management practices, gender equality discussions and actions.
        </div>""", unsafe_allow_html=True)
        c1,c2 = st.columns(2)
        with c1: st.plotly_chart(make_comparison_bar(data['goals_defined'],'Category','Defined Goals', height=350), use_container_width=True)
        with c2: st.plotly_chart(make_comparison_bar(data['goals_stated'],'Goal','Stated Goals', height=350, orientation='h'), use_container_width=True)
        c3,c4 = st.columns(2)
        with c3: st.plotly_chart(make_comparison_bar(data['rights'],'Right','Right Entitlements', height=400, orientation='h'), use_container_width=True)
        with c4: st.plotly_chart(make_comparison_bar(data['responsibilities'],'Responsibility','Govt Responsibilities', height=400, orientation='h'), use_container_width=True)
        st.markdown("---")
        c5,c6 = st.columns(2)
        with c5: st.plotly_chart(make_comparison_bar(data['board_roles'],'Category','Board Role Clarity', height=380, orientation='h'), use_container_width=True)
        with c6: st.plotly_chart(make_comparison_bar(data['guidelines'],'Category','Guidelines Status', height=380, orientation='h'), use_container_width=True)
        c7,c8 = st.columns(2)
        with c7: st.plotly_chart(make_comparison_bar(data['meetings'],'Category','Meeting Frequency', height=350, orientation='h'), use_container_width=True)
        with c8: st.plotly_chart(make_comparison_bar(data['women_leadership'],'Category',"Women's Leadership", height=350, orientation='h'), use_container_width=True)

        # Management practices
        mp = data['mgmt_practices']
        fig_mp = go.Figure()
        fig_mp.add_trace(go.Bar(y=mp['Practice'], x=mp['Agree_Baseline']*100, name='Agree+SA (BL)', orientation='h', marker_color=COLORS['baseline']))
        fig_mp.add_trace(go.Bar(y=mp['Practice'], x=mp['Agree_Midline']*100, name='Agree+SA (ML)', orientation='h', marker_color=COLORS['midline']))
        fig_mp.add_trace(go.Bar(y=mp['Practice'], x=mp['StronglyAgree_Baseline']*100, name='SA (BL)', orientation='h', marker_color=COLORS['baseline_light']))
        fig_mp.add_trace(go.Bar(y=mp['Practice'], x=mp['StronglyAgree_Midline']*100, name='SA (ML)', orientation='h', marker_color=COLORS['midline_light']))
        fig_mp.update_layout(barmode='group', height=500, xaxis_title="%",
                             legend=dict(orientation='h', yanchor='bottom', y=1.02),
                             font=dict(size=13, color='#333'), title_font=dict(size=16, color='#222'),
                             plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)')
        st.plotly_chart(fig_mp, use_container_width=True)
        st.markdown("---")
        c9,c10 = st.columns(2)
        with c9:
            st.plotly_chart(make_comparison_bar(data['ge_discussion'],'Category','GE Discussion', height=350), use_container_width=True)
            st.plotly_chart(make_comparison_bar(data['ge_actions'],'Category','GE Actions Agreed', height=350), use_container_width=True)
        with c10:
            st.plotly_chart(make_comparison_bar(data['ge_topics'],'Topic','GE Topics', height=350, orientation='h'), use_container_width=True)
            st.plotly_chart(make_comparison_bar(data['ge_completion'],'Category','GE Action Completion', height=350), use_container_width=True)

    # ---- TAB 4: TRAINING & ASSETS ----
    with tab4:
        st.markdown("""<div class="section-narrative">
        <strong>Training & Assets:</strong> Training coverage, topics, and assets/inputs received.
        </div>""", unsafe_allow_html=True)
        c1,c2 = st.columns(2)
        with c1: st.plotly_chart(make_comparison_bar(data['training_coverage'],'Category','Training Coverage', height=400), use_container_width=True)
        with c2:
            tc = data['training_coverage'].copy(); tc['Baseline']*=100; tc['Midline']*=100
            fig_tc = go.Figure()
            for i,(_,row) in enumerate(tc.iterrows()):
                fig_tc.add_trace(go.Bar(x=['Baseline','Midline'], y=[row['Baseline'],row['Midline']],
                                        name=row['Category'], marker_color=['#E53935','#FB8C00','#FDD835','#43A047'][i],
                                        text=[f"{row['Baseline']:.1f}%",f"{row['Midline']:.1f}%"], textposition='inside'))
            fig_tc.update_layout(title="Coverage (Stacked)", barmode='stack', height=400, yaxis_title="%",
                                 legend=dict(orientation='h', yanchor='bottom', y=1.02),
                                 font=dict(size=13, color='#333'), title_font=dict(size=16, color='#222'),
                                 plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)')
            st.plotly_chart(fig_tc, use_container_width=True)

        if show_change:
            st.plotly_chart(make_delta_bar(data['training_topics'],'Topic','Training Topics Change (pp)', height=500), use_container_width=True)
        else:
            st.plotly_chart(make_comparison_bar(data['training_topics'],'Topic','Training Topics', height=600, orientation='h'), use_container_width=True)
        st.markdown("---")
        c3,c4 = st.columns(2)
        with c3: st.plotly_chart(make_comparison_bar(data['assets_received'],'Category','Assets Received', height=350), use_container_width=True)
        with c4: st.plotly_chart(make_comparison_bar(data['asset_types'],'Asset','Asset Types', height=400, orientation='h'), use_container_width=True)

    # ---- TAB 5: FOREST CONDITION & THREATS ----
    with tab5:
        st.markdown("""<div class="section-narrative">
        <strong>Forest Condition & Threats:</strong> Perceived condition, changes, threat levels, harvesting patterns.
        </div>""", unsafe_allow_html=True)
        tp = st.radio("Timepoint", ["Baseline","Midline"], horizontal=True, key="fc_tp")
        fc = data['forest_condition']
        cols = [f'{tp}_Good',f'{tp}_Medium',f'{tp}_Poor']
        fig = make_stacked_bar(fc,'Characteristic', cols, [COLORS['good'],COLORS['medium'],COLORS['poor']],
                               f'Forest Condition — {tp}', height=400)
        for i,l in enumerate(['Good','Medium','Poor']): fig.data[i].name = l
        st.plotly_chart(fig, use_container_width=True)
        st.markdown("---")
        tp2 = st.radio("Timepoint", ["Baseline","Midline"], horizontal=True, key="fch_tp")
        fch = data['forest_change']
        cols2 = [f'{tp2}_Decrease',f'{tp2}_NoChange',f'{tp2}_Increase']
        fig2 = make_stacked_bar(fch,'Characteristic', cols2, [COLORS['danger'],COLORS['no_change'],COLORS['good']],
                                f'Perceived Changes — {tp2}', height=400)
        for i,l in enumerate(['Decrease','No Change','Increase']): fig2.data[i].name = l
        st.plotly_chart(fig2, use_container_width=True)
        st.markdown("---")
        c1,c2 = st.columns(2)
        with c1:
            fig_t = make_stacked_bar(data['threats'],'Threat', ['Baseline_Low','Baseline_Medium','Baseline_High'],
                                     [COLORS['good'],COLORS['medium'],COLORS['danger']], 'Threats — Baseline', height=450, orientation='h')
            for i,l in enumerate(['Low','Medium','High']): fig_t.data[i].name = l
            st.plotly_chart(fig_t, use_container_width=True)
        with c2:
            fig_t2 = make_stacked_bar(data['threats'],'Threat', ['Midline_Low','Midline_Medium','Midline_High'],
                                      [COLORS['good'],COLORS['medium'],COLORS['danger']], 'Threats — Midline', height=450, orientation='h')
            for i,l in enumerate(['Low','Medium','High']): fig_t2.data[i].name = l
            st.plotly_chart(fig_t2, use_container_width=True)

        # Threat change trends
        st.markdown('---')
        _section_header('', 'Threat Change Trends', 'Section E')
        tc1, tc2 = st.columns(2)
        with tc1:
            fig_tc = make_stacked_bar(data['threat_changes'], 'Threat',
                                      ['Baseline_Decrease','Baseline_NoChange','Baseline_Increase'],
                                      [COLORS['good'],'#9E9E9E',COLORS['danger']],
                                      'Threat Changes — Baseline', height=450, orientation='h')
            for i,l in enumerate(['Decrease','No Change','Increase']): fig_tc.data[i].name = l
            st.plotly_chart(fig_tc, use_container_width=True)
        with tc2:
            fig_tc2 = make_stacked_bar(data['threat_changes'], 'Threat',
                                       ['Midline_Decrease','Midline_NoChange','Midline_Increase'],
                                       [COLORS['good'],'#9E9E9E',COLORS['danger']],
                                       'Threat Changes — Midline', height=450, orientation='h')
            for i,l in enumerate(['Decrease','No Change','Increase']): fig_tc2.data[i].name = l
            st.plotly_chart(fig_tc2, use_container_width=True)

        # Harvest amounts
        st.markdown('---')
        _section_header('', 'Harvest Amounts by Product', 'Section E')
        hc1, hc2 = st.columns(2)
        with hc1:
            fig_ha = make_stacked_bar(data['harvest_amount'], 'Product',
                                      ['Baseline_None','Baseline_Medium','Baseline_Substantial'],
                                      ['#9E9E9E',COLORS['medium'],COLORS['good']],
                                      'Harvest Amount — Baseline', height=450, orientation='h')
            for i,l in enumerate(['None','Medium','Substantial']): fig_ha.data[i].name = l
            st.plotly_chart(fig_ha, use_container_width=True)
        with hc2:
            fig_ha2 = make_stacked_bar(data['harvest_amount'], 'Product',
                                       ['Midline_None','Midline_Medium','Midline_Substantial'],
                                       ['#9E9E9E',COLORS['medium'],COLORS['good']],
                                       'Harvest Amount — Midline', height=450, orientation='h')
            for i,l in enumerate(['None','Medium','Substantial']): fig_ha2.data[i].name = l
            st.plotly_chart(fig_ha2, use_container_width=True)

        # Harvest change trends
        _section_header('', 'Harvest Change Trends', 'Section E')
        ht1, ht2 = st.columns(2)
        with ht1:
            fig_hc = make_stacked_bar(data['harvest_changes'], 'Product',
                                      ['Baseline_Decrease','Baseline_NoChange','Baseline_Increase'],
                                      [COLORS['danger'],'#9E9E9E',COLORS['good']],
                                      'Harvest Changes — Baseline', height=450, orientation='h')
            for i,l in enumerate(['Decrease','No Change','Increase']): fig_hc.data[i].name = l
            st.plotly_chart(fig_hc, use_container_width=True)
        with ht2:
            fig_hc2 = make_stacked_bar(data['harvest_changes'], 'Product',
                                       ['Midline_Decrease','Midline_NoChange','Midline_Increase'],
                                       [COLORS['danger'],'#9E9E9E',COLORS['good']],
                                       'Harvest Changes — Midline', height=450, orientation='h')
            for i,l in enumerate(['Decrease','No Change','Increase']): fig_hc2.data[i].name = l
            st.plotly_chart(fig_hc2, use_container_width=True)

    # ---- TAB 6: INCOME & AGROFORESTRY ----
    with tab6:
        st.markdown("""<div class="section-narrative">
        <strong>Income & Agroforestry:</strong> Income generation, sources, uses, agroforestry practices, support, and challenges.
        </div>""", unsafe_allow_html=True)
        c1,c2,c3 = st.columns(3)
        with c1: st.plotly_chart(make_comparison_bar(data['income_gen'],'Category','Income Generation', height=350), use_container_width=True)
        with c2: st.plotly_chart(make_comparison_bar(data['income_sources'],'Source','Income Sources', height=400, orientation='h'), use_container_width=True)
        with c3: st.plotly_chart(make_comparison_bar(data['income_use'],'Use','Income Use', height=400, orientation='h'), use_container_width=True)
        st.markdown("---")
        c4,c5 = st.columns(2)
        with c4: st.plotly_chart(make_comparison_bar(data['agroforestry'],'Category','Agroforestry Practice', height=350), use_container_width=True)
        with c5: st.plotly_chart(make_comparison_bar(data['af_types'],'Practice','AF Types', height=350, orientation='h'), use_container_width=True)
        c6,c7 = st.columns(2)
        with c6: st.plotly_chart(make_comparison_bar(data['af_objectives'],'Objective','AF Objectives', height=400, orientation='h'), use_container_width=True)
        with c7: st.plotly_chart(make_comparison_bar(data['af_training'],'Category','AF Training', height=350), use_container_width=True)
        st.markdown("---")
        c8,c9 = st.columns(2)
        with c8:
            fn = make_delta_bar if show_change else make_comparison_bar
            kw = {'title':'Change in AF Support (pp)'} if show_change else {'title':'AF Support Types', 'height':380, 'orientation':'h'}
            st.plotly_chart(fn(data['af_support'],'Support',**kw), use_container_width=True)
        with c9:
            fn = make_delta_bar if show_change else make_comparison_bar
            kw = {'title':'Change in AF Challenges (pp)'} if show_change else {'title':'AF Challenges', 'height':380, 'orientation':'h'}
            st.plotly_chart(fn(data['af_challenges'],'Challenge',**kw), use_container_width=True)
        c10,c11 = st.columns(2)
        with c10: st.plotly_chart(make_comparison_bar(data['af_reinvest'],'Category','AF Reinvestment', height=380, orientation='h'), use_container_width=True)
        with c11: st.plotly_chart(make_comparison_bar(data['af_potential'],'Category','Livelihood Potential', height=350), use_container_width=True)


# ============================================================================
# MEN SURVEY TAB RENDERERS
# ============================================================================

def render_men_tab1(m):
    """Tab 1: Household Characteristics."""
    st.markdown("""<div class="section-narrative">
    <strong>Household Profile:</strong> Demographics of surveyed men — location type (Marine vs
    Terrestrial), education levels, and main household economic activities such as agriculture,
    fishing, and small business.
    </div>""", unsafe_allow_html=True)

    _quick_nav_pills(['Demographics', 'Economic Activities'])

    c1, c2 = st.columns(2)
    with c1:
        st.plotly_chart(make_comparison_bar(m['location'], 'Category',
                        'Men by Location Type', height=300), use_container_width=True)
    with c2:
        st.plotly_chart(make_comparison_bar(m['education'], 'Category',
                        'Education Level', height=350), use_container_width=True)

    st.markdown("---")
    _section_header('', 'Economic Activities', 'Section A')
    st.plotly_chart(make_comparison_bar(m['main_econ'], 'Activity',
                    'Main HH Economic Activity', height=500, orientation='h'), use_container_width=True)


def render_men_tab2(m):
    """Tab 2: Climate Change & NbS Knowledge."""
    st.markdown("""<div class="section-narrative">
    <strong>Climate & NbS Awareness:</strong> Men's knowledge of climate change, ability to define it,
    perceived environmental and livelihood effects, awareness of Nature-based Solutions (NbS),
    examples cited, and perceived benefits.
    </div>""", unsafe_allow_html=True)

    _quick_nav_pills(['Climate Awareness', 'CC Effects', 'NbS Knowledge', 'NbS Benefits'])

    c1, c2 = st.columns(2)
    with c1:
        st.plotly_chart(make_comparison_bar(m['cc_heard'], 'Response',
                        'Heard of Climate Change', height=300), use_container_width=True)
    with c2:
        st.plotly_chart(make_comparison_bar(m['cc_define'], 'Response',
                        'Ability to Define Climate Change', height=350), use_container_width=True)

    _section_header('', 'CC Effects', 'Section B')
    c3, c4 = st.columns(2)
    with c3:
        st.plotly_chart(make_comparison_bar(m['cc_env_effects'], 'Effect',
                        'Perceived Environmental Effects of CC', height=500, orientation='h'),
                        use_container_width=True)
    with c4:
        st.plotly_chart(make_comparison_bar(m['cc_livelihood_effects'], 'Effect',
                        'Perceived Livelihood Effects of CC', height=500, orientation='h'),
                        use_container_width=True)

    st.markdown("---")
    _section_header('', 'NbS Knowledge', 'Section B')
    c5, c6 = st.columns(2)
    with c5:
        st.plotly_chart(make_comparison_bar(m['nbs_heard'], 'Response',
                        'Heard of NbS', height=300), use_container_width=True)
    with c6:
        st.plotly_chart(make_comparison_bar(m['nbs_define'], 'Response',
                        'Ability to Define NbS', height=350), use_container_width=True)

    c7, c8 = st.columns(2)
    with c7:
        st.plotly_chart(make_comparison_bar(m['nbs_examples'], 'Example',
                        'NbS Examples Men Can Cite', height=450, orientation='h'), use_container_width=True)
    with c8:
        st.plotly_chart(make_comparison_bar(m['nbs_benefits'], 'Benefit',
                        'Perceived Benefits of NbS', height=500, orientation='h'), use_container_width=True)


def render_men_tab3(m):
    """Tab 3: Support for Women in NbS Participation."""
    st.markdown("""<div class="section-narrative">
    <strong>Support for Women in NbS:</strong> Men's awareness of and support for female household
    members' participation in mangrove restoration, seaweed farming, and forest management.
    Includes types of support provided such as encouragement, household chore support, and
    material assistance.
    </div>""", unsafe_allow_html=True)

    _quick_nav_pills(['Mangrove Restoration', 'Seaweed Farming', 'Forest Management'])

    for module, prefix, icon in [('Mangrove Restoration', 'mangrove', ''),
                                  ('Seaweed Farming', 'seaweed', ''),
                                  ('Forest Management', 'forest', '')]:
        _section_header(icon, module, 'NbS Module')
        ca, cb = st.columns(2)
        with ca:
            st.plotly_chart(make_comparison_bar(m[f'{prefix}_heard'], 'Response',
                            f'Awareness of {module}', height=300), use_container_width=True)
            st.plotly_chart(make_comparison_bar(m[f'{prefix}_current'], 'Response',
                            f'Female HH Currently Involved', height=280), use_container_width=True)
        with cb:
            st.plotly_chart(make_comparison_bar(m[f'{prefix}_ever'], 'Response',
                            f'Female HH Ever Participated', height=300), use_container_width=True)
            st.plotly_chart(make_comparison_bar(m[f'{prefix}_support'], 'Response',
                            f'Men Supporting Female HH Members', height=300), use_container_width=True)

        _section_header('', f'Type of Support — {module}', 'Support')
        st.plotly_chart(make_comparison_bar(m[f'{prefix}_support_type'], 'Type',
                        f'Support Type for {module}', height=400, orientation='h'),
                        use_container_width=True)
        st.markdown("---")


def render_men_tab4(m):
    """Tab 4: Roles, Responsibilities & Time Use."""
    st.markdown("""<div class="section-narrative">
    <strong>Roles & Time Use:</strong> Gender norms around household roles as reported by men —
    who SHOULD do tasks vs. who DOES them (joint, men only, women only). Time use patterns
    across unpaid care, productive work, community conservation, personal development, and leisure.
    </div>""", unsafe_allow_html=True)

    _quick_nav_pills(['Roles: Norms vs Practice', 'Gendered Roles', 'Time Use'])

    _section_header('', 'Roles & Responsibilities: Norms vs Practice', 'Section C')
    st.plotly_chart(make_two_col_bar(m['roles_should_joint'], m['roles_does_joint'],
                    'Should be Joint', 'Actually Joint', 'Role',
                    'Roles: Should be Joint vs. Actually Joint (Men Reporting)', height=500),
                    use_container_width=True)

    _section_header('', 'Gendered Roles', 'Section C')
    rw_c1, rw_c2 = st.columns(2)
    with rw_c1:
        st.plotly_chart(make_comparison_bar(m['roles_should_women'], 'Role',
                        'Should be Done by Women Only', height=500, orientation='h'),
                        use_container_width=True)
    with rw_c2:
        st.plotly_chart(make_comparison_bar(m['roles_should_men'], 'Role',
                        'Should be Done by Men Only', height=500, orientation='h'),
                        use_container_width=True)

    rd_c1, rd_c2 = st.columns(2)
    with rd_c1:
        st.plotly_chart(make_comparison_bar(m['roles_does_men'], 'Role',
                        'Actually Done by Myself (Men)', height=500, orientation='h'),
                        use_container_width=True)
    with rd_c2:
        st.plotly_chart(make_comparison_bar(m['roles_does_women'], 'Role',
                        'Actually Done by My Spouse (Women)', height=500, orientation='h'),
                        use_container_width=True)

    st.markdown('---')
    _section_header('', 'Time Use (Average Hours per Day)', 'Section C')
    ts = m['time_summary']
    fig_time = go.Figure()
    fig_time.add_trace(go.Bar(x=ts['Category'], y=ts['Baseline'], name='Baseline',
                              marker_color=COLORS['baseline'],
                              text=ts['Baseline'].apply(lambda x: f"{x:.1f}h"), textposition='auto'))
    fig_time.add_trace(go.Bar(x=ts['Category'], y=ts['Midline'], name='Midline',
                              marker_color=COLORS['midline'],
                              text=ts['Midline'].apply(lambda x: f"{x:.1f}h"), textposition='auto'))
    fig_time.update_layout(title="Average Hours per Day by Category (Men)",
                           barmode='group', height=450, yaxis_title='Hours',
                           legend=dict(orientation='h', yanchor='bottom', y=1.02),
                           font=dict(size=13, color='#333'), title_font=dict(size=16, color='#222'),
                           plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)')
    st.plotly_chart(fig_time, use_container_width=True)

    c1, c2 = st.columns(2)
    with c1:
        tu = m['time_unpaid']
        fig_uc = go.Figure()
        fig_uc.add_trace(go.Bar(y=tu['Activity'], x=tu['Baseline'], name='Baseline',
                                orientation='h', marker_color=COLORS['baseline']))
        fig_uc.add_trace(go.Bar(y=tu['Activity'], x=tu['Midline'], name='Midline',
                                orientation='h', marker_color=COLORS['midline']))
        fig_uc.update_layout(title='Unpaid Care Work (Hours/Day)', barmode='group', height=350,
                             xaxis_title='Hours', legend=dict(orientation='h', yanchor='bottom', y=1.02),
                             font=dict(size=13, color='#333'), title_font=dict(size=16, color='#222'),
                             plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)')
        st.plotly_chart(fig_uc, use_container_width=True)
    with c2:
        tp = m['time_productive']
        fig_pw = go.Figure()
        fig_pw.add_trace(go.Bar(y=tp['Activity'], x=tp['Baseline'], name='Baseline',
                                orientation='h', marker_color=COLORS['baseline']))
        fig_pw.add_trace(go.Bar(y=tp['Activity'], x=tp['Midline'], name='Midline',
                                orientation='h', marker_color=COLORS['midline']))
        fig_pw.update_layout(title='Productive Work (Hours/Day)', barmode='group', height=350,
                             xaxis_title='Hours', legend=dict(orientation='h', yanchor='bottom', y=1.02),
                             font=dict(size=13, color='#333'), title_font=dict(size=16, color='#222'),
                             plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)')
        st.plotly_chart(fig_pw, use_container_width=True)

    tc = m['time_conservation']
    fig_tc = go.Figure()
    fig_tc.add_trace(go.Bar(y=tc['Activity'], x=tc['Baseline'], name='Baseline',
                            orientation='h', marker_color=COLORS['baseline']))
    fig_tc.add_trace(go.Bar(y=tc['Activity'], x=tc['Midline'], name='Midline',
                            orientation='h', marker_color=COLORS['midline']))
    fig_tc.update_layout(title='Community Conservation Work (Hours/Day)', barmode='group', height=300,
                         xaxis_title='Hours', legend=dict(orientation='h', yanchor='bottom', y=1.02),
                         font=dict(size=13, color='#333'), title_font=dict(size=16, color='#222'),
                         plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)')
    st.plotly_chart(fig_tc, use_container_width=True)

    # Total hours summary metrics
    tt_c1, tt_c2, tt_c3 = st.columns(3)
    tut = m['time_unpaid_total']
    tpt = m['time_productive_total']
    tct = m['time_conservation_total']
    with tt_c1:
        st.metric('Unpaid Care (BL / ML)',
                  f"{tut['Baseline'].values[0]:.1f}h / {tut['Midline'].values[0]:.1f}h")
    with tt_c2:
        st.metric('Productive (BL / ML)',
                  f"{tpt['Baseline'].values[0]:.1f}h / {tpt['Midline'].values[0]:.1f}h")
    with tt_c3:
        st.metric('Conservation (BL / ML)',
                  f"{tct['Baseline'].values[0]:.1f}h / {tct['Midline'].values[0]:.1f}h")


def render_men_tab5(m):
    """Tab 5: Decision-Making."""
    st.markdown("""<div class="section-narrative">
    <strong>Decision-Making:</strong> Men's perspectives on who SHOULD and who DOES make key
    household decisions — joint, men only, or women only. Covers decisions on business, finances,
    education, household purchases, women's mobility and conservation work.
    </div>""", unsafe_allow_html=True)

    _quick_nav_pills(['Norms vs Experience', 'Gendered Decisions'])

    _section_header('', 'Decision-Making: Norms vs Experience', 'Section D')
    st.plotly_chart(make_two_col_bar(m['decision_should_joint'], m['decision_does_joint'],
                    'Should be Joint', 'Actually Joint', 'Decision',
                    'Decision-Making: Norms vs. Experience (Joint %)', height=550),
                    use_container_width=True)

    _section_header('', 'Gendered Decisions', 'Section D')
    dw_c1, dw_c2 = st.columns(2)
    with dw_c1:
        st.plotly_chart(make_comparison_bar(m['decision_should_women'], 'Decision',
                        'Should be Decided by Women Only', height=550, orientation='h'),
                        use_container_width=True)
    with dw_c2:
        st.plotly_chart(make_comparison_bar(m['decision_should_men'], 'Decision',
                        'Should be Decided by Men Only', height=550, orientation='h'),
                        use_container_width=True)

    st.markdown("---")
    _section_header('', 'Actual Decision-Making Practice', 'Section D')
    dm_c1, dm_c2 = st.columns(2)
    with dm_c1:
        st.plotly_chart(make_comparison_bar(m['decision_does_men'], 'Decision',
                        'Actually Decided by Myself (Men)', height=550, orientation='h'),
                        use_container_width=True)
    with dm_c2:
        st.plotly_chart(make_comparison_bar(m['decision_does_women'], 'Decision',
                        'Actually Decided by My Spouse (Women)', height=550, orientation='h'),
                        use_container_width=True)


def render_men_tab6(m):
    """Tab 6: Social Norms."""
    st.markdown("""<div class="section-narrative">
    <strong>Social Norms:</strong> Men's agreement with social norms statements about gender roles,
    economic control, domestic work, ecosystem participation, emotional expression, and attitudes
    toward women's rights and independence. Includes both Agree/StronglyAgree and StronglyAgree
    only proportions.
    </div>""", unsafe_allow_html=True)

    _quick_nav_pills(['Agree/Strongly Agree', 'Strongly Agree Only', 'Change'])

    _section_header('', 'Social Norms — Agree or Strongly Agree', 'Section E')
    st.plotly_chart(make_comparison_bar(m['socialnorms_agree'], 'Norm',
                    'Men Agreeing/Strongly Agreeing with Social Norms',
                    height=600, orientation='h'), use_container_width=True)

    st.markdown("---")
    _section_header('', 'Social Norms — Strongly Agree Only', 'Section E')
    st.plotly_chart(make_comparison_bar(m['socialnorms_strong'], 'Norm',
                    'Men Strongly Agreeing with Social Norms',
                    height=600, orientation='h'), use_container_width=True)

    st.markdown("---")
    _section_header('', 'Change in Social Norms (pp)', 'Section E')
    st.plotly_chart(make_delta_bar(m['socialnorms_agree'], 'Norm',
                    'Change in Agreement (Baseline to Midline)',
                    height=550), use_container_width=True)


# ============================================================================
# GJJ KAP WOMEN — CHART HELPER & TAB RENDERERS
# ============================================================================

def _gjj_bar(df, cat_col, title, y_label="Percentage (%)", multiply=True,
             height=450, orientation='v'):
    """Baseline vs Endline grouped bar chart (uses Endline column instead of Midline)."""
    cb = COLORS["baseline"]
    cm = COLORS["midline"]
    plot_df = df.copy()
    if multiply:
        for c in ['Baseline', 'Endline']:
            plot_df[c] = pd.to_numeric(plot_df[c], errors='coerce').fillna(0) * 100
    if orientation == 'h':
        plot_df['_sort'] = plot_df[['Baseline', 'Endline']].max(axis=1)
        plot_df = plot_df.sort_values('_sort', ascending=True).drop(columns='_sort')
        fig = go.Figure()
        fig.add_trace(go.Bar(y=plot_df[cat_col], x=plot_df['Baseline'], name='Baseline',
                             orientation='h', marker_color=cb,
                             text=plot_df['Baseline'].apply(lambda x: f"{x:.1f}%"), textposition='auto'))
        fig.add_trace(go.Bar(y=plot_df[cat_col], x=plot_df['Endline'], name='Endline',
                             orientation='h', marker_color=cm,
                             text=plot_df['Endline'].apply(lambda x: f"{x:.1f}%"), textposition='auto'))
        fig.update_layout(title=title, barmode='group', height=height,
                          xaxis_title=y_label, legend=dict(orientation='h', yanchor='bottom', y=1.02),
                          yaxis=dict(categoryorder='trace'))
    else:
        fig = go.Figure()
        fig.add_trace(go.Bar(x=plot_df[cat_col], y=plot_df['Baseline'], name='Baseline',
                             marker_color=cb,
                             text=plot_df['Baseline'].apply(lambda x: f"{x:.1f}%"), textposition='auto'))
        fig.add_trace(go.Bar(x=plot_df[cat_col], y=plot_df['Endline'], name='Endline',
                             marker_color=cm,
                             text=plot_df['Endline'].apply(lambda x: f"{x:.1f}%"), textposition='auto'))
        fig.update_layout(title=title, barmode='group', height=height, yaxis_title=y_label,
                          legend=dict(orientation='h', yanchor='bottom', y=1.02))
    fig.update_layout(
        font=dict(size=13, color='#333'),
        title_font=dict(size=16, color='#222'),
        plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
        margin=dict(l=20, r=20, t=60, b=20),
    )
    return fig


def _gjj_delta_bar(df, cat_col, title, multiply=True, height=400):
    """Baseline to Endline change bar (horizontal, green/red)."""
    plot_df = df.copy()
    factor = 100 if multiply else 1
    plot_df['Change'] = (pd.to_numeric(plot_df['Endline'], errors='coerce').fillna(0)
                         - pd.to_numeric(plot_df['Baseline'], errors='coerce').fillna(0)) * factor
    plot_df = plot_df.sort_values('Change')
    colors = [COLORS['good'] if v >= 0 else COLORS['danger'] for v in plot_df['Change']]
    fig = go.Figure()
    fig.add_trace(go.Bar(y=plot_df[cat_col], x=plot_df['Change'], orientation='h',
                         marker_color=colors,
                         text=plot_df['Change'].apply(lambda x: f"{x:+.1f}pp"), textposition='auto'))
    fig.update_layout(title=title, height=height, xaxis_title="Change (pp)",
                      yaxis=dict(categoryorder='trace'),
                      font=dict(size=13, color='#333'),
                      title_font=dict(size=16, color='#222'),
                      plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                      margin=dict(l=20, r=20, t=60, b=20))
    return fig


def _gjj_stacked_bar(df, cat_col, columns, colors_list, title,
                     y_label="Percentage (%)", height=400, multiply=True):
    """Stacked horizontal bar chart for Likert / frequency distributions."""
    plot_df = df.copy()
    if multiply:
        for c in columns:
            plot_df[c] = pd.to_numeric(plot_df[c], errors='coerce').fillna(0) * 100
    fig = go.Figure()
    for col, color in zip(columns, colors_list):
        fig.add_trace(go.Bar(
            y=plot_df[cat_col], x=plot_df[col], name=col, orientation='h',
            marker_color=color,
            text=plot_df[col].apply(lambda x: f"{x:.1f}%" if x > 3 else ""),
            textposition='inside',
        ))
    fig.update_layout(title=title, barmode='stack', height=height,
                      xaxis_title=y_label,
                      legend=dict(orientation='h', yanchor='bottom', y=1.02),
                      yaxis=dict(categoryorder='array', categoryarray=list(plot_df[cat_col])),
                      font=dict(size=13, color='#333'),
                      title_font=dict(size=16, color='#222'),
                      plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                      margin=dict(l=20, r=20, t=60, b=20))
    return fig


def _short_label(s, max_len=55):
    """Truncate long statement text for chart axis readability."""
    if isinstance(s, str) and len(s) > max_len:
        return s[:max_len].rstrip() + '...'
    return s


# ---- Tab 1: SELF — Confidence, Self-Worth & Compassion ----
def render_gjj_tab1(g):
    """GJJ KAP Women — Tab 1: SELF (Self-Esteem, Self-Compassion, Confidence)."""
    st.markdown("""<div class="section-narrative">
    <strong>Self-Efficacy &amp; Self-Beliefs:</strong> Women's agreement with self-esteem statements
    (strengths, feelings matter, gender equality) across baseline and endline. Tracks shifts in
    <em>Strongly Agree</em> rates and self-compassion frequency to assess programme impact on
    women's confidence and self-worth.
    </div>""", unsafe_allow_html=True)

    _quick_nav_pills(['Likert Breakdown', 'Strongly Agree Shift', 'Agreement vs Disagreement', 'Self-Compassion'])

    # --- KPIs: pp increase in Strongly Agree ---
    sa = g['self_strongly_agree']
    kpi_cols = st.columns(len(sa))
    for i, row in sa.iterrows():
        bl = row['Baseline'] if isinstance(row['Baseline'], (int, float)) else 0
        el = row['Endline'] if isinstance(row['Endline'], (int, float)) else 0
        change = (el - bl) * 100
        label = _short_label(row['Statement'], 40)
        kpi_cols[i].markdown(f"""<div class="kpi-card">
            <h3>{label}</h3>
            <div class="value">{el*100:.1f}%</div>
            <div class="delta-{'positive' if change>=0 else 'negative'}">{change:+.1f}pp Strongly Agree</div>
        </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # --- Likert stacked bars: Baseline vs Endline side by side ---
    _section_header('', 'Likert Breakdown — Baseline', 'SELF')
    likert_cols = ['Strongly agree', 'Agree', 'Disagree', 'Strongly disagree']
    likert_colors = ['#2E7D32', '#81C784', '#EF9A9A', '#C62828']
    bl_likert = g['self_baseline_likert'].copy()
    bl_likert['Statement'] = bl_likert['Statement'].apply(lambda s: _short_label(s))
    st.plotly_chart(
        _gjj_stacked_bar(bl_likert, 'Statement', likert_cols, likert_colors,
                         'SELF Statements — Baseline Likert', height=300),
        use_container_width=True)

    _section_header('', 'Likert Breakdown — Endline', 'SELF')
    el_likert = g['self_endline_likert'].copy()
    el_likert['Statement'] = el_likert['Statement'].apply(lambda s: _short_label(s))
    st.plotly_chart(
        _gjj_stacked_bar(el_likert, 'Statement', likert_cols, likert_colors,
                         'SELF Statements — Endline Likert', height=300),
        use_container_width=True)

    # --- Strongly Agree change ---
    st.markdown("---")
    _section_header('', 'Strongly Agree Shift (Baseline vs Endline)', 'SELF')
    sa_chart = sa.copy()
    sa_chart['Statement'] = sa_chart['Statement'].apply(lambda s: _short_label(s))
    st.plotly_chart(_gjj_bar(sa_chart, 'Statement', 'Strongly Agree: Baseline vs Endline',
                             height=350, orientation='h'), use_container_width=True)

    # --- Agreement vs Disagreement ---
    st.markdown("---")
    _section_header('', 'Agreement vs Disagreement', 'SELF')
    agr = g['self_agreement'].copy()
    agr['Statement'] = agr['Statement'].apply(lambda s: _short_label(s))
    agr_plot = agr.rename(columns={'Agreement_BL': 'Agree BL', 'Agreement_EL': 'Agree EL',
                                   'Disagreement_BL': 'Disagree BL', 'Disagreement_EL': 'Disagree EL'})
    agr_cols = ['Agree BL', 'Agree EL', 'Disagree BL', 'Disagree EL']
    agr_colors = ['#5B8DB8', '#2E7D32', '#EF9A9A', '#C62828']
    st.plotly_chart(
        _gjj_stacked_bar(agr_plot, 'Statement', agr_cols, agr_colors,
                         'Agreement vs Disagreement (Baseline & Endline)', height=300,
                         multiply=True),
        use_container_width=True)

    # --- Self-compassion frequency ---
    st.markdown("---")
    _section_header('', 'Self-Compassion Frequency', 'SELF')
    st.plotly_chart(_gjj_bar(g['self_compassion'], 'Category',
                             'Self-Compassion: "I nurture myself with kind self-talk..."',
                             height=350), use_container_width=True)


# ---- Tab 2: RELATIONAL WELLBEING ----
def render_gjj_tab2(g):
    """GJJ KAP Women — Tab 2: Relational Wellbeing."""
    st.markdown("""<div class="section-narrative">
    <strong>Relational Dynamics:</strong> Women's experience of partner support, respect, and
    communication — from laughing together to comfort voicing disagreement. Tracks frequency shifts
    (Always/Frequently vs Rarely/Never) from baseline to endline, plus relational equality and
    GJJ community support agreement.
    </div>""", unsafe_allow_html=True)

    _quick_nav_pills(['Frequency (BL)', 'Frequency (EL)', 'Always/Frequently vs Rarely/Never',
                       'Difference', 'Relational Equality'])

    # Short labels for axis readability
    def _shorten_rel(df):
        out = df.copy()
        out['Statement'] = out['Statement'].apply(lambda s: _short_label(s, 50))
        return out

    # --- Baseline frequency ---
    _section_header('', 'Relational Frequency — Baseline', 'Relational')
    freq_cols = ['Always', 'Frequently', 'Sometimes', 'Rarely', 'Never']
    freq_colors = ['#1B5E20', '#43A047', '#FDD835', '#EF9A9A', '#C62828']
    st.plotly_chart(
        _gjj_stacked_bar(_shorten_rel(g['rel_baseline_freq']), 'Statement',
                         freq_cols, freq_colors,
                         'Relational Circumstances — Baseline Frequency', height=450),
        use_container_width=True)

    # --- Endline frequency ---
    _section_header('', 'Relational Frequency — Endline', 'Relational')
    st.plotly_chart(
        _gjj_stacked_bar(_shorten_rel(g['rel_endline_freq']), 'Statement',
                         freq_cols, freq_colors,
                         'Relational Circumstances — Endline Frequency', height=450),
        use_container_width=True)

    # --- Always/Frequently vs Rarely/Never ---
    st.markdown("---")
    _section_header('', 'Always/Frequently vs Rarely/Never', 'Relational')
    af_rn = g['rel_af_rn'].copy()
    af_rn['Statement'] = af_rn['Statement'].apply(lambda s: _short_label(s, 50))

    # Build two side-by-side comparison charts
    c1, c2 = st.columns(2)
    with c1:
        af_df = af_rn[['Statement', 'AF_Baseline', 'AF_Endline']].rename(
            columns={'AF_Baseline': 'Baseline', 'AF_Endline': 'Endline'})
        st.plotly_chart(_gjj_bar(af_df, 'Statement', 'Always / Frequently',
                                 height=450, orientation='h'), use_container_width=True)
    with c2:
        rn_df = af_rn[['Statement', 'RN_Baseline', 'RN_Endline']].rename(
            columns={'RN_Baseline': 'Baseline', 'RN_Endline': 'Endline'})
        st.plotly_chart(_gjj_bar(rn_df, 'Statement', 'Rarely / Never',
                                 height=450, orientation='h'), use_container_width=True)

    # --- Difference chart ---
    st.markdown("---")
    _section_header('', 'Difference: Always/Frequently Change (pp)', 'Relational')
    dif_df = af_rn[['Statement', 'DIF_AF']].copy()
    dif_df['DIF_AF'] = pd.to_numeric(dif_df['DIF_AF'], errors='coerce').fillna(0) * 100
    dif_df = dif_df.sort_values('DIF_AF')
    colors = [COLORS['good'] if v >= 0 else COLORS['danger'] for v in dif_df['DIF_AF']]
    fig_dif = go.Figure()
    fig_dif.add_trace(go.Bar(y=dif_df['Statement'], x=dif_df['DIF_AF'], orientation='h',
                             marker_color=colors,
                             text=dif_df['DIF_AF'].apply(lambda x: f"{x:+.1f}pp"), textposition='auto'))
    fig_dif.update_layout(title='Change in Always/Frequently (Baseline to Endline)', height=400,
                          xaxis_title='Change (pp)', yaxis=dict(categoryorder='trace'),
                          font=dict(size=13, color='#333'), title_font=dict(size=16, color='#222'),
                          plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                          margin=dict(l=20, r=20, t=60, b=20))
    st.plotly_chart(fig_dif, use_container_width=True)

    # --- Relational equality agreement statements ---
    st.markdown("---")
    _section_header('', 'Relational Equality Statements', 'Relational')
    likert_cols = ['Strongly agree', 'Agree', 'Disagree', 'Strongly disagree']
    likert_colors = ['#2E7D32', '#81C784', '#EF9A9A', '#C62828']

    c3, c4 = st.columns(2)
    with c3:
        bl_agr = g['rel_agree_baseline'].copy()
        bl_agr['Statement'] = bl_agr['Statement'].apply(lambda s: _short_label(s, 45))
        st.plotly_chart(
            _gjj_stacked_bar(bl_agr, 'Statement', likert_cols, likert_colors,
                             'Relational Equality — Baseline', height=280),
            use_container_width=True)
    with c4:
        el_agr = g['rel_agree_endline'].copy()
        el_agr['Statement'] = el_agr['Statement'].apply(lambda s: _short_label(s, 45))
        st.plotly_chart(
            _gjj_stacked_bar(el_agr, 'Statement', likert_cols, likert_colors,
                             'Relational Equality — Endline', height=280),
            use_container_width=True)

    # Agreement summary
    _section_header('', 'Agreement vs Disagreement Summary', 'Relational')
    summ = g['rel_agreement_summary'].copy()
    summ['Statement'] = summ['Statement'].apply(lambda s: _short_label(s, 50))
    summ_plot = summ.rename(columns={'Agreement_BL': 'Agree BL', 'Agreement_EL': 'Agree EL',
                                     'Disagreement_BL': 'Disagree BL', 'Disagreement_EL': 'Disagree EL'})
    st.plotly_chart(
        _gjj_stacked_bar(summ_plot, 'Statement',
                         ['Agree BL', 'Agree EL', 'Disagree BL', 'Disagree EL'],
                         ['#5B8DB8', '#2E7D32', '#EF9A9A', '#C62828'],
                         'Agreement vs Disagreement (Baseline & Endline)', height=280),
        use_container_width=True)


# ---- Tab 3: Gender Transformation — Shared Responsibility ----
def render_gjj_tab3(g):
    """GJJ KAP Women — Tab 3: Shared Responsibility (Household Chores)."""
    st.markdown("""<div class="section-narrative">
    <strong>Shared Responsibility — Gender Transformation:</strong> Tracks husbands' support with
    traditionally female household chores, how long the support has been provided, frequency,
    types of chores completed, hours saved for women, and frequency of husbands enabling women
    time for rest, conservation work, or business.
    </div>""", unsafe_allow_html=True)

    _quick_nav_pills(['Chore Support', 'Duration & Frequency', 'Chore Types',
                       'Hours Saved', 'Women Self-Time'])

    # --- KPIs ---
    yes_bl = g['shared_chores_yn'].loc[g['shared_chores_yn']['Response'] == 'Yes', 'Baseline'].values
    yes_el = g['shared_chores_yn'].loc[g['shared_chores_yn']['Response'] == 'Yes', 'Endline'].values
    yes_bl = float(yes_bl[0]) if len(yes_bl) > 0 else 0
    yes_el = float(yes_el[0]) if len(yes_el) > 0 else 0
    chg = (yes_el - yes_bl) * 100

    disc_yes_bl = g['chore_discussed'].loc[g['chore_discussed']['Response'] == 'Yes', 'Baseline'].values
    disc_yes_el = g['chore_discussed'].loc[g['chore_discussed']['Response'] == 'Yes', 'Endline'].values
    disc_bl = float(disc_yes_bl[0]) if len(disc_yes_bl) > 0 else 0
    disc_el = float(disc_yes_el[0]) if len(disc_yes_el) > 0 else 0
    disc_chg = (disc_el - disc_bl) * 100

    comp_yes_bl = g['chore_completed'].loc[g['chore_completed']['Response'] == 'Yes', 'Baseline'].values
    comp_yes_el = g['chore_completed'].loc[g['chore_completed']['Response'] == 'Yes', 'Endline'].values
    comp_bl = float(comp_yes_bl[0]) if len(comp_yes_bl) > 0 else 0
    comp_el = float(comp_yes_el[0]) if len(comp_yes_el) > 0 else 0
    comp_chg = (comp_el - comp_bl) * 100

    kc1, kc2, kc3 = st.columns(3)
    kc1.markdown(f"""<div class="kpi-card">
        <h3>Husband Supports Chores</h3>
        <div class="value">{yes_el*100:.1f}%</div>
        <div class="delta-{'positive' if chg>=0 else 'negative'}">{chg:+.1f}pp from BL</div>
    </div>""", unsafe_allow_html=True)
    kc2.markdown(f"""<div class="kpi-card">
        <h3>Discussed Sharing Chores</h3>
        <div class="value">{disc_el*100:.1f}%</div>
        <div class="delta-{'positive' if disc_chg>=0 else 'negative'}">{disc_chg:+.1f}pp from BL</div>
    </div>""", unsafe_allow_html=True)
    kc3.markdown(f"""<div class="kpi-card">
        <h3>Completed Significant Chore</h3>
        <div class="value">{comp_el*100:.1f}%</div>
        <div class="delta-{'positive' if comp_chg>=0 else 'negative'}">{comp_chg:+.1f}pp from BL</div>
    </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # --- Husband supports chores Y/N ---
    _section_header('', 'Husband Supports Household Chores', 'Shared Responsibility')
    c1, c2 = st.columns(2)
    with c1:
        st.plotly_chart(_gjj_bar(g['shared_chores_yn'], 'Response',
                                 'Husband Supports Chores (Yes/No)', height=300),
                        use_container_width=True)
    with c2:
        st.plotly_chart(_gjj_bar(g['chore_discussed'], 'Response',
                                 'Discussed Sharing Unpaid Chores', height=300),
                        use_container_width=True)

    # --- Duration & Frequency ---
    st.markdown("---")
    _section_header('', 'Duration & Frequency of Support', 'Shared Responsibility')
    c3, c4 = st.columns(2)
    with c3:
        st.plotly_chart(_gjj_bar(g['chore_duration'], 'Category',
                                 'Time Since Husband Supporting Chores',
                                 height=400, orientation='h'), use_container_width=True)
    with c4:
        st.plotly_chart(_gjj_bar(g['chore_frequency'], 'Category',
                                 'Frequency Husbands Carry Out Chores',
                                 height=400, orientation='h'), use_container_width=True)

    # --- Who started conversation + completed chore ---
    st.markdown("---")
    c5, c6 = st.columns(2)
    with c5:
        st.plotly_chart(_gjj_bar(g['chore_initiator'], 'Person',
                                 'Person Who Started Conversation', height=300),
                        use_container_width=True)
    with c6:
        st.plotly_chart(_gjj_bar(g['chore_completed'], 'Response',
                                 'Husband Completed Significant Chore', height=300),
                        use_container_width=True)

    # --- Chore types ---
    st.markdown("---")
    _section_header('', 'Types of Chores Completed by Husband', 'Shared Responsibility')
    st.plotly_chart(_gjj_bar(g['chore_type'], 'Chore',
                             'Type of Household Chore Completed', height=350, orientation='h'),
                    use_container_width=True)

    # Change in chore types
    st.plotly_chart(_gjj_delta_bar(g['chore_type'], 'Chore',
                                   'Change in Chore Types (Baseline to Endline)', height=300),
                    use_container_width=True)

    # --- Hours saved ---
    st.markdown("---")
    _section_header('', 'Hours Saved for Women', 'Shared Responsibility')
    st.plotly_chart(_gjj_bar(g['hours_saved'], 'Hours',
                             'Hours Saved Due to Husband Support', height=380),
                    use_container_width=True)

    # --- Husband support for women self-time ---
    st.markdown("---")
    _section_header('', 'Women Self-Time Support', 'Shared Responsibility')
    st.plotly_chart(_gjj_bar(g['support_self_time'], 'Category',
                             'Husband Supports Women Taking Time for Themselves',
                             height=380), use_container_width=True)


# ---- Tab 4: Shared Power & Decision-Making ----
def render_gjj_tab4(g):
    """GJJ KAP Women — Tab 4: Shared Power & Decision-Making."""
    st.markdown("""<div class="section-narrative">
    <strong>Shared Power — Decision-Making:</strong> Household conversations about decision-making,
    types of important decisions taken in the past 6 months, who made them (husband alone, joint,
    wife alone), and whether women report having an equal say in joint decisions.
    </div>""", unsafe_allow_html=True)

    _quick_nav_pills(['Decision Conversations', 'Decision Types', 'Who Decides', 'Equal Say'])

    # --- KPIs ---
    conv_yes_bl = g['decision_conversations'].loc[
        g['decision_conversations']['Response'] == 'Yes', 'Baseline'].values
    conv_yes_el = g['decision_conversations'].loc[
        g['decision_conversations']['Response'] == 'Yes', 'Endline'].values
    conv_bl = float(conv_yes_bl[0]) if len(conv_yes_bl) > 0 else 0
    conv_el = float(conv_yes_el[0]) if len(conv_yes_el) > 0 else 0

    joint_bl = g['decision_maker'].loc[g['decision_maker']['Person'] == 'Joint', 'Baseline'].values
    joint_el = g['decision_maker'].loc[g['decision_maker']['Person'] == 'Joint', 'Endline'].values
    jt_bl = float(joint_bl[0]) if len(joint_bl) > 0 else 0
    jt_el = float(joint_el[0]) if len(joint_el) > 0 else 0

    eq_yes_bl = g['equal_say'].loc[g['equal_say']['Response'] == 'Yes', 'Baseline'].values
    eq_yes_el = g['equal_say'].loc[g['equal_say']['Response'] == 'Yes', 'Endline'].values
    eq_bl = float(eq_yes_bl[0]) if len(eq_yes_bl) > 0 else 0
    eq_el = float(eq_yes_el[0]) if len(eq_yes_el) > 0 else 0

    kc1, kc2, kc3 = st.columns(3)
    kc1.markdown(f"""<div class="kpi-card">
        <h3>Decision Conversations</h3>
        <div class="value">{conv_el*100:.1f}%</div>
        <div class="delta-{'positive' if conv_el>=conv_bl else 'negative'}">{(conv_el-conv_bl)*100:+.1f}pp</div>
    </div>""", unsafe_allow_html=True)
    kc2.markdown(f"""<div class="kpi-card">
        <h3>Joint Decision-Making</h3>
        <div class="value">{jt_el*100:.1f}%</div>
        <div class="delta-{'positive' if jt_el>=jt_bl else 'negative'}">{(jt_el-jt_bl)*100:+.1f}pp</div>
    </div>""", unsafe_allow_html=True)
    kc3.markdown(f"""<div class="kpi-card">
        <h3>Women Equal Say</h3>
        <div class="value">{eq_el*100:.1f}%</div>
        <div class="delta-{'positive' if eq_el>=eq_bl else 'negative'}">{(eq_el-eq_bl)*100:+.1f}pp</div>
    </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # --- Decision conversations & decisions made ---
    _section_header('', 'Decision Conversations & Decisions Made', 'Shared Power')
    c1, c2 = st.columns(2)
    with c1:
        st.plotly_chart(_gjj_bar(g['decision_conversations'], 'Response',
                                 'Conversations to Change Decision-Making', height=300),
                        use_container_width=True)
    with c2:
        st.plotly_chart(_gjj_bar(g['decisions_made'], 'Response',
                                 'Made Important Decisions (Past 6 Months)', height=300),
                        use_container_width=True)

    # --- Types of decisions ---
    st.markdown("---")
    _section_header('', 'Types of Decisions Taken', 'Shared Power')
    st.plotly_chart(_gjj_bar(g['decision_types'], 'Decision',
                             'Type of Decision Taken (Past 6 Months)',
                             height=500, orientation='h'), use_container_width=True)

    st.plotly_chart(_gjj_delta_bar(g['decision_types'], 'Decision',
                                   'Change in Decision Types (Baseline to Endline)',
                                   height=450), use_container_width=True)

    # --- Who makes decisions ---
    st.markdown("---")
    _section_header('', 'Who Makes Decisions', 'Shared Power')
    st.plotly_chart(_gjj_bar(g['decision_maker'], 'Person',
                             'Person Making the Decision', height=300),
                    use_container_width=True)

    # --- Equal say ---
    st.markdown("---")
    _section_header('', 'Equal Say in Joint Decisions', 'Shared Power')
    st.plotly_chart(_gjj_bar(g['equal_say'], 'Response',
                             'Women Reporting Equal Say in Joint Decisions', height=300),
                    use_container_width=True)


# ---- Tab 5: Autonomy & Leadership Support ----
def render_gjj_tab5(g):
    """GJJ KAP Women — Tab 5: Autonomy & Leadership."""
    st.markdown("""<div class="section-narrative">
    <strong>Autonomy &amp; Leadership:</strong> Whether women hide money to save, and the degree to
    which husbands would support women becoming community leaders or starting/growing their own
    businesses. Tracks shifts from baseline to endline across a 6-point certainty scale (Definitely
    to Definitely Not).
    </div>""", unsafe_allow_html=True)

    _quick_nav_pills(['Hiding Money', 'Leadership Support', 'Business Support'])

    # --- KPIs: Definitely support leadership + Definitely support business ---
    lead_def_bl = g['support_leader'].loc[
        g['support_leader']['Category'].str.strip().str.lower() == 'definitely', 'Baseline'].values
    lead_def_el = g['support_leader'].loc[
        g['support_leader']['Category'].str.strip().str.lower() == 'definitely', 'Endline'].values
    l_bl = float(lead_def_bl[0]) if len(lead_def_bl) > 0 else 0
    l_el = float(lead_def_el[0]) if len(lead_def_el) > 0 else 0

    biz_def_bl = g['support_business'].loc[
        g['support_business']['Category'].str.strip().str.lower() == 'definitely', 'Baseline'].values
    biz_def_el = g['support_business'].loc[
        g['support_business']['Category'].str.strip().str.lower() == 'definitely', 'Endline'].values
    b_bl = float(biz_def_bl[0]) if len(biz_def_bl) > 0 else 0
    b_el = float(biz_def_el[0]) if len(biz_def_el) > 0 else 0

    # Hiding money "Always" — lower is better
    hide_alw_bl = g['hide_money'].loc[
        g['hide_money']['Category'].str.strip().str.lower() == 'always', 'Baseline'].values
    hide_alw_el = g['hide_money'].loc[
        g['hide_money']['Category'].str.strip().str.lower() == 'always', 'Endline'].values
    h_bl = float(hide_alw_bl[0]) if len(hide_alw_bl) > 0 else 0
    h_el = float(hide_alw_el[0]) if len(hide_alw_el) > 0 else 0
    hide_chg = (h_el - h_bl) * 100

    kc1, kc2, kc3 = st.columns(3)
    kc1.markdown(f"""<div class="kpi-card">
        <h3>Hide Money "Always"</h3>
        <div class="value">{h_el*100:.1f}%</div>
        <div class="delta-{'positive' if hide_chg<=0 else 'negative'}">{hide_chg:+.1f}pp (lower = better)</div>
    </div>""", unsafe_allow_html=True)
    kc2.markdown(f"""<div class="kpi-card">
        <h3>Definitely Support Leader</h3>
        <div class="value">{l_el*100:.1f}%</div>
        <div class="delta-{'positive' if l_el>=l_bl else 'negative'}">{(l_el-l_bl)*100:+.1f}pp</div>
    </div>""", unsafe_allow_html=True)
    kc3.markdown(f"""<div class="kpi-card">
        <h3>Definitely Support Business</h3>
        <div class="value">{b_el*100:.1f}%</div>
        <div class="delta-{'positive' if b_el>=b_bl else 'negative'}">{(b_el-b_bl)*100:+.1f}pp</div>
    </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # --- Hiding money ---
    _section_header('', 'Hiding Money from Husband to Save', 'Autonomy')
    st.plotly_chart(_gjj_bar(g['hide_money'], 'Category',
                             'Women Hiding Money (Frequency)', height=380),
                    use_container_width=True)

    # --- Leadership support ---
    st.markdown("---")
    _section_header('', 'Husband Support for Women Becoming Leaders', 'Leadership')
    st.plotly_chart(_gjj_bar(g['support_leader'], 'Category',
                             'Support if Wife Wanted to Become Community Leader',
                             height=380), use_container_width=True)
    st.plotly_chart(_gjj_delta_bar(g['support_leader'], 'Category',
                                   'Change in Leadership Support (Baseline to Endline)',
                                   height=350), use_container_width=True)

    # --- Business support ---
    st.markdown("---")
    _section_header('', 'Husband Support for Women Starting/Growing Business', 'Leadership')
    st.plotly_chart(_gjj_bar(g['support_business'], 'Category',
                             'Support if Wife Started/Grew Own Business',
                             height=380), use_container_width=True)
    st.plotly_chart(_gjj_delta_bar(g['support_business'], 'Category',
                                   'Change in Business Support (Baseline to Endline)',
                                   height=350), use_container_width=True)


# ============================================================================
# GJJ KAP MEN — TAB RENDERERS
# ============================================================================

# ---- Men Tab 1: SELF ----
def render_gjj_men_tab1(g):
    """GJJ KAP Men — Tab 1: SELF (Self-Esteem, Self-Compassion, Confidence)."""
    st.markdown("""<div class="section-narrative">
    <strong>Self-Efficacy &amp; Self-Beliefs:</strong> Men's agreement with self-esteem statements
    (responsibility for feelings, learning from mistakes, gender equality) across baseline and
    endline. Tracks shifts in agreement rates and self-compassion frequency to assess programme
    impact on men's confidence and responsibility.
    </div>""", unsafe_allow_html=True)

    _quick_nav_pills(['Likert Breakdown', 'Agreement vs Disagreement', 'Self-Compassion'])

    # --- KPIs: pp increase in Agreement (BL → EL) ---
    agr = g['self_agreement']
    kpi_cols = st.columns(len(agr))
    for i, row in agr.iterrows():
        bl = float(row['Agreement_BL']) if isinstance(row['Agreement_BL'], (int, float)) else 0
        el = float(row['Agreement_EL']) if isinstance(row['Agreement_EL'], (int, float)) else 0
        change = (el - bl) * 100
        label = _short_label(row['Statement'], 40)
        kpi_cols[i].markdown(f"""<div class="kpi-card">
            <h3>{label}</h3>
            <div class="value">{el*100:.1f}%</div>
            <div class="delta-{'positive' if change>=0 else 'negative'}">{change:+.1f}pp Agreement</div>
        </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # --- Likert stacked bars: Baseline vs Endline ---
    _section_header('', 'Likert Breakdown — Baseline', 'SELF')
    likert_cols = ['Strongly agree', 'Agree', 'Disagree', 'Strongly disagree']
    likert_colors = ['#2E7D32', '#81C784', '#EF9A9A', '#C62828']
    bl_likert = g['self_baseline_likert'].copy()
    bl_likert['Statement'] = bl_likert['Statement'].apply(lambda s: _short_label(s))
    st.plotly_chart(
        _gjj_stacked_bar(bl_likert, 'Statement', likert_cols, likert_colors,
                         'SELF Statements — Baseline Likert', height=300),
        use_container_width=True)

    _section_header('', 'Likert Breakdown — Endline', 'SELF')
    el_likert = g['self_endline_likert'].copy()
    el_likert['Statement'] = el_likert['Statement'].apply(lambda s: _short_label(s))
    st.plotly_chart(
        _gjj_stacked_bar(el_likert, 'Statement', likert_cols, likert_colors,
                         'SELF Statements — Endline Likert', height=300),
        use_container_width=True)

    # --- Agreement vs Disagreement ---
    st.markdown("---")
    _section_header('', 'Agreement vs Disagreement', 'SELF')
    agr_copy = g['self_agreement'].copy()
    agr_copy['Statement'] = agr_copy['Statement'].apply(lambda s: _short_label(s))
    agr_plot = agr_copy.rename(columns={
        'Agreement_BL': 'Agree BL', 'Agreement_EL': 'Agree EL',
        'Disagreement_BL': 'Disagree BL', 'Disagreement_EL': 'Disagree EL'})
    agr_cols = ['Agree BL', 'Agree EL', 'Disagree BL', 'Disagree EL']
    agr_colors = ['#5B8DB8', '#2E7D32', '#EF9A9A', '#C62828']
    st.plotly_chart(
        _gjj_stacked_bar(agr_plot, 'Statement', agr_cols, agr_colors,
                         'Agreement vs Disagreement (Baseline & Endline)', height=300,
                         multiply=True),
        use_container_width=True)

    # --- Self-compassion frequency ---
    st.markdown("---")
    _section_header('', 'Self-Compassion Frequency', 'SELF')
    st.plotly_chart(_gjj_bar(g['self_compassion'], 'Category',
                             'Self-Compassion: Frequency of Nurturing Self-Talk',
                             height=350), use_container_width=True)


# ---- Men Tab 2: RELATIONAL WELLBEING ----
def render_gjj_men_tab2(g):
    """GJJ KAP Men — Tab 2: Relational Wellbeing."""
    st.markdown("""<div class="section-narrative">
    <strong>Relational Dynamics:</strong> Men's experience of partner support, respect, and
    communication — from laughing together to comfort in sharing feelings. Tracks frequency shifts
    (Always/Frequently vs Rarely/Never) from baseline to endline, plus men's perception of
    GJJ community support for their gender justice journey.
    </div>""", unsafe_allow_html=True)

    _quick_nav_pills(['Frequency (BL)', 'Frequency (EL)', 'Always/Frequently vs Rarely/Never',
                       'GJJ Journey Support'])

    def _shorten_rel(df):
        out = df.copy()
        out['Statement'] = out['Statement'].apply(lambda s: _short_label(s, 50))
        return out

    # --- Baseline frequency ---
    _section_header('', 'Relational Frequency — Baseline', 'Relational')
    freq_cols = ['Always', 'Frequently', 'Sometimes', 'Rarely', 'Never']
    freq_colors = ['#1B5E20', '#43A047', '#FDD835', '#EF9A9A', '#C62828']
    st.plotly_chart(
        _gjj_stacked_bar(_shorten_rel(g['rel_baseline_freq']), 'Statement',
                         freq_cols, freq_colors,
                         'Relational Circumstances — Baseline Frequency', height=450),
        use_container_width=True)

    # --- Endline frequency ---
    _section_header('', 'Relational Frequency — Endline', 'Relational')
    st.plotly_chart(
        _gjj_stacked_bar(_shorten_rel(g['rel_endline_freq']), 'Statement',
                         freq_cols, freq_colors,
                         'Relational Circumstances — Endline Frequency', height=450),
        use_container_width=True)

    # --- Always/Frequently vs Rarely/Never ---
    st.markdown("---")
    _section_header('', 'Always/Frequently vs Rarely/Never', 'Relational')
    af_rn = g['rel_af_rn'].copy()
    af_rn['Statement'] = af_rn['Statement'].apply(lambda s: _short_label(s, 50))

    c1, c2 = st.columns(2)
    with c1:
        af_df = af_rn[['Statement', 'AF_Baseline', 'AF_Endline']].rename(
            columns={'AF_Baseline': 'Baseline', 'AF_Endline': 'Endline'})
        st.plotly_chart(_gjj_bar(af_df, 'Statement', 'Always / Frequently',
                                 height=450, orientation='h'), use_container_width=True)
    with c2:
        rn_df = af_rn[['Statement', 'RN_Baseline', 'RN_Endline']].rename(
            columns={'RN_Baseline': 'Baseline', 'RN_Endline': 'Endline'})
        st.plotly_chart(_gjj_bar(rn_df, 'Statement', 'Rarely / Never',
                                 height=450, orientation='h'), use_container_width=True)

    # --- Difference chart (computed, since Men file lacks DIF columns) ---
    st.markdown("---")
    _section_header('', 'Difference: Always/Frequently Change (pp)', 'Relational')
    dif_df = af_rn[['Statement']].copy()
    dif_df['DIF_AF'] = (pd.to_numeric(af_rn['AF_Endline'], errors='coerce').fillna(0)
                        - pd.to_numeric(af_rn['AF_Baseline'], errors='coerce').fillna(0)) * 100
    dif_df = dif_df.sort_values('DIF_AF')
    colors = [COLORS['good'] if v >= 0 else COLORS['danger'] for v in dif_df['DIF_AF']]
    fig_dif = go.Figure()
    fig_dif.add_trace(go.Bar(y=dif_df['Statement'], x=dif_df['DIF_AF'], orientation='h',
                             marker_color=colors,
                             text=dif_df['DIF_AF'].apply(lambda x: f"{x:+.1f}pp"), textposition='auto'))
    fig_dif.update_layout(title='Change in Always/Frequently (Baseline to Endline)', height=400,
                          xaxis_title='Change (pp)', yaxis=dict(categoryorder='trace'),
                          font=dict(size=13, color='#333'), title_font=dict(size=16, color='#222'),
                          plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                          margin=dict(l=20, r=20, t=60, b=20))
    st.plotly_chart(fig_dif, use_container_width=True)

    # --- GJJ Journey Support (Baseline only) ---
    st.markdown("---")
    _section_header('', 'GJJ Journey Support — Men Supported & Encouraged', 'Relational')
    gjj_sup = g['gjj_journey_support'].copy()
    fig_gjj = go.Figure()
    fig_gjj.add_trace(go.Bar(
        x=gjj_sup['Category'], y=[v * 100 for v in gjj_sup['Baseline']],
        marker_color=['#2E7D32', '#81C784', '#EF9A9A', '#C62828'],
        text=[f"{v*100:.1f}%" for v in gjj_sup['Baseline']], textposition='auto'))
    fig_gjj.update_layout(
        title='Men Supported on Gender Justice Journey (Baseline)',
        yaxis_title='Percentage (%)', height=380,
        font=dict(size=13, color='#333'), title_font=dict(size=16, color='#222'),
        plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
        margin=dict(l=20, r=20, t=60, b=20))
    st.plotly_chart(fig_gjj, use_container_width=True)
    st.info("ℹ️ GJJ Journey Support data is available for baseline only.")


# ---- Men Tab 3: Shared Responsibility ----
def render_gjj_men_tab3(g):
    """GJJ KAP Men — Tab 3: Shared Responsibility (Household Chores)."""
    st.markdown("""<div class="section-narrative">
    <strong>Shared Responsibility — Gender Transformation:</strong> Tracks men's self-reported
    support with traditionally female household chores, duration of support, frequency, whether
    chore-sharing has been discussed, and whether men encourage wives to make time for themselves.
    </div>""", unsafe_allow_html=True)

    _quick_nav_pills(['Chore Support', 'Duration & Frequency', 'Discussion & Self-Time'])

    # --- KPIs ---
    yes_bl = g['shared_chores_yn'].loc[g['shared_chores_yn']['Response'] == 'Yes', 'Baseline'].values
    yes_el = g['shared_chores_yn'].loc[g['shared_chores_yn']['Response'] == 'Yes', 'Endline'].values
    yes_bl = float(yes_bl[0]) if len(yes_bl) > 0 else 0
    yes_el = float(yes_el[0]) if len(yes_el) > 0 else 0
    chg = (yes_el - yes_bl) * 100

    disc_yes_bl = g['chore_discussed'].loc[g['chore_discussed']['Response'] == 'Yes', 'Baseline'].values
    disc_yes_el = g['chore_discussed'].loc[g['chore_discussed']['Response'] == 'Yes', 'Endline'].values
    disc_bl = float(disc_yes_bl[0]) if len(disc_yes_bl) > 0 else 0
    disc_el = float(disc_yes_el[0]) if len(disc_yes_el) > 0 else 0
    disc_chg = (disc_el - disc_bl) * 100

    self_yes_bl = g['support_self_time'].loc[g['support_self_time']['Response'] == 'Yes', 'Baseline'].values
    self_yes_el = g['support_self_time'].loc[g['support_self_time']['Response'] == 'Yes', 'Endline'].values
    st_bl = float(self_yes_bl[0]) if len(self_yes_bl) > 0 else 0
    st_el = float(self_yes_el[0]) if len(self_yes_el) > 0 else 0
    st_chg = (st_el - st_bl) * 100

    kc1, kc2, kc3 = st.columns(3)
    kc1.markdown(f"""<div class="kpi-card">
        <h3>Supports Household Chores</h3>
        <div class="value">{yes_el*100:.1f}%</div>
        <div class="delta-{'positive' if chg>=0 else 'negative'}">{chg:+.1f}pp from BL</div>
    </div>""", unsafe_allow_html=True)
    kc2.markdown(f"""<div class="kpi-card">
        <h3>Discussed Sharing Chores</h3>
        <div class="value">{disc_el*100:.1f}%</div>
        <div class="delta-{'positive' if disc_chg>=0 else 'negative'}">{disc_chg:+.1f}pp from BL</div>
    </div>""", unsafe_allow_html=True)
    kc3.markdown(f"""<div class="kpi-card">
        <h3>Encourages Wife Self-Time</h3>
        <div class="value">{st_el*100:.1f}%</div>
        <div class="delta-{'positive' if st_chg>=0 else 'negative'}">{st_chg:+.1f}pp from BL</div>
    </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # --- Supports chores Y/N + Discussed ---
    _section_header('', 'Supports Household Chores & Discussion', 'Shared Responsibility')
    c1, c2 = st.columns(2)
    with c1:
        st.plotly_chart(_gjj_bar(g['shared_chores_yn'], 'Response',
                                 'Men Supporting Household Chores (Yes/No)', height=300),
                        use_container_width=True)
    with c2:
        st.plotly_chart(_gjj_bar(g['chore_discussed'], 'Response',
                                 'Discussed Sharing Unpaid Chores', height=300),
                        use_container_width=True)

    # --- Duration & Frequency ---
    st.markdown("---")
    _section_header('', 'Duration & Frequency of Support', 'Shared Responsibility')
    c3, c4 = st.columns(2)
    with c3:
        st.plotly_chart(_gjj_bar(g['chore_duration'], 'Category',
                                 'Time Since Supporting Household Chores',
                                 height=400, orientation='h'), use_container_width=True)
    with c4:
        st.plotly_chart(_gjj_bar(g['chore_frequency'], 'Category',
                                 'Frequency of Carrying Out Household Chores',
                                 height=400, orientation='h'), use_container_width=True)

    # --- Encourages wife self-time ---
    st.markdown("---")
    _section_header('', 'Encouraging Wife to Make Time for Herself', 'Shared Responsibility')
    st.plotly_chart(_gjj_bar(g['support_self_time'], 'Response',
                             'Encourages Wife/Partner to Take Time for Themselves',
                             height=300), use_container_width=True)


# ---- Men Tab 4: Shared Power ----
def render_gjj_men_tab4(g):
    """GJJ KAP Men — Tab 4: Shared Power & Decision-Making."""
    st.markdown("""<div class="section-narrative">
    <strong>Shared Power — Decision-Making:</strong> Household conversations about decision-making
    and whether important decisions have been made in the past 6 months, as reported by men.
    </div>""", unsafe_allow_html=True)

    _quick_nav_pills(['Decision Conversations', 'Decisions Made'])

    # --- KPIs ---
    conv_yes_bl = g['decision_conversations'].loc[
        g['decision_conversations']['Response'] == 'Yes', 'Baseline'].values
    conv_yes_el = g['decision_conversations'].loc[
        g['decision_conversations']['Response'] == 'Yes', 'Endline'].values
    conv_bl = float(conv_yes_bl[0]) if len(conv_yes_bl) > 0 else 0
    conv_el = float(conv_yes_el[0]) if len(conv_yes_el) > 0 else 0

    dec_yes_bl = g['decisions_made'].loc[
        g['decisions_made']['Response'] == 'Yes', 'Baseline'].values
    dec_yes_el = g['decisions_made'].loc[
        g['decisions_made']['Response'] == 'Yes', 'Endline'].values
    dec_bl = float(dec_yes_bl[0]) if len(dec_yes_bl) > 0 else 0
    dec_el = float(dec_yes_el[0]) if len(dec_yes_el) > 0 else 0

    kc1, kc2 = st.columns(2)
    kc1.markdown(f"""<div class="kpi-card">
        <h3>Decision Conversations</h3>
        <div class="value">{conv_el*100:.1f}%</div>
        <div class="delta-{'positive' if conv_el>=conv_bl else 'negative'}">{(conv_el-conv_bl)*100:+.1f}pp</div>
    </div>""", unsafe_allow_html=True)
    kc2.markdown(f"""<div class="kpi-card">
        <h3>Decisions Made</h3>
        <div class="value">{dec_el*100:.1f}%</div>
        <div class="delta-{'positive' if dec_el>=dec_bl else 'negative'}">{(dec_el-dec_bl)*100:+.1f}pp</div>
    </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # --- Decision conversations & decisions made ---
    _section_header('', 'Decision Conversations & Decisions Made', 'Shared Power')
    c1, c2 = st.columns(2)
    with c1:
        st.plotly_chart(_gjj_bar(g['decision_conversations'], 'Response',
                                 'Conversations to Change Decision-Making', height=300),
                        use_container_width=True)
    with c2:
        st.plotly_chart(_gjj_bar(g['decisions_made'], 'Response',
                                 'Made Important Decisions (Past 6 Months)', height=300),
                        use_container_width=True)


# ---- Men Tab 5: Leadership & Business Support ----
def render_gjj_men_tab5(g):
    """GJJ KAP Men — Tab 5: Leadership & Business Support."""
    st.markdown("""<div class="section-narrative">
    <strong>Leadership &amp; Business Support:</strong> Men's willingness to support their wife
    becoming a community leader or starting/growing a business. Tracks frequency of support from
    baseline to endline across multiple response categories.
    </div>""", unsafe_allow_html=True)

    _quick_nav_pills(['Leadership Support', 'Business Support'])

    # --- KPIs: Always support leadership + Definitely support business ---
    lead_alw_bl = g['support_leader'].loc[
        g['support_leader']['Category'].str.strip().str.lower() == 'always', 'Baseline'].values
    lead_alw_el = g['support_leader'].loc[
        g['support_leader']['Category'].str.strip().str.lower() == 'always', 'Endline'].values
    l_bl = float(lead_alw_bl[0]) if len(lead_alw_bl) > 0 else 0
    l_el = float(lead_alw_el[0]) if len(lead_alw_el) > 0 else 0

    biz_def_bl = g['support_business'].loc[
        g['support_business']['Category'].str.strip().str.lower() == 'definitely', 'Baseline'].values
    biz_def_el = g['support_business'].loc[
        g['support_business']['Category'].str.strip().str.lower() == 'definitely', 'Endline'].values
    b_bl = float(biz_def_bl[0]) if len(biz_def_bl) > 0 else 0
    b_el = float(biz_def_el[0]) if len(biz_def_el) > 0 else 0

    kc1, kc2 = st.columns(2)
    kc1.markdown(f"""<div class="kpi-card">
        <h3>"Always" Support Leader</h3>
        <div class="value">{l_el*100:.1f}%</div>
        <div class="delta-{'positive' if l_el>=l_bl else 'negative'}">{(l_el-l_bl)*100:+.1f}pp</div>
    </div>""", unsafe_allow_html=True)
    kc2.markdown(f"""<div class="kpi-card">
        <h3>"Definitely" Support Business</h3>
        <div class="value">{b_el*100:.1f}%</div>
        <div class="delta-{'positive' if b_el>=b_bl else 'negative'}">{(b_el-b_bl)*100:+.1f}pp</div>
    </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # --- Leadership support ---
    _section_header('', 'Support for Wife Becoming Community Leader', 'Leadership')
    st.plotly_chart(_gjj_bar(g['support_leader'], 'Category',
                             'Would Support Wife Becoming Community Leader',
                             height=380), use_container_width=True)
    st.plotly_chart(_gjj_delta_bar(g['support_leader'], 'Category',
                                   'Change in Leadership Support (Baseline to Endline)',
                                   height=350), use_container_width=True)

    # --- Business support ---
    st.markdown("---")
    _section_header('', 'Support for Wife Starting/Growing Business', 'Leadership')
    st.plotly_chart(_gjj_bar(g['support_business'], 'Category',
                             'Would Support Wife Starting/Growing Own Business',
                             height=380), use_container_width=True)
    st.plotly_chart(_gjj_delta_bar(g['support_business'], 'Category',
                                   'Change in Business Support (Baseline to Endline)',
                                   height=350), use_container_width=True)


# ============================================================================
# FOREST TRAINING RENDERER
# ============================================================================

def render_forest_training_tabs(t_data, timepoint_filter='Combined'):
    """Render the Forest Training Pre/Post Knowledge Assessment tabs."""

    thresholds = t_data['thresholds']
    scores = t_data['scores']
    questions = t_data['questions']

    tab1, tab2, tab3 = st.tabs([
        "Overview of Training Outcomes",
        "Knowledge Improvements by Question",
        "Domain-Level Grouping",
    ])

    # ==================================================================
    # TAB 1 — Overview
    # ==================================================================
    with tab1:
        st.markdown("""<div class="section-narrative">
        <strong>Training Outcomes Overview:</strong> Key performance indicators from the
        forest conservation knowledge assessment, comparing Pre-training (Baseline) with
        Post-training (Endline) results. The ≥70% threshold is the main PMF reporting
        standard for adequate knowledge.
        </div>""", unsafe_allow_html=True)

        # ---- KPI metrics ----
        bl_scores = scores[scores['Timepoint'] == 'Baseline'].iloc[0]
        el_scores = scores[scores['Timepoint'] == 'Endline'].iloc[0]

        # ≥70 threshold
        t70 = thresholds[thresholds['Threshold'].str.contains('70')]
        t70_bl = float(t70[t70['Timepoint'] == 'Baseline']['Proportion'].values[0]) if len(t70[t70['Timepoint'] == 'Baseline']) else 0
        t70_el = float(t70[t70['Timepoint'] == 'Endline']['Proportion'].values[0]) if len(t70[t70['Timepoint'] == 'Endline']) else 0

        k1, k2, k3, k4 = st.columns(4)
        k1.markdown(f"""<div class="kpi-card">
            <h3>≥70% Pass Rate (PMF)</h3>
            <div class="value">{t70_el*100:.1f}%</div>
            <div class="delta-{'positive' if t70_el>=t70_bl else 'negative'}">{(t70_el-t70_bl)*100:+.1f}pp vs Baseline</div>
        </div>""", unsafe_allow_html=True)
        k2.markdown(f"""<div class="kpi-card">
            <h3>Average Score</h3>
            <div class="value">{el_scores['AverageScore']:.1f}%</div>
            <div class="delta-{'positive' if el_scores['AverageScore']>=bl_scores['AverageScore'] else 'negative'}">{el_scores['AverageScore']-bl_scores['AverageScore']:+.1f}pp vs Baseline</div>
        </div>""", unsafe_allow_html=True)
        k3.markdown(f"""<div class="kpi-card">
            <h3>Baseline Respondents</h3>
            <div class="value">{int(bl_scores['Respondents']):,}</div>
            <div class="delta-neutral">Pre-training</div>
        </div>""", unsafe_allow_html=True)
        k4.markdown(f"""<div class="kpi-card">
            <h3>Endline Respondents</h3>
            <div class="value">{int(el_scores['Respondents']):,}</div>
            <div class="delta-positive">Post-training</div>
        </div>""", unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        # ---- Pass-rate chart by threshold ----
        _section_header('', 'Pass Rates by Threshold', 'Pre vs Post')

        if timepoint_filter == 'Combined':
            plot_thresh = thresholds.copy()
        else:
            plot_thresh = thresholds[thresholds['Timepoint'] == timepoint_filter].copy()

        plot_thresh['Pct'] = plot_thresh['Proportion'] * 100
        fig_thresh = go.Figure()
        for tp in plot_thresh['Timepoint'].unique():
            sub = plot_thresh[plot_thresh['Timepoint'] == tp]
            color = COLORS['baseline'] if tp == 'Baseline' else COLORS['midline']
            fig_thresh.add_trace(go.Bar(
                x=sub['Threshold'], y=sub['Pct'], name=tp,
                marker_color=color,
                text=sub['Pct'].apply(lambda x: f"{x:.1f}%"), textposition='auto',
            ))
        fig_thresh.update_layout(
            title='Percentage of Participants Achieving Score Thresholds',
            barmode='group', height=420, yaxis_title='% of Participants',
            legend=dict(orientation='h', yanchor='bottom', y=1.02, x=0.5, xanchor='center'),
            font=dict(size=13, color='#333'),
            title_font=dict(size=16, color='#222'),
            plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
            margin=dict(l=20, r=20, t=60, b=20),
        )
        fig_thresh.add_hline(y=70, line_dash='dash', line_color='#E65100',
                             annotation_text='PMF Threshold (70%)', annotation_position='top left')
        st.plotly_chart(fig_thresh, use_container_width=True)

        # ---- Standardized scores comparison ----
        _section_header('', 'Standardized Scores', 'Avg / Max / Min')

        col_a, col_b = st.columns(2)
        with col_a:
            # Average score comparison
            fig_avg = go.Figure()
            for _, row in scores.iterrows():
                color = COLORS['baseline'] if row['Timepoint'] == 'Baseline' else COLORS['midline']
                fig_avg.add_trace(go.Bar(
                    x=[row['Timepoint']], y=[row['AverageScore']],
                    name=row['Timepoint'], marker_color=color,
                    text=[f"{row['AverageScore']:.1f}%"], textposition='auto',
                    width=0.5,
                ))
            fig_avg.update_layout(
                title='Average Test Score', height=350,
                yaxis_title='Score (%)', showlegend=False,
                font=dict(size=13, color='#333'),
                title_font=dict(size=15, color='#222'),
                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                margin=dict(l=20, r=20, t=50, b=20),
                yaxis=dict(range=[0, 105]),
            )
            st.plotly_chart(fig_avg, use_container_width=True)

        with col_b:
            # Min score comparison
            fig_minmax = go.Figure()
            for _, row in scores.iterrows():
                color = COLORS['baseline'] if row['Timepoint'] == 'Baseline' else COLORS['midline']
                fig_minmax.add_trace(go.Bar(
                    x=[f"{row['Timepoint']} Min", f"{row['Timepoint']} Max"],
                    y=[row['MinScore'], row['MaxScore']],
                    name=row['Timepoint'], marker_color=color,
                    text=[f"{row['MinScore']:.1f}", f"{row['MaxScore']:.1f}"],
                    textposition='auto', width=0.5,
                ))
            fig_minmax.update_layout(
                title='Score Range (Min & Max)', height=350,
                yaxis_title='Score (%)', showlegend=False,
                font=dict(size=13, color='#333'),
                title_font=dict(size=15, color='#222'),
                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                margin=dict(l=20, r=20, t=50, b=20),
                yaxis=dict(range=[0, 105]),
            )
            st.plotly_chart(fig_minmax, use_container_width=True)

        # Data table
        with st.expander("View Scores Data Table"):
            st.dataframe(scores.style.format({
                'AverageScore': '{:.2f}%', 'MaxScore': '{:.1f}', 'MinScore': '{:.1f}',
                'Respondents': '{:,.0f}'
            }), use_container_width=True)

    # ==================================================================
    # TAB 2 — Knowledge Improvements by Question
    # ==================================================================
    with tab2:
        st.markdown("""<div class="section-narrative">
        <strong>Question-Level Analysis:</strong> Performance on each of the 21 knowledge
        test questions, comparing Pre-training (Baseline) with Post-training (Endline)
        results. Use the keyword filter to focus on specific topics.
        </div>""", unsafe_allow_html=True)

        if questions.empty:
            st.warning("No question-level data available.")
        else:
            # Keyword filter
            keyword = st.text_input(
                "Filter questions by keyword",
                placeholder="e.g. gender, agroforestry, PFM, climate",
                help="Type a keyword to filter questions (case-insensitive)"
            )
            filtered_q = questions.copy()
            if keyword:
                filtered_q = filtered_q[filtered_q['QuestionText'].str.contains(
                    keyword, case=False, na=False)]
                if filtered_q.empty:
                    st.info(f"No questions matching '{keyword}'.")
                    filtered_q = questions.copy()

            filtered_q['Baseline_Pct'] = filtered_q['Baseline'] * 100
            filtered_q['Endline_Pct'] = filtered_q['Endline'] * 100
            filtered_q['Change'] = filtered_q['Endline_Pct'] - filtered_q['Baseline_Pct']
            filtered_q['Label'] = 'Q' + filtered_q['QuestionNumber'].astype(str)
            filtered_q = filtered_q.sort_values('QuestionNumber')

            _section_header('', 'Baseline vs Endline by Question', 'All 21 Questions')

            # Horizontal grouped bar chart
            plot_q = filtered_q.sort_values('QuestionNumber', ascending=True)
            fig_q = go.Figure()
            fig_q.add_trace(go.Bar(
                y=plot_q['Label'], x=plot_q['Baseline_Pct'],
                name='Baseline', orientation='h',
                marker_color=COLORS['baseline'],
                text=plot_q['Baseline_Pct'].apply(lambda x: f"{x:.1f}%"),
                textposition='auto',
            ))
            fig_q.add_trace(go.Bar(
                y=plot_q['Label'], x=plot_q['Endline_Pct'],
                name='Endline', orientation='h',
                marker_color=COLORS['midline'],
                text=plot_q['Endline_Pct'].apply(lambda x: f"{x:.1f}%"),
                textposition='auto',
            ))
            fig_q.update_layout(
                title='% Correct by Question (Baseline vs Endline)',
                barmode='group', height=max(500, len(plot_q) * 35),
                xaxis_title='% Correct', yaxis=dict(autorange='reversed'),
                legend=dict(orientation='h', yanchor='bottom', y=1.02, x=0.5, xanchor='center'),
                font=dict(size=12, color='#333'),
                title_font=dict(size=16, color='#222'),
                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                margin=dict(l=60, r=20, t=60, b=20),
            )
            st.plotly_chart(fig_q, use_container_width=True)

            # Change bar chart
            _section_header('', 'Knowledge Improvement (Endline − Baseline)', 'Top Gains')
            change_q = filtered_q.sort_values('Change', ascending=True)
            colors_change = [COLORS['good'] if c > 5 else (COLORS['danger'] if c < 0 else '#FF9800')
                             for c in change_q['Change']]
            fig_change = go.Figure(go.Bar(
                y=change_q['Label'],
                x=change_q['Change'],
                orientation='h',
                marker_color=colors_change,
                text=change_q['Change'].apply(lambda x: f"{x:+.1f}pp"),
                textposition='auto',
            ))
            fig_change.update_layout(
                title='Change in % Correct (Endline − Baseline)',
                height=max(500, len(change_q) * 30),
                xaxis_title='Percentage Point Change',
                yaxis=dict(autorange='reversed'),
                font=dict(size=12, color='#333'),
                title_font=dict(size=16, color='#222'),
                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                margin=dict(l=60, r=20, t=60, b=20),
            )
            fig_change.add_vline(x=0, line_color='#666', line_width=1)
            st.plotly_chart(fig_change, use_container_width=True)

            # Hover-expandable question text table
            with st.expander("View Full Question Text & Scores"):
                display_df = filtered_q[['QuestionNumber', 'QuestionText',
                                          'Baseline_Pct', 'Endline_Pct', 'Change']].copy()
                display_df.columns = ['Q#', 'Question', 'Baseline %', 'Endline %', 'Change (pp)']
                st.dataframe(display_df.style.format({
                    'Baseline %': '{:.1f}', 'Endline %': '{:.1f}', 'Change (pp)': '{:+.1f}'
                }).applymap(
                    lambda v: 'color: #2E7D32; font-weight:700' if isinstance(v, (int, float)) and v > 5
                    else ('color: #C62828; font-weight:700' if isinstance(v, (int, float)) and v < 0
                          else ''),
                    subset=['Change (pp)']
                ), use_container_width=True, height=600)

    # ==================================================================
    # TAB 3 — Domain-Level Grouping
    # ==================================================================
    with tab3:
        st.markdown("""<div class="section-narrative">
        <strong>Thematic Domain Analysis:</strong> Questions grouped into thematic areas
        with average Baseline and Endline performance per domain. This helps identify
        which knowledge areas improved most and which need further attention.
        </div>""", unsafe_allow_html=True)

        if questions.empty:
            st.warning("No question-level data available for domain grouping.")
        else:
            domain_rows = []
            for domain, q_nums in FOREST_TRAINING_DOMAINS.items():
                dq = questions[questions['QuestionNumber'].isin(q_nums)]
                if dq.empty:
                    continue
                bl_avg = dq['Baseline'].mean() * 100
                el_avg = dq['Endline'].mean() * 100
                domain_rows.append({
                    'Domain': domain,
                    'Questions': len(dq),
                    'Baseline_Avg': round(bl_avg, 1),
                    'Endline_Avg': round(el_avg, 1),
                    'Change': round(el_avg - bl_avg, 1),
                })
            domain_df = pd.DataFrame(domain_rows)

            if domain_df.empty:
                st.info("Domain grouping could not be computed.")
            else:
                # KPI: best and weakest domains
                best = domain_df.loc[domain_df['Change'].idxmax()]
                weakest = domain_df.loc[domain_df['Change'].idxmin()]
                dc1, dc2 = st.columns(2)
                dc1.markdown(f"""<div class="kpi-card">
                    <h3>Highest Improvement</h3>
                    <div class="value">{best['Domain']}</div>
                    <div class="delta-positive">{best['Change']:+.1f}pp gain</div>
                </div>""", unsafe_allow_html=True)
                dc2.markdown(f"""<div class="kpi-card">
                    <h3>Needs Most Attention</h3>
                    <div class="value">{weakest['Domain']}</div>
                    <div class="delta-{'negative' if weakest['Change']<5 else 'positive'}">{weakest['Change']:+.1f}pp</div>
                </div>""", unsafe_allow_html=True)

                st.markdown("<br>", unsafe_allow_html=True)

                # Grouped bar by domain
                _section_header('', 'Average Score by Domain', 'Pre vs Post')
                dom_sorted = domain_df.sort_values('Change', ascending=True)
                fig_dom = go.Figure()
                fig_dom.add_trace(go.Bar(
                    y=dom_sorted['Domain'], x=dom_sorted['Baseline_Avg'],
                    name='Baseline', orientation='h',
                    marker_color=COLORS['baseline'],
                    text=dom_sorted['Baseline_Avg'].apply(lambda x: f"{x:.1f}%"),
                    textposition='auto',
                ))
                fig_dom.add_trace(go.Bar(
                    y=dom_sorted['Domain'], x=dom_sorted['Endline_Avg'],
                    name='Endline', orientation='h',
                    marker_color=COLORS['midline'],
                    text=dom_sorted['Endline_Avg'].apply(lambda x: f"{x:.1f}%"),
                    textposition='auto',
                ))
                fig_dom.update_layout(
                    title='Average % Correct by Thematic Domain',
                    barmode='group', height=420,
                    xaxis_title='% Correct',
                    legend=dict(orientation='h', yanchor='bottom', y=1.02, x=0.5, xanchor='center'),
                    font=dict(size=13, color='#333'),
                    title_font=dict(size=16, color='#222'),
                    plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                    margin=dict(l=20, r=20, t=60, b=20),
                )
                st.plotly_chart(fig_dom, use_container_width=True)

                # Improvement bar
                colors_dom = [COLORS['good'] if c > 10 else (COLORS['danger'] if c < 0 else '#FF9800')
                              for c in dom_sorted['Change']]
                fig_dom_chg = go.Figure(go.Bar(
                    y=dom_sorted['Domain'], x=dom_sorted['Change'],
                    orientation='h', marker_color=colors_dom,
                    text=dom_sorted['Change'].apply(lambda x: f"{x:+.1f}pp"),
                    textposition='auto',
                ))
                fig_dom_chg.update_layout(
                    title='Knowledge Improvement by Domain (pp)',
                    height=380, xaxis_title='Percentage Point Change',
                    font=dict(size=13, color='#333'),
                    title_font=dict(size=16, color='#222'),
                    plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                    margin=dict(l=20, r=20, t=60, b=20),
                )
                fig_dom_chg.add_vline(x=0, line_color='#666', line_width=1)
                st.plotly_chart(fig_dom_chg, use_container_width=True)

                # Data table
                with st.expander("View Domain Data Table"):
                    st.dataframe(domain_df.style.format({
                        'Baseline_Avg': '{:.1f}%', 'Endline_Avg': '{:.1f}%', 'Change': '{:+.1f}pp'
                    }), use_container_width=True)


# ============================================================================
# MANGROVE TRAINING RENDERER
# ============================================================================

def render_mangrove_training_tabs(m_data, timepoint_filter='Combined'):
    """Render the Mangrove Training Pre/Post Knowledge Assessment tabs."""

    thresholds = m_data['thresholds']
    scores = m_data['scores']
    adequate_county = m_data['adequate_county']
    adequate_sex = m_data['adequate_sex']
    sex_threshold = m_data.get('sex_threshold', pd.DataFrame())

    tab1, tab2, tab3 = st.tabs([
        "Overview & Knowledge Gains",
        "County-Level Performance",
        "Sex Disaggregation",
    ])

    # ==================================================================
    # TAB 1 — Overview & Knowledge Gains
    # ==================================================================
    with tab1:
        st.markdown("""<div class="section-narrative">
        <strong>Mangrove Training Overview:</strong> Key knowledge assessment results
        comparing Pre-Test (Baseline) with Post-Test (Endline) performance on mangrove
        restoration. The \u226560 points threshold indicates adequate knowledge.
        </div>""", unsafe_allow_html=True)

        # ---- KPI cards ----
        all_county = adequate_county[adequate_county['County'] == 'All']
        pre_60 = float(all_county[all_county['Timepoint'] == 'Pre-Test']['Value'].values[0]) if len(all_county[all_county['Timepoint'] == 'Pre-Test']) else 0
        post_60 = float(all_county[all_county['Timepoint'] == 'Post-Test']['Value'].values[0]) if len(all_county[all_county['Timepoint'] == 'Post-Test']) else 0
        change_60 = (post_60 - pre_60) * 100

        all_scores = scores[scores['County'] == 'All']
        avg_score = float(all_scores['AvgScore'].values[0]) if len(all_scores) else 0
        total_n = int(all_scores['Respondents'].values[0]) if len(all_scores) else 0

        k1, k2, k3, k4 = st.columns(4)
        k1.markdown(f"""<div class="kpi-card">
            <h3>\u226560 Pass Rate (Post)</h3>
            <div class="value">{post_60*100:.1f}%</div>
            <div class="delta-{'positive' if change_60>0 else 'negative'}">{change_60:+.1f}pp vs Pre</div>
        </div>""", unsafe_allow_html=True)
        k2.markdown(f"""<div class="kpi-card">
            <h3>\u226560 Pass Rate (Pre)</h3>
            <div class="value">{pre_60*100:.1f}%</div>
            <div class="delta-neutral">Baseline</div>
        </div>""", unsafe_allow_html=True)
        k3.markdown(f"""<div class="kpi-card">
            <h3>Average Score</h3>
            <div class="value">{avg_score:.1f}%</div>
            <div class="delta-neutral">Overall</div>
        </div>""", unsafe_allow_html=True)
        k4.markdown(f"""<div class="kpi-card">
            <h3>Total Respondents</h3>
            <div class="value">{total_n:,}</div>
            <div class="delta-neutral">All counties</div>
        </div>""", unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        # ---- Pass thresholds by county ----
        _section_header('', 'Pass Rates by Threshold & County', 'Pre-Test')

        thr_plot = thresholds.copy()
        thr_plot['Pct'] = thr_plot['Value'] * 100
        fig_thr = go.Figure()
        county_colors = {'Kilifi': COLORS['baseline'], 'Kwale': COLORS['midline'], 'All': '#FF9800'}
        for county in thr_plot['County'].unique():
            sub = thr_plot[thr_plot['County'] == county]
            fig_thr.add_trace(go.Bar(
                x=sub['Threshold'], y=sub['Pct'], name=county,
                marker_color=county_colors.get(county, '#666'),
                text=sub['Pct'].apply(lambda x: f"{x:.1f}%"), textposition='auto',
            ))
        fig_thr.update_layout(
            title='Percentage of Participants Achieving Score Thresholds by County',
            barmode='group', height=420, yaxis_title='% of Participants',
            legend=dict(orientation='h', yanchor='bottom', y=1.02, x=0.5, xanchor='center'),
            font=dict(size=13, color='#333'),
            title_font=dict(size=16, color='#222'),
            plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
            margin=dict(l=20, r=20, t=60, b=20),
        )
        st.plotly_chart(fig_thr, use_container_width=True)

        # ---- Adequate knowledge Pre vs Post (all) ----
        _section_header('', 'Adequate Knowledge (\u226560%) Pre vs Post', 'Overall')
        all_pre_post = adequate_county[adequate_county['County'] == 'All'].copy()
        all_pre_post['Pct'] = all_pre_post['Value'] * 100
        fig_pp = go.Figure()
        for _, row in all_pre_post.iterrows():
            color = COLORS['baseline'] if row['Timepoint'] == 'Pre-Test' else COLORS['midline']
            fig_pp.add_trace(go.Bar(
                x=[row['Timepoint']], y=[row['Pct']],
                name=row['Timepoint'], marker_color=color,
                text=[f"{row['Pct']:.1f}%"], textposition='auto', width=0.5,
            ))
        fig_pp.update_layout(
            title='Overall Adequate Knowledge (\u226560 points)',
            height=350, yaxis_title='% of Participants', showlegend=False,
            font=dict(size=13, color='#333'),
            title_font=dict(size=15, color='#222'),
            plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
            margin=dict(l=20, r=20, t=50, b=20), yaxis=dict(range=[0, 105]),
        )
        st.plotly_chart(fig_pp, use_container_width=True)

    # ==================================================================
    # TAB 2 — County-Level Performance
    # ==================================================================
    with tab2:
        st.markdown("""<div class="section-narrative">
        <strong>County Comparison:</strong> Comparing Kilifi and Kwale counties on
        mangrove restoration knowledge test performance, including pre-to-post improvements
        and score distributions.
        </div>""", unsafe_allow_html=True)

        # ---- Adequate knowledge by county Pre vs Post ----
        _section_header('', 'Adequate Knowledge (\u226560%) by County', 'Pre vs Post')

        county_plot = adequate_county.copy()
        county_plot['Pct'] = county_plot['Value'] * 100

        if timepoint_filter == 'Combined':
            plot_data = county_plot
        elif timepoint_filter == 'Baseline':
            plot_data = county_plot[county_plot['Timepoint'] == 'Pre-Test']
        else:
            plot_data = county_plot[county_plot['Timepoint'] == 'Post-Test']

        fig_county = go.Figure()
        for tp in plot_data['Timepoint'].unique():
            sub = plot_data[plot_data['Timepoint'] == tp]
            color = COLORS['baseline'] if tp == 'Pre-Test' else COLORS['midline']
            fig_county.add_trace(go.Bar(
                x=sub['County'], y=sub['Pct'], name=tp,
                marker_color=color,
                text=sub['Pct'].apply(lambda x: f"{x:.1f}%"), textposition='auto',
            ))
        fig_county.update_layout(
            title='Adequate Knowledge by County (Pre vs Post)',
            barmode='group', height=420, yaxis_title='% of Participants',
            legend=dict(orientation='h', yanchor='bottom', y=1.02, x=0.5, xanchor='center'),
            font=dict(size=13, color='#333'),
            title_font=dict(size=16, color='#222'),
            plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
            margin=dict(l=20, r=20, t=60, b=20), yaxis=dict(range=[0, 105]),
        )
        st.plotly_chart(fig_county, use_container_width=True)

        # ---- Change (pp) by county ----
        change_data = []
        for county in adequate_county['County'].unique():
            c_data = adequate_county[adequate_county['County'] == county]
            pre = c_data[c_data['Timepoint'] == 'Pre-Test']['Value'].values
            post = c_data[c_data['Timepoint'] == 'Post-Test']['Value'].values
            if len(pre) and len(post):
                change_data.append({'County': county, 'Change': (float(post[0]) - float(pre[0])) * 100})
        if change_data:
            ch_df = pd.DataFrame(change_data)
            ch_colors = [COLORS['good'] if c > 0 else COLORS['danger'] for c in ch_df['Change']]
            fig_ch = go.Figure(go.Bar(
                x=ch_df['County'], y=ch_df['Change'],
                marker_color=ch_colors,
                text=ch_df['Change'].apply(lambda x: f"{x:+.1f}pp"), textposition='auto',
            ))
            fig_ch.update_layout(
                title='Knowledge Improvement by County (pp)',
                height=350, yaxis_title='Percentage Point Change', showlegend=False,
                font=dict(size=13, color='#333'),
                title_font=dict(size=15, color='#222'),
                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                margin=dict(l=20, r=20, t=50, b=20),
            )
            fig_ch.add_hline(y=0, line_color='#666', line_width=1)
            st.plotly_chart(fig_ch, use_container_width=True)

        # ---- Standardized scores by county ----
        _section_header('', 'Standardized Scores by County', 'Avg / Max / Min')

        sc1, sc2 = st.columns(2)
        with sc1:
            fig_avg_c = go.Figure()
            for _, row in scores.iterrows():
                color = county_colors.get(row['County'], '#666')
                fig_avg_c.add_trace(go.Bar(
                    x=[row['County']], y=[row['AvgScore']],
                    name=row['County'], marker_color=color,
                    text=[f"{row['AvgScore']:.1f}%"], textposition='auto', width=0.5,
                ))
            fig_avg_c.update_layout(
                title='Average Score by County', height=350,
                yaxis_title='Score (%)', showlegend=False,
                font=dict(size=13, color='#333'),
                title_font=dict(size=15, color='#222'),
                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                margin=dict(l=20, r=20, t=50, b=20), yaxis=dict(range=[0, 105]),
            )
            st.plotly_chart(fig_avg_c, use_container_width=True)

        with sc2:
            fig_range = go.Figure()
            for _, row in scores.iterrows():
                color = county_colors.get(row['County'], '#666')
                fig_range.add_trace(go.Bar(
                    x=[f"{row['County']} Min", f"{row['County']} Max"],
                    y=[row['MinScore'], row['MaxScore']],
                    name=row['County'], marker_color=color,
                    text=[f"{row['MinScore']:.1f}", f"{row['MaxScore']:.1f}"],
                    textposition='auto', width=0.5,
                ))
            fig_range.update_layout(
                title='Score Range by County (Min & Max)', height=350,
                yaxis_title='Score (%)', showlegend=False,
                font=dict(size=13, color='#333'),
                title_font=dict(size=15, color='#222'),
                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                margin=dict(l=20, r=20, t=50, b=20), yaxis=dict(range=[0, 105]),
            )
            st.plotly_chart(fig_range, use_container_width=True)

        # Respondent breakdown
        _section_header('', 'Respondents by County', '')
        resp_data = scores[['County', 'Respondents']].copy()
        fig_resp = go.Figure(go.Bar(
            x=resp_data['County'], y=resp_data['Respondents'],
            marker_color=[county_colors.get(c, '#666') for c in resp_data['County']],
            text=resp_data['Respondents'].apply(lambda x: f"{x:,}"), textposition='auto',
        ))
        fig_resp.update_layout(
            title='Number of Respondents by County', height=320,
            yaxis_title='Count', showlegend=False,
            font=dict(size=13, color='#333'),
            title_font=dict(size=15, color='#222'),
            plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
            margin=dict(l=20, r=20, t=50, b=20),
        )
        st.plotly_chart(fig_resp, use_container_width=True)

        # Data table
        with st.expander("View Scores Data Table"):
            st.dataframe(scores.style.format({
                'AvgScore': '{:.2f}%', 'MaxScore': '{:.1f}', 'MinScore': '{:.1f}',
                'Respondents': '{:,.0f}'
            }), use_container_width=True)

    # ==================================================================
    # TAB 3 — Sex Disaggregation
    # ==================================================================
    with tab3:
        st.markdown("""<div class="section-narrative">
        <strong>Sex Disaggregation:</strong> Comparing male and female participants'
        adequate knowledge levels before and after training, highlighting improvements
        and any gender gaps in mangrove restoration knowledge.
        </div>""", unsafe_allow_html=True)

        # ---- Adequate knowledge by sex Pre vs Post ----
        _section_header('', 'Adequate Knowledge (\u226560%) by Sex', 'Pre vs Post')

        sex_plot = adequate_sex.copy()
        sex_plot['Pct'] = sex_plot['Value'] * 100

        fig_sex = go.Figure()
        sex_colors = {'Pre-Test': COLORS['baseline'], 'Post-Test': COLORS['midline']}
        for tp in ['Pre-Test', 'Post-Test']:
            sub = sex_plot[sex_plot['Timepoint'] == tp]
            fig_sex.add_trace(go.Bar(
                x=sub['Sex'], y=sub['Pct'], name=tp,
                marker_color=sex_colors[tp],
                text=sub['Pct'].apply(lambda x: f"{x:.1f}%"), textposition='auto',
            ))
        fig_sex.update_layout(
            title='Adequate Knowledge by Sex (Pre vs Post)',
            barmode='group', height=420, yaxis_title='% of Participants',
            legend=dict(orientation='h', yanchor='bottom', y=1.02, x=0.5, xanchor='center'),
            font=dict(size=13, color='#333'),
            title_font=dict(size=16, color='#222'),
            plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
            margin=dict(l=20, r=20, t=60, b=20), yaxis=dict(range=[0, 105]),
        )
        st.plotly_chart(fig_sex, use_container_width=True)

        # ---- Improvement by sex ----
        sex_change = []
        for sex in adequate_sex['Sex'].unique():
            s_data = adequate_sex[adequate_sex['Sex'] == sex]
            pre = s_data[s_data['Timepoint'] == 'Pre-Test']['Value'].values
            post = s_data[s_data['Timepoint'] == 'Post-Test']['Value'].values
            if len(pre) and len(post):
                sex_change.append({'Sex': sex, 'Change': (float(post[0]) - float(pre[0])) * 100})
        if sex_change:
            sch_df = pd.DataFrame(sex_change)
            sch_colors = [COLORS['good'] if c > 0 else COLORS['danger'] for c in sch_df['Change']]
            fig_sch = go.Figure(go.Bar(
                x=sch_df['Sex'], y=sch_df['Change'],
                marker_color=sch_colors,
                text=sch_df['Change'].apply(lambda x: f"{x:+.1f}pp"), textposition='auto',
            ))
            fig_sch.update_layout(
                title='Knowledge Improvement by Sex (pp)',
                height=350, yaxis_title='Percentage Point Change', showlegend=False,
                font=dict(size=13, color='#333'),
                title_font=dict(size=15, color='#222'),
                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                margin=dict(l=20, r=20, t=50, b=20),
            )
            fig_sch.add_hline(y=0, line_color='#666', line_width=1)
            st.plotly_chart(fig_sch, use_container_width=True)

        # ---- Gender gap visualisation ----
        _section_header('', 'Gender Gap Analysis', 'Post-Training')
        sex_post = adequate_sex[adequate_sex['Timepoint'] == 'Post-Test'].copy()
        sex_post['Pct'] = sex_post['Value'] * 100
        male_post = sex_post[sex_post['Sex'] == 'Male']['Pct'].values
        female_post = sex_post[sex_post['Sex'] == 'Female']['Pct'].values
        if len(male_post) and len(female_post):
            gap = float(female_post[0]) - float(male_post[0])
            gap_color = COLORS['good'] if gap >= 0 else COLORS['danger']
            gc1, gc2, gc3 = st.columns(3)
            gc1.markdown(f"""<div class="kpi-card">
                <h3>Male (Post)</h3>
                <div class="value">{float(male_post[0]):.1f}%</div>
            </div>""", unsafe_allow_html=True)
            gc2.markdown(f"""<div class="kpi-card">
                <h3>Female (Post)</h3>
                <div class="value">{float(female_post[0]):.1f}%</div>
            </div>""", unsafe_allow_html=True)
            gc3.markdown(f"""<div class="kpi-card">
                <h3>Gender Gap (F-M)</h3>
                <div class="value" style="color:{gap_color}">{gap:+.1f}pp</div>
            </div>""", unsafe_allow_html=True)

        # ---- Sex-level threshold from pre-test (if available) ----
        if not sex_threshold.empty:
            st.markdown("<br>", unsafe_allow_html=True)
            _section_header('', 'Pre-Test \u226560% Pass by Sex', 'Baseline')
            sex_thr_plot = sex_threshold.copy()
            sex_thr_plot['Pct'] = sex_thr_plot['Value_60'] * 100
            fig_st = go.Figure(go.Bar(
                x=sex_thr_plot['Sex'], y=sex_thr_plot['Pct'],
                marker_color=[COLORS['baseline'], '#E91E63'],
                text=sex_thr_plot['Pct'].apply(lambda x: f"{x:.1f}%"), textposition='auto',
            ))
            fig_st.update_layout(
                title='Pre-Test \u226560% Pass Rate by Sex',
                height=320, yaxis_title='% of Participants', showlegend=False,
                font=dict(size=13, color='#333'),
                title_font=dict(size=15, color='#222'),
                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                margin=dict(l=20, r=20, t=50, b=20), yaxis=dict(range=[0, 105]),
            )
            st.plotly_chart(fig_st, use_container_width=True)

        # Data tables
        with st.expander("View Adequate Knowledge Data"):
            st.markdown("**By Sex (Pre/Post)**")
            display_sex = adequate_sex.copy()
            display_sex['Pct'] = display_sex['Value'] * 100
            st.dataframe(display_sex[['Sex', 'Timepoint', 'Pct']].rename(
                columns={'Pct': '% Adequate'}), use_container_width=True)


# ============================================================================
# SEAWEED PRODUCTION & CHALLENGES RENDERER — 5 TABS (ENHANCED)
# ============================================================================

def render_seaweed_tabs(sw_df, group_filter=None, casual_filter='All',
                        challenge_filter=None, min_total_kg=0,
                        min_achievement_pct=0):
    """Render the Seaweed Production & Challenges module with 5 enhanced tabs."""
    # ---- Apply sidebar filters ----
    df = sw_df.copy()
    if group_filter:
        df = df[df['Group'].isin(group_filter)]
    if casual_filter != 'All':
        df = df[df['Casual_Workers'].str.lower() == casual_filter.lower()]
    if challenge_filter:
        flag_map = {
            'Transport': 'flag_transport', 'Market Access': 'flag_market',
            'Disease': 'flag_disease', 'Equipment': 'flag_equipment',
            'Storage': 'flag_storage', 'Labour': 'flag_labour',
            'Sand / Tide': 'flag_sand_tide',
        }
        for cf in challenge_filter:
            fc = flag_map.get(cf) if isinstance(cf, str) and not cf.startswith('flag_') else cf
            if fc and fc in df.columns:
                df = df[df[fc] == True]
    if min_total_kg > 0:
        df = df[df['Total_KG'] >= min_total_kg]
    if min_achievement_pct > 0:
        df = df[df['Ropes_Achievement_pct'] >= min_achievement_pct]

    if len(df) == 0:
        st.warning("No data matches the current filters. Please adjust the sidebar filters.")
        return

    agg = prepare_seaweed_aggregates(df)
    grp_df = agg['group_summary']
    ch_df = agg['challenge_counts']
    ov = agg['overall']

    # Color palette used across tabs
    palette = [COLORS['baseline'], COLORS['midline'], '#FF9800', '#9C27B0',
               '#00BCD4', '#795548', '#E91E63', '#607D8B']
    group_colors_map = {}
    for i, g in enumerate(sorted(df['Group'].dropna().unique())):
        group_colors_map[g] = palette[i % len(palette)]

    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "Overview & KPIs",
        "Group Performance",
        "Production & Yields",
        "Challenges & Constraints",
        "Map View",
    ])

    # ==================================================================
    # TAB 1 — Overview & KPIs (ENHANCED)
    # ==================================================================
    with tab1:
        st.markdown("""<div class="section-narrative">
        <strong>Seaweed Production Overview:</strong> High-level indicators summarising
        production volume, rope deployment, farmer participation, workforce composition,
        target achievement, and production gaps across all seaweed farming groups.
        </div>""", unsafe_allow_html=True)

        # ---- KPI cards — Row 1 (5 cards) ----
        k1, k2, k3, k4, k5 = st.columns(5)
        k1.markdown(f"""<div class="kpi-card">
            <h3>Total Production</h3>
            <div class="value">{ov['total_kg']:,.1f} kg</div>
            <div class="delta-neutral">{ov['dried_kg']:,.0f} dried + {ov['wet_kg']:,.0f} wet</div>
        </div>""", unsafe_allow_html=True)
        k2.markdown(f"""<div class="kpi-card">
            <h3>Seaweed Farmers</h3>
            <div class="value">{ov['n_farmers']:,}</div>
            <div class="delta-neutral">{ov['n_groups']} groups</div>
        </div>""", unsafe_allow_html=True)
        k3.markdown(f"""<div class="kpi-card">
            <h3>Ropes in Ocean</h3>
            <div class="value">{ov['ropes_ocean']:,.0f}</div>
            <div class="delta-neutral">of {ov['ropes_total']:,.0f} total</div>
        </div>""", unsafe_allow_html=True)
        k4.markdown(f"""<div class="kpi-card">
            <h3>Avg Prod / Rope</h3>
            <div class="value">{ov['avg_prod_per_rope']:.2f} kg</div>
            <div class="delta-neutral">{ov['avg_production_per_farmer']:.1f} kg/farmer</div>
        </div>""", unsafe_allow_html=True)
        k5.markdown(f"""<div class="kpi-card">
            <h3>Casual Workers</h3>
            <div class="value">{ov['casual_pct']:.1f}%</div>
            <div class="delta-neutral">of farmers</div>
        </div>""", unsafe_allow_html=True)

        # ---- KPI cards — Row 2 (4 cards) ----
        r1, r2, r3, r4 = st.columns(4)
        ach_trend = 'positive' if ov['avg_achievement_pct'] >= 70 else 'neutral'
        r1.markdown(f"""<div class="kpi-card">
            <h3>Target Achievement</h3>
            <div class="value">{ov['avg_achievement_pct']:.1f}%</div>
            <div class="delta-{ach_trend}">{ov['pct_meeting_target']:.0f}% meet target</div>
        </div>""", unsafe_allow_html=True)
        r2.markdown(f"""<div class="kpi-card">
            <h3>Dried / Wet Ratio</h3>
            <div class="value">{ov['dried_wet_ratio']:.2f}</div>
            <div class="delta-neutral">{ov['dried_kg']:,.0f} vs {ov['wet_kg']:,.0f}</div>
        </div>""", unsafe_allow_html=True)
        r3.markdown(f"""<div class="kpi-card">
            <h3>Rope Gap</h3>
            <div class="value">{ov['gap_total']:,.0f}</div>
            <div class="delta-negative">ropes needed</div>
        </div>""", unsafe_allow_html=True)
        r4.markdown(f"""<div class="kpi-card">
            <h3>Multi-Challenge</h3>
            <div class="value">{ov['multi_challenge_pct']:.1f}%</div>
            <div class="delta-{'negative' if ov['multi_challenge_pct'] > 30 else 'neutral'}">face 2+ challenges</div>
        </div>""", unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        # ---- Two-column: Production bar + Group share pie ----
        col_bar, col_pie = st.columns([3, 2])
        with col_bar:
            _section_header('', 'Total Production by Group (kg)', '2025')
            top_prod = grp_df.sort_values('Total_KG', ascending=False)
            fig_tp = go.Figure(go.Bar(
                x=top_prod['Group'], y=top_prod['Total_KG'],
                marker_color=[group_colors_map.get(g, COLORS['baseline']) for g in top_prod['Group']],
                text=top_prod['Total_KG'].apply(lambda v: f"{v:,.0f}"),
                textposition='auto',
            ))
            fig_tp.update_layout(
                title='Total Seaweed Production by Group (kg)',
                height=420, yaxis_title='Total KG',
                font=dict(size=13, color='#333'),
                title_font=dict(size=16, color='#222'),
                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                margin=dict(l=20, r=20, t=60, b=20),
                xaxis_tickangle=-25,
            )
            st.plotly_chart(fig_tp, use_container_width=True)

        with col_pie:
            _section_header('', 'Production Share by Group', '')
            fig_pie = go.Figure(go.Pie(
                labels=grp_df['Group'], values=grp_df['Total_KG'],
                hole=0.4,
                marker=dict(colors=[group_colors_map.get(g, '#999') for g in grp_df['Group']]),
                textinfo='label+percent',
                textposition='auto',
                hovertemplate='<b>%{label}</b><br>%{value:,.0f} kg (%{percent})<extra></extra>',
            ))
            fig_pie.update_layout(
                title='Production Share (%)',
                height=420, showlegend=False,
                font=dict(size=12, color='#333'),
                title_font=dict(size=15, color='#222'),
                margin=dict(l=10, r=10, t=60, b=10),
            )
            st.plotly_chart(fig_pie, use_container_width=True)

        # ---- Ropes in Ocean vs Target ----
        _section_header('', 'Ropes in Ocean vs Target by Group', '')
        fig_rt = go.Figure()
        fig_rt.add_trace(go.Bar(
            x=grp_df['Group'], y=grp_df['Ropes_Ocean'],
            name='Ropes in Ocean', marker_color=COLORS['baseline'],
            text=grp_df['Ropes_Ocean'].apply(lambda v: f"{v:,.0f}"), textposition='auto',
        ))
        fig_rt.add_trace(go.Bar(
            x=grp_df['Group'], y=grp_df['Target_Ropes'],
            name='Target Ropes', marker_color=COLORS['midline'],
            text=grp_df['Target_Ropes'].apply(lambda v: f"{v:,.0f}"), textposition='auto',
        ))
        fig_rt.update_layout(
            barmode='group', height=420, yaxis_title='Ropes',
            legend=dict(orientation='h', yanchor='bottom', y=1.02, x=0.5, xanchor='center'),
            font=dict(size=13, color='#333'),
            title_font=dict(size=16, color='#222'),
            plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
            margin=dict(l=20, r=20, t=60, b=20), xaxis_tickangle=-25,
        )
        st.plotly_chart(fig_rt, use_container_width=True)

        # ---- Achievement traffic-light ----
        _section_header('', 'Target Achievement by Group', 'Avg % of target ropes')
        ach_df = grp_df[['Group', 'Avg_Achievement_pct']].copy()
        ach_colors = []
        for v in ach_df['Avg_Achievement_pct']:
            if v >= 90:
                ach_colors.append(COLORS['good'])
            elif v >= 60:
                ach_colors.append('#FF9800')
            else:
                ach_colors.append(COLORS['danger'])
        fig_ach = go.Figure(go.Bar(
            x=ach_df['Group'], y=ach_df['Avg_Achievement_pct'],
            marker_color=ach_colors,
            text=ach_df['Avg_Achievement_pct'].apply(lambda v: f"{v:.1f}%"),
            textposition='auto',
        ))
        fig_ach.update_layout(
            title='Average Ropes Target Achievement by Group',
            height=380, yaxis_title='% of Target',
            font=dict(size=13, color='#333'),
            title_font=dict(size=15, color='#222'),
            plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
            margin=dict(l=20, r=20, t=50, b=20),
            xaxis_tickangle=-25,
        )
        fig_ach.add_hline(y=100, line_dash='dash', line_color='#2E7D32', line_width=1,
                          annotation_text='Target 100%', annotation_position='top left')
        st.plotly_chart(fig_ach, use_container_width=True)

        # ---- Gap Analysis Waterfall ----
        _section_header('', 'Rope Gap Analysis by Group', 'Target vs Actual shortfall')
        gap_df = grp_df.sort_values('Gap', ascending=False)
        fig_gap = go.Figure(go.Waterfall(
            x=gap_df['Group'],
            y=gap_df['Gap'],
            text=gap_df['Gap'].apply(lambda v: f"{v:,.0f}"),
            textposition='auto',
            connector=dict(line=dict(color='#ccc', width=1)),
            increasing=dict(marker=dict(color=COLORS['danger'])),
            decreasing=dict(marker=dict(color=COLORS['good'])),
            totals=dict(marker=dict(color=COLORS['baseline'])),
        ))
        fig_gap.update_layout(
            title='Rope Gap by Group (Target − Actual)',
            height=380, yaxis_title='Gap (Ropes)',
            font=dict(size=13, color='#333'), title_font=dict(size=15, color='#222'),
            plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
            margin=dict(l=20, r=20, t=50, b=20), xaxis_tickangle=-25,
        )
        st.plotly_chart(fig_gap, use_container_width=True)

    # ==================================================================
    # TAB 2 — Group Performance & Target Achievement (ENHANCED)
    # ==================================================================
    with tab2:
        st.markdown("""<div class="section-narrative">
        <strong>Group Performance:</strong> Multi-dimensional comparison of rope deployment,
        target achievement, production efficiency, and workforce composition across seaweed
        farming groups. Includes radar profiling and member-level performance analysis.
        </div>""", unsafe_allow_html=True)

        # ---- Radar chart: multi-metric group comparison ----
        _section_header('', 'Group Performance Radar', 'Normalised 0–100')
        radar_metrics = ['Avg_Achievement_pct', 'Avg_Prod_per_Rope', 'Members',
                         'Total_KG', 'Ropes_Total']
        radar_labels = ['Target Achievement', 'Prod/Rope', 'Members', 'Total Production', 'Total Ropes']
        if len(grp_df) >= 2:
            # Normalise each metric 0-100
            radar_norm = grp_df[radar_metrics].copy()
            for c in radar_metrics:
                cmax = radar_norm[c].max()
                if cmax > 0:
                    radar_norm[c] = (radar_norm[c] / cmax * 100).round(1)
            fig_radar = go.Figure()
            for idx, row in grp_df.iterrows():
                vals = [radar_norm.loc[idx, c] for c in radar_metrics]
                vals.append(vals[0])   # close polygon
                fig_radar.add_trace(go.Scatterpolar(
                    r=vals,
                    theta=radar_labels + [radar_labels[0]],
                    name=row['Group'],
                    fill='toself',
                    opacity=0.3,
                    line=dict(color=group_colors_map.get(row['Group'], '#999')),
                ))
            fig_radar.update_layout(
                polar=dict(radialaxis=dict(visible=True, range=[0, 100])),
                height=500, showlegend=True,
                legend=dict(orientation='h', yanchor='top', y=-0.05, x=0.5, xanchor='center'),
                font=dict(size=12, color='#333'), title_font=dict(size=16, color='#222'),
                title='Multi-Metric Group Comparison',
                margin=dict(l=40, r=40, t=60, b=40),
            )
            st.plotly_chart(fig_radar, use_container_width=True)

        # ---- Grouped bar: Ropes_Total vs Target vs Required ----
        _section_header('', 'Ropes: Total vs Target vs Required', '')
        fig_rtr = go.Figure()
        fig_rtr.add_trace(go.Bar(x=grp_df['Group'], y=grp_df['Ropes_Total'],
                                  name='Ropes Total', marker_color=COLORS['baseline']))
        fig_rtr.add_trace(go.Bar(x=grp_df['Group'], y=grp_df['Target_Ropes'],
                                  name='Target Ropes', marker_color=COLORS['midline']))
        fig_rtr.add_trace(go.Bar(x=grp_df['Group'], y=grp_df['Ropes_Required'],
                                  name='Ropes Required', marker_color='#FF9800'))
        fig_rtr.update_layout(
            barmode='group', height=420, yaxis_title='Ropes',
            legend=dict(orientation='h', yanchor='bottom', y=1.02, x=0.5, xanchor='center'),
            font=dict(size=13, color='#333'), title_font=dict(size=16, color='#222'),
            plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
            margin=dict(l=20, r=20, t=60, b=20), xaxis_tickangle=-25,
        )
        st.plotly_chart(fig_rtr, use_container_width=True)

        # ---- Production per rope by group (box + violin toggle) ----
        _section_header('', 'Production per Rope (kg) Distribution', '')
        chart_type = st.radio("Chart type", ["Box Plot", "Violin Plot"],
                              horizontal=True, key='sw_box_violin_toggle')
        fig_ppr = go.Figure()
        for grp_name in sorted(df['Group'].dropna().unique()):
            sub = df[df['Group'] == grp_name]
            valid = sub[sub['Production_per_rope_kg'] > 0]['Production_per_rope_kg']
            if len(valid):
                if chart_type == "Box Plot":
                    fig_ppr.add_trace(go.Box(y=valid, name=grp_name, boxmean=True,
                                             marker_color=group_colors_map.get(grp_name, '#999')))
                else:
                    fig_ppr.add_trace(go.Violin(y=valid, name=grp_name, box_visible=True,
                                                meanline_visible=True,
                                                fillcolor=group_colors_map.get(grp_name, '#999'),
                                                opacity=0.6, line_color='#333'))
        fig_ppr.update_layout(
            title=f'Production per Rope by Group ({chart_type})',
            height=420, yaxis_title='kg per Rope', showlegend=False,
            font=dict(size=13, color='#333'), title_font=dict(size=15, color='#222'),
            plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
            margin=dict(l=20, r=20, t=50, b=20),
        )
        st.plotly_chart(fig_ppr, use_container_width=True)

        # ---- Member-level performance scatter ----
        _section_header('', 'Member-Level Performance', 'Achievement % vs Production')
        fig_mem = go.Figure()
        for g in sorted(df['Group'].dropna().unique()):
            sub = df[df['Group'] == g]
            fig_mem.add_trace(go.Scatter(
                x=sub['Ropes_Achievement_pct'], y=sub['Total_KG'],
                mode='markers', name=g,
                marker=dict(size=9, color=group_colors_map.get(g, '#999'),
                            line=dict(width=0.5, color='white')),
                hovertemplate=(
                    '<b>%{customdata[0]}</b><br>Group: %{customdata[1]}<br>'
                    'Achievement: %{x:.1f}%<br>Production: %{y:.1f} kg<extra></extra>'),
                customdata=sub[['Member', 'Group']].values,
            ))
        fig_mem.update_layout(
            title='Member Performance: Target Achievement vs Production',
            height=480, xaxis_title='Ropes Achievement (%)', yaxis_title='Total Production (kg)',
            legend=dict(orientation='h', yanchor='bottom', y=1.02, x=0.5, xanchor='center'),
            font=dict(size=13, color='#333'), title_font=dict(size=16, color='#222'),
            plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
            margin=dict(l=20, r=20, t=60, b=20),
        )
        fig_mem.add_vline(x=100, line_dash='dash', line_color='#2E7D32', line_width=1,
                          annotation_text='Target 100%')
        st.plotly_chart(fig_mem, use_container_width=True)

        # ---- Target achievement breakdown per group table + chart ----
        _section_header('', 'Target Achievement Breakdown', 'Per group summary')
        ach_rows = []
        for grp_name in sorted(df['Group'].dropna().unique()):
            sub = df[df['Group'] == grp_name]
            n = len(sub)
            if n == 0:
                continue
            meeting = len(sub[sub['Ropes_Total'] >= sub['Target_Ropes']])
            below50 = len(sub[sub['Ropes_Achievement_pct'] < 50])
            avg_ach = sub['Ropes_Achievement_pct'].mean()
            avg_prod = sub['Total_KG'].mean()
            ach_rows.append({
                'Group': grp_name, 'Members': n,
                'Avg Achievement (%)': round(avg_ach, 1),
                '≥ Target (%)': round(meeting / n * 100, 1),
                '< 50% Target (%)': round(below50 / n * 100, 1),
                'Avg Production (kg)': round(avg_prod, 1),
            })
        if ach_rows:
            ach_table = pd.DataFrame(ach_rows)
            try:
                styled = ach_table.style.format({
                    'Avg Achievement (%)': '{:.1f}%', '≥ Target (%)': '{:.1f}%',
                    '< 50% Target (%)': '{:.1f}%', 'Avg Production (kg)': '{:,.1f}',
                }).background_gradient(subset=['Avg Achievement (%)'], cmap='RdYlGn', vmin=0, vmax=100)
            except ImportError:
                styled = ach_table.style.format({
                    'Avg Achievement (%)': '{:.1f}%', '≥ Target (%)': '{:.1f}%',
                    '< 50% Target (%)': '{:.1f}%', 'Avg Production (kg)': '{:,.1f}',
                })
            st.dataframe(styled, use_container_width=True)

            # Stacked bar showing meeting vs not meeting
            fig_ab = go.Figure()
            fig_ab.add_trace(go.Bar(
                x=ach_table['Group'], y=ach_table['≥ Target (%)'],
                name='≥ Target', marker_color=COLORS['good'],
            ))
            fig_ab.add_trace(go.Bar(
                x=ach_table['Group'], y=ach_table['< 50% Target (%)'],
                name='< 50% Target', marker_color=COLORS['danger'],
            ))
            fig_ab.update_layout(
                barmode='group', height=380, yaxis_title='% of Members',
                legend=dict(orientation='h', yanchor='bottom', y=1.02, x=0.5, xanchor='center'),
                font=dict(size=13, color='#333'), title_font=dict(size=15, color='#222'),
                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                margin=dict(l=20, r=20, t=50, b=20), xaxis_tickangle=-25,
            )
            st.plotly_chart(fig_ab, use_container_width=True)

        # ---- Top / Bottom 10 members ----
        _section_header('', 'Top & Bottom Performers', 'By Total Production')
        col_top, col_bot = st.columns(2)
        ranked = df[['Member', 'Group', 'Total_KG', 'Ropes_Achievement_pct']].sort_values(
            'Total_KG', ascending=False).reset_index(drop=True)
        with col_top:
            st.markdown("**Top 10 Producers**")
            top10 = ranked.head(10).copy()
            top10.index = range(1, len(top10) + 1)
            st.dataframe(top10.style.format({
                'Total_KG': '{:,.1f}', 'Ropes_Achievement_pct': '{:.1f}%',
            }), use_container_width=True)
        with col_bot:
            st.markdown("**Bottom 10 Producers**")
            bot10 = ranked.tail(10).copy()
            bot10.index = range(1, len(bot10) + 1)
            st.dataframe(bot10.style.format({
                'Total_KG': '{:,.1f}', 'Ropes_Achievement_pct': '{:.1f}%',
            }), use_container_width=True)

    # ==================================================================
    # TAB 3 — Production & Yields (ENHANCED)
    # ==================================================================
    with tab3:
        st.markdown("""<div class="section-narrative">
        <strong>Production Analysis:</strong> Exploring the relationship between rope
        deployment and seaweed yields, production efficiency, Pareto analysis of farmer
        contributions, and dried vs wet production breakdown.
        </div>""", unsafe_allow_html=True)

        # ---- Scatter: Ropes_Total vs Total_KG with trendline ----
        _section_header('', 'Ropes vs Production', 'With trendline')
        fig_sc = go.Figure()
        for g in sorted(df['Group'].dropna().unique()):
            sub = df[df['Group'] == g]
            sizes = sub['Casual_Workers'].apply(
                lambda v: 14 if str(v).lower() == 'yes' else 8)
            fig_sc.add_trace(go.Scatter(
                x=sub['Ropes_Total'], y=sub['Total_KG'],
                mode='markers', name=g,
                marker=dict(size=sizes, color=group_colors_map[g],
                            line=dict(width=0.5, color='white')),
                hovertemplate=(
                    '<b>%{customdata[0]}</b><br>Group: %{customdata[1]}<br>'
                    'Ropes: %{x:.0f}<br>Production: %{y:.1f} kg<extra></extra>'),
                customdata=sub[['Member', 'Group']].values,
            ))

        # Add OLS trendline
        valid_scatter = df[(df['Ropes_Total'] > 0) & (df['Total_KG'] > 0)]
        if len(valid_scatter) > 3:
            x_vals = valid_scatter['Ropes_Total'].values
            y_vals = valid_scatter['Total_KG'].values
            z = np.polyfit(x_vals, y_vals, 1)
            p = np.poly1d(z)
            x_line = np.linspace(x_vals.min(), x_vals.max(), 50)
            fig_sc.add_trace(go.Scatter(
                x=x_line, y=p(x_line), mode='lines',
                name=f'Trend (slope={z[0]:.2f})',
                line=dict(color='#333', width=2, dash='dash'),
                showlegend=True,
            ))
            # Compute R²
            ss_res = ((y_vals - p(x_vals)) ** 2).sum()
            ss_tot = ((y_vals - y_vals.mean()) ** 2).sum()
            r_sq = 1 - ss_res / ss_tot if ss_tot > 0 else 0
            fig_sc.add_annotation(
                x=0.02, y=0.98, xref='paper', yref='paper',
                text=f'R² = {r_sq:.3f}', showarrow=False,
                font=dict(size=13, color='#333'),
                bgcolor='rgba(255,255,255,0.8)', bordercolor='#ccc',
            )

        fig_sc.update_layout(
            title='Ropes (Total) vs Production (kg)',
            height=480, xaxis_title='Ropes Total', yaxis_title='Total KG',
            legend=dict(orientation='h', yanchor='bottom', y=1.02, x=0.5, xanchor='center'),
            font=dict(size=13, color='#333'), title_font=dict(size=16, color='#222'),
            plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
            margin=dict(l=20, r=20, t=60, b=20),
        )
        st.plotly_chart(fig_sc, use_container_width=True)

        # ---- Pareto Analysis: Top farmers contributing 80% of production ----
        _section_header('', 'Pareto Analysis', 'Which farmers contribute most?')
        pareto = df[['Member', 'Group', 'Total_KG']].sort_values('Total_KG', ascending=False).reset_index(drop=True)
        pareto['Cumulative_KG'] = pareto['Total_KG'].cumsum()
        pareto['Cumulative_Pct'] = (pareto['Cumulative_KG'] / pareto['Total_KG'].sum() * 100).round(1)
        pareto['Farmer_Pct'] = ((pareto.index + 1) / len(pareto) * 100).round(1)

        n_80 = len(pareto[pareto['Cumulative_Pct'] <= 80]) + 1
        pct_farmers_80 = round(n_80 / len(pareto) * 100, 1)

        fig_pareto = go.Figure()
        fig_pareto.add_trace(go.Bar(
            x=list(range(1, len(pareto) + 1)),
            y=pareto['Total_KG'],
            name='Production (kg)',
            marker_color=COLORS['baseline'],
            opacity=0.7,
        ))
        fig_pareto.add_trace(go.Scatter(
            x=list(range(1, len(pareto) + 1)),
            y=pareto['Cumulative_Pct'],
            name='Cumulative %',
            yaxis='y2',
            line=dict(color=COLORS['midline'], width=2.5),
            mode='lines',
        ))
        fig_pareto.add_hline(y=80, line_dash='dot', line_color='red', line_width=1.5,
                             annotation_text='80% threshold', annotation_position='top right',
                             yref='y2')
        fig_pareto.update_layout(
            title=f'Pareto: {pct_farmers_80:.0f}% of farmers ({n_80}) produce 80% of output',
            height=420,
            xaxis_title='Farmers (ranked by production)',
            yaxis_title='Production (kg)',
            yaxis2=dict(title='Cumulative %', overlaying='y', side='right', range=[0, 105]),
            legend=dict(orientation='h', yanchor='bottom', y=1.02, x=0.5, xanchor='center'),
            font=dict(size=12, color='#333'), title_font=dict(size=14, color='#222'),
            plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
            margin=dict(l=20, r=50, t=60, b=20),
        )
        st.plotly_chart(fig_pareto, use_container_width=True)

        # ---- Production per rope histogram ----
        _section_header('', 'Production per Rope Distribution', '')
        sel_grp = st.selectbox('Select Group', ['All Groups'] + sorted(df['Group'].dropna().unique().tolist()),
                               key='sw_ppr_hist_grp')
        if sel_grp == 'All Groups':
            hist_data = df[df['Production_per_rope_kg'] > 0]['Production_per_rope_kg']
        else:
            hist_data = df[(df['Group'] == sel_grp) & (df['Production_per_rope_kg'] > 0)]['Production_per_rope_kg']

        if len(hist_data):
            fig_hist = go.Figure(go.Histogram(
                x=hist_data, nbinsx=25,
                marker_color=COLORS['baseline'],
                opacity=0.85,
            ))
            avg_val = hist_data.mean()
            fig_hist.add_vline(x=avg_val, line_dash='dash', line_color='red', line_width=1.5,
                               annotation_text=f'Mean: {avg_val:.2f}')
            fig_hist.update_layout(
                title=f'Production per Rope — {sel_grp}',
                height=370, xaxis_title='kg / Rope', yaxis_title='Count',
                font=dict(size=13, color='#333'), title_font=dict(size=15, color='#222'),
                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                margin=dict(l=20, r=20, t=50, b=20),
            )
            st.plotly_chart(fig_hist, use_container_width=True)
        else:
            st.info("No data available for the selected filter.")

        # ---- Dried vs Wet by group ----
        _section_header('', 'Dried vs Wet Seaweed by Group', '')
        col_stack, col_ratio = st.columns([3, 2])
        with col_stack:
            fig_dw = go.Figure()
            fig_dw.add_trace(go.Bar(x=grp_df['Group'], y=grp_df['Dried_KG'],
                                     name='Dried KG', marker_color=COLORS['baseline']))
            fig_dw.add_trace(go.Bar(x=grp_df['Group'], y=grp_df['Wet_KG'],
                                     name='Wet KG', marker_color='#00BCD4'))
            fig_dw.update_layout(
                barmode='stack', height=420,
                title='Dried vs Wet Seaweed Production',
                yaxis_title='KG',
                legend=dict(orientation='h', yanchor='bottom', y=1.02, x=0.5, xanchor='center'),
                font=dict(size=13, color='#333'), title_font=dict(size=16, color='#222'),
                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                margin=dict(l=20, r=20, t=60, b=20), xaxis_tickangle=-25,
            )
            st.plotly_chart(fig_dw, use_container_width=True)
        with col_ratio:
            # Dried/Wet ratio per group
            grp_ratio = grp_df[['Group']].copy()
            grp_ratio['Dried_Wet_Ratio'] = (grp_df['Dried_KG'] / grp_df['Wet_KG'].replace(0, np.nan)).round(2).fillna(0)
            grp_ratio = grp_ratio.sort_values('Dried_Wet_Ratio', ascending=True)
            fig_ratio = go.Figure(go.Bar(
                x=grp_ratio['Dried_Wet_Ratio'], y=grp_ratio['Group'],
                orientation='h', marker_color=COLORS['midline'],
                text=grp_ratio['Dried_Wet_Ratio'].apply(lambda v: f"{v:.2f}"),
                textposition='auto',
            ))
            fig_ratio.update_layout(
                title='Dried / Wet Ratio by Group',
                height=420, xaxis_title='Ratio (Dried ÷ Wet)',
                font=dict(size=12, color='#333'), title_font=dict(size=14, color='#222'),
                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                margin=dict(l=20, r=20, t=60, b=20),
            )
            fig_ratio.add_vline(x=1, line_dash='dash', line_color='#888', line_width=1,
                                annotation_text='Equal')
            st.plotly_chart(fig_ratio, use_container_width=True)

    # ==================================================================
    # TAB 4 — Challenges & Constraints (ENHANCED)
    # ==================================================================
    with tab4:
        st.markdown("""<div class="section-narrative">
        <strong>Challenges Profile:</strong> Comprehensive analysis of barriers facing seaweed
        farmers — prevalence across groups, co-occurrence patterns, impact on production,
        and qualitative free-text analysis.
        </div>""", unsafe_allow_html=True)

        # ---- Overall challenge bar chart ----
        _section_header('', 'Challenge Prevalence', '% of farmers reporting each challenge')
        if not ch_df.empty:
            ch_sorted = ch_df.sort_values('Pct', ascending=True)
            fig_ch = go.Figure(go.Bar(
                x=ch_sorted['Pct'], y=ch_sorted['Challenge'],
                orientation='h', marker_color=COLORS['baseline'],
                text=ch_sorted.apply(lambda r: f"{r['Pct']:.1f}% ({r['Count']})", axis=1),
                textposition='auto',
            ))
            fig_ch.update_layout(
                title='Prevalence of Challenges Among Seaweed Farmers',
                height=400, xaxis_title='% of Farmers',
                font=dict(size=13, color='#333'), title_font=dict(size=16, color='#222'),
                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                margin=dict(l=20, r=20, t=60, b=20),
            )
            st.plotly_chart(fig_ch, use_container_width=True)

        # ---- Per-group challenge heatmap ----
        _section_header('', 'Challenges by Group', 'Heatmap')
        challenge_labels = ['Transport', 'Market Access', 'Disease', 'Equipment',
                            'Storage', 'Labour', 'Sand / Tide']
        pct_cols = [f'pct_{lbl}' for lbl in challenge_labels]
        available_cols = [c for c in pct_cols if c in grp_df.columns]
        if available_cols:
            heat_data = grp_df.set_index('Group')[available_cols].copy()
            heat_data.columns = [c.replace('pct_', '') for c in heat_data.columns]
            fig_hm = go.Figure(go.Heatmap(
                z=heat_data.values,
                x=heat_data.columns.tolist(),
                y=heat_data.index.tolist(),
                colorscale='YlOrRd',
                text=heat_data.values,
                texttemplate='%{text:.0f}%',
                hovertemplate='Group: %{y}<br>Challenge: %{x}<br>% Affected: %{z:.1f}%<extra></extra>',
            ))
            fig_hm.update_layout(
                title='Challenge Prevalence by Group (%)',
                height=max(350, len(heat_data) * 55 + 120),
                font=dict(size=13, color='#333'), title_font=dict(size=15, color='#222'),
                margin=dict(l=20, r=20, t=60, b=20),
            )
            st.plotly_chart(fig_hm, use_container_width=True)

        # ---- Challenge Co-occurrence Matrix ----
        _section_header('', 'Challenge Co-occurrence', 'How often challenges appear together')
        flag_cols = ['flag_transport', 'flag_market', 'flag_disease',
                     'flag_equipment', 'flag_storage', 'flag_labour', 'flag_sand_tide']
        flag_labels_map = {
            'flag_transport': 'Transport', 'flag_market': 'Market',
            'flag_disease': 'Disease', 'flag_equipment': 'Equipment',
            'flag_storage': 'Storage', 'flag_labour': 'Labour',
            'flag_sand_tide': 'Sand/Tide',
        }
        available_flags = [f for f in flag_cols if f in df.columns]
        if len(available_flags) >= 2:
            flag_df = df[available_flags].fillna(False).astype(int)
            cooccur = flag_df.T.dot(flag_df)
            # Normalise by row total to get %
            diag = np.diag(cooccur.values).copy()
            diag[diag == 0] = 1
            cooccur_pct = (cooccur.values / diag[:, None] * 100).round(1)
            labels_co = [flag_labels_map.get(f, f) for f in available_flags]
            fig_co = go.Figure(go.Heatmap(
                z=cooccur_pct,
                x=labels_co, y=labels_co,
                colorscale='Blues',
                text=cooccur_pct,
                texttemplate='%{text:.0f}%',
                hovertemplate='%{y} → %{x}: %{z:.1f}%<extra></extra>',
            ))
            fig_co.update_layout(
                title='Challenge Co-occurrence (% of row challenge also having column challenge)',
                height=max(380, len(labels_co) * 55 + 100),
                font=dict(size=12, color='#333'), title_font=dict(size=14, color='#222'),
                margin=dict(l=20, r=20, t=60, b=20),
            )
            st.plotly_chart(fig_co, use_container_width=True)

        # ---- Challenge Burden vs Production ----
        _section_header('', 'Challenge Burden vs Production', 'Number of challenges per farmer')
        if available_flags:
            df_ch_burden = df.copy()
            df_ch_burden['n_challenges'] = df_ch_burden[available_flags].fillna(False).astype(int).sum(axis=1)
            burden_agg = df_ch_burden.groupby('n_challenges').agg(
                n_farmers=('Member', 'count'),
                avg_production=('Total_KG', 'mean'),
                avg_achievement=('Ropes_Achievement_pct', 'mean'),
            ).reset_index()

            col_bur1, col_bur2 = st.columns(2)
            with col_bur1:
                fig_bur = go.Figure()
                fig_bur.add_trace(go.Bar(
                    x=burden_agg['n_challenges'], y=burden_agg['n_farmers'],
                    name='Farmers', marker_color=COLORS['baseline'],
                    text=burden_agg['n_farmers'], textposition='auto',
                ))
                fig_bur.update_layout(
                    title='Farmers by Number of Challenges',
                    height=380, xaxis_title='Number of Challenges', yaxis_title='Farmers',
                    font=dict(size=12, color='#333'), title_font=dict(size=14, color='#222'),
                    plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                    margin=dict(l=20, r=20, t=50, b=20),
                )
                st.plotly_chart(fig_bur, use_container_width=True)
            with col_bur2:
                fig_impact = go.Figure()
                fig_impact.add_trace(go.Scatter(
                    x=burden_agg['n_challenges'], y=burden_agg['avg_production'],
                    mode='lines+markers', name='Avg Production (kg)',
                    marker=dict(size=12, color=COLORS['baseline']),
                    line=dict(width=2.5),
                ))
                fig_impact.add_trace(go.Scatter(
                    x=burden_agg['n_challenges'], y=burden_agg['avg_achievement'],
                    mode='lines+markers', name='Avg Achievement (%)',
                    yaxis='y2',
                    marker=dict(size=12, color=COLORS['midline']),
                    line=dict(width=2.5, dash='dot'),
                ))
                fig_impact.update_layout(
                    title='Challenge Impact on Performance',
                    height=380,
                    xaxis_title='Number of Challenges',
                    yaxis_title='Avg Production (kg)',
                    yaxis2=dict(title='Avg Achievement (%)', overlaying='y', side='right'),
                    legend=dict(orientation='h', yanchor='bottom', y=1.02, x=0.5, xanchor='center'),
                    font=dict(size=12, color='#333'), title_font=dict(size=14, color='#222'),
                    plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                    margin=dict(l=20, r=50, t=50, b=20),
                )
                st.plotly_chart(fig_impact, use_container_width=True)

        # ---- Free-text challenge word frequency ----
        _section_header('', 'Challenge Keyword Frequency', 'From free-text responses')
        ch_texts = df[df['Challenges_str'].notna() & (df['Challenges_str'].str.strip() != '')]
        if len(ch_texts):
            all_words = ' '.join(ch_texts['Challenges_str'].str.lower()).replace(',', ' ').replace('.', ' ')
            words = re.findall(r'\b[a-z]{3,}\b', all_words)
            stop_words = {'the', 'and', 'for', 'that', 'with', 'are', 'this', 'from', 'but',
                          'not', 'have', 'has', 'was', 'were', 'been', 'being', 'very', 'too',
                          'also', 'they', 'them', 'their', 'there', 'what', 'which', 'when',
                          'where', 'how', 'why', 'all', 'each', 'every', 'some', 'any', 'few'}
            filtered = [w for w in words if w not in stop_words]
            word_counts = Counter(filtered).most_common(20)
            if word_counts:
                wc_df = pd.DataFrame(word_counts, columns=['Word', 'Count']).sort_values('Count', ascending=True)
                fig_wc = go.Figure(go.Bar(
                    x=wc_df['Count'], y=wc_df['Word'],
                    orientation='h', marker_color=COLORS['baseline'],
                    text=wc_df['Count'], textposition='auto',
                ))
                fig_wc.update_layout(
                    title='Top 20 Challenge Keywords (from free-text)',
                    height=max(350, len(wc_df) * 25 + 80),
                    xaxis_title='Frequency',
                    font=dict(size=12, color='#333'), title_font=dict(size=14, color='#222'),
                    plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                    margin=dict(l=20, r=20, t=50, b=20),
                )
                st.plotly_chart(fig_wc, use_container_width=True)

        # ---- Free-text challenges table ----
        _section_header('', 'Reported Challenges (Free Text)', '')
        ch_text_df = df[df['Challenges_str'].notna() & (df['Challenges_str'].str.strip() != '')][
            ['Member', 'Group', 'Challenges_str']].copy()
        ch_text_df = ch_text_df.rename(columns={'Challenges_str': 'Challenge Description'})
        if len(ch_text_df):
            with st.expander(f"View Free-Text Challenges ({len(ch_text_df)} entries)", expanded=False):
                st.dataframe(ch_text_df, use_container_width=True, height=400)
        else:
            st.info("No free-text challenge descriptions available for current filters.")

    # ==================================================================
    # TAB 5 — Map View (ENHANCED)
    # ==================================================================
    with tab5:
        st.markdown("""<div class="section-narrative">
        <strong>Spatial Distribution:</strong> Geographic location of seaweed farming members.
        Point size reflects total production; colour represents farming group.
        Toggle between production and challenge density views.
        </div>""", unsafe_allow_html=True)

        map_df = df.dropna(subset=['x', 'y']).copy()
        map_df = map_df[(map_df['x'] != 0) & (map_df['y'] != 0)]

        if len(map_df) > 0:
            # Map mode selector
            map_mode = st.radio("Map View Mode",
                                ["Production (size by kg)", "Challenge Density (size by # challenges)"],
                                horizontal=True, key='sw_map_mode')

            if available_flags:
                map_df['n_challenges'] = map_df[available_flags].fillna(False).astype(int).sum(axis=1)
            else:
                map_df['n_challenges'] = 0

            # Compute sizes
            if map_mode.startswith("Production"):
                max_val = max(map_df['Total_KG'].max(), 1)
                map_df['_size'] = (map_df['Total_KG'] / max_val * 25).clip(lower=4)
                size_label = 'Production (kg)'
            else:
                max_val = max(map_df['n_challenges'].max(), 1)
                map_df['_size'] = (map_df['n_challenges'] / max_val * 25).clip(lower=4)
                size_label = '# Challenges'

            fig_map = go.Figure()
            for g in sorted(map_df['Group'].dropna().unique()):
                sub = map_df[map_df['Group'] == g]
                fig_map.add_trace(go.Scattermapbox(
                    lat=sub['y'], lon=sub['x'],
                    mode='markers',
                    marker=dict(
                        size=sub['_size'],
                        color=group_colors_map.get(g, '#666'),
                        opacity=0.8,
                    ),
                    name=g,
                    hovertemplate=(
                        '<b>%{customdata[0]}</b><br>Group: %{customdata[1]}<br>'
                        f'Production: %{{customdata[2]:.1f}} kg<br>'
                        f'Ropes: %{{customdata[3]:.0f}}<br>'
                        f'Challenges: %{{customdata[4]}}<extra></extra>'),
                    customdata=sub[['Member', 'Group', 'Total_KG', 'Ropes_Total', 'n_challenges']].values,
                ))

            center_lat = map_df['y'].mean()
            center_lon = map_df['x'].mean()
            fig_map.update_layout(
                mapbox=dict(
                    style='open-street-map',
                    center=dict(lat=center_lat, lon=center_lon),
                    zoom=10,
                ),
                height=600,
                legend=dict(orientation='h', yanchor='bottom', y=1.02, x=0.5, xanchor='center'),
                margin=dict(l=0, r=0, t=30, b=0),
                title=f'Seaweed Farmer Locations — {size_label}',
                title_font=dict(size=16, color='#222'),
            )
            st.plotly_chart(fig_map, use_container_width=True)

            # ---- Group cluster summary ----
            _section_header('', 'Group Location Summary', '')
            loc_summary = map_df.groupby('Group').agg(
                Farmers=('Member', 'count'),
                Avg_Lat=('y', 'mean'),
                Avg_Lon=('x', 'mean'),
                Total_KG=('Total_KG', 'sum'),
                Avg_Challenges=('n_challenges', 'mean'),
            ).round(3).reset_index()
            st.dataframe(loc_summary.style.format({
                'Avg_Lat': '{:.4f}', 'Avg_Lon': '{:.4f}',
                'Total_KG': '{:,.0f}', 'Avg_Challenges': '{:.1f}',
            }), use_container_width=True)
        else:
            st.info("No valid geographic coordinates available for the current filter selection.")

        # Data table
        with st.expander("View Raw Data Table"):
            display_cols = ['Member', 'Group', 'Ropes_Ocean', 'Ropes_Home', 'Ropes_Total',
                            'Target_Ropes', 'Ropes_Achievement_pct', 'Dried_KG', 'Wet_KG',
                            'Total_KG', 'Production_per_rope_kg', 'Casual_Workers']
            available = [c for c in display_cols if c in df.columns]
            st.dataframe(df[available].style.format({
                'Ropes_Achievement_pct': '{:.1f}%',
                'Production_per_rope_kg': '{:.2f}',
                'Total_KG': '{:,.1f}',
                'Dried_KG': '{:,.1f}',
                'Wet_KG': '{:,.1f}',
            }), use_container_width=True)


# ============================================================================
# MAIN
# ============================================================================

# ============================================================================
# INSIGHTS TAB — Automated insights across both datasets
# ============================================================================

def _insight_card(title, body, trend="neutral"):
    """Render a styled insight card with optional trend indicator."""
    color_map = {"positive": "#2E7D32", "negative": "#C62828", "neutral": "#37474F", "warning": "#E65100"}
    bg_map = {"positive": "#E8F5E9", "negative": "#FFEBEE", "neutral": "#ECEFF1", "warning": "#FFF3E0"}
    border = color_map.get(trend, color_map["neutral"])
    bg = bg_map.get(trend, bg_map["neutral"])
    st.markdown(f"""
    <div style="background:{bg}; border-left:5px solid {border}; border-radius:0 10px 10px 0;
                padding:1.2rem 1.5rem; margin-bottom:1rem; box-shadow:0 1px 6px rgba(0,0,0,0.06);">
        <h4 style="margin:0 0 0.4rem; color:{border}; font-size:1.05rem; font-weight:700;">{title}</h4>
        <p style="margin:0; color:#333; font-size:0.93rem; line-height:1.6;">{body}</p>
    </div>""", unsafe_allow_html=True)


def _pp(bl, ml, multiply=True):
    """Calculate percentage-point change."""
    factor = 100 if multiply else 1
    if isinstance(bl, (int, float)) and isinstance(ml, (int, float)):
        return round((ml - bl) * factor, 1)
    return 0.0


# ============================================================================
# PROJECT OUTPUTS & ACTIVITY INDICATORS — Data Loader & Renderer
# ============================================================================

def _safe_num(val, default=0.0):
    """Safely convert a cell value to float, treating errors/blanks as default."""
    if val is None or str(val).strip() in ('', 'ND', 'N/A', '#VALUE!', '#DIV/0!'):
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def _safe_int(val, default=0):
    """Safely convert a cell value to int."""
    n = _safe_num(val, default=float(default))
    return int(round(n))


@st.cache_data(show_spinner=False)
def load_project_outputs(filepath):
    """Load all 4 sheets from Project Outputs.xlsx and return structured dicts.

    Uses dynamic row scanning to handle layout differences between sheets
    (e.g., Forest has extra blank rows before the modules section).
    """
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)

    def _cell(ws, r, c):
        return ws.cell(r, c).value

    def _find_row(ws, label_fragment, start=1, end=50, col=2):
        """Find the first row where column *col* contains *label_fragment*."""
        for r in range(start, min(end + 1, ws.max_row + 1)):
            v = str(_cell(ws, r, col) or '')
            if label_fragment.lower() in v.lower():
                return r
        return None

    def _read_scalar_row(ws, row):
        """Read a single-value row: Y2(C), Y3_SA(F), Target(I), Change(L), Pct(M)."""
        return {
            'Y2': _safe_int(_cell(ws, row, 3)),
            'Y3_SA': _safe_int(_cell(ws, row, 6)),
            'Target': _safe_int(_cell(ws, row, 9)),
            'Change': _safe_int(_cell(ws, row, 12)),
            'Pct_Achieved': round(_safe_num(_cell(ws, row, 13)) * 100, 1),
        }

    def _read_sex_row(ws, row):
        """Read a sex-disaggregated data row (F/M/All for Y2, Y3, Target)."""
        return {
            'Y2_F': _safe_int(_cell(ws, row, 3)),
            'Y2_M': _safe_int(_cell(ws, row, 4)),
            'Y2_All': _safe_int(_cell(ws, row, 5)),
            'Y3_F': _safe_int(_cell(ws, row, 6)),
            'Y3_M': _safe_int(_cell(ws, row, 7)),
            'Y3_All': _safe_int(_cell(ws, row, 8)),
            'Target_F': _safe_int(_cell(ws, row, 9)),
            'Target_M': _safe_int(_cell(ws, row, 10)),
            'Target_All': _safe_int(_cell(ws, row, 11)),
            'Change': _safe_int(_cell(ws, row, 12)),
            'Pct_Achieved': round(_safe_num(_cell(ws, row, 13)) * 100, 1),
        }

    def _read_county_row(ws, row):
        """Read a county-disaggregated data row (Kilifi/Kwale/All)."""
        return {
            'Y2_Kilifi': _safe_int(_cell(ws, row, 3)),
            'Y2_Kwale': _safe_int(_cell(ws, row, 4)),
            'Y2_All': _safe_int(_cell(ws, row, 5)),
            'Y3_Kilifi': _safe_int(_cell(ws, row, 6)),
            'Y3_Kwale': _safe_int(_cell(ws, row, 7)),
            'Y3_All': _safe_int(_cell(ws, row, 8)),
            'Target_Kilifi': _safe_int(_cell(ws, row, 9)),
            'Target_Kwale': _safe_int(_cell(ws, row, 10)),
            'Target_All': _safe_int(_cell(ws, row, 11)),
            'Change': _safe_int(_cell(ws, row, 12)),
            'Pct_Achieved': round(_safe_num(_cell(ws, row, 13)) * 100, 1),
        }

    results = {}

    for sheet_name, module_key, strict_label in [
        ('Mangrove', 'mangrove', '3 and above'),
        ('Seaweed', 'seaweed', '3 and above'),
        ('Forest', 'forestry', '5m and above'),
        ('GJJ', 'gjj', '3 and above'),
    ]:
        ws = wb[sheet_name]
        data = {}

        # --- Output description (always row 8) ---
        data['output_desc'] = str(_cell(ws, 8, 2) or '').replace('\xa0', ' ').strip()

        # --- Dynamically find key rows ---
        groups_row = _find_row(ws, 'Number of Groups', start=10, end=20)
        members_row = _find_row(ws, 'Members Registered', start=14, end=25)
        modules_row = _find_row(ws, 'Number of Modules Imparted', start=18, end=30)

        # Find strict threshold section rows by scanning after "STRICT THRESHOLD"
        strict_section = _find_row(ws, 'STRICT THRESHOLD', start=20, end=30)
        # First "Members Trained" after strict section = strict vs target
        strict_target_row = _find_row(ws, 'Members Trained', start=(strict_section or 22) + 1, end=40)
        # Second "Members Trained" = strict vs registered
        strict_reg_row = _find_row(ws, 'Members Trained', start=(strict_target_row or 26) + 1, end=40) if strict_target_row else None
        # Third = by county
        strict_county_row = _find_row(ws, 'Members Trained', start=(strict_reg_row or 30) + 1, end=45) if strict_reg_row else None

        # --- Registration: Groups ---
        data['groups'] = {
            'Y2': _safe_int(_cell(ws, groups_row, 3)) if groups_row else 0,
            'Y3_SA': _safe_int(_cell(ws, groups_row, 6)) if groups_row else 0,
            'Target': _safe_int(_cell(ws, groups_row, 9)) if groups_row else 0,
            'Change': _safe_int(_cell(ws, groups_row, 12)) if groups_row else 0,
            'Pct_Achieved': round(_safe_num(_cell(ws, groups_row, 13)) * 100, 1) if groups_row else 0,
            'Notes': str(_cell(ws, groups_row, 15) or '').strip() if groups_row else '',
        }

        # --- Members Registered (sex-disaggregated) ---
        data['members'] = _read_sex_row(ws, members_row) if members_row else {
            k: 0 for k in ['Y2_F','Y2_M','Y2_All','Y3_F','Y3_M','Y3_All',
                           'Target_F','Target_M','Target_All','Change','Pct_Achieved']}

        # --- Modules Imparted ---
        if modules_row:
            data['modules_imparted'] = _read_scalar_row(ws, modules_row)
        else:
            data['modules_imparted'] = {'Y2': 0, 'Y3_SA': 0, 'Target': 0, 'Change': 0, 'Pct_Achieved': 0}

        # --- Strict threshold — vs Target ---
        data['strict_target'] = _read_sex_row(ws, strict_target_row) if strict_target_row else {
            k: 0 for k in ['Y2_F','Y2_M','Y2_All','Y3_F','Y3_M','Y3_All',
                           'Target_F','Target_M','Target_All','Change','Pct_Achieved']}

        # --- Strict threshold — vs Registered ---
        if strict_reg_row:
            sr = _read_sex_row(ws, strict_reg_row)
            # Rename Target_* to Reg_* for clarity
            data['strict_registered'] = {
                'Y2_F': sr['Y2_F'], 'Y2_M': sr['Y2_M'], 'Y2_All': sr['Y2_All'],
                'Y3_F': sr['Y3_F'], 'Y3_M': sr['Y3_M'], 'Y3_All': sr['Y3_All'],
                'Reg_F': sr['Target_F'], 'Reg_M': sr['Target_M'], 'Reg_All': sr['Target_All'],
                'Change': sr['Change'], 'Pct_Achieved': sr['Pct_Achieved'],
            }
        else:
            data['strict_registered'] = {k: 0 for k in ['Y2_F','Y2_M','Y2_All','Y3_F','Y3_M','Y3_All',
                                                         'Reg_F','Reg_M','Reg_All','Change','Pct_Achieved']}

        # --- Strict threshold — by County ---
        data['strict_county'] = _read_county_row(ws, strict_county_row) if strict_county_row else {
            k: 0 for k in ['Y2_Kilifi','Y2_Kwale','Y2_All','Y3_Kilifi','Y3_Kwale','Y3_All',
                           'Target_Kilifi','Target_Kwale','Target_All','Change','Pct_Achieved']}

        # --- Soft threshold: Module completion tables ---
        def _parse_module_table(start_row, max_modules=12):
            """Parse a module completion table starting at given row."""
            modules = []
            if start_row is None:
                return modules
            for offset in range(max_modules):
                r = start_row + offset
                if r > ws.max_row:
                    break
                mod_name = _cell(ws, r, 2)
                if mod_name is None or str(mod_name).strip() == '':
                    break
                mod_str = str(mod_name).strip()
                if mod_str.lower().startswith('grand total') or mod_str.lower().startswith('reference'):
                    if mod_str.lower().startswith('grand total'):
                        modules.append({
                            'Module': 'Grand Total',
                            'Target': _safe_int(_cell(ws, r, 3)),
                            'Registered': _safe_int(_cell(ws, r, 4)),
                            'Completed': round(_safe_num(_cell(ws, r, 5)), 1),
                            'Target_Pct': round(_safe_num(_cell(ws, r, 6)) * 100, 1),
                            'Reg_Pct': round(_safe_num(_cell(ws, r, 7)) * 100, 1),
                        })
                    continue
                modules.append({
                    'Module': mod_str,
                    'Target': _safe_int(_cell(ws, r, 3)),
                    'Registered': _safe_int(_cell(ws, r, 4)),
                    'Completed': _safe_int(_cell(ws, r, 5)),
                    'Target_Pct': round(_safe_num(_cell(ws, r, 6)) * 100, 1),
                    'Reg_Pct': round(_safe_num(_cell(ws, r, 7)) * 100, 1),
                })
            return modules

        # Find the soft threshold section and the first "Module 1" row for All
        soft_section = _find_row(ws, 'SOFT THRESHOLD', start=30, end=45)
        all_table_start = _find_row(ws, 'Module 1', start=(soft_section or 36), end=55)
        data['soft_all'] = pd.DataFrame(_parse_module_table(all_table_start))

        # Find Male and Female table start rows
        male_start = None
        female_start = None
        search_start = (all_table_start or 40) + 5
        for r in range(search_start, min(ws.max_row + 1, 80)):
            val = str(_cell(ws, r, 2) or '')
            if 'Male)' in val and 'Female' not in val:
                # Find Module 1 after this header
                male_start = _find_row(ws, 'Module 1', start=r + 1, end=r + 5)
            elif 'Female)' in val:
                female_start = _find_row(ws, 'Module 1', start=r + 1, end=r + 5)

        data['soft_male'] = pd.DataFrame(_parse_module_table(male_start)) if male_start else pd.DataFrame()
        data['soft_female'] = pd.DataFrame(_parse_module_table(female_start)) if female_start else pd.DataFrame()

        # Forestry Reference - Y2 row
        if module_key == 'forestry' and all_table_start:
            for r in range(all_table_start, min(ws.max_row + 1, all_table_start + 15)):
                val = str(_cell(ws, r, 2) or '')
                if val.strip().lower().startswith('reference'):
                    data['soft_all_y2_ref'] = {
                        'Target': _safe_int(_cell(ws, r, 3)),
                        'Registered': _safe_int(_cell(ws, r, 4)),
                        'Completed': _safe_int(_cell(ws, r, 5)),
                        'Target_Pct': round(_safe_num(_cell(ws, r, 6)) * 100, 1),
                        'Reg_Pct': round(_safe_num(_cell(ws, r, 7)) * 100, 1),
                    }
                    break

        results[module_key] = data

    wb.close()
    return results


def prepare_project_outputs_tables(po_data):
    """Convert raw project outputs into tidy long-format DataFrames for charting.

    Returns a dict with:
      'summary': Cross-module summary DataFrame
      'members_sex': Sex-disaggregated members table
      'strict_training': Strict threshold training summary
      'soft_completion': Long-format soft threshold module completion
    """
    # --- Helper: extract Grand Total row from soft DataFrame ---
    def _soft_grand_total(df):
        if df is not None and not df.empty and 'Grand Total' in df['Module'].values:
            return df[df['Module'] == 'Grand Total'].iloc[0]
        return None

    # --- Cross-module summary (using soft threshold for training) ---
    summary_rows = []
    soft_train_rows = []
    module_labels = {'mangrove': 'Mangrove', 'seaweed': 'Seaweed',
                     'forestry': 'Forestry', 'gjj': 'GJJ'}
    for key, label in module_labels.items():
        d = po_data[key]
        gt_all = _soft_grand_total(d.get('soft_all'))
        gt_m = _soft_grand_total(d.get('soft_male'))
        gt_f = _soft_grand_total(d.get('soft_female'))
        trained_y3 = int(gt_all['Completed']) if gt_all is not None else 0
        trained_target = int(gt_all['Target']) if gt_all is not None else 0
        trained_pct = float(gt_all['Target_Pct']) if gt_all is not None else 0.0

        summary_rows.append({
            'Module': label,
            'Groups_Y2': d['groups']['Y2'],
            'Groups_Y3': d['groups']['Y3_SA'],
            'Groups_Target': d['groups']['Target'],
            'Groups_Pct': d['groups']['Pct_Achieved'],
            'Groups_Change': d['groups']['Change'],
            'Members_Y2': d['members']['Y2_All'],
            'Members_Y3': d['members']['Y3_All'],
            'Members_Target': d['members']['Target_All'],
            'Members_Pct': d['members']['Pct_Achieved'],
            'Members_Change': d['members']['Change'],
            'Modules_Y2': d['modules_imparted']['Y2'],
            'Modules_Y3': d['modules_imparted']['Y3_SA'],
            'Modules_Target': d['modules_imparted']['Target'],
            'Modules_Pct': d['modules_imparted']['Pct_Achieved'],
            'Trained_Y3': trained_y3,
            'Trained_Target': trained_target,
            'Trained_Pct': trained_pct,
        })
        soft_train_rows.append({
            'Module': label,
            'Y3_Female': int(gt_f['Completed']) if gt_f is not None else 0,
            'Y3_Male': int(gt_m['Completed']) if gt_m is not None else 0,
            'Y3_All': trained_y3,
            'Target_All': trained_target,
            'Pct_Achieved': trained_pct,
        })
    summary = pd.DataFrame(summary_rows)
    soft_training = pd.DataFrame(soft_train_rows)

    # --- Sex-disaggregated members ---
    mem_rows = []
    for key, label in module_labels.items():
        m = po_data[key]['members']
        for period, prefix in [('Y2', 'Y2'), ('Y3 (SA)', 'Y3')]:
            mem_rows.append({'Module': label, 'Period': period,
                             'Female': m[f'{prefix}_F'], 'Male': m[f'{prefix}_M'],
                             'Total': m[f'{prefix}_All']})
    members_sex = pd.DataFrame(mem_rows)

    # --- Long-format soft completion ---
    soft_rows = []
    for key, label in module_labels.items():
        for sex_key, sex_label in [('soft_all', 'All'), ('soft_male', 'Male'), ('soft_female', 'Female')]:
            df = po_data[key].get(sex_key)
            if df is not None and not df.empty:
                for _, row in df.iterrows():
                    soft_rows.append({
                        'Module_Area': label,
                        'Sex': sex_label,
                        'Module': row['Module'],
                        'Target': row['Target'],
                        'Registered': row['Registered'],
                        'Completed': row['Completed'],
                        'Target_Pct': row['Target_Pct'],
                        'Reg_Pct': row['Reg_Pct'],
                    })
    soft_completion = pd.DataFrame(soft_rows)

    return {
        'summary': summary,
        'members_sex': members_sex,
        'soft_training': soft_training,
        'soft_completion': soft_completion,
    }


def render_project_outputs_tabs(po_data):
    """Render the Project Outputs & Activity Indicators module with 5 tabs."""
    tables = prepare_project_outputs_tables(po_data)
    summary = tables['summary']
    members_sex = tables['members_sex']
    soft_training = tables['soft_training']
    soft_completion = tables['soft_completion']

    tabs = st.tabs([
        "\U0001F4CA High-Level KPIs",
        "\U0001F333 Mangrove Outputs",
        "\U0001F33F Seaweed Outputs",
        "\U0001F332 Forestry Outputs",
        "\U0001F465 GJJ Outputs",
    ])

    # ================================================================
    # TAB 1 — High-Level KPIs (Cross-Module)
    # ================================================================
    with tabs[0]:
        # Per-module KPI cards
        _section_header('', 'Module-Level Progress', 'Y3 (SA) vs Target')
        for _, row in summary.iterrows():
            mod = row['Module']
            mc1, mc2, mc3, mc4, mc5 = st.columns(5)
            mc1.markdown(f"""<div class="kpi-card"><h3>{mod}</h3><div class="value" style="font-size:1rem;">&nbsp;</div></div>""", unsafe_allow_html=True)
            g_color = COLORS['good'] if row['Groups_Pct'] >= 100 else '#FF9800'
            mc2.markdown(f"""<div class="kpi-card"><h3>Groups</h3><div class="value">{row['Groups_Y3']:,}</div>
                <p style="color:{g_color}">{row['Groups_Pct']:.1f}% of target</p></div>""", unsafe_allow_html=True)
            m_color = COLORS['good'] if row['Members_Pct'] >= 100 else '#FF9800'
            mc3.markdown(f"""<div class="kpi-card"><h3>Members</h3><div class="value">{row['Members_Y3']:,}</div>
                <p style="color:{m_color}">{row['Members_Pct']:.1f}% of target</p></div>""", unsafe_allow_html=True)
            mod_color = COLORS['good'] if row['Modules_Pct'] >= 100 else '#FF9800'
            mc4.markdown(f"""<div class="kpi-card"><h3>Modules</h3><div class="value">{row['Modules_Y3']}</div>
                <p style="color:{mod_color}">{row['Modules_Pct']:.1f}% of target</p></div>""", unsafe_allow_html=True)
            t_color = COLORS['good'] if row['Trained_Pct'] >= 60 else COLORS['danger']
            mc5.markdown(f"""<div class="kpi-card"><h3>Trained (Soft)</h3><div class="value">{row['Trained_Y3']:,}</div>
                <p style="color:{t_color}">{row['Trained_Pct']:.1f}% achieved</p></div>""", unsafe_allow_html=True)

        st.markdown("---")

        # Bar charts
        _section_header('', 'Cross-Module Comparisons', 'Charts')
        chart_col1, chart_col2 = st.columns(2)

        with chart_col1:
            # Groups: Y3 vs Target
            fig_grp = go.Figure()
            fig_grp.add_trace(go.Bar(
                x=summary['Module'], y=summary['Groups_Y3'],
                name='Y3 (SA)', marker_color=COLORS['midline'],
                text=summary['Groups_Y3'], textposition='auto',
            ))
            fig_grp.add_trace(go.Bar(
                x=summary['Module'], y=summary['Groups_Target'],
                name='Target', marker_color='rgba(0,0,0,0.15)',
                text=summary['Groups_Target'], textposition='auto',
            ))
            fig_grp.update_layout(
                title="Groups: Y3 (SA) vs Target", barmode='group', height=380,
                yaxis_title="Number of Groups",
                legend=dict(orientation='h', yanchor='bottom', y=1.02, x=0.5, xanchor='center'),
                font=dict(size=13, color='#333'), title_font=dict(size=16, color='#222'),
                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                margin=dict(l=20, r=20, t=60, b=20),
            )
            st.plotly_chart(fig_grp, use_container_width=True)

        with chart_col2:
            # Members Registered: Y3 vs Target
            fig_mem = go.Figure()
            fig_mem.add_trace(go.Bar(
                x=summary['Module'], y=summary['Members_Y3'],
                name='Y3 (SA)', marker_color=COLORS['midline'],
                text=summary['Members_Y3'].apply(lambda v: f"{v:,}"), textposition='auto',
            ))
            fig_mem.add_trace(go.Bar(
                x=summary['Module'], y=summary['Members_Target'],
                name='Target', marker_color='rgba(0,0,0,0.15)',
                text=summary['Members_Target'].apply(lambda v: f"{v:,}"), textposition='auto',
            ))
            fig_mem.update_layout(
                title="Members Registered: Y3 (SA) vs Target", barmode='group', height=380,
                yaxis_title="Members",
                legend=dict(orientation='h', yanchor='bottom', y=1.02, x=0.5, xanchor='center'),
                font=dict(size=13, color='#333'), title_font=dict(size=16, color='#222'),
                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                margin=dict(l=20, r=20, t=60, b=20),
            )
            st.plotly_chart(fig_mem, use_container_width=True)

        chart_col3, chart_col4 = st.columns(2)

        with chart_col3:
            # % Achieved by module (gauge-like horizontal bar)
            fig_pct = go.Figure()
            pct_colors = [COLORS['good'] if v >= 100 else ('#FF9800' if v >= 60 else COLORS['danger'])
                          for v in summary['Members_Pct']]
            fig_pct.add_trace(go.Bar(
                y=summary['Module'], x=summary['Members_Pct'],
                orientation='h', marker_color=pct_colors,
                text=summary['Members_Pct'].apply(lambda v: f"{v:.1f}%"), textposition='auto',
            ))
            fig_pct.add_vline(x=100, line_dash='dash', line_color='#999', line_width=1.5,
                             annotation_text='Target (100%)', annotation_position='top right')
            fig_pct.update_layout(
                title="Members Registration % Achieved", height=350,
                xaxis_title="% of Target Achieved", yaxis=dict(categoryorder='total ascending'),
                font=dict(size=13, color='#333'), title_font=dict(size=16, color='#222'),
                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                margin=dict(l=20, r=20, t=60, b=20),
            )
            st.plotly_chart(fig_pct, use_container_width=True)

        with chart_col4:
            # Y2 → Y3 Change by module
            fig_chg = go.Figure()
            chg_colors = [COLORS['good'] if v >= 0 else COLORS['danger'] for v in summary['Members_Change']]
            fig_chg.add_trace(go.Bar(
                y=summary['Module'], x=summary['Members_Change'],
                orientation='h', marker_color=chg_colors,
                text=summary['Members_Change'].apply(lambda v: f"{v:+,}"), textposition='auto',
            ))
            fig_chg.update_layout(
                title="Members Change: Y2 → Y3 (SA)", height=350,
                xaxis_title="Change (count)",
                yaxis=dict(categoryorder='total ascending'),
                font=dict(size=13, color='#333'), title_font=dict(size=16, color='#222'),
                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                margin=dict(l=20, r=20, t=60, b=20),
            )
            st.plotly_chart(fig_chg, use_container_width=True)

        # Training progress: Soft threshold bar
        _section_header('', 'Training Progress (Soft Threshold)', 'Y3 (SA)')
        fig_train = go.Figure()
        fig_train.add_trace(go.Bar(
            x=soft_training['Module'], y=soft_training['Y3_All'],
            name='Trained (Y3)', marker_color=COLORS['midline'],
            text=soft_training['Y3_All'].apply(lambda v: f"{v:,}"), textposition='auto',
        ))
        fig_train.add_trace(go.Bar(
            x=soft_training['Module'], y=soft_training['Target_All'],
            name='Target', marker_color='rgba(0,0,0,0.15)',
            text=soft_training['Target_All'].apply(lambda v: f"{v:,}"), textposition='auto',
        ))
        fig_train.update_layout(
            title="Soft Threshold Training: Y3 (SA) vs Target", barmode='group', height=380,
            yaxis_title="Members Trained",
            legend=dict(orientation='h', yanchor='bottom', y=1.02, x=0.5, xanchor='center'),
            font=dict(size=13, color='#333'), title_font=dict(size=16, color='#222'),
            plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
            margin=dict(l=20, r=20, t=60, b=20),
        )
        st.plotly_chart(fig_train, use_container_width=True)

        # Full summary table
        with st.expander("View Full Cross-Module Summary Table"):
            display = summary.copy()
            display.columns = [c.replace('_', ' ') for c in display.columns]
            st.dataframe(display, use_container_width=True, hide_index=True)

    # ================================================================
    # TAB 2-5 — Module-Specific Outputs
    # ================================================================
    def _render_module_tab(tab, module_key, module_label):
        """Render a single module output tab."""
        d = po_data[module_key]

        # Extract soft Grand Total for KPI
        def _sgt(df):
            if df is not None and not df.empty and 'Grand Total' in df['Module'].values:
                return df[df['Module'] == 'Grand Total'].iloc[0]
            return None
        sgt_all = _sgt(d.get('soft_all'))
        sgt_m = _sgt(d.get('soft_male'))
        sgt_f = _sgt(d.get('soft_female'))
        soft_trained = int(sgt_all['Completed']) if sgt_all is not None else 0
        soft_target = int(sgt_all['Target']) if sgt_all is not None else 0
        soft_pct = float(sgt_all['Target_Pct']) if sgt_all is not None else 0.0

        with tab:
            _section_header('', f'{module_label} Output & Activity Summary', 'SAR Y3')

            # --- KPI Row ---
            k1, k2, k3, k4 = st.columns(4)
            g = d['groups']
            g_pct_color = COLORS['good'] if g['Pct_Achieved'] >= 100 else '#FF9800'
            k1.markdown(f"""<div class="kpi-card">
                <h3>Groups (Y3)</h3>
                <div class="value">{g['Y3_SA']:,}</div>
                <p>Target: {g['Target']:,} | <span style="color:{g_pct_color}">{g['Pct_Achieved']:.1f}%</span></p>
            </div>""", unsafe_allow_html=True)

            m = d['members']
            m_pct_color = COLORS['good'] if m['Pct_Achieved'] >= 100 else '#FF9800'
            k2.markdown(f"""<div class="kpi-card">
                <h3>Members (Y3)</h3>
                <div class="value">{m['Y3_All']:,}</div>
                <p>Target: {m['Target_All']:,} | <span style="color:{m_pct_color}">{m['Pct_Achieved']:.1f}%</span></p>
            </div>""", unsafe_allow_html=True)

            mi = d['modules_imparted']
            mi_color = COLORS['good'] if mi['Pct_Achieved'] >= 100 else '#FF9800'
            k3.markdown(f"""<div class="kpi-card">
                <h3>Modules Imparted</h3>
                <div class="value">{mi['Y3_SA']} / {mi['Target']}</div>
                <p style="color:{mi_color}">{mi['Pct_Achieved']:.1f}% achieved</p>
            </div>""", unsafe_allow_html=True)

            st_color = COLORS['good'] if soft_pct >= 60 else COLORS['danger']
            k4.markdown(f"""<div class="kpi-card">
                <h3>Trained (Soft)</h3>
                <div class="value">{soft_trained:,}</div>
                <p style="color:{st_color}">{soft_pct:.1f}% of target</p>
            </div>""", unsafe_allow_html=True)

            if g['Notes']:
                st.info(f"**Note:** {g['Notes']}")

            st.markdown("---")

            # --- Registration Progress Section ---
            _section_header('', 'Registration Progress', 'Y2 → Y3 (SA)')
            reg_col1, reg_col2 = st.columns(2)

            with reg_col1:
                # Groups bar: Y2, Y3, Target
                fig_g = go.Figure()
                for val, lbl, clr in [(g['Y2'], 'Y2', COLORS['baseline']),
                                       (g['Y3_SA'], 'Y3 (SA)', COLORS['midline']),
                                       (g['Target'], 'Target', 'rgba(0,0,0,0.15)')]:
                    fig_g.add_trace(go.Bar(x=[lbl], y=[val], name=lbl, marker_color=clr,
                                           text=[f"{val:,}"], textposition='auto', showlegend=False))
                fig_g.update_layout(
                    title=f"{module_label}: Groups Progress", height=330,
                    yaxis_title="Groups", barmode='group',
                    font=dict(size=13, color='#333'), title_font=dict(size=16, color='#222'),
                    plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                    margin=dict(l=20, r=20, t=60, b=20),
                )
                st.plotly_chart(fig_g, use_container_width=True)

            with reg_col2:
                # Members by sex: grouped bar Y2 vs Y3
                fig_ms = go.Figure()
                cats = ['Female', 'Male', 'All']
                y2_vals = [m['Y2_F'], m['Y2_M'], m['Y2_All']]
                y3_vals = [m['Y3_F'], m['Y3_M'], m['Y3_All']]
                tgt_vals = [m['Target_F'], m['Target_M'], m['Target_All']]
                fig_ms.add_trace(go.Bar(x=cats, y=y2_vals, name='Y2', marker_color=COLORS['baseline'],
                                        text=[f"{v:,}" for v in y2_vals], textposition='auto'))
                fig_ms.add_trace(go.Bar(x=cats, y=y3_vals, name='Y3 (SA)', marker_color=COLORS['midline'],
                                        text=[f"{v:,}" for v in y3_vals], textposition='auto'))
                fig_ms.add_trace(go.Bar(x=cats, y=tgt_vals, name='Target', marker_color='rgba(0,0,0,0.15)',
                                        text=[f"{v:,}" for v in tgt_vals], textposition='auto'))
                fig_ms.update_layout(
                    title=f"{module_label}: Members by Sex (Y2 → Y3)", barmode='group', height=330,
                    yaxis_title="Members",
                    legend=dict(orientation='h', yanchor='bottom', y=1.02, x=0.5, xanchor='center'),
                    font=dict(size=13, color='#333'), title_font=dict(size=16, color='#222'),
                    plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                    margin=dict(l=20, r=20, t=60, b=20),
                )
                st.plotly_chart(fig_ms, use_container_width=True)

            st.markdown("---")

            # --- Training Section (Soft Threshold) ---
            _section_header('', f'Training Achievement (Soft Threshold)', 'Completion')
            tr_col1, tr_col2 = st.columns(2)

            with tr_col1:
                # Soft training by sex
                fig_soft_sex = go.Figure()
                sex_cats = ['Female', 'Male', 'All']
                trained_vals = [
                    int(sgt_f['Completed']) if sgt_f is not None else 0,
                    int(sgt_m['Completed']) if sgt_m is not None else 0,
                    soft_trained,
                ]
                target_vals = [soft_target] * 3  # same overall target
                fig_soft_sex.add_trace(go.Bar(x=sex_cats, y=trained_vals, name='Completed (Y3)',
                                              marker_color=COLORS['midline'],
                                              text=[f"{v:,}" for v in trained_vals], textposition='auto'))
                fig_soft_sex.add_trace(go.Bar(x=sex_cats, y=target_vals, name='Target',
                                              marker_color='rgba(0,0,0,0.15)',
                                              text=[f"{v:,}" for v in target_vals], textposition='auto'))
                fig_soft_sex.update_layout(
                    title=f"{module_label}: Soft Training by Sex", barmode='group', height=350,
                    yaxis_title="Members",
                    legend=dict(orientation='h', yanchor='bottom', y=1.02, x=0.5, xanchor='center'),
                    font=dict(size=13, color='#333'), title_font=dict(size=16, color='#222'),
                    plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                    margin=dict(l=20, r=20, t=60, b=20),
                )
                st.plotly_chart(fig_soft_sex, use_container_width=True)

            with tr_col2:
                # Soft threshold: Completed vs Registered vs Target
                fig_overview = go.Figure()
                soft_reg = int(sgt_all['Registered']) if sgt_all is not None else 0
                overview_cats = ['Target', 'Registered', 'Completed']
                overview_vals = [soft_target, soft_reg, soft_trained]
                overview_colors = ['rgba(0,0,0,0.15)', COLORS['baseline'], COLORS['midline']]
                fig_overview.add_trace(go.Bar(
                    x=overview_cats, y=overview_vals,
                    marker_color=overview_colors,
                    text=[f"{v:,}" for v in overview_vals], textposition='auto',
                    showlegend=False,
                ))
                fig_overview.update_layout(
                    title=f"{module_label}: Training Pipeline", height=350,
                    yaxis_title="Members",
                    font=dict(size=13, color='#333'), title_font=dict(size=16, color='#222'),
                    plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                    margin=dict(l=20, r=20, t=60, b=20),
                )
                st.plotly_chart(fig_overview, use_container_width=True)

            st.markdown("---")

            # --- Soft Threshold Module Completion ---
            _section_header('', 'Module Completion (Soft Threshold)', 'By Module')

            # Filter soft completion for this module
            sc = soft_completion[soft_completion['Module_Area'] == module_label].copy()
            sc_actual = sc[sc['Module'] != 'Grand Total'].copy()

            if not sc_actual.empty:
                # Chart: completion % by module for All
                sc_all = sc_actual[sc_actual['Sex'] == 'All']
                if not sc_all.empty:
                    fig_soft = go.Figure()
                    fig_soft.add_trace(go.Bar(
                        x=sc_all['Module'], y=sc_all['Target_Pct'],
                        name='% of Target', marker_color=COLORS['midline'],
                        text=sc_all['Target_Pct'].apply(lambda v: f"{v:.1f}%"), textposition='auto',
                    ))
                    fig_soft.add_trace(go.Bar(
                        x=sc_all['Module'], y=sc_all['Reg_Pct'],
                        name='% of Registered', marker_color=COLORS['baseline'],
                        text=sc_all['Reg_Pct'].apply(lambda v: f"{v:.1f}%"), textposition='auto',
                    ))
                    fig_soft.add_hline(y=100, line_dash='dash', line_color='#999', line_width=1)
                    fig_soft.update_layout(
                        title=f"{module_label}: Module Completion Rates (All)", barmode='group', height=380,
                        yaxis_title="Completion %",
                        legend=dict(orientation='h', yanchor='bottom', y=1.02, x=0.5, xanchor='center'),
                        font=dict(size=13, color='#333'), title_font=dict(size=16, color='#222'),
                        plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                        margin=dict(l=20, r=20, t=60, b=20),
                    )
                    st.plotly_chart(fig_soft, use_container_width=True)

                # Gender comparison if we have both male and female data
                sc_m = sc_actual[sc_actual['Sex'] == 'Male']
                sc_f = sc_actual[sc_actual['Sex'] == 'Female']
                has_male = not sc_m.empty and sc_m['Completed'].sum() > 0
                has_female = not sc_f.empty and sc_f['Completed'].sum() > 0

                if has_male or has_female:
                    gen_col1, gen_col2 = st.columns(2)

                    if has_female:
                        with gen_col1:
                            fig_sf = go.Figure()
                            fig_sf.add_trace(go.Bar(
                                x=sc_f['Module'], y=sc_f['Reg_Pct'],
                                marker_color='#E91E63',
                                text=sc_f['Reg_Pct'].apply(lambda v: f"{v:.1f}%"), textposition='auto',
                            ))
                            fig_sf.update_layout(
                                title=f"{module_label}: Female Completion (% of Registered)", height=350,
                                yaxis_title="Completion %", showlegend=False,
                                font=dict(size=13, color='#333'), title_font=dict(size=15, color='#222'),
                                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                                margin=dict(l=20, r=20, t=60, b=20),
                            )
                            st.plotly_chart(fig_sf, use_container_width=True)

                    if has_male:
                        target_col = gen_col2 if has_female else gen_col1
                        with target_col:
                            fig_sm = go.Figure()
                            fig_sm.add_trace(go.Bar(
                                x=sc_m['Module'], y=sc_m['Reg_Pct'],
                                marker_color='#1976D2',
                                text=sc_m['Reg_Pct'].apply(lambda v: f"{v:.1f}%"), textposition='auto',
                            ))
                            fig_sm.update_layout(
                                title=f"{module_label}: Male Completion (% of Registered)", height=350,
                                yaxis_title="Completion %", showlegend=False,
                                font=dict(size=13, color='#333'), title_font=dict(size=15, color='#222'),
                                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                                margin=dict(l=20, r=20, t=60, b=20),
                            )
                            st.plotly_chart(fig_sm, use_container_width=True)

                    if not has_male:
                        st.info(f"No male training data available for {module_label} — all training is female-only.")

                # Absolute numbers table
                with st.expander(f"View {module_label} Module Completion Data"):
                    display = sc_actual[['Sex', 'Module', 'Target', 'Registered', 'Completed',
                                         'Target_Pct', 'Reg_Pct']].copy()
                    display.columns = ['Sex', 'Module', 'Target', 'Registered', 'Completed',
                                       'Target (%)', 'Registered (%)']
                    st.dataframe(display, use_container_width=True, hide_index=True)
            else:
                st.warning(f"No soft-threshold module completion data available for {module_label}.")

    # Render each module tab
    _render_module_tab(tabs[1], 'mangrove', 'Mangrove')
    _render_module_tab(tabs[2], 'seaweed', 'Seaweed')
    _render_module_tab(tabs[3], 'forestry', 'Forestry')
    _render_module_tab(tabs[4], 'gjj', 'GJJ')


# ============================================================================
# VSLA FUNCTIONALITY — Data Loader & Renderer
# ============================================================================

@st.cache_data(show_spinner=False)
def load_vsla_data(filepath):
    """Load VSLA Functionality Excel and return a structured dict.

    Parses "Results (Across Qs)" sheet with dynamic row-scanning.
    Sections: A (Membership/Meetings), B (Savings), C (Social Fund), D (Loans).
    """
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[VSLA_SHEET]

    def _c(r, c):
        v = ws.cell(r, c).value
        if v is None:
            return 0
        if isinstance(v, str):
            v = v.strip()
            if v in ('', '-', 'N/A', 'ND'):
                return 0
            try:
                return float(v)
            except ValueError:
                return v
        return v

    def _num(v):
        if v is None or (isinstance(v, str) and v.strip() in ('', '-', 'N/A', 'ND')):
            return 0
        try:
            return float(v)
        except (ValueError, TypeError):
            return 0

    def _county_quarter(start):
        """Read a simple county x quarter table (3 rows × Q2/Q3/Q4)."""
        rows = []
        for r in range(start, start + 3):
            rows.append({
                'county': str(_c(r, 2)),
                'Q2': _num(_c(r, 3)),
                'Q3': _num(_c(r, 4)),
                'Q4': _num(_c(r, 5)),
            })
        return rows

    def _sumavg(header_row):
        """Read county table with Sum/Avg for Q2/Q3/Q4."""
        rows = []
        for r in range(header_row + 1, header_row + 4):
            rows.append({
                'county': str(_c(r, 2)),
                'Q2': {'sum': _num(_c(r, 3)), 'average': _num(_c(r, 4))},
                'Q3': {'sum': _num(_c(r, 5)), 'average': _num(_c(r, 6))},
                'Q4': {'sum': _num(_c(r, 7)), 'average': _num(_c(r, 8))},
            })
        return rows

    def _bands(start, end):
        """Read a band distribution table."""
        bands = []
        for r in range(start, end + 1):
            label = _c(r, 2)
            if isinstance(label, (int, float)):
                label = str(label)
            bands.append({
                'band': str(label),
                'Q2': _num(_c(r, 3)),
                'Q3': _num(_c(r, 4)),
                'Q4': _num(_c(r, 5)),
            })
        return bands

    data = {
        'meta': {
            'title': str(_c(2, 2) or 'VSLA Functionality'),
            'date': str(_c(3, 2) or ''),
            'outcome': str(_c(4, 2) or ''),
        },
        'vslasAssessed': _county_quarter(8),
        'membership': {
            'female': _sumavg(15),
            'male': _sumavg(21),
            'all': _sumavg(27),
        },
        'membersLeft': _county_quarter(34),
        'percentLeftBands': _bands(40, 46),
        'meetings': {
            'frequency': _county_quarter(50),
            'attendance': _county_quarter(56),
        },
        'savings': {
            'membersSaving': _sumavg(65),
            'proportionBands': _bands(72, 76),
            'value': _sumavg(80),
            'valueBands': _bands(87, 91),
        },
        'socialFund': {
            'percentageWithFund': _county_quarter(97),
            'value': _sumavg(103),
            'valueBands': _bands(110, 115),
        },
        'loans': {
            'disbursed': {
                'count': _sumavg(122),
                'countBands': _bands(129, 135),
                'value': _sumavg(139),
                'valueBands': _bands(146, 151),
            },
            'repaid': {
                'count': _sumavg(155),
                'countBands': _bands(162, 166),
                'value': _sumavg(170),
                'valueBands': _bands(177, 182),
            },
            'interest': {
                'value': _sumavg(186),
                'valueBands': _bands(193, 197),
            },
            'behaviour': [],
            'failingToPay': _county_quarter(208),
            'useDistribution': _bands(214, 218),
        },
    }

    # Loan behaviour (special layout)
    for r in range(202, 205):
        data['loans']['behaviour'].append({
            'county': str(_c(r, 2)),
            'avgProportionRepaid': {'Q2': _num(_c(r, 3)), 'Q3': _num(_c(r, 4)), 'Q4': _num(_c(r, 5))},
            'avgValueRepaid': {'Q2': _num(_c(r, 6)), 'Q3': _num(_c(r, 7)), 'Q4': _num(_c(r, 8))},
            'avgROI': {'Q2': _num(_c(r, 9)), 'Q3': _num(_c(r, 10)), 'Q4': _num(_c(r, 11))},
        })

    wb.close()
    return data


# ── VSLA helper functions ──

def _vsla_get_cq(rows, county, quarter):
    """Get a value from a county-quarter list for a given county & quarter."""
    target = county if county != 'All Counties' else 'All'
    for r in rows:
        if r['county'] == target:
            return r.get(quarter, 0)
    return 0


def _vsla_get_sa(rows, county, quarter, field='sum'):
    """Get sum or average from a SumAvg list."""
    target = county if county != 'All Counties' else 'All'
    for r in rows:
        if r['county'] == target:
            q = r.get(quarter, {})
            if isinstance(q, dict):
                return q.get(field, 0)
    return 0


def _vsla_trend(rows, county, field='sum'):
    """Get Q2→Q4 trend for sum/avg rows."""
    target = county if county != 'All Counties' else 'All'
    for r in rows:
        if r['county'] == target:
            vals = []
            for q in ['Q2', 'Q3', 'Q4']:
                v = r.get(q, {})
                if isinstance(v, dict):
                    vals.append(v.get(field, 0))
                else:
                    vals.append(v)
            return vals
    return [0, 0, 0]


def _vsla_cq_trend(rows, county):
    """Get Q2→Q4 trend for simple county-quarter rows."""
    target = county if county != 'All Counties' else 'All'
    for r in rows:
        if r['county'] == target:
            return [r.get('Q2', 0), r.get('Q3', 0), r.get('Q4', 0)]
    return [0, 0, 0]


def _vsla_fmt_kes(v):
    """Format KES amount."""
    if abs(v) >= 1_000_000:
        return f"KES {v / 1_000_000:,.1f}M"
    if abs(v) >= 1_000:
        return f"KES {v / 1_000:,.0f}K"
    return f"KES {v:,.0f}"


def _vsla_fmt_pct(v):
    """Format percentage (0–1 → %)."""
    return f"{v * 100:.0f}%" if abs(v) <= 1 else f"{v:.1f}%"


def _vsla_band_data(bands, quarter):
    """Get band values for a specific quarter."""
    return [(b['band'], b.get(quarter, 0)) for b in bands]


def _vsla_plotly_layout(title, height=400, **kwargs):
    """Standard VSLA chart layout."""
    layout = dict(
        title=dict(text=title, font=dict(size=16, color='#222')),
        height=height,
        font=dict(size=13, color='#333'),
        plot_bgcolor='rgba(0,0,0,0)',
        paper_bgcolor='rgba(0,0,0,0)',
        margin=dict(l=20, r=20, t=60, b=20),
        legend=dict(orientation='h', yanchor='bottom', y=1.02,
                    x=0.5, xanchor='center'),
    )
    layout.update(kwargs)
    return layout


def render_vsla_tabs(vsla_data, county_filter='All Counties', quarter_filter='Q4'):
    """Render the VSLA Functionality module with 5 tabs."""

    d = vsla_data
    c = county_filter
    q = quarter_filter
    Q_LABELS = {'Q2': 'Q2 2025', 'Q3': 'Q3 2025', 'Q4': 'Q4 2025'}
    ql = Q_LABELS.get(q, q)
    QUARTERS = ['Q2', 'Q3', 'Q4']

    tabs = st.tabs([
        "📊 Overview",
        "👥 Membership & Meetings",
        "💰 Savings & Social Fund",
        "🏦 Loans & Repayment",
        "📈 County Comparison",
    ])

    # ================================================================
    # TAB 1 — Overview
    # ================================================================
    with tabs[0]:
        _section_header('', 'VSLA Functionality Overview', ql)

        # KPI cards
        vslas = _vsla_get_cq(d['vslasAssessed'], c, q)
        total_members = _vsla_get_sa(d['membership']['all'], c, q, 'sum')
        pct_women = 0
        if total_members > 0:
            female_m = _vsla_get_sa(d['membership']['female'], c, q, 'sum')
            pct_women = female_m / total_members * 100
        meeting_freq = _vsla_get_cq(d['meetings']['frequency'], c, q)
        attendance = _vsla_get_cq(d['meetings']['attendance'], c, q)
        total_savings = _vsla_get_sa(d['savings']['value'], c, q, 'sum')
        loans_disbursed = _vsla_get_sa(d['loans']['disbursed']['value'], c, q, 'sum')
        loans_repaid = _vsla_get_sa(d['loans']['repaid']['value'], c, q, 'sum')

        # ROI
        roi_val = 0
        for r in d['loans']['behaviour']:
            target = c if c != 'All Counties' else 'All'
            if r['county'] == target:
                roi_val = r['avgROI'].get(q, 0)
                break

        c1, c2, c3, c4, c5, c6 = st.columns(6)
        c1.markdown(f"""<div class="kpi-card"><h3>VSLAs Assessed</h3>
            <div class="value">{int(vslas):,}</div></div>""", unsafe_allow_html=True)
        c2.markdown(f"""<div class="kpi-card"><h3>Total Members</h3>
            <div class="value">{int(total_members):,}</div>
            <p style="color:#666">{pct_women:.0f}% women</p></div>""", unsafe_allow_html=True)
        m_color = COLORS['good'] if attendance >= 0.7 else COLORS['danger']
        c3.markdown(f"""<div class="kpi-card"><h3>Meeting Frequency</h3>
            <div class="value">{meeting_freq:.1f}</div>
            <p style="color:{m_color}">Attendance: {_vsla_fmt_pct(attendance)}</p></div>""", unsafe_allow_html=True)
        c4.markdown(f"""<div class="kpi-card"><h3>Total Savings</h3>
            <div class="value">{_vsla_fmt_kes(total_savings)}</div></div>""", unsafe_allow_html=True)
        c5.markdown(f"""<div class="kpi-card"><h3>Loans Disbursed</h3>
            <div class="value">{_vsla_fmt_kes(loans_disbursed)}</div>
            <p style="color:#666">Repaid: {_vsla_fmt_kes(loans_repaid)}</p></div>""", unsafe_allow_html=True)
        roi_color = COLORS['good'] if roi_val >= 0.05 else '#FF9800'
        c6.markdown(f"""<div class="kpi-card"><h3>Average ROI</h3>
            <div class="value">{_vsla_fmt_pct(roi_val)}</div></div>""", unsafe_allow_html=True)

        st.markdown("---")

        # Trend Charts — 2×2 grid
        _section_header('', 'Quarterly Trends', 'Q2 → Q4')
        ch1, ch2 = st.columns(2)

        with ch1:
            # Members trend
            m_trend = _vsla_trend(d['membership']['all'], c, 'sum')
            fig = go.Figure()
            fig.add_trace(go.Scatter(
                x=QUARTERS, y=m_trend, mode='lines+markers+text',
                name='Total Members', line=dict(color=COLORS['midline'], width=3),
                text=[f"{int(v):,}" for v in m_trend], textposition='top center',
                marker=dict(size=10),
            ))
            fig.update_layout(**_vsla_plotly_layout('Total Members — Q2→Q4', height=350))
            st.plotly_chart(fig, use_container_width=True)

        with ch2:
            # Savings trend
            s_trend = _vsla_trend(d['savings']['value'], c, 'sum')
            fig = go.Figure()
            fig.add_trace(go.Scatter(
                x=QUARTERS, y=s_trend, mode='lines+markers+text',
                name='Total Savings', line=dict(color='#1B5E20', width=3),
                text=[_vsla_fmt_kes(v) for v in s_trend], textposition='top center',
                marker=dict(size=10),
            ))
            fig.update_layout(**_vsla_plotly_layout('Total Savings (KES) — Q2→Q4', height=350))
            st.plotly_chart(fig, use_container_width=True)

        ch3, ch4 = st.columns(2)

        with ch3:
            # Loans disbursed vs repaid
            ld_trend = _vsla_trend(d['loans']['disbursed']['value'], c, 'sum')
            lr_trend = _vsla_trend(d['loans']['repaid']['value'], c, 'sum')
            fig = go.Figure()
            fig.add_trace(go.Bar(
                x=QUARTERS, y=ld_trend, name='Disbursed',
                marker_color=COLORS['midline'],
                text=[_vsla_fmt_kes(v) for v in ld_trend], textposition='auto',
            ))
            fig.add_trace(go.Scatter(
                x=QUARTERS, y=lr_trend, name='Repaid', mode='lines+markers',
                line=dict(color=COLORS['good'], width=3),
                marker=dict(size=10),
            ))
            fig.update_layout(**_vsla_plotly_layout('Loans Disbursed vs Repaid (KES)', height=350))
            st.plotly_chart(fig, use_container_width=True)

        with ch4:
            # Social fund coverage trend
            sf_trend = _vsla_cq_trend(d['socialFund']['percentageWithFund'], c)
            fig = go.Figure()
            fig.add_trace(go.Scatter(
                x=QUARTERS, y=[v * 100 for v in sf_trend], mode='lines+markers+text',
                name='% with Fund', line=dict(color='#FF9800', width=3),
                text=[f"{v * 100:.0f}%" for v in sf_trend], textposition='top center',
                marker=dict(size=10),
            ))
            fig.update_layout(**_vsla_plotly_layout('% VSLAs with Social Fund — Q2→Q4', height=350,
                                                     yaxis_title='%'))
            st.plotly_chart(fig, use_container_width=True)

        st.markdown("---")

        # Distribution charts
        _section_header('', 'Key Distributions', ql)
        dc1, dc2 = st.columns(2)

        with dc1:
            loan_use = _vsla_band_data(d['loans']['useDistribution'], q)
            lu_labels = [b[0] for b in loan_use]
            lu_vals = [b[1] * 100 for b in loan_use]
            fig = go.Figure(data=[go.Pie(
                labels=lu_labels, values=lu_vals,
                hole=0.45, textinfo='label+percent',
                marker=dict(colors=['#1B5E20', '#0D47A1', '#FF9800', '#E53935', '#7B1FA2']),
            )])
            fig.update_layout(**_vsla_plotly_layout(f'Loan Use Distribution — {ql}', height=380))
            st.plotly_chart(fig, use_container_width=True)

        with dc2:
            sav_bands = _vsla_band_data(d['savings']['proportionBands'], q)
            fig = go.Figure()
            fig.add_trace(go.Bar(
                x=[b[0] for b in sav_bands], y=[b[1] * 100 for b in sav_bands],
                name='% of VSLAs', marker_color=COLORS['midline'],
                text=[f"{b[1] * 100:.0f}%" for b in sav_bands], textposition='auto',
            ))
            fig.update_layout(**_vsla_plotly_layout(f'Members Saving Proportion Bands — {ql}', height=380,
                                                     yaxis_title='% of VSLAs'))
            st.plotly_chart(fig, use_container_width=True)

    # ================================================================
    # TAB 2 — Membership & Meetings
    # ================================================================
    with tabs[1]:
        _section_header('', 'Membership Overview', ql)

        female_sum = _vsla_get_sa(d['membership']['female'], c, q, 'sum')
        male_sum = _vsla_get_sa(d['membership']['male'], c, q, 'sum')
        all_sum = _vsla_get_sa(d['membership']['all'], c, q, 'sum')
        left_count = _vsla_get_cq(d['membersLeft'], c, q)

        k1, k2, k3, k4 = st.columns(4)
        k1.markdown(f"""<div class="kpi-card"><h3>Total Members</h3>
            <div class="value">{int(all_sum):,}</div></div>""", unsafe_allow_html=True)
        k2.markdown(f"""<div class="kpi-card"><h3>Female Members</h3>
            <div class="value">{int(female_sum):,}</div>
            <p style="color:#666">{female_sum/max(all_sum,1)*100:.0f}% of total</p></div>""", unsafe_allow_html=True)
        k3.markdown(f"""<div class="kpi-card"><h3>Male Members</h3>
            <div class="value">{int(male_sum):,}</div></div>""", unsafe_allow_html=True)
        k4.markdown(f"""<div class="kpi-card"><h3>Members Left</h3>
            <div class="value">{int(left_count):,}</div></div>""", unsafe_allow_html=True)

        st.markdown("---")

        # Membership trend by gender
        _section_header('', 'Membership Trends by Gender', 'Q2 → Q4')
        f_trend = _vsla_trend(d['membership']['female'], c, 'sum')
        m_trend_g = _vsla_trend(d['membership']['male'], c, 'sum')
        a_trend = _vsla_trend(d['membership']['all'], c, 'sum')

        fig = go.Figure()
        fig.add_trace(go.Bar(x=QUARTERS, y=f_trend, name='Female',
                             marker_color='#E91E63', text=[f"{int(v):,}" for v in f_trend], textposition='auto'))
        fig.add_trace(go.Bar(x=QUARTERS, y=m_trend_g, name='Male',
                             marker_color='#1565C0', text=[f"{int(v):,}" for v in m_trend_g], textposition='auto'))
        fig.update_layout(**_vsla_plotly_layout('Members by Gender — Q2→Q4', height=400, barmode='group',
                                                 yaxis_title='Members'))
        st.plotly_chart(fig, use_container_width=True)

        # Gender split donut
        gc1, gc2 = st.columns(2)
        with gc1:
            fig = go.Figure(data=[go.Pie(
                labels=['Female', 'Male'], values=[female_sum, male_sum],
                hole=0.5, marker=dict(colors=['#E91E63', '#1565C0']),
                textinfo='label+percent+value',
            )])
            fig.update_layout(**_vsla_plotly_layout(f'Gender Split — {ql}', height=350))
            st.plotly_chart(fig, use_container_width=True)

        with gc2:
            # % left bands
            left_bands = _vsla_band_data(d['percentLeftBands'], q)
            fig = go.Figure()
            fig.add_trace(go.Bar(
                x=[b[0] for b in left_bands], y=[b[1] * 100 for b in left_bands],
                name='% of VSLAs', marker_color='#E53935',
                text=[f"{b[1] * 100:.0f}%" for b in left_bands], textposition='auto',
            ))
            fig.update_layout(**_vsla_plotly_layout(f'Members Left Proportion Bands — {ql}', height=350,
                                                     yaxis_title='% of VSLAs'))
            st.plotly_chart(fig, use_container_width=True)

        st.markdown("---")

        # Meetings
        _section_header('', 'Meetings & Attendance', ql)
        mk1, mk2 = st.columns(2)
        freq_trend = _vsla_cq_trend(d['meetings']['frequency'], c)
        att_trend = _vsla_cq_trend(d['meetings']['attendance'], c)

        with mk1:
            fig = go.Figure()
            fig.add_trace(go.Bar(
                x=QUARTERS, y=freq_trend, name='Avg Meetings/Month',
                marker_color=COLORS['midline'],
                text=[f"{v:.1f}" for v in freq_trend], textposition='auto',
            ))
            fig.update_layout(**_vsla_plotly_layout('Average Meeting Frequency — Q2→Q4', height=350,
                                                     yaxis_title='Meetings per Month'))
            st.plotly_chart(fig, use_container_width=True)

        with mk2:
            fig = go.Figure()
            fig.add_trace(go.Scatter(
                x=QUARTERS, y=[v * 100 for v in att_trend], mode='lines+markers+text',
                name='Attendance %', line=dict(color=COLORS['good'], width=3),
                text=[f"{v * 100:.0f}%" for v in att_trend], textposition='top center',
                marker=dict(size=10),
            ))
            fig.update_layout(**_vsla_plotly_layout('Meeting Attendance Rate — Q2→Q4', height=350,
                                                     yaxis_title='% Attendance'))
            st.plotly_chart(fig, use_container_width=True)

    # ================================================================
    # TAB 3 — Savings & Social Fund
    # ================================================================
    with tabs[2]:
        _section_header('', 'Savings Analysis', ql)

        sav_sum = _vsla_get_sa(d['savings']['value'], c, q, 'sum')
        sav_avg = _vsla_get_sa(d['savings']['value'], c, q, 'average')
        members_saving = _vsla_get_sa(d['savings']['membersSaving'], c, q, 'sum')
        high_savers = 0
        for b in d['savings']['proportionBands']:
            if '81-100' in b['band']:
                high_savers = b.get(q, 0)
                break

        sk1, sk2, sk3, sk4 = st.columns(4)
        sk1.markdown(f"""<div class="kpi-card"><h3>Members Saving</h3>
            <div class="value">{int(members_saving):,}</div></div>""", unsafe_allow_html=True)
        sk2.markdown(f"""<div class="kpi-card"><h3>Total Savings</h3>
            <div class="value">{_vsla_fmt_kes(sav_sum)}</div></div>""", unsafe_allow_html=True)
        sk3.markdown(f"""<div class="kpi-card"><h3>Avg per VSLA</h3>
            <div class="value">{_vsla_fmt_kes(sav_avg)}</div></div>""", unsafe_allow_html=True)
        sk4.markdown(f"""<div class="kpi-card"><h3>High Savers (81-100%)</h3>
            <div class="value">{_vsla_fmt_pct(high_savers)}</div></div>""", unsafe_allow_html=True)

        st.markdown("---")

        sc1, sc2 = st.columns(2)
        with sc1:
            s_trend = _vsla_trend(d['savings']['value'], c, 'sum')
            fig = go.Figure()
            fig.add_trace(go.Scatter(
                x=QUARTERS, y=s_trend, mode='lines+markers+text',
                name='Total Savings', line=dict(color='#1B5E20', width=3),
                text=[_vsla_fmt_kes(v) for v in s_trend], textposition='top center',
                marker=dict(size=10),
            ))
            fig.update_layout(**_vsla_plotly_layout('Total Savings Value (KES) — Q2→Q4', height=350))
            st.plotly_chart(fig, use_container_width=True)

        with sc2:
            savg_trend = _vsla_trend(d['savings']['value'], c, 'average')
            fig = go.Figure()
            fig.add_trace(go.Scatter(
                x=QUARTERS, y=savg_trend, mode='lines+markers+text',
                name='Avg Savings', line=dict(color='#0D47A1', width=3),
                text=[_vsla_fmt_kes(v) for v in savg_trend], textposition='top center',
                marker=dict(size=10),
            ))
            fig.update_layout(**_vsla_plotly_layout('Average Savings per VSLA (KES) — Q2→Q4', height=350))
            st.plotly_chart(fig, use_container_width=True)

        # Savings bands
        sc3, sc4 = st.columns(2)
        with sc3:
            sp_bands = _vsla_band_data(d['savings']['proportionBands'], q)
            fig = go.Figure(data=[go.Pie(
                labels=[b[0] for b in sp_bands], values=[b[1] * 100 for b in sp_bands],
                hole=0.45, textinfo='label+percent',
            )])
            fig.update_layout(**_vsla_plotly_layout(f'Savings Proportion Bands — {ql}', height=380))
            st.plotly_chart(fig, use_container_width=True)

        with sc4:
            sv_bands = _vsla_band_data(d['savings']['valueBands'], q)
            fig = go.Figure()
            fig.add_trace(go.Bar(
                x=[b[0] for b in sv_bands], y=[b[1] * 100 for b in sv_bands],
                marker_color='#1B5E20',
                text=[f"{b[1] * 100:.0f}%" for b in sv_bands], textposition='auto',
            ))
            fig.update_layout(**_vsla_plotly_layout(f'Savings Amount Bands — {ql}', height=380,
                                                     yaxis_title='% of VSLAs'))
            st.plotly_chart(fig, use_container_width=True)

        # Savings data table
        _section_header('', 'Savings by County & Quarter', '')
        tbl_data = []
        for county_name in ['Kilifi', 'Kwale', 'All']:
            row = {'County': county_name}
            for qq in QUARTERS:
                s = _vsla_get_sa(d['savings']['value'], county_name if county_name != 'All' else 'All Counties', qq, 'sum')
                a = _vsla_get_sa(d['savings']['value'], county_name if county_name != 'All' else 'All Counties', qq, 'average')
                row[f'{qq} Total'] = _vsla_fmt_kes(s)
                row[f'{qq} Avg'] = _vsla_fmt_kes(a)
            tbl_data.append(row)
        st.dataframe(pd.DataFrame(tbl_data), use_container_width=True, hide_index=True)

        st.markdown("---")

        # Social Fund section
        _section_header('', 'Social Fund Analysis', ql)

        pct_fund = _vsla_get_cq(d['socialFund']['percentageWithFund'], c, q)
        sf_sum = _vsla_get_sa(d['socialFund']['value'], c, q, 'sum')
        sf_avg = _vsla_get_sa(d['socialFund']['value'], c, q, 'average')

        sfk1, sfk2, sfk3 = st.columns(3)
        sfk1.markdown(f"""<div class="kpi-card"><h3>VSLAs with Social Fund</h3>
            <div class="value">{_vsla_fmt_pct(pct_fund)}</div></div>""", unsafe_allow_html=True)
        sfk2.markdown(f"""<div class="kpi-card"><h3>Total Social Fund</h3>
            <div class="value">{_vsla_fmt_kes(sf_sum)}</div></div>""", unsafe_allow_html=True)
        sfk3.markdown(f"""<div class="kpi-card"><h3>Avg per VSLA</h3>
            <div class="value">{_vsla_fmt_kes(sf_avg)}</div></div>""", unsafe_allow_html=True)

        sf1, sf2 = st.columns(2)
        with sf1:
            sf_pct_trend = _vsla_cq_trend(d['socialFund']['percentageWithFund'], c)
            fig = go.Figure()
            fig.add_trace(go.Scatter(
                x=QUARTERS, y=[v * 100 for v in sf_pct_trend], mode='lines+markers+text',
                name='% with Fund', line=dict(color='#FF9800', width=3),
                text=[f"{v * 100:.0f}%" for v in sf_pct_trend], textposition='top center',
                marker=dict(size=10),
            ))
            fig.update_layout(**_vsla_plotly_layout('% VSLAs with Social Fund — Q2→Q4', height=350,
                                                     yaxis_title='%'))
            st.plotly_chart(fig, use_container_width=True)

        with sf2:
            sf_val_trend = _vsla_trend(d['socialFund']['value'], c, 'sum')
            fig = go.Figure()
            fig.add_trace(go.Scatter(
                x=QUARTERS, y=sf_val_trend, mode='lines+markers+text',
                name='Fund Value', line=dict(color='#1B5E20', width=3),
                text=[_vsla_fmt_kes(v) for v in sf_val_trend], textposition='top center',
                marker=dict(size=10),
            ))
            fig.update_layout(**_vsla_plotly_layout('Total Social Fund Value (KES) — Q2→Q4', height=350))
            st.plotly_chart(fig, use_container_width=True)

        # Social fund bands
        sf_bands = _vsla_band_data(d['socialFund']['valueBands'], q)
        fig = go.Figure()
        fig.add_trace(go.Bar(
            x=[b[0] for b in sf_bands], y=[b[1] * 100 for b in sf_bands],
            marker_color='#FF9800',
            text=[f"{b[1] * 100:.0f}%" for b in sf_bands], textposition='auto',
        ))
        fig.update_layout(**_vsla_plotly_layout(f'Social Fund Amount Bands — {ql}', height=380,
                                                 yaxis_title='% of VSLAs'))
        st.plotly_chart(fig, use_container_width=True)

    # ================================================================
    # TAB 4 — Loans & Repayment
    # ================================================================
    with tabs[3]:
        _section_header('', 'Loans Overview', ql)

        ld_count = _vsla_get_sa(d['loans']['disbursed']['count'], c, q, 'sum')
        ld_val = _vsla_get_sa(d['loans']['disbursed']['value'], c, q, 'sum')
        lr_count = _vsla_get_sa(d['loans']['repaid']['count'], c, q, 'sum')
        lr_val = _vsla_get_sa(d['loans']['repaid']['value'], c, q, 'sum')
        interest_val = _vsla_get_sa(d['loans']['interest']['value'], c, q, 'sum')
        failing_pct = _vsla_get_cq(d['loans']['failingToPay'], c, q)

        lk1, lk2, lk3 = st.columns(3)
        lk1.markdown(f"""<div class="kpi-card"><h3>Loans Disbursed</h3>
            <div class="value">{int(ld_count):,}</div>
            <p style="color:#666">Value: {_vsla_fmt_kes(ld_val)}</p></div>""", unsafe_allow_html=True)
        lk2.markdown(f"""<div class="kpi-card"><h3>Loans Repaid</h3>
            <div class="value">{int(lr_count):,}</div>
            <p style="color:#666">Value: {_vsla_fmt_kes(lr_val)}</p></div>""", unsafe_allow_html=True)
        repay_rate = lr_val / max(ld_val, 1) * 100
        rr_color = COLORS['good'] if repay_rate >= 70 else COLORS['danger']
        lk3.markdown(f"""<div class="kpi-card"><h3>Repayment Rate</h3>
            <div class="value" style="color:{rr_color}">{repay_rate:.0f}%</div>
            <p style="color:#666">Interest: {_vsla_fmt_kes(interest_val)}</p></div>""", unsafe_allow_html=True)

        lk4, lk5, lk6 = st.columns(3)
        lk4.markdown(f"""<div class="kpi-card"><h3>Avg ROI</h3>
            <div class="value">{_vsla_fmt_pct(roi_val)}</div></div>""", unsafe_allow_html=True)
        fail_color = COLORS['good'] if failing_pct <= 0.1 else COLORS['danger']
        lk5.markdown(f"""<div class="kpi-card"><h3>Failing to Pay</h3>
            <div class="value" style="color:{fail_color}">{_vsla_fmt_pct(failing_pct)}</div></div>""",
            unsafe_allow_html=True)
        lk6.markdown(f"""<div class="kpi-card"><h3>Avg Value Repaid</h3>
            <div class="value">{_vsla_fmt_kes(_vsla_get_sa(d['loans']['repaid']['value'], c, q, 'average'))}</div></div>""",
            unsafe_allow_html=True)

        st.markdown("---")

        # Loan trends
        _section_header('', 'Loan Trends', 'Q2 → Q4')
        lt1, lt2 = st.columns(2)

        with lt1:
            ld_trend = _vsla_trend(d['loans']['disbursed']['value'], c, 'sum')
            lr_trend = _vsla_trend(d['loans']['repaid']['value'], c, 'sum')
            fig = go.Figure()
            fig.add_trace(go.Bar(
                x=QUARTERS, y=ld_trend, name='Disbursed',
                marker_color=COLORS['midline'],
                text=[_vsla_fmt_kes(v) for v in ld_trend], textposition='auto',
            ))
            fig.add_trace(go.Bar(
                x=QUARTERS, y=lr_trend, name='Repaid',
                marker_color=COLORS['good'],
                text=[_vsla_fmt_kes(v) for v in lr_trend], textposition='auto',
            ))
            fig.update_layout(**_vsla_plotly_layout('Loan Value: Disbursed vs Repaid — Q2→Q4', height=400,
                                                     barmode='group', yaxis_title='KES'))
            st.plotly_chart(fig, use_container_width=True)

        with lt2:
            int_trend = _vsla_trend(d['loans']['interest']['value'], c, 'sum')
            fig = go.Figure()
            fig.add_trace(go.Scatter(
                x=QUARTERS, y=int_trend, mode='lines+markers+text',
                name='Interest', line=dict(color='#FF9800', width=3),
                text=[_vsla_fmt_kes(v) for v in int_trend], textposition='top center',
                marker=dict(size=10),
            ))
            fig.update_layout(**_vsla_plotly_layout('Interest Income (KES) — Q2→Q4', height=400))
            st.plotly_chart(fig, use_container_width=True)

        # Loan distribution charts
        st.markdown("---")
        _section_header('', 'Loan Distributions', ql)
        ld1, ld2 = st.columns(2)

        with ld1:
            loan_use = _vsla_band_data(d['loans']['useDistribution'], q)
            fig = go.Figure(data=[go.Pie(
                labels=[b[0] for b in loan_use], values=[b[1] * 100 for b in loan_use],
                hole=0.45, textinfo='label+percent',
                marker=dict(colors=['#1B5E20', '#0D47A1', '#FF9800', '#E53935', '#7B1FA2']),
            )])
            fig.update_layout(**_vsla_plotly_layout(f'Loan Use Purpose — {ql}', height=380))
            st.plotly_chart(fig, use_container_width=True)

        with ld2:
            lv_bands = _vsla_band_data(d['loans']['disbursed']['valueBands'], q)
            fig = go.Figure()
            fig.add_trace(go.Bar(
                x=[b[0] for b in lv_bands], y=[b[1] * 100 for b in lv_bands],
                marker_color=COLORS['midline'],
                text=[f"{b[1] * 100:.0f}%" for b in lv_bands], textposition='auto',
            ))
            fig.update_layout(**_vsla_plotly_layout(f'Loan Value Bands — {ql}', height=380,
                                                     yaxis_title='% of VSLAs'))
            st.plotly_chart(fig, use_container_width=True)

        # ROI behaviour table
        _section_header('', 'Loan Behaviour by County', '')
        beh_rows = []
        for br in d['loans']['behaviour']:
            beh_rows.append({
                'County': br['county'],
                'Avg % Repaid (Q2)': f"{br['avgProportionRepaid']['Q2'] * 100:.0f}%",
                'Avg % Repaid (Q3)': f"{br['avgProportionRepaid']['Q3'] * 100:.0f}%",
                'Avg % Repaid (Q4)': f"{br['avgProportionRepaid']['Q4'] * 100:.0f}%",
                'Avg ROI (Q2)': f"{br['avgROI']['Q2'] * 100:.1f}%",
                'Avg ROI (Q3)': f"{br['avgROI']['Q3'] * 100:.1f}%",
                'Avg ROI (Q4)': f"{br['avgROI']['Q4'] * 100:.1f}%",
            })
        st.dataframe(pd.DataFrame(beh_rows), use_container_width=True, hide_index=True)

    # ================================================================
    # TAB 5 — County Comparison
    # ================================================================
    with tabs[4]:
        _section_header('', 'Kilifi vs Kwale Comparison', ql)

        compare_indicators = [
            ('VSLAs Assessed', lambda cn: _vsla_get_cq(d['vslasAssessed'], cn, q)),
            ('Total Members', lambda cn: _vsla_get_sa(d['membership']['all'], cn, q, 'sum')),
            ('Female Members', lambda cn: _vsla_get_sa(d['membership']['female'], cn, q, 'sum')),
            ('Total Savings (KES)', lambda cn: _vsla_get_sa(d['savings']['value'], cn, q, 'sum')),
            ('Avg Savings / VSLA', lambda cn: _vsla_get_sa(d['savings']['value'], cn, q, 'average')),
            ('Social Fund %', lambda cn: _vsla_get_cq(d['socialFund']['percentageWithFund'], cn, q) * 100),
            ('Loans Disbursed (KES)', lambda cn: _vsla_get_sa(d['loans']['disbursed']['value'], cn, q, 'sum')),
            ('Loans Repaid (KES)', lambda cn: _vsla_get_sa(d['loans']['repaid']['value'], cn, q, 'sum')),
            ('Meeting Attendance', lambda cn: _vsla_get_cq(d['meetings']['attendance'], cn, q) * 100),
        ]

        comp_data = []
        for label, fn in compare_indicators:
            kilifi_val = fn('Kilifi')
            kwale_val = fn('Kwale')
            comp_data.append({
                'Indicator': label,
                'Kilifi': kilifi_val,
                'Kwale': kwale_val,
                'Difference': kilifi_val - kwale_val,
            })

        # Grouped horizontal bar
        fig = go.Figure()
        indicator_labels = [r['Indicator'] for r in comp_data]
        kilifi_vals = [r['Kilifi'] for r in comp_data]
        kwale_vals = [r['Kwale'] for r in comp_data]

        fig.add_trace(go.Bar(
            y=indicator_labels, x=kilifi_vals, name='Kilifi',
            orientation='h', marker_color='#1B5E20',
        ))
        fig.add_trace(go.Bar(
            y=indicator_labels, x=kwale_vals, name='Kwale',
            orientation='h', marker_color='#0D47A1',
        ))
        fig.update_layout(**_vsla_plotly_layout(f'County Comparison — {ql}', height=500,
                                                 barmode='group',
                                                 yaxis=dict(categoryorder='array',
                                                            categoryarray=list(reversed(indicator_labels)))))
        st.plotly_chart(fig, use_container_width=True)

        # Comparison table
        st.markdown("---")
        _section_header('', 'Detailed Comparison Table', '')
        comp_df = pd.DataFrame(comp_data)
        # Format large numbers
        for col in ['Kilifi', 'Kwale', 'Difference']:
            comp_df[col] = comp_df[col].apply(lambda v: f"{v:,.0f}")
        st.dataframe(comp_df, use_container_width=True, hide_index=True)

        # Per-indicator trend comparison
        st.markdown("---")
        _section_header('', 'Trend Comparison: Kilifi vs Kwale', 'Q2 → Q4')

        tc1, tc2 = st.columns(2)
        with tc1:
            ki_sav = _vsla_trend(d['savings']['value'], 'Kilifi', 'sum')
            kw_sav = _vsla_trend(d['savings']['value'], 'Kwale', 'sum')
            fig = go.Figure()
            fig.add_trace(go.Scatter(x=QUARTERS, y=ki_sav, name='Kilifi',
                                     mode='lines+markers', line=dict(color='#1B5E20', width=3)))
            fig.add_trace(go.Scatter(x=QUARTERS, y=kw_sav, name='Kwale',
                                     mode='lines+markers', line=dict(color='#0D47A1', width=3)))
            fig.update_layout(**_vsla_plotly_layout('Savings Trend — Kilifi vs Kwale', height=350))
            st.plotly_chart(fig, use_container_width=True)

        with tc2:
            ki_loan = _vsla_trend(d['loans']['disbursed']['value'], 'Kilifi', 'sum')
            kw_loan = _vsla_trend(d['loans']['disbursed']['value'], 'Kwale', 'sum')
            fig = go.Figure()
            fig.add_trace(go.Scatter(x=QUARTERS, y=ki_loan, name='Kilifi',
                                     mode='lines+markers', line=dict(color='#1B5E20', width=3)))
            fig.add_trace(go.Scatter(x=QUARTERS, y=kw_loan, name='Kwale',
                                     mode='lines+markers', line=dict(color='#0D47A1', width=3)))
            fig.update_layout(**_vsla_plotly_layout('Loans Disbursed Trend — Kilifi vs Kwale', height=350))
            st.plotly_chart(fig, use_container_width=True)


def _generate_forestry_insights(data):
    """Generate automated insights from forestry data."""
    insights = []
    try:
        _gen_forestry_insights_inner(data, insights)
    except Exception as e:
        insights.append(("Insight Generation Note",
                         f"Some forestry insights could not be generated: {e}",
                         "neutral"))
    return insights


def _gen_forestry_insights_inner(data, insights):
    ft = data['functionality_threshold']
    bl_60 = ft.loc[ft['Timepoint'] == 'Baseline', 'Functional_60_pct'].values[0]
    ml_60 = ft.loc[ft['Timepoint'] == 'Midline', 'Functional_60_pct'].values[0]
    change_60 = _pp(bl_60, ml_60, multiply=isinstance(bl_60, float) and bl_60 <= 1)
    if change_60 > 0:
        insights.append(("Functionality Threshold Improved",
                         f"Groups meeting the 60% functionality threshold increased by {change_60:+.1f} pp "
                         f"(Baseline: {bl_60*100 if bl_60 <= 1 else bl_60:.1f}% to Midline: {ml_60*100 if ml_60 <= 1 else ml_60:.1f}%). "
                         f"This signals stronger institutional capacity at community level.",
                         "positive"))
    elif change_60 < 0:
        insights.append(("Functionality Threshold Declined",
                         f"Groups meeting the 60% functionality threshold dropped by {change_60:+.1f} pp. "
                         f"This may require targeted institutional strengthening.",
                         "negative"))

    # 2. Domain scores
    fd = data['functionality_domain']
    for domain in ['Management', 'Gender', 'Effectiveness']:
        bl_d = fd.loc[fd['Timepoint'] == 'Baseline', domain].values[0]
        ml_d = fd.loc[fd['Timepoint'] == 'Midline', domain].values[0]
        change_d = round(ml_d - bl_d, 1) if isinstance(bl_d, (int, float)) and isinstance(ml_d, (int, float)) else 0
        if abs(change_d) >= 3:
            trend = "positive" if change_d > 0 else "negative"
            insights.append((f"{domain} Domain: {change_d:+.1f} Points",
                             f"The {domain} domain score moved from {bl_d:.1f} to {ml_d:.1f} "
                             f"({change_d:+.1f} pts). "
                             f"{'This is a meaningful improvement.' if change_d > 0 else 'This decline warrants attention.'}",
                             trend))

    # 3. Women leadership
    wl = data['women_leadership']
    sig_bl = wl.loc[wl['Category'] == 'Significant Leadership', 'Baseline'].values[0]
    sig_ml = wl.loc[wl['Category'] == 'Significant Leadership', 'Midline'].values[0]
    wl_change = _pp(sig_bl, sig_ml, multiply=isinstance(sig_bl, float) and sig_bl <= 1)
    if abs(wl_change) >= 2:
        trend = "positive" if wl_change > 0 else "warning"
        insights.append(("Women in Significant Leadership",
                         f"The share of groups with significant women's leadership changed by {wl_change:+.1f} pp "
                         f"({sig_bl*100 if sig_bl <= 1 else sig_bl:.1f}% to {sig_ml*100 if sig_ml <= 1 else sig_ml:.1f}%). "
                         f"Gender equity in forest governance is {'advancing' if wl_change > 0 else 'a continuing challenge'}.",
                         trend))

    # 4. Training coverage
    tc = data['training_coverage']
    most_bl = tc.loc[tc['Category'] == 'Most Members', 'Baseline'].values[0]
    most_ml = tc.loc[tc['Category'] == 'Most Members', 'Midline'].values[0]
    tc_change = _pp(most_bl, most_ml, multiply=isinstance(most_bl, float) and most_bl <= 1)
    insights.append(("Training Coverage (Most Members)",
                     f"Groups where most members received training: "
                     f"{most_bl*100 if most_bl <= 1 else most_bl:.1f}% (BL) to {most_ml*100 if most_ml <= 1 else most_ml:.1f}% (ML), "
                     f"a change of {tc_change:+.1f} pp. "
                     f"{'Capacity building is expanding.' if tc_change > 0 else 'Training outreach needs scaling up.'}",
                     "positive" if tc_change > 0 else "warning"))

    # 5. Income generation
    ig = data['income_gen']
    yes_bl = ig.loc[ig['Category'] == 'Yes', 'Baseline'].values[0]
    yes_ml = ig.loc[ig['Category'] == 'Yes', 'Midline'].values[0]
    ig_change = _pp(yes_bl, yes_ml, multiply=isinstance(yes_bl, float) and yes_bl <= 1)
    insights.append(("Forest-Based Income Generation",
                     f"Groups generating income from forests: "
                     f"{yes_bl*100 if yes_bl <= 1 else yes_bl:.1f}% (BL) to {yes_ml*100 if yes_ml <= 1 else yes_ml:.1f}% (ML) "
                     f"({ig_change:+.1f} pp). "
                     f"{'Economic sustainability is strengthening.' if ig_change > 0 else 'Livelihood diversification needs attention.'}",
                     "positive" if ig_change > 0 else "warning"))

    # 6. Forest condition — good rated
    fc = data['forest_condition']
    good_cols_bl = [c for c in fc.columns if 'Good' in c and 'Baseline' in c]
    good_cols_ml = [c for c in fc.columns if 'Good' in c and 'Midline' in c]
    if good_cols_bl and good_cols_ml:
        avg_good_bl = fc[good_cols_bl[0]].mean()
        avg_good_ml = fc[good_cols_ml[0]].mean()
        fc_change = _pp(avg_good_bl, avg_good_ml, multiply=isinstance(avg_good_bl, float) and avg_good_bl <= 1)
        insights.append(("Forest Condition (Good Rating)",
                         f"Average 'Good' forest condition rating across all characteristics: "
                         f"{avg_good_bl*100 if avg_good_bl <= 1 else avg_good_bl:.1f}% (BL) to "
                         f"{avg_good_ml*100 if avg_good_ml <= 1 else avg_good_ml:.1f}% (ML) ({fc_change:+.1f} pp). "
                         f"{'Forests are perceived as healthier.' if fc_change > 0 else 'Forest health perceptions have not improved.'}",
                         "positive" if fc_change > 0 else "warning"))

    # 7. Agroforestry adoption
    af = data['agroforestry']
    af_yes_bl = af.loc[af['Category'] == 'Yes', 'Baseline'].values[0]
    af_yes_ml = af.loc[af['Category'] == 'Yes', 'Midline'].values[0]
    af_change = _pp(af_yes_bl, af_yes_ml, multiply=isinstance(af_yes_bl, float) and af_yes_bl <= 1)
    insights.append(("Agroforestry Adoption",
                     f"Groups practicing agroforestry: "
                     f"{af_yes_bl*100 if af_yes_bl <= 1 else af_yes_bl:.1f}% (BL) to "
                     f"{af_yes_ml*100 if af_yes_ml <= 1 else af_yes_ml:.1f}% (ML) ({af_change:+.1f} pp). "
                     f"{'Adoption is growing, supporting landscape restoration.' if af_change > 0 else 'More incentives may be needed to drive adoption.'}",
                     "positive" if af_change > 0 else "neutral"))

    return insights


def _generate_women_insights(w):
    """Generate automated insights from women survey data."""
    insights = []
    try:
        _gen_women_insights_inner(w, insights)
    except Exception as e:
        insights.append(("Insight Generation Note",
                         f"Some women survey insights could not be generated: {e}",
                         "neutral"))
    return insights


def _gen_women_insights_inner(w, insights):
    # 1. Climate change awareness
    cc = w['cc_heard']
    cc_yes_bl = cc.loc[cc['Response'] == 'Yes', 'Baseline'].values[0]
    cc_yes_ml = cc.loc[cc['Response'] == 'Yes', 'Midline'].values[0]
    cc_change = _pp(cc_yes_bl, cc_yes_ml)
    insights.append(("Climate Change Awareness",
                     f"Women who have heard of climate change: "
                     f"{cc_yes_bl*100:.1f}% (BL) to {cc_yes_ml*100:.1f}% (ML) ({cc_change:+.1f} pp). "
                     f"{'Awareness is growing across communities.' if cc_change > 0 else 'Additional outreach on climate literacy is needed.'}",
                     "positive" if cc_change > 0 else "warning"))

    # 2. NbS awareness
    nbs = w['nbs_heard']
    nbs_yes_bl = nbs.loc[nbs['Response'] == 'Yes', 'Baseline'].values[0]
    nbs_yes_ml = nbs.loc[nbs['Response'] == 'Yes', 'Midline'].values[0]
    nbs_change = _pp(nbs_yes_bl, nbs_yes_ml)
    insights.append(("Nature-based Solutions Awareness",
                     f"Women aware of NbS: {nbs_yes_bl*100:.1f}% (BL) to {nbs_yes_ml*100:.1f}% (ML) "
                     f"({nbs_change:+.1f} pp). "
                     f"{'NbS concepts are reaching more women.' if nbs_change > 0 else 'NbS messaging needs strengthening.'}",
                     "positive" if nbs_change > 0 else "warning"))

    # 3. Savings behavior
    ps = w['personal_saving']
    ps_yes_bl = ps.loc[ps['Response'] == 'Yes', 'Baseline'].values[0]
    ps_yes_ml = ps.loc[ps['Response'] == 'Yes', 'Midline'].values[0]
    ps_change = _pp(ps_yes_bl, ps_yes_ml)
    insights.append(("Personal Savings",
                     f"Women with personal savings: {ps_yes_bl*100:.1f}% (BL) to {ps_yes_ml*100:.1f}% (ML) "
                     f"({ps_change:+.1f} pp). "
                     f"Financial resilience is {'building' if ps_change > 0 else 'a key area for intervention'}.",
                     "positive" if ps_change > 0 else "warning"))

    # 4. Decision-making influence — average across all decisions
    di = w['decision_influence']
    avg_bl = di['Baseline'].mean()
    avg_ml = di['Midline'].mean()
    di_change = _pp(avg_bl, avg_ml)
    insights.append(("Decision-Making Influence",
                     f"Average share of women reporting 'large extent' influence on HH decisions: "
                     f"{avg_bl*100:.1f}% (BL) to {avg_ml*100:.1f}% (ML) ({di_change:+.1f} pp). "
                     + ("Women's agency is strengthening." if di_change > 0 else "Empowerment efforts need scaling."),
                     "positive" if di_change > 0 else "warning"))

    # 5. Life skills — average agree/strongly agree
    ls = w['lifeskills_agree']
    ls_avg_bl = ls['Baseline'].mean()
    ls_avg_ml = ls['Midline'].mean()
    ls_change = _pp(ls_avg_bl, ls_avg_ml)
    insights.append(("Life Skills Confidence",
                     f"Average agree/strongly agree on life skills statements: "
                     f"{ls_avg_bl*100:.1f}% (BL) to {ls_avg_ml*100:.1f}% (ML) ({ls_change:+.1f} pp). "
                     f"{'Self-esteem and leadership confidence are growing.' if ls_change > 0 else 'Soft skills programming needs reinforcement.'}",
                     "positive" if ls_change > 0 else "neutral"))

    # 6. Social norms — want DECREASES (lower = better for harmful norms)
    sn = w['socialnorms_agree']
    harmful_norms = sn[~sn['Norm'].str.contains('Planting Crops|Restoring Ecosystems|Express Emotions', regex=True)]
    sn_avg_bl = harmful_norms['Baseline'].mean()
    sn_avg_ml = harmful_norms['Midline'].mean()
    sn_change = _pp(sn_avg_bl, sn_avg_ml)
    trend = "positive" if sn_change < 0 else ("negative" if sn_change > 2 else "neutral")
    insights.append(("Harmful Social Norms",
                     f"Average agreement with harmful gender norms: "
                     f"{sn_avg_bl*100:.1f}% (BL) to {sn_avg_ml*100:.1f}% (ML) ({sn_change:+.1f} pp). "
                     f"{'Positive shift — harmful norms are weakening.' if sn_change < 0 else 'Norms remain entrenched; continued dialogue is critical.'}",
                     trend))

    # 7. Coping strategies — top 3 most used at midline
    cop = w['coping'].copy()
    cop['ML_pct'] = cop['Midline'].apply(lambda x: x * 100 if isinstance(x, (int, float)) and x <= 1 else x)
    top3 = cop.nlargest(3, 'ML_pct')
    top3_text = ', '.join([f"{row['Strategy']} ({row['ML_pct']:.1f}%)" for _, row in top3.iterrows()])
    insights.append(("Top Coping Strategies (Midline)",
                     f"The three most common coping strategies at midline are: {top3_text}. "
                     f"High reliance on meal-skipping or asset sales signals food insecurity concerns.",
                     "warning"))

    # 8. Time use
    ts = w['time_summary']
    unpaid_bl = ts.loc[ts['Category'] == 'Unpaid Care Work', 'Baseline'].values[0]
    unpaid_ml = ts.loc[ts['Category'] == 'Unpaid Care Work', 'Midline'].values[0]
    unpaid_diff = round(unpaid_ml - unpaid_bl, 2) if isinstance(unpaid_bl, (int, float)) else 0
    insights.append(("Women's Unpaid Care Work",
                     f"Average daily unpaid care work: {unpaid_bl:.1f}h (BL) to {unpaid_ml:.1f}h (ML) "
                     f"({unpaid_diff:+.1f}h change). "
                     f"{'Care burden is decreasing — more time for productive activities.' if unpaid_diff < 0 else 'Care burden persists; labour-sharing interventions may help.'}",
                     "positive" if unpaid_diff < 0 else "warning"))

    # 9. Disaster preparedness
    pk = w['prep_knowledge']
    pk_yes_bl = pk.loc[pk['Response'] == 'Yes', 'Baseline'].values[0]
    pk_yes_ml = pk.loc[pk['Response'] == 'Yes', 'Midline'].values[0]
    pk_change = _pp(pk_yes_bl, pk_yes_ml)
    insights.append(("Disaster Preparedness Knowledge",
                     f"Women aware of disaster preparedness plans: "
                     f"{pk_yes_bl*100:.1f}% (BL) to {pk_yes_ml*100:.1f}% (ML) ({pk_change:+.1f} pp). "
                     f"{'Communities are becoming better prepared.' if pk_change > 0 else 'DRR awareness remains a gap.'}",
                     "positive" if pk_change > 0 else "warning"))

    return insights


def _generate_men_insights(m):
    """Generate automated insights from men survey data."""
    insights = []
    try:
        _gen_men_insights_inner(m, insights)
    except Exception as e:
        insights.append(("Insight Generation Note",
                         f"Some men survey insights could not be generated: {e}",
                         "neutral"))
    return insights


def _gen_men_insights_inner(m, insights):
    # 1. Climate change awareness
    cc = m['cc_heard']
    cc_yes_bl = cc.loc[cc['Response'] == 'Yes', 'Baseline'].values[0]
    cc_yes_ml = cc.loc[cc['Response'] == 'Yes', 'Midline'].values[0]
    cc_change = _pp(cc_yes_bl, cc_yes_ml)
    insights.append(("Men's CC Awareness",
                     f"Men who have heard of climate change: "
                     f"{cc_yes_bl*100:.1f}% (BL) to {cc_yes_ml*100:.1f}% (ML) ({cc_change:+.1f} pp). "
                     f"{'Awareness is growing among men.' if cc_change > 0 else 'Climate literacy programmes should target men as well.'}",
                     "positive" if cc_change > 0 else "warning"))

    # 2. NbS awareness
    nbs = m['nbs_heard']
    nbs_yes_bl = nbs.loc[nbs['Response'] == 'Yes', 'Baseline'].values[0]
    nbs_yes_ml = nbs.loc[nbs['Response'] == 'Yes', 'Midline'].values[0]
    nbs_change = _pp(nbs_yes_bl, nbs_yes_ml)
    insights.append(("Men's NbS Awareness",
                     f"Men aware of NbS: {nbs_yes_bl*100:.1f}% (BL) to {nbs_yes_ml*100:.1f}% (ML) "
                     f"({nbs_change:+.1f} pp). "
                     f"{'NbS concepts are reaching more men.' if nbs_change > 0 else 'NbS outreach to men needs strengthening.'}",
                     "positive" if nbs_change > 0 else "warning"))

    # 3. Joint roles (should) — average
    rj = m['roles_should_joint']
    rj_avg_bl = rj['Baseline'].mean()
    rj_avg_ml = rj['Midline'].mean()
    rj_change = _pp(rj_avg_bl, rj_avg_ml)
    insights.append(("Men's Attitude: Joint Roles",
                     f"Average share of men saying household roles SHOULD be joint: "
                     f"{rj_avg_bl*100:.1f}% (BL) to {rj_avg_ml*100:.1f}% (ML) ({rj_change:+.1f} pp). "
                     + ("Men increasingly support equitable household labour sharing." if rj_change > 0
                        else "Attitudes toward joint roles remain largely unchanged."),
                     "positive" if rj_change > 0 else "neutral"))

    # 4. Joint decisions (should) — average
    dj = m['decision_should_joint']
    dj_avg_bl = dj['Baseline'].mean()
    dj_avg_ml = dj['Midline'].mean()
    dj_change = _pp(dj_avg_bl, dj_avg_ml)
    insights.append(("Men's Attitude: Joint Decisions",
                     f"Average share of men saying household decisions SHOULD be joint: "
                     f"{dj_avg_bl*100:.1f}% (BL) to {dj_avg_ml*100:.1f}% (ML) ({dj_change:+.1f} pp). "
                     + ("More men are embracing shared decision-making norms." if dj_change > 0
                        else "Further engagement on equitable decision-making is needed."),
                     "positive" if dj_change > 0 else "neutral"))

    # 5. Social norms — harmful norms (want DECREASES)
    sn = m['socialnorms_agree']
    harmful = sn[~sn['Norm'].str.contains('Planting Crops|Restoring Ecosystems|Express Emotions', regex=True)]
    sn_avg_bl = harmful['Baseline'].mean()
    sn_avg_ml = harmful['Midline'].mean()
    sn_change = _pp(sn_avg_bl, sn_avg_ml)
    trend = "positive" if sn_change < 0 else ("negative" if sn_change > 2 else "neutral")
    insights.append(("Men's Harmful Social Norms",
                     f"Average agreement with harmful gender norms among men: "
                     f"{sn_avg_bl*100:.1f}% (BL) to {sn_avg_ml*100:.1f}% (ML) ({sn_change:+.1f} pp). "
                     f"{'Positive shift — harmful norms are weakening among men.' if sn_change < 0 else 'Norms among men remain entrenched; continued dialogue is critical.'}",
                     trend))

    # 6. Support for women in NbS — Mangrove + Seaweed + Forest (Yes)
    support_sum = 0
    support_count = 0
    for key in ['mangrove_support', 'seaweed_support', 'forest_support']:
        df = m[key]
        yes_bl = df.loc[df['Response'] == 'Yes', 'Baseline'].values[0] if len(df[df['Response'] == 'Yes']) else 0
        yes_ml = df.loc[df['Response'] == 'Yes', 'Midline'].values[0] if len(df[df['Response'] == 'Yes']) else 0
        support_sum += _pp(yes_bl, yes_ml)
        support_count += 1
    avg_support_change = support_sum / max(support_count, 1)
    insights.append(("Men's Support for Women in NbS",
                     f"Average change in men saying 'Yes' to supporting women's NbS participation "
                     f"(mangrove, seaweed, forest): {avg_support_change:+.1f} pp. "
                     + ("Men are increasingly supporting women's conservation roles." if avg_support_change > 0
                        else "Men's active support for women's NbS participation needs strengthening."),
                     "positive" if avg_support_change > 0 else "warning"))

    # 7. Time use — men's unpaid care work
    ts = m['time_summary']
    unpaid_bl = ts.loc[ts['Category'] == 'Unpaid Care Work', 'Baseline'].values[0]
    unpaid_ml = ts.loc[ts['Category'] == 'Unpaid Care Work', 'Midline'].values[0]
    unpaid_diff = round(unpaid_ml - unpaid_bl, 2) if isinstance(unpaid_bl, (int, float)) else 0
    care_msg = ("Men are taking on more care responsibilities \u2014 positive for gender equity."
                if unpaid_diff > 0
                else "Men's care contribution remains low; household labour-sharing needs attention.")
    insights.append(("Men's Unpaid Care Work",
                     f"Average daily unpaid care work by men: {unpaid_bl:.1f}h (BL) to {unpaid_ml:.1f}h (ML) "
                     f"({unpaid_diff:+.1f}h change). {care_msg}",
                     "positive" if unpaid_diff > 0 else "warning"))

    return insights


def _generate_gjj_insights(g):
    """Generate automated insights from GJJ KAP Women (Baseline / Endline) data."""
    insights = []
    try:
        _gen_gjj_insights_inner(g, insights)
    except Exception as e:
        insights.append(("GJJ Insight Generation Note",
                         f"Some GJJ KAP Women insights could not be generated: {e}",
                         "neutral"))
    return insights


def _gen_gjj_insights_inner(g, insights):
    # 1. Self-esteem — Strongly Agree shift (average across 3 statements)
    sa = g['self_strongly_agree']
    sa_bl_avg = sa['Baseline'].mean()
    sa_el_avg = sa['Endline'].mean()
    sa_change = (sa_el_avg - sa_bl_avg) * 100
    insights.append(("GJJ: Self-Esteem (Strongly Agree)",
                     f"Average 'Strongly Agree' with self-esteem statements: "
                     f"{sa_bl_avg*100:.1f}% (BL) to {sa_el_avg*100:.1f}% (EL) ({sa_change:+.1f} pp). "
                     + ("Women's self-belief is strengthening." if sa_change > 0
                        else "Self-esteem reinforcement programmes need attention."),
                     "positive" if sa_change > 0 else "warning"))

    # 2. Self-compassion — frequency shift
    comp = g['self_compassion']
    # Use 'Always' or first category as key indicator
    comp_bl = comp['Baseline'].iloc[0] if len(comp) else 0
    comp_el = comp['Endline'].iloc[0] if len(comp) else 0
    comp_change = (float(comp_el) - float(comp_bl)) * 100
    comp_label = comp['Category'].iloc[0] if len(comp) else 'Self-care'
    insights.append(("GJJ: Self-Compassion",
                     f"'{comp_label}' self-compassion frequency: "
                     f"{float(comp_bl)*100:.1f}% (BL) to {float(comp_el)*100:.1f}% (EL) ({comp_change:+.1f} pp). "
                     + ("More women are practising kind self-talk." if comp_change > 0
                        else "Self-compassion practices need continued reinforcement."),
                     "positive" if comp_change > 0 else "neutral"))

    # 3. Relational wellbeing — Always/Frequently average
    af_rn = g['rel_af_rn']
    af_bl_avg = af_rn['AF_Baseline'].mean()
    af_el_avg = af_rn['AF_Endline'].mean()
    af_change = (float(af_el_avg) - float(af_bl_avg)) * 100
    insights.append(("GJJ: Relational Wellbeing",
                     f"Average 'Always/Frequently' for relational wellbeing: "
                     f"{float(af_bl_avg)*100:.1f}% (BL) to {float(af_el_avg)*100:.1f}% (EL) ({af_change:+.1f} pp). "
                     + ("Partner support and communication are improving." if af_change > 0
                        else "Relational dynamics need more GJJ programme focus."),
                     "positive" if af_change > 0 else "warning"))

    # 4. Shared responsibility — husband supports chores (Yes)
    chores = g['shared_chores_yn']
    yes_rows = chores[chores['Response'].str.strip().str.lower() == 'yes']
    if len(yes_rows):
        ch_bl = float(yes_rows['Baseline'].values[0])
        ch_el = float(yes_rows['Endline'].values[0])
        ch_change = (ch_el - ch_bl) * 100
        insights.append(("GJJ: Husband Supports Chores",
                         f"Women reporting husband supports chores: "
                         f"{ch_bl*100:.1f}% (BL) to {ch_el*100:.1f}% (EL) ({ch_change:+.1f} pp). "
                         + ("More husbands are sharing household responsibilities." if ch_change > 0
                            else "Shared housework remains a challenge area."),
                         "positive" if ch_change > 0 else "warning"))

    # 5. Decision-making — conversations to change decisions (Yes)
    dec_conv = g['decision_conversations']
    dec_yes = dec_conv[dec_conv['Response'].str.strip().str.lower() == 'yes']
    if len(dec_yes):
        dc_bl = float(dec_yes['Baseline'].values[0])
        dc_el = float(dec_yes['Endline'].values[0])
        dc_change = (dc_el - dc_bl) * 100
        insights.append(("GJJ: Decision Conversations",
                         f"Women who had conversations to change decision-making: "
                         f"{dc_bl*100:.1f}% (BL) to {dc_el*100:.1f}% (EL) ({dc_change:+.1f} pp). "
                         + ("Women are exercising more agency in decision-making." if dc_change > 0
                            else "Empowering women to negotiate household decisions needs strengthening."),
                         "positive" if dc_change > 0 else "warning"))

    # 6. Equal say in joint decisions (Yes)
    eq = g['equal_say']
    eq_yes = eq[eq['Response'].str.strip().str.lower() == 'yes']
    if len(eq_yes):
        eq_bl = float(eq_yes['Baseline'].values[0])
        eq_el = float(eq_yes['Endline'].values[0])
        eq_change = (eq_el - eq_bl) * 100
        insights.append(("GJJ: Equal Say in Decisions",
                         f"Women reporting equal say: "
                         f"{eq_bl*100:.1f}% (BL) to {eq_el*100:.1f}% (EL) ({eq_change:+.1f} pp). "
                         + ("Joint decision-making is becoming more equitable." if eq_change > 0
                            else "Barriers to equal decision-making participation persist."),
                         "positive" if eq_change > 0 else "warning"))

    # 7. Autonomy — husband supports leadership
    sup_ldr = g['support_leader']
    # Use 'Very supportive' or first category
    if len(sup_ldr):
        sl_bl = float(sup_ldr['Baseline'].iloc[0])
        sl_el = float(sup_ldr['Endline'].iloc[0])
        sl_change = (sl_el - sl_bl) * 100
        sl_label = sup_ldr['Category'].iloc[0] if 'Category' in sup_ldr.columns else 'Very supportive'
        insights.append(("GJJ: Husband Supports Leadership",
                         f"'{sl_label}' for women becoming leaders: "
                         f"{sl_bl*100:.1f}% (BL) to {sl_el*100:.1f}% (EL) ({sl_change:+.1f} pp). "
                         + ("Husbands are increasingly supportive of women's leadership." if sl_change > 0
                            else "Male allyship for women's leadership needs more focus."),
                         "positive" if sl_change > 0 else "warning"))

    return insights


# ============================================================================
# GJJ KAP MEN — INSIGHTS GENERATOR
# ============================================================================

def _generate_gjj_men_insights(g):
    """Generate automated insights from GJJ KAP Men (Baseline / Endline) data."""
    insights = []
    try:
        _gen_gjj_men_insights_inner(g, insights)
    except Exception as e:
        insights.append(("GJJ Men Insight Generation Note",
                         f"Some GJJ KAP Men insights could not be generated: {e}",
                         "neutral"))
    return insights


def _gen_gjj_men_insights_inner(g, insights):
    # 1. Self-esteem — Agreement shift (average across 3 statements)
    agr = g['self_agreement']
    agr_bl_avg = agr['Agreement_BL'].mean()
    agr_el_avg = agr['Agreement_EL'].mean()
    agr_change = (float(agr_el_avg) - float(agr_bl_avg)) * 100
    insights.append(("GJJ Men: Self-Esteem (Agreement)",
                     f"Average agreement with self-esteem statements: "
                     f"{float(agr_bl_avg)*100:.1f}% (BL) to {float(agr_el_avg)*100:.1f}% (EL) ({agr_change:+.1f} pp). "
                     + ("Men's self-responsibility and gender equality beliefs are growing." if agr_change > 0
                        else "Self-esteem reinforcement for men needs attention."),
                     "positive" if agr_change > 0 else "warning"))

    # 2. Self-compassion — frequency shift (first category = Frequently)
    comp = g['self_compassion']
    comp_bl = float(comp['Baseline'].iloc[0]) if len(comp) else 0
    comp_el = float(comp['Endline'].iloc[0]) if len(comp) else 0
    comp_change = (comp_el - comp_bl) * 100
    comp_label = comp['Category'].iloc[0] if len(comp) else 'Self-care'
    insights.append(("GJJ Men: Self-Compassion",
                     f"'{comp_label}' self-compassion frequency: "
                     f"{comp_bl*100:.1f}% (BL) to {comp_el*100:.1f}% (EL) ({comp_change:+.1f} pp). "
                     + ("More men are practising kind self-talk." if comp_change > 0
                        else "Self-compassion practices for men need continued reinforcement."),
                     "positive" if comp_change > 0 else "neutral"))

    # 3. Relational wellbeing — Always/Frequently average
    af_rn = g['rel_af_rn']
    af_bl_avg = float(af_rn['AF_Baseline'].mean())
    af_el_avg = float(af_rn['AF_Endline'].mean())
    af_change = (af_el_avg - af_bl_avg) * 100
    insights.append(("GJJ Men: Relational Wellbeing",
                     f"Average 'Always/Frequently' for relational wellbeing: "
                     f"{af_bl_avg*100:.1f}% (BL) to {af_el_avg*100:.1f}% (EL) ({af_change:+.1f} pp). "
                     + ("Partner support and communication are improving among men." if af_change > 0
                        else "Relational dynamics for men need more GJJ focus."),
                     "positive" if af_change > 0 else "warning"))

    # 4. Shared responsibility — supports chores (Yes)
    chores = g['shared_chores_yn']
    yes_rows = chores[chores['Response'].str.strip().str.lower() == 'yes']
    if len(yes_rows):
        ch_bl = float(yes_rows['Baseline'].values[0])
        ch_el = float(yes_rows['Endline'].values[0])
        ch_change = (ch_el - ch_bl) * 100
        insights.append(("GJJ Men: Supports Household Chores",
                         f"Men reporting they support household chores: "
                         f"{ch_bl*100:.1f}% (BL) to {ch_el*100:.1f}% (EL) ({ch_change:+.1f} pp). "
                         + ("More men are actively sharing household responsibilities." if ch_change > 0
                            else "Men's participation in household chores needs strengthening."),
                         "positive" if ch_change > 0 else "warning"))

    # 5. Decision-making — conversations to change decisions (Yes)
    dec_conv = g['decision_conversations']
    dec_yes = dec_conv[dec_conv['Response'].str.strip().str.lower() == 'yes']
    if len(dec_yes):
        dc_bl = float(dec_yes['Baseline'].values[0])
        dc_el = float(dec_yes['Endline'].values[0])
        dc_change = (dc_el - dc_bl) * 100
        insights.append(("GJJ Men: Decision Conversations",
                         f"Men who had conversations to change decision-making: "
                         f"{dc_bl*100:.1f}% (BL) to {dc_el*100:.1f}% (EL) ({dc_change:+.1f} pp). "
                         + ("More men are engaging in shared decision-making discussions." if dc_change > 0
                            else "Encouraging men's participation in joint decision-making needs attention."),
                         "positive" if dc_change > 0 else "warning"))

    # 6. Leadership — 'Always' support wife becoming leader
    sup_ldr = g['support_leader']
    alw_rows = sup_ldr[sup_ldr['Category'].str.strip().str.lower() == 'always']
    if len(alw_rows):
        sl_bl = float(alw_rows['Baseline'].values[0])
        sl_el = float(alw_rows['Endline'].values[0])
        sl_change = (sl_el - sl_bl) * 100
        insights.append(("GJJ Men: Supports Wife Leadership",
                         f"Men 'Always' supporting wife becoming leader: "
                         f"{sl_bl*100:.1f}% (BL) to {sl_el*100:.1f}% (EL) ({sl_change:+.1f} pp). "
                         + ("Male allyship for women's leadership is strengthening." if sl_change > 0
                            else "More men need to be encouraged to support women's leadership."),
                         "positive" if sl_change > 0 else "warning"))

    # 7. Business — 'Definitely' support wife starting business
    sup_biz = g['support_business']
    def_rows = sup_biz[sup_biz['Category'].str.strip().str.lower() == 'definitely']
    if len(def_rows):
        sb_bl = float(def_rows['Baseline'].values[0])
        sb_el = float(def_rows['Endline'].values[0])
        sb_change = (sb_el - sb_bl) * 100
        insights.append(("GJJ Men: Supports Wife Business",
                         f"Men 'Definitely' supporting wife's business: "
                         f"{sb_bl*100:.1f}% (BL) to {sb_el*100:.1f}% (EL) ({sb_change:+.1f} pp). "
                         + ("Men are increasingly supportive of women's entrepreneurship." if sb_change > 0
                            else "Encouraging men to support women's business ventures needs focus."),
                         "positive" if sb_change > 0 else "warning"))

    return insights


def _generate_forest_training_insights(t_data):
    """Generate automated insights from Forest Training Pre/Post data."""
    insights = []
    if t_data is None:
        return insights
    try:
        thresholds = t_data['thresholds']
        scores = t_data['scores']
        questions = t_data['questions']

        # 1. PMF threshold (≥70%) improvement
        t70 = thresholds[thresholds['Threshold'].str.contains('70')]
        if len(t70) >= 2:
            bl_70 = float(t70[t70['Timepoint'] == 'Baseline']['Proportion'].values[0]) * 100
            el_70 = float(t70[t70['Timepoint'] == 'Endline']['Proportion'].values[0]) * 100
            change_70 = el_70 - bl_70
            insights.append(("Forest Training: PMF Pass Rate (≥70%)",
                             f"Participants scoring ≥70%: {bl_70:.1f}% (Pre) → {el_70:.1f}% (Post), "
                             f"a {change_70:+.1f} pp improvement. "
                             + ("Training dramatically increased knowledge adequacy."
                                if change_70 > 20 else "Steady improvement in knowledge adequacy."),
                             "positive" if change_70 > 0 else "warning"))

        # 2. Average score change
        if len(scores) >= 2:
            bl_avg = float(scores[scores['Timepoint'] == 'Baseline']['AverageScore'].values[0])
            el_avg = float(scores[scores['Timepoint'] == 'Endline']['AverageScore'].values[0])
            avg_change = el_avg - bl_avg
            insights.append(("Forest Training: Average Score",
                             f"Average score rose from {bl_avg:.1f}% to {el_avg:.1f}% "
                             f"({avg_change:+.1f} pp). "
                             + ("Strong knowledge gains across participants."
                                if avg_change > 15 else "Moderate improvement in overall scores."),
                             "positive" if avg_change > 0 else "warning"))

        # 3. Minimum score improvement
        if len(scores) >= 2:
            bl_min = float(scores[scores['Timepoint'] == 'Baseline']['MinScore'].values[0])
            el_min = float(scores[scores['Timepoint'] == 'Endline']['MinScore'].values[0])
            insights.append(("Forest Training: Weakest Performers",
                             f"Minimum score improved from {bl_min:.1f}% to {el_min:.1f}%. "
                             + ("Even the weakest performers gained significantly."
                                if el_min > bl_min + 10 else "Some participants still need additional support."),
                             "positive" if el_min > bl_min else "warning"))

        # 4. Best-improving question
        if not questions.empty:
            questions_c = questions.copy()
            questions_c['Change'] = (questions_c['Endline'] - questions_c['Baseline']) * 100
            best_q = questions_c.loc[questions_c['Change'].idxmax()]
            insights.append(("Forest Training: Biggest Knowledge Gain",
                             f"Q{int(best_q['QuestionNumber'])} ({best_q['QuestionText'][:80]}…) "
                             f"improved by {best_q['Change']:+.1f} pp. "
                             "This topic resonated most with participants.",
                             "positive"))

            # 5. Weakest endline question
            weakest_q = questions_c.loc[questions_c['Endline'].idxmin()]
            insights.append(("Forest Training: Area Needing Attention",
                             f"Q{int(weakest_q['QuestionNumber'])} ({weakest_q['QuestionText'][:80]}…) "
                             f"had the lowest post-training score ({weakest_q['Endline']*100:.1f}%). "
                             "Consider reinforcing this topic in future training.",
                             "warning" if weakest_q['Endline'] < 0.85 else "positive"))

        # 6. Domain-level best improvement
        if not questions.empty:
            best_domain, best_gain = None, -999
            for domain, q_nums in FOREST_TRAINING_DOMAINS.items():
                dq = questions[questions['QuestionNumber'].isin(q_nums)]
                if dq.empty:
                    continue
                gain = (dq['Endline'].mean() - dq['Baseline'].mean()) * 100
                if gain > best_gain:
                    best_gain = gain
                    best_domain = domain
            if best_domain:
                insights.append(("Forest Training: Strongest Domain",
                                 f"The '{best_domain}' domain showed the highest avg improvement "
                                 f"of {best_gain:+.1f} pp across its questions.",
                                 "positive"))

    except Exception as e:
        insights.append(("Insight Generation Note",
                         f"Some Forest Training insights could not be generated: {e}",
                         "neutral"))
    return insights


# ---- Mangrove Training insights ----
def _generate_mangrove_training_insights(mg_data):
    """Generate automated insights for Mangrove Training Pre/Post data."""
    insights = []
    try:
        adequate_county = mg_data['adequate_county']
        adequate_sex = mg_data['adequate_sex']
        scores = mg_data['scores']

        # 1. Overall pass-rate improvement
        all_county = adequate_county[adequate_county['County'] == 'All']
        pre_val = float(all_county[all_county['Timepoint'] == 'Pre-Test']['Value'].values[0]) * 100
        post_val = float(all_county[all_county['Timepoint'] == 'Post-Test']['Value'].values[0]) * 100
        change = post_val - pre_val
        sentiment = 'positive' if change > 20 else ('positive' if change > 0 else 'negative')
        insights.append((
            "Overall Knowledge Improvement",
            f"Adequate mangrove restoration knowledge (≥60%) increased from {pre_val:.1f}% "
            f"to {post_val:.1f}% ({change:+.1f}pp). {'Strong' if change > 30 else 'Meaningful'} "
            f"effect observed from training intervention.",
            sentiment))

        # 2. County comparison
        counties = [c for c in adequate_county['County'].unique() if c != 'All']
        county_changes = {}
        for c in counties:
            c_data = adequate_county[adequate_county['County'] == c]
            c_pre = float(c_data[c_data['Timepoint'] == 'Pre-Test']['Value'].values[0]) * 100
            c_post = float(c_data[c_data['Timepoint'] == 'Post-Test']['Value'].values[0]) * 100
            county_changes[c] = c_post - c_pre
        if county_changes:
            best = max(county_changes, key=county_changes.get)
            worst = min(county_changes, key=county_changes.get)
            insights.append((
                "County Performance Gap",
                f"{best} showed the largest improvement ({county_changes[best]:+.1f}pp), "
                f"while {worst} improved by {county_changes[worst]:+.1f}pp. "
                f"{'Consider targeted follow-up for ' + worst + '.' if county_changes[worst] < county_changes[best] - 10 else 'Both counties responded well to training.'}",
                'positive' if county_changes[worst] > 15 else 'neutral'))

        # 3. Gender analysis
        sex_all = adequate_sex[adequate_sex['Sex'] != 'All']
        male_data = sex_all[sex_all['Sex'] == 'Male']
        female_data = sex_all[sex_all['Sex'] == 'Female']
        if len(male_data) and len(female_data):
            m_post = float(male_data[male_data['Timepoint'] == 'Post-Test']['Value'].values[0]) * 100
            f_post = float(female_data[female_data['Timepoint'] == 'Post-Test']['Value'].values[0]) * 100
            m_pre = float(male_data[male_data['Timepoint'] == 'Pre-Test']['Value'].values[0]) * 100
            f_pre = float(female_data[female_data['Timepoint'] == 'Pre-Test']['Value'].values[0]) * 100
            gender_gap = f_post - m_post
            insights.append((
                "Gender Equity in Knowledge",
                f"Post-training: Female {f_post:.1f}% vs Male {m_post:.1f}% "
                f"(gap: {gender_gap:+.1f}pp). Female improvement: {f_post-f_pre:+.1f}pp, "
                f"Male improvement: {m_post-m_pre:+.1f}pp.",
                'positive' if abs(gender_gap) < 10 else 'neutral'))

        # 4. Score distribution
        all_scores = scores[scores['County'] == 'All']
        if len(all_scores):
            avg = float(all_scores['AvgScore'].values[0])
            max_s = float(all_scores['MaxScore'].values[0])
            min_s = float(all_scores['MinScore'].values[0])
            spread = max_s - min_s
            insights.append((
                "Score Distribution",
                f"Average score: {avg:.1f}%, range {min_s:.1f}%–{max_s:.1f}% "
                f"(spread: {spread:.1f}pp). "
                f"{'Wide variation suggests need for differentiated support.' if spread > 60 else 'Moderate variation in participant performance.'}",
                'neutral'))

        # 5. Respondent coverage
        total_n = int(all_scores['Respondents'].values[0]) if len(all_scores) else 0
        insights.append((
            "Assessment Coverage",
            f"A total of {total_n:,} participants completed the mangrove "
            f"restoration knowledge assessment across all counties.",
            'neutral'))

    except Exception as e:
        insights.append(("Insight Generation Note",
                         f"Some Mangrove Training insights could not be generated: {e}",
                         "neutral"))
    return insights


# ---- Seaweed Production insights ----
def _generate_seaweed_insights(sw_df):
    """Generate automated insights for the Seaweed Production & Challenges dataset."""
    insights = []
    try:
        agg = prepare_seaweed_aggregates(sw_df)
        ov = agg['overall']
        grp_df = agg['group_summary']
        ch_df = agg['challenge_counts']

        # 1. Production headline
        insights.append((
            "Total Seaweed Production",
            f"Across {ov['n_groups']} groups and {ov['n_farmers']:,} farmers, "
            f"total production is {ov['total_kg']:,.1f} kg "
            f"({ov['dried_kg']:,.0f} kg dried, {ov['wet_kg']:,.0f} kg wet). "
            f"Average production per rope is {ov['avg_prod_per_rope']:.2f} kg "
            f"and average per farmer is {ov['avg_production_per_farmer']:.1f} kg.",
            'neutral'))

        # 2. Top-producing group
        if len(grp_df):
            top_grp = grp_df.loc[grp_df['Total_KG'].idxmax()]
            insights.append((
                "Highest-Producing Group",
                f"{top_grp['Group']} leads production with {top_grp['Total_KG']:,.0f} kg "
                f"from {int(top_grp['Members'])} members ({top_grp['Ropes_Total']:,.0f} ropes). "
                f"Average achievement: {top_grp['Avg_Achievement_pct']:.1f}%.",
                'positive'))

        # 3. Target achievement
        avg_ach = sw_df['Ropes_Achievement_pct'].mean()
        pct_meeting = len(sw_df[sw_df['Ropes_Total'] >= sw_df['Target_Ropes']]) / max(len(sw_df), 1) * 100
        sentiment = 'positive' if pct_meeting > 60 else ('neutral' if pct_meeting > 30 else 'warning')
        insights.append((
            "Target Achievement",
            f"Average rope target achievement is {avg_ach:.1f}%. "
            f"{pct_meeting:.1f}% of farmers meet or exceed their target ropes. "
            f"{'Strong progress toward targets.' if pct_meeting > 60 else 'Many farmers still below rope targets — consider support interventions.'}",
            sentiment))

        # 4. Top challenge
        if not ch_df.empty:
            top_ch = ch_df.loc[ch_df['Pct'].idxmax()]
            insights.append((
                "Primary Challenge",
                f"The most common challenge is '{top_ch['Challenge']}', affecting "
                f"{top_ch['Pct']:.1f}% of farmers ({int(top_ch['Count'])} members). "
                f"Programme support should prioritise this barrier.",
                'warning'))

        # 5. Casual workers
        insights.append((
            "Casual Labour Usage",
            f"{ov['casual_pct']:.1f}% of seaweed farmers report using casual workers. "
            f"{'This is a significant proportion — workforce development may be needed.' if ov['casual_pct'] > 30 else 'Relatively low reliance on casual labour.'}",
            'neutral'))

        # 6. Group productivity variation
        if len(grp_df) >= 2:
            best_eff = grp_df.loc[grp_df['Avg_Prod_per_Rope'].idxmax()]
            worst_eff = grp_df.loc[grp_df['Avg_Prod_per_Rope'].idxmin()]
            insights.append((
                "Productivity Variation",
                f"Production efficiency varies from {worst_eff['Avg_Prod_per_Rope']:.2f} kg/rope "
                f"({worst_eff['Group']}) to {best_eff['Avg_Prod_per_Rope']:.2f} kg/rope "
                f"({best_eff['Group']}). Knowledge sharing between groups could reduce this gap.",
                'neutral'))

        # 7. Rope Gap analysis
        insights.append((
            "Rope Gap Analysis",
            f"Total rope gap across all groups: {ov['gap_total']:,.0f} ropes. "
            f"Average ropes per farmer: {ov['avg_ropes_per_farmer']:.1f}. "
            f"{'Significant rope shortfall — scaling up rope acquisition is critical.' if ov['gap_total'] > 500 else 'Moderate gap — targeted rope distribution can close it.'}",
            'warning' if ov['gap_total'] > 500 else 'neutral'))

        # 8. Dried vs Wet ratio interpretation
        dw = ov['dried_wet_ratio']
        dw_msg = ("Production is balanced between dried and wet seaweed."
                  if 0.8 <= dw <= 1.2
                  else ("Dried seaweed dominates — good market readiness."
                        if dw > 1.2
                        else "Wet seaweed dominates — post-harvest drying capacity may need strengthening."))
        insights.append((
            "Dried vs Wet Balance",
            f"Overall dried-to-wet ratio is {dw:.2f}. {dw_msg} "
            f"Total dried: {ov['dried_kg']:,.0f} kg, wet: {ov['wet_kg']:,.0f} kg.",
            'positive' if dw > 1.0 else 'neutral'))

        # 9. Multi-challenge burden
        mc_pct = ov['multi_challenge_pct']
        insights.append((
            "Multi-Challenge Burden",
            f"{mc_pct:.1f}% of farmers face 2 or more simultaneous challenges. "
            f"{'This high burden requires integrated, multi-sector support interventions.' if mc_pct > 30 else 'Most farmers face single or no challenges — focused interventions can help.'}",
            'warning' if mc_pct > 30 else 'neutral'))

        # 10. Pareto observation
        ranked = sw_df.sort_values('Total_KG', ascending=False)
        cumul = ranked['Total_KG'].cumsum()
        total_prod = ranked['Total_KG'].sum()
        if total_prod > 0:
            n_80 = int((cumul / total_prod <= 0.8).sum()) + 1
            pct_80 = round(n_80 / max(len(ranked), 1) * 100, 1)
            insights.append((
                "Pareto Concentration",
                f"The top {n_80} farmers ({pct_80}%) contribute 80% of total production. "
                f"{'Output is highly concentrated — consider equity-focused scaling.' if pct_80 < 30 else 'Production is reasonably distributed across farmers.'}",
                'neutral' if pct_80 >= 30 else 'warning'))

    except Exception as e:
        insights.append(("Insight Generation Note",
                         f"Some Seaweed insights could not be generated: {e}",
                         "neutral"))
    return insights


def _generate_cross_cutting_insights(f_data, w_data, m_data=None, gjj_data=None):
    """Generate insights that span all datasets."""
    insights = []
    try:
        _gen_cross_cutting_inner(f_data, w_data, m_data, insights, gjj_data)
    except Exception as e:
        insights.append(("Insight Generation Note",
                         f"Some cross-cutting insights could not be generated: {e}",
                         "neutral"))
    return insights


def _gen_cross_cutting_inner(f_data, w_data, m_data, insights, gjj_data=None):
    # 1. Governance + Empowerment linkage
    fd = f_data['functionality_domain']
    gender_bl = fd.loc[fd['Timepoint'] == 'Baseline', 'Gender'].values[0]
    gender_ml = fd.loc[fd['Timepoint'] == 'Midline', 'Gender'].values[0]
    gender_change = round(gender_ml - gender_bl, 1)

    di = w_data['decision_influence']
    avg_infl_bl = di['Baseline'].mean() * 100
    avg_infl_ml = di['Midline'].mean() * 100
    infl_change = round(avg_infl_ml - avg_infl_bl, 1)

    insights.append(("Governance & Empowerment Linkage",
                     f"Forestry gender domain score changed by {gender_change:+.1f} pts while women's "
                     f"decision-making influence shifted by {infl_change:+.1f} pp. "
                     + ("Both are moving in the right direction — community and household empowerment reinforce each other."
                        if gender_change > 0 and infl_change > 0
                        else "Progress at one level has not fully translated to the other — integrated programming may help."),
                     "positive" if gender_change > 0 and infl_change > 0 else "neutral"))

    # 2. Environment + NbS
    nbs_yes_ml = w_data['nbs_heard'].loc[w_data['nbs_heard']['Response'] == 'Yes', 'Midline'].values[0] * 100
    fc = f_data['forest_condition']
    avg_good_ml = fc['Midline_Good'].mean()
    avg_good_ml_pct = avg_good_ml * 100 if avg_good_ml <= 1 else avg_good_ml
    insights.append(("Environmental Knowledge & Forest Health",
                     f"At midline, {nbs_yes_ml:.1f}% of women know about NbS, while {avg_good_ml_pct:.1f}% "
                     f"of forests are rated 'Good' condition. Growing environmental literacy can translate "
                     f"into better community-based conservation outcomes.",
                     "neutral"))

    # 3. Training + Life Skills
    tc = f_data['training_coverage']
    most_ml = tc.loc[tc['Category'] == 'Most Members', 'Midline'].values[0]
    most_ml_pct = most_ml * 100 if most_ml <= 1 else most_ml

    ls = w_data['lifeskills_agree']
    ls_avg_ml = ls['Midline'].mean() * 100
    insights.append(("Training Reach & Life Skills",
                     f"Groups where most members are trained: {most_ml_pct:.1f}%. "
                     f"Average women's life skills confidence: {ls_avg_ml:.1f}%. "
                     f"Investing in both community-level training and individual skills development "
                     f"creates compounding empowerment effects.",
                     "positive" if most_ml_pct > 30 and ls_avg_ml > 50 else "neutral"))

    # 4. Economic resilience
    ig = f_data['income_gen']
    ig_yes_ml = ig.loc[ig['Category'] == 'Yes', 'Midline'].values[0]
    ig_yes_ml_pct = ig_yes_ml * 100 if ig_yes_ml <= 1 else ig_yes_ml

    ps = w_data['personal_saving']
    ps_yes_ml = ps.loc[ps['Response'] == 'Yes', 'Midline'].values[0] * 100
    insights.append(("Economic Resilience Snapshot",
                     f"At midline, {ig_yes_ml_pct:.1f}% of forestry groups generate income and "
                     f"{ps_yes_ml:.1f}% of women have personal savings. Linking group-level income "
                     f"activities with individual savings programs can strengthen household resilience.",
                     "positive" if ig_yes_ml_pct > 50 and ps_yes_ml > 40 else "neutral"))

    # 5. Men-Women alignment on social norms (only if men data is available)
    if m_data is not None:
        w_sn = w_data['socialnorms_agree']
        w_harmful = w_sn[~w_sn['Norm'].str.contains('Planting Crops|Restoring Ecosystems|Express Emotions', regex=True)]
        w_sn_change = _pp(w_harmful['Baseline'].mean(), w_harmful['Midline'].mean())
        m_sn = m_data['socialnorms_agree']
        m_harmful = m_sn[~m_sn['Norm'].str.contains('Planting Crops|Restoring Ecosystems|Express Emotions', regex=True)]
        m_sn_change = _pp(m_harmful['Baseline'].mean(), m_harmful['Midline'].mean())
        both_improving = w_sn_change < 0 and m_sn_change < 0
        insights.append(("Gender Norms: Men vs Women Alignment",
                         f"Harmful norms agreement changed by {w_sn_change:+.1f} pp among women "
                         f"and {m_sn_change:+.1f} pp among men. "
                         + ("Both groups show declining support for harmful norms — a strong signal of programme impact."
                            if both_improving
                            else "Norms are shifting unevenly between men and women — targeted messaging for each group may help."),
                         "positive" if both_improving else "neutral"))

        # 6. Men's support for women + women's actual participation
        m_support_pcts = []
        for key in ['mangrove_support', 'seaweed_support', 'forest_support']:
            df = m_data[key]
            yes_ml = df.loc[df['Response'] == 'Yes', 'Midline'].values[0] if len(df[df['Response'] == 'Yes']) else 0
            m_support_pcts.append(yes_ml * 100)
        avg_support = sum(m_support_pcts) / len(m_support_pcts)
        insights.append(("Men's Support & Women's Conservation Participation",
                         f"On average, {avg_support:.1f}% of men say they support women's NbS participation "
                         f"at midline. Aligning men's declared support with actual resource allocation "
                         f"and time-sharing remains a key programming challenge.",
                         "positive" if avg_support > 50 else "neutral"))

        # 7. Decision-making alignment: who should vs who does
        m_dj_should_avg = m_data['decision_should_joint']['Midline'].mean() * 100
        m_dj_does_avg = m_data['decision_does_joint']['Midline'].mean() * 100
        gap = round(m_dj_should_avg - m_dj_does_avg, 1)
        insights.append(("Decision-Making: Attitudes vs Practice (Men)",
                         f"At midline, {m_dj_should_avg:.1f}% of men say decisions should be joint, "
                         f"but only {m_dj_does_avg:.1f}% report decisions actually being joint "
                         f"(gap: {gap:+.1f} pp). "
                         + ("A gap persists between progressive attitudes and actual behaviour." if gap > 5
                            else "Attitudes and practice are closely aligned — encouraging."),
                         "warning" if gap > 5 else "positive"))

    # 8. GJJ KAP Women — Self-esteem + Decision-making linkage
    if gjj_data is not None:
        try:
            sa = gjj_data['self_strongly_agree']
            sa_el_avg = float(sa['Endline'].mean()) * 100
            eq = gjj_data['equal_say']
            eq_yes = eq[eq['Response'].str.strip().str.lower() == 'yes']
            eq_el = float(eq_yes['Endline'].values[0]) * 100 if len(eq_yes) else 0
            insights.append(("GJJ Self-Esteem & Decision-Making",
                             f"Women's 'Strongly Agree' with self-esteem: {sa_el_avg:.1f}% (EL). "
                             f"Equal say in decisions: {eq_el:.1f}% (EL). "
                             "GJJ programme's self-efficacy work is translating into greater agency in household decisions.",
                             "positive" if sa_el_avg > 50 and eq_el > 50 else "neutral"))

            # Shared responsibility + Women's time use
            chores = gjj_data['shared_chores_yn']
            ch_yes = chores[chores['Response'].str.strip().str.lower() == 'yes']
            ch_el = float(ch_yes['Endline'].values[0]) * 100 if len(ch_yes) else 0
            insights.append(("GJJ Shared Responsibility & Empowerment",
                             f"At endline, {ch_el:.1f}% of women report husbands supporting chores. "
                             "When combined with broader programme indicators, shared responsibility "
                             "contributes to women's time freedom for economic and conservation activities.",
                             "positive" if ch_el > 60 else "neutral"))
        except Exception:
            pass

    return insights


def _build_indicator_table(f_data, w_data, m_data=None, gjj_data=None, gjj_men_data=None, ft_data=None, mg_data=None, sw_data=None):
    """Build a master table of key indicators with BL/ML values for trend charts."""
    rows = []

    # Forestry indicators
    ft = f_data['functionality_threshold']
    bl_60 = ft.loc[ft['Timepoint'] == 'Baseline', 'Functional_60_pct'].values[0]
    ml_60 = ft.loc[ft['Timepoint'] == 'Midline', 'Functional_60_pct'].values[0]
    rows.append({'Indicator': 'Functionality >=60%', 'Dataset': 'Forestry',
                 'Baseline': round(bl_60 * 100 if bl_60 <= 1 else bl_60, 1),
                 'Midline': round(ml_60 * 100 if ml_60 <= 1 else ml_60, 1)})

    for domain in ['Management', 'Gender', 'Effectiveness', 'Overall']:
        fd = f_data['functionality_domain']
        b = fd.loc[fd['Timepoint'] == 'Baseline', domain].values[0]
        m = fd.loc[fd['Timepoint'] == 'Midline', domain].values[0]
        rows.append({'Indicator': f'{domain} Score', 'Dataset': 'Forestry',
                     'Baseline': round(b, 1), 'Midline': round(m, 1)})

    ig = f_data['income_gen']
    b_ig = ig.loc[ig['Category'] == 'Yes', 'Baseline'].values[0]
    m_ig = ig.loc[ig['Category'] == 'Yes', 'Midline'].values[0]
    rows.append({'Indicator': 'Income Generation', 'Dataset': 'Forestry',
                 'Baseline': round(b_ig * 100 if b_ig <= 1 else b_ig, 1),
                 'Midline': round(m_ig * 100 if m_ig <= 1 else m_ig, 1)})

    af = f_data['agroforestry']
    b_af = af.loc[af['Category'] == 'Yes', 'Baseline'].values[0]
    m_af = af.loc[af['Category'] == 'Yes', 'Midline'].values[0]
    rows.append({'Indicator': 'Agroforestry', 'Dataset': 'Forestry',
                 'Baseline': round(b_af * 100 if b_af <= 1 else b_af, 1),
                 'Midline': round(m_af * 100 if m_af <= 1 else m_af, 1)})

    # Women indicators
    for label, df_key, resp_col, resp_val in [
        ('CC Awareness', 'cc_heard', 'Response', 'Yes'),
        ('NbS Awareness', 'nbs_heard', 'Response', 'Yes'),
        ('Personal Savings', 'personal_saving', 'Response', 'Yes'),
        ('Prep Knowledge', 'prep_knowledge', 'Response', 'Yes'),
    ]:
        df = w_data[df_key]
        b_w = df.loc[df[resp_col] == resp_val, 'Baseline'].values[0]
        m_w = df.loc[df[resp_col] == resp_val, 'Midline'].values[0]
        rows.append({'Indicator': label, 'Dataset': 'Women',
                     'Baseline': round(b_w * 100, 1), 'Midline': round(m_w * 100, 1)})

    ls = w_data['lifeskills_agree']
    rows.append({'Indicator': 'Life Skills (avg)', 'Dataset': 'Women',
                 'Baseline': round(ls['Baseline'].mean() * 100, 1),
                 'Midline': round(ls['Midline'].mean() * 100, 1)})

    di = w_data['decision_influence']
    rows.append({'Indicator': 'Decision Influence', 'Dataset': 'Women',
                 'Baseline': round(di['Baseline'].mean() * 100, 1),
                 'Midline': round(di['Midline'].mean() * 100, 1)})

    # Men indicators (if available)
    if m_data is not None:
        for label, df_key, resp_col, resp_val in [
            ('CC Awareness (Men)', 'cc_heard', 'Response', 'Yes'),
            ('NbS Awareness (Men)', 'nbs_heard', 'Response', 'Yes'),
        ]:
            df = m_data[df_key]
            b_m = df.loc[df[resp_col] == resp_val, 'Baseline'].values[0]
            m_m = df.loc[df[resp_col] == resp_val, 'Midline'].values[0]
            rows.append({'Indicator': label, 'Dataset': 'Men',
                         'Baseline': round(b_m * 100, 1), 'Midline': round(m_m * 100, 1)})

        rj = m_data['roles_should_joint']
        rows.append({'Indicator': 'Joint Roles (Should)', 'Dataset': 'Men',
                     'Baseline': round(rj['Baseline'].mean() * 100, 1),
                     'Midline': round(rj['Midline'].mean() * 100, 1)})

        dj = m_data['decision_should_joint']
        rows.append({'Indicator': 'Joint Decisions (Should)', 'Dataset': 'Men',
                     'Baseline': round(dj['Baseline'].mean() * 100, 1),
                     'Midline': round(dj['Midline'].mean() * 100, 1)})

        m_sn = m_data['socialnorms_agree']
        m_harmful = m_sn[~m_sn['Norm'].str.contains(
            'Planting Crops|Restoring Ecosystems|Express Emotions', regex=True)]
        rows.append({'Indicator': 'Harmful Norms (Men)', 'Dataset': 'Men',
                     'Baseline': round(m_harmful['Baseline'].mean() * 100, 1),
                     'Midline': round(m_harmful['Midline'].mean() * 100, 1)})

        # Men's support for women in NbS (average across mangrove, seaweed, forest)
        support_bl, support_ml = [], []
        for key in ['mangrove_support', 'seaweed_support', 'forest_support']:
            df = m_data[key]
            yes_rows = df[df['Response'] == 'Yes']
            try:
                bl_v = float(yes_rows['Baseline'].values[0]) if len(yes_rows) else 0.0
            except (ValueError, TypeError):
                bl_v = 0.0
            try:
                ml_v = float(yes_rows['Midline'].values[0]) if len(yes_rows) else 0.0
            except (ValueError, TypeError):
                ml_v = 0.0
            if pd.isna(bl_v): bl_v = 0.0
            if pd.isna(ml_v): ml_v = 0.0
            support_bl.append(bl_v)
            support_ml.append(ml_v)
        rows.append({'Indicator': 'NbS Support for Women', 'Dataset': 'Men',
                     'Baseline': round(sum(support_bl) / len(support_bl) * 100, 1),
                     'Midline': round(sum(support_ml) / len(support_ml) * 100, 1)})

    # GJJ KAP Women indicators (if available) — uses Baseline/Endline mapped to Baseline/Midline columns
    if gjj_data is not None:
        try:
            # Self-esteem — Strongly Agree average
            sa = gjj_data['self_strongly_agree']
            rows.append({'Indicator': 'Self-Esteem (Strongly Agree)', 'Dataset': 'GJJ Women',
                         'Baseline': round(float(sa['Baseline'].mean()) * 100, 1),
                         'Midline': round(float(sa['Endline'].mean()) * 100, 1)})

            # Relational wellbeing — Always/Frequently average
            af_rn = gjj_data['rel_af_rn']
            rows.append({'Indicator': 'Relational Wellbeing (AF)', 'Dataset': 'GJJ Women',
                         'Baseline': round(float(af_rn['AF_Baseline'].mean()) * 100, 1),
                         'Midline': round(float(af_rn['AF_Endline'].mean()) * 100, 1)})

            # Shared responsibility — husband supports chores (Yes)
            chores = gjj_data['shared_chores_yn']
            ch_yes = chores[chores['Response'].str.strip().str.lower() == 'yes']
            if len(ch_yes):
                rows.append({'Indicator': 'Husband Supports Chores', 'Dataset': 'GJJ Women',
                             'Baseline': round(float(ch_yes['Baseline'].values[0]) * 100, 1),
                             'Midline': round(float(ch_yes['Endline'].values[0]) * 100, 1)})

            # Decision conversations — Yes
            dec_conv = gjj_data['decision_conversations']
            dec_yes = dec_conv[dec_conv['Response'].str.strip().str.lower() == 'yes']
            if len(dec_yes):
                rows.append({'Indicator': 'Decision Conversations', 'Dataset': 'GJJ Women',
                             'Baseline': round(float(dec_yes['Baseline'].values[0]) * 100, 1),
                             'Midline': round(float(dec_yes['Endline'].values[0]) * 100, 1)})

            # Equal say — Yes
            eq = gjj_data['equal_say']
            eq_yes = eq[eq['Response'].str.strip().str.lower() == 'yes']
            if len(eq_yes):
                rows.append({'Indicator': 'Equal Say in Decisions', 'Dataset': 'GJJ Women',
                             'Baseline': round(float(eq_yes['Baseline'].values[0]) * 100, 1),
                             'Midline': round(float(eq_yes['Endline'].values[0]) * 100, 1)})
        except Exception:
            pass

    # GJJ KAP Men indicators (if available)
    if gjj_men_data is not None:
        try:
            agr = gjj_men_data['self_agreement']
            rows.append({'Indicator': 'Self-Esteem Agreement (Men)', 'Dataset': 'GJJ Men',
                         'Baseline': round(float(agr['Agreement_BL'].mean()) * 100, 1),
                         'Midline': round(float(agr['Agreement_EL'].mean()) * 100, 1)})

            af_rn = gjj_men_data['rel_af_rn']
            rows.append({'Indicator': 'Relational Wellbeing AF (Men)', 'Dataset': 'GJJ Men',
                         'Baseline': round(float(af_rn['AF_Baseline'].mean()) * 100, 1),
                         'Midline': round(float(af_rn['AF_Endline'].mean()) * 100, 1)})

            chores = gjj_men_data['shared_chores_yn']
            ch_yes = chores[chores['Response'].str.strip().str.lower() == 'yes']
            if len(ch_yes):
                rows.append({'Indicator': 'Supports Chores (Men)', 'Dataset': 'GJJ Men',
                             'Baseline': round(float(ch_yes['Baseline'].values[0]) * 100, 1),
                             'Midline': round(float(ch_yes['Endline'].values[0]) * 100, 1)})

            dc = gjj_men_data['decision_conversations']
            dc_yes = dc[dc['Response'].str.strip().str.lower() == 'yes']
            if len(dc_yes):
                rows.append({'Indicator': 'Decision Conversations (Men)', 'Dataset': 'GJJ Men',
                             'Baseline': round(float(dc_yes['Baseline'].values[0]) * 100, 1),
                             'Midline': round(float(dc_yes['Endline'].values[0]) * 100, 1)})

            lead = gjj_men_data['support_leader']
            alw = lead[lead['Category'].str.strip().str.lower() == 'always']
            if len(alw):
                rows.append({'Indicator': 'Support Leadership (Men)', 'Dataset': 'GJJ Men',
                             'Baseline': round(float(alw['Baseline'].values[0]) * 100, 1),
                             'Midline': round(float(alw['Endline'].values[0]) * 100, 1)})
        except Exception:
            pass

    # Forest Training indicators (if available)
    if ft_data is not None:
        try:
            thresholds = ft_data['thresholds']
            scores = ft_data['scores']
            questions = ft_data['questions']

            # Pass rate >=70%
            t70 = thresholds[thresholds['Threshold'].str.contains('70')]
            if len(t70) >= 2:
                rows.append({'Indicator': 'Pass Rate \u226570%', 'Dataset': 'Forest Training',
                             'Baseline': round(float(t70[t70['Timepoint']=='Baseline']['Proportion'].values[0]) * 100, 1),
                             'Midline': round(float(t70[t70['Timepoint']=='Endline']['Proportion'].values[0]) * 100, 1)})

            # Pass rate >=80%
            t80 = thresholds[thresholds['Threshold'].str.contains('80')]
            if len(t80) >= 2:
                rows.append({'Indicator': 'Pass Rate \u226580%', 'Dataset': 'Forest Training',
                             'Baseline': round(float(t80[t80['Timepoint']=='Baseline']['Proportion'].values[0]) * 100, 1),
                             'Midline': round(float(t80[t80['Timepoint']=='Endline']['Proportion'].values[0]) * 100, 1)})

            # Average score
            if len(scores) >= 2:
                rows.append({'Indicator': 'Average Test Score', 'Dataset': 'Forest Training',
                             'Baseline': round(float(scores[scores['Timepoint']=='Baseline']['AverageScore'].values[0]), 1),
                             'Midline': round(float(scores[scores['Timepoint']=='Endline']['AverageScore'].values[0]), 1)})

            # Min score
            if len(scores) >= 2:
                rows.append({'Indicator': 'Minimum Score', 'Dataset': 'Forest Training',
                             'Baseline': round(float(scores[scores['Timepoint']=='Baseline']['MinScore'].values[0]), 1),
                             'Midline': round(float(scores[scores['Timepoint']=='Endline']['MinScore'].values[0]), 1)})

            # Avg % correct across all questions
            if not questions.empty:
                rows.append({'Indicator': 'Avg % Correct (All Qs)', 'Dataset': 'Forest Training',
                             'Baseline': round(float(questions['Baseline'].mean()) * 100, 1),
                             'Midline': round(float(questions['Endline'].mean()) * 100, 1)})
        except Exception:
            pass

    # ---- Mangrove Training indicators ----
    if mg_data is not None:
        try:
            mg_county = mg_data['adequate_county']
            mg_scores = mg_data['scores']
            mg_all = mg_county[mg_county['County'] == 'All']
            mg_pre = mg_all[mg_all['Timepoint'] == 'Pre-Test']['Value'].values
            mg_post = mg_all[mg_all['Timepoint'] == 'Post-Test']['Value'].values
            if len(mg_pre) and len(mg_post):
                rows.append({'Indicator': 'Adequate Knowledge (≥60%)', 'Dataset': 'Mangrove Training',
                             'Baseline': round(float(mg_pre[0]) * 100, 1),
                             'Midline': round(float(mg_post[0]) * 100, 1)})
            mg_all_scores = mg_scores[mg_scores['County'] == 'All']
            if len(mg_all_scores):
                rows.append({'Indicator': 'Average Test Score', 'Dataset': 'Mangrove Training',
                             'Baseline': 0.0,
                             'Midline': round(float(mg_all_scores['AvgScore'].values[0]), 1)})
                rows.append({'Indicator': 'Total Respondents', 'Dataset': 'Mangrove Training',
                             'Baseline': 0,
                             'Midline': int(mg_all_scores['Respondents'].values[0])})
        except Exception:
            pass

    # ---- Seaweed Production indicators ----
    if sw_data is not None:
        try:
            sw_agg = prepare_seaweed_aggregates(sw_data)
            ov = sw_agg['overall']
            avg_ach = sw_data['Ropes_Achievement_pct'].mean()
            rows.append({'Indicator': 'Total Production (kg)', 'Dataset': 'Seaweed',
                         'Baseline': 0.0, 'Midline': round(ov['total_kg'], 1)})
            rows.append({'Indicator': 'Avg Target Achievement (%)', 'Dataset': 'Seaweed',
                         'Baseline': 0.0, 'Midline': round(avg_ach, 1)})
            rows.append({'Indicator': 'Avg Production/Rope (kg)', 'Dataset': 'Seaweed',
                         'Baseline': 0.0, 'Midline': round(ov['avg_prod_per_rope'], 2)})
        except Exception:
            pass

    return rows


def _make_slope_chart(data_tuples, title):
    """Create a slope / parallel coordinates chart showing BL→ML trends.

    data_tuples: list of (label, baseline_pct, midline_pct)
    """
    fig = go.Figure()
    for label, bl, ml in data_tuples:
        change = ml - bl
        color = COLORS['good'] if change >= 0 else COLORS['danger']
        # Line connecting BL to ML
        fig.add_trace(go.Scatter(
            x=['Baseline', 'Midline'],
            y=[bl, ml],
            mode='lines+markers+text',
            line=dict(color=color, width=2.5),
            marker=dict(size=9, color=color, line=dict(width=1.5, color='white')),
            text=[f"{bl:.1f}%", f"{ml:.1f}%"],
            textposition=['middle left', 'middle right'],
            textfont=dict(size=10, color=color),
            name=label,
            hovertemplate=f'<b>{label}</b><br>%{{x}}: %{{y:.1f}}%<br>Change: {change:+.1f}pp<extra></extra>'
        ))
    fig.update_layout(
        title=title,
        height=380,
        xaxis=dict(tickvals=['Baseline', 'Midline'], tickfont=dict(size=13, color='#333')),
        yaxis_title='Percentage (%)',
        legend=dict(orientation='v', yanchor='top', y=1, x=1.02, font=dict(size=10)),
        font=dict(size=13, color='#333'),
        title_font=dict(size=16, color='#222'),
        plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
        margin=dict(l=50, r=120, t=60, b=20),
    )
    return fig


def render_insights_tab(f_data, w_data, m_data=None, gjj_data=None, gjj_men_data=None, ft_data=None, mg_data=None, sw_data=None):
    """Render the Insights tab with automated analysis across all datasets."""

    st.markdown("""<div class="section-narrative">
    <strong>Automated Insights:</strong> This tab generates data-driven insights by analyzing
    trends, changes, and patterns across the Forestry Conservation Groups, Women's Survey,
    Men's Survey, GJJ KAP Women, GJJ KAP Men, Forest Training, Mangrove Training, and
    Seaweed Production datasets. Insights are automatically derived from data comparisons.
    </div>""", unsafe_allow_html=True)

    # Summary counters
    f_insights = _generate_forestry_insights(f_data)
    w_insights = _generate_women_insights(w_data)
    m_insights = _generate_men_insights(m_data) if m_data is not None else []
    gjj_insights = _generate_gjj_insights(gjj_data) if gjj_data is not None else []
    gjj_men_insights = _generate_gjj_men_insights(gjj_men_data) if gjj_men_data is not None else []
    ft_insights = _generate_forest_training_insights(ft_data) if ft_data is not None else []
    mg_insights = _generate_mangrove_training_insights(mg_data) if mg_data is not None else []
    sw_insights = _generate_seaweed_insights(sw_data) if sw_data is not None else []
    cc_insights = _generate_cross_cutting_insights(f_data, w_data, m_data, gjj_data)

    all_insights = f_insights + w_insights + m_insights + gjj_insights + gjj_men_insights + ft_insights + mg_insights + sw_insights + cc_insights
    positive_count = sum(1 for _, _, t in all_insights if t == "positive")
    warning_count = sum(1 for _, _, t in all_insights if t in ("warning", "negative"))
    neutral_count = sum(1 for _, _, t in all_insights if t == "neutral")
    total = len(all_insights)

    # KPI summary
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Total Insights", total)
    c2.metric("Positive Trends", positive_count)
    c3.metric("Areas of Concern", warning_count)
    c4.metric("Cross-cutting", len(cc_insights))

    st.markdown("---")
    nav_items = ['Trend Overview', 'Forestry Insights', 'Women Survey Insights']
    if m_data is not None:
        nav_items.append('Men Survey Insights')
    if gjj_data is not None:
        nav_items.append('GJJ KAP Women Insights')
    if gjj_men_data is not None:
        nav_items.append('GJJ KAP Men Insights')
    if ft_data is not None:
        nav_items.append('Forest Training Insights')
    if mg_data is not None:
        nav_items.append('Mangrove Training Insights')
    if sw_data is not None:
        nav_items.append('Seaweed Production Insights')
    nav_items.extend(['Cross-Cutting Insights', 'Change Heatmap', 'Recommendations'])
    _quick_nav_pills(nav_items)

    # ====================================================================
    # TREND OVERVIEW — Donut + Dumbbell Chart
    # ====================================================================
    _section_header('', 'Trend Overview', 'At a Glance')

    # --- Build master indicator table used across multiple charts ---
    indicator_rows = _build_indicator_table(f_data, w_data, m_data, gjj_data, gjj_men_data, ft_data, mg_data, sw_data)
    ind_df = pd.DataFrame(indicator_rows)
    # Exclude Seaweed from trend/quadrant/heatmap charts — its metrics are raw
    # values (kg, counts) not percentages, and have no Baseline comparison,
    # which distorts the scale of charts designed for % indicators.
    ind_df = ind_df[ind_df['Dataset'] != 'Seaweed'].reset_index(drop=True)
    # Also exclude raw-count indicators that distort percentage-based charts
    ind_df = ind_df[ind_df['Indicator'] != 'Total Respondents'].reset_index(drop=True)
    ind_df['Change'] = round(ind_df['Midline'] - ind_df['Baseline'], 1)
    ind_df['Direction'] = ind_df['Change'].apply(
        lambda x: 'Improving' if x > 0.5 else ('Declining' if x < -0.5 else 'Stable'))
    # Descriptive label combining dataset + indicator for chart readability
    ind_df['Label'] = ind_df['Dataset'] + ': ' + ind_df['Indicator']

    # --- Row 1: Donut (trend distribution) + Dumbbell overview ---
    ov_col1, ov_col2 = st.columns([1, 2])

    with ov_col1:
        # Trend classification donut
        trend_counts = ind_df['Direction'].value_counts()
        donut_colors = {'Improving': COLORS['good'], 'Declining': COLORS['danger'], 'Stable': '#9E9E9E'}
        fig_donut = go.Figure(go.Pie(
            labels=trend_counts.index,
            values=trend_counts.values,
            hole=0.55,
            marker=dict(colors=[donut_colors.get(d, '#888') for d in trend_counts.index]),
            textinfo='label+value',
            textfont=dict(size=13),
            hovertemplate='%{label}: %{value} indicators<extra></extra>'
        ))
        fig_donut.update_layout(
            title="Indicator Trend Distribution",
            height=350,
            font=dict(size=13, color='#333'),
            title_font=dict(size=16, color='#222'),
            showlegend=False,
            margin=dict(l=10, r=10, t=50, b=10),
            annotations=[dict(text=f"{len(ind_df)}<br>Indicators", x=0.5, y=0.5,
                               font=dict(size=16, color='#333', family='Arial'),
                               showarrow=False)]
        )
        st.plotly_chart(fig_donut, use_container_width=True)

    with ov_col2:
        # Dumbbell / slope chart: Baseline dot → Midline dot with connecting line
        dumb_df = ind_df.sort_values('Change', ascending=True).reset_index(drop=True)
        fig_dumb = go.Figure()

        # Connecting lines (drawn first so dots sit on top)
        for i, row in dumb_df.iterrows():
            line_color = COLORS['good'] if row['Change'] >= 0 else COLORS['danger']
            fig_dumb.add_trace(go.Scatter(
                x=[row['Baseline'], row['Midline']],
                y=[row['Label'], row['Label']],
                mode='lines',
                line=dict(color=line_color, width=2.5),
                showlegend=False,
                hoverinfo='skip'
            ))
        # Baseline dots
        fig_dumb.add_trace(go.Scatter(
            x=dumb_df['Baseline'], y=dumb_df['Label'],
            mode='markers',
            marker=dict(size=10, color=COLORS['baseline'], symbol='circle',
                        line=dict(width=1.5, color='white')),
            name='Baseline',
            hovertemplate='%{y}<br>Baseline: %{x:.1f}%<extra></extra>'
        ))
        # Midline dots
        fig_dumb.add_trace(go.Scatter(
            x=dumb_df['Midline'], y=dumb_df['Label'],
            mode='markers',
            marker=dict(size=10, color=COLORS['midline'], symbol='diamond',
                        line=dict(width=1.5, color='white')),
            name='Midline',
            hovertemplate='%{y}<br>Midline: %{x:.1f}%<extra></extra>'
        ))
        fig_dumb.update_layout(
            title="Baseline to Midline Shift (Dumbbell Chart)",
            height=max(420, len(dumb_df) * 30),
            xaxis_title="Percentage (%)",
            yaxis=dict(categoryorder='trace'),
            legend=dict(orientation='h', yanchor='bottom', y=1.02, x=0.5, xanchor='center'),
            font=dict(size=13, color='#333'),
            title_font=dict(size=16, color='#222'),
            plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
            margin=dict(l=20, r=20, t=60, b=20),
        )
        st.plotly_chart(fig_dumb, use_container_width=True)

    st.markdown("---")

    # ====================================================================
    # FORESTRY INSIGHTS + Domain Radar Trend
    # ====================================================================
    _section_header('', 'Forestry Conservation Insights', 'Community Level')

    # Forestry Domain Radar — BL vs ML overlay
    fd = f_data['functionality_domain']
    domains = ['Management', 'Gender', 'Effectiveness', 'Overall']
    bl_vals = [fd.loc[fd['Timepoint'] == 'Baseline', d].values[0] for d in domains]
    ml_vals = [fd.loc[fd['Timepoint'] == 'Midline', d].values[0] for d in domains]

    rad_col1, rad_col2 = st.columns([1, 1])
    with rad_col1:
        fig_radar = go.Figure()
        fig_radar.add_trace(go.Scatterpolar(
            r=bl_vals + [bl_vals[0]],
            theta=domains + [domains[0]],
            name='Baseline',
            fill='toself',
            fillcolor='rgba(66,133,244,0.12)',
            line=dict(color=COLORS['baseline'], width=2.5),
            marker=dict(size=7)
        ))
        fig_radar.add_trace(go.Scatterpolar(
            r=ml_vals + [ml_vals[0]],
            theta=domains + [domains[0]],
            name='Midline',
            fill='toself',
            fillcolor='rgba(52,168,83,0.12)',
            line=dict(color=COLORS['midline'], width=2.5),
            marker=dict(size=7)
        ))
        fig_radar.update_layout(
            title="Domain Score Trend (Radar)",
            polar=dict(radialaxis=dict(visible=True, range=[0, max(max(bl_vals), max(ml_vals)) * 1.15],
                                        tickfont=dict(size=11))),
            showlegend=True,
            legend=dict(orientation='h', yanchor='bottom', y=-0.15, x=0.5, xanchor='center'),
            height=380,
            font=dict(size=13, color='#333'),
            title_font=dict(size=16, color='#222'),
            margin=dict(l=60, r=60, t=60, b=40),
        )
        st.plotly_chart(fig_radar, use_container_width=True)

    with rad_col2:
        # Forestry slope chart — selected key indicators
        f_slope_data = []
        # Functionality threshold
        ft = f_data['functionality_threshold']
        bl_60 = ft.loc[ft['Timepoint'] == 'Baseline', 'Functional_60_pct'].values[0]
        ml_60 = ft.loc[ft['Timepoint'] == 'Midline', 'Functional_60_pct'].values[0]
        f_slope_data.append(('Functionality >=60%', bl_60 * 100 if bl_60 <= 1 else bl_60,
                             ml_60 * 100 if ml_60 <= 1 else ml_60))
        # Women leadership
        wl = f_data['women_leadership']
        sig_bl = wl.loc[wl['Category'] == 'Significant Leadership', 'Baseline'].values[0]
        sig_ml = wl.loc[wl['Category'] == 'Significant Leadership', 'Midline'].values[0]
        f_slope_data.append(('Women Leadership', sig_bl * 100 if sig_bl <= 1 else sig_bl,
                             sig_ml * 100 if sig_ml <= 1 else sig_ml))
        # Training coverage
        tc = f_data['training_coverage']
        tc_bl = tc.loc[tc['Category'] == 'Most Members', 'Baseline'].values[0]
        tc_ml = tc.loc[tc['Category'] == 'Most Members', 'Midline'].values[0]
        f_slope_data.append(('Training (Most)', tc_bl * 100 if tc_bl <= 1 else tc_bl,
                             tc_ml * 100 if tc_ml <= 1 else tc_ml))
        # Income gen
        ig = f_data['income_gen']
        ig_bl = ig.loc[ig['Category'] == 'Yes', 'Baseline'].values[0]
        ig_ml = ig.loc[ig['Category'] == 'Yes', 'Midline'].values[0]
        f_slope_data.append(('Income Gen', ig_bl * 100 if ig_bl <= 1 else ig_bl,
                             ig_ml * 100 if ig_ml <= 1 else ig_ml))
        # Agroforestry
        af = f_data['agroforestry']
        af_bl = af.loc[af['Category'] == 'Yes', 'Baseline'].values[0]
        af_ml = af.loc[af['Category'] == 'Yes', 'Midline'].values[0]
        f_slope_data.append(('Agroforestry', af_bl * 100 if af_bl <= 1 else af_bl,
                             af_ml * 100 if af_ml <= 1 else af_ml))

        fig_fslope = _make_slope_chart(f_slope_data, "Forestry Key Indicator Trends")
        st.plotly_chart(fig_fslope, use_container_width=True)

    # Insight cards
    for title, body, trend in f_insights:
        _insight_card(title, body, trend)

    st.markdown("---")

    # ====================================================================
    # WOMEN SURVEY INSIGHTS + Empowerment Trend Charts
    # ====================================================================
    _section_header('', "Women's Survey Insights", 'Household Level')

    # Women empowerment slope chart + life skills domain trend
    we_col1, we_col2 = st.columns([1, 1])

    with we_col1:
        # Women's key indicators slope chart
        w_slope_data = []
        for label, df_key, resp_col, resp_val in [
            ('CC Awareness', 'cc_heard', 'Response', 'Yes'),
            ('NbS Awareness', 'nbs_heard', 'Response', 'Yes'),
            ('Personal Savings', 'personal_saving', 'Response', 'Yes'),
            ('Prep Knowledge', 'prep_knowledge', 'Response', 'Yes'),
        ]:
            df = w_data[df_key]
            b_w = df.loc[df[resp_col] == resp_val, 'Baseline'].values[0] * 100
            m_w = df.loc[df[resp_col] == resp_val, 'Midline'].values[0] * 100
            w_slope_data.append((label, b_w, m_w))

        # Decision influence avg
        di = w_data['decision_influence']
        w_slope_data.append(('Decision Influence', di['Baseline'].mean() * 100, di['Midline'].mean() * 100))

        # Life skills avg
        ls = w_data['lifeskills_agree']
        w_slope_data.append(('Life Skills', ls['Baseline'].mean() * 100, ls['Midline'].mean() * 100))

        fig_wslope = _make_slope_chart(w_slope_data, "Women's Key Indicator Trends")
        st.plotly_chart(fig_wslope, use_container_width=True)

    with we_col2:
        # Life skills by domain — grouped bar with trend lines
        ls_all = w_data['lifeskills_agree'].copy()
        # Attempt to group by domain if Statement column contains domain hints
        domain_map = {}
        for _, row in ls_all.iterrows():
            stmt = str(row.get('Statement', ''))
            if any(k in stmt.lower() for k in ['confident', 'proud', 'worth', 'self']):
                domain_map[stmt] = 'Self Esteem'
            elif any(k in stmt.lower() for k in ['future', 'plan', 'goal', 'aspir']):
                domain_map[stmt] = 'Aspirations'
            else:
                domain_map[stmt] = 'Leadership'

        # Build domain averages
        ls_all['Domain'] = ls_all['Statement'].map(domain_map)
        domain_avg = ls_all.groupby('Domain')[['Baseline', 'Midline']].mean().reset_index()

        # Add communication and conflict as extra domains
        comm = w_data['communication_agree']
        conf = w_data['conflict_agree']
        extra_domains = [
            {'Domain': 'Communication', 'Baseline': comm['Baseline'].mean(), 'Midline': comm['Midline'].mean()},
            {'Domain': 'Conflict Res.', 'Baseline': conf['Baseline'].mean(), 'Midline': conf['Midline'].mean()},
        ]
        domain_avg = pd.concat([domain_avg, pd.DataFrame(extra_domains)], ignore_index=True)
        domain_avg['BL_pct'] = domain_avg['Baseline'] * 100
        domain_avg['ML_pct'] = domain_avg['Midline'] * 100
        domain_avg['Change'] = round(domain_avg['ML_pct'] - domain_avg['BL_pct'], 1)

        fig_dom = go.Figure()
        fig_dom.add_trace(go.Bar(
            x=domain_avg['Domain'], y=domain_avg['BL_pct'],
            name='Baseline', marker_color=COLORS['baseline'],
            text=domain_avg['BL_pct'].apply(lambda x: f"{x:.1f}%"), textposition='auto'
        ))
        fig_dom.add_trace(go.Bar(
            x=domain_avg['Domain'], y=domain_avg['ML_pct'],
            name='Midline', marker_color=COLORS['midline'],
            text=domain_avg['ML_pct'].apply(lambda x: f"{x:.1f}%"), textposition='auto'
        ))
        # Trend line connecting midline values
        fig_dom.add_trace(go.Scatter(
            x=domain_avg['Domain'], y=domain_avg['ML_pct'],
            mode='lines+markers',
            name='Midline Trend',
            line=dict(color=COLORS['card_value'], width=2.5, dash='dot'),
            marker=dict(size=8, symbol='diamond'),
            yaxis='y'
        ))
        fig_dom.update_layout(
            title="Skills Domain Trends (BL vs ML)",
            barmode='group', height=380,
            yaxis_title='Agree/Strongly Agree (%)',
            legend=dict(orientation='h', yanchor='bottom', y=1.02, x=0.5, xanchor='center'),
            font=dict(size=13, color='#333'),
            title_font=dict(size=16, color='#222'),
            plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
            margin=dict(l=20, r=20, t=60, b=20),
        )
        st.plotly_chart(fig_dom, use_container_width=True)

    # Insight cards
    for title, body, trend in w_insights:
        _insight_card(title, body, trend)

    st.markdown("---")

    # ====================================================================
    # MEN SURVEY INSIGHTS + Key Indicator Trends
    # ====================================================================
    if m_data is not None and m_insights:
        _section_header('', "Men's Survey Insights", 'Household Level')

        me_col1, me_col2 = st.columns([1, 1])
        with me_col1:
            # Men's key indicators slope chart
            m_slope_data = []
            m_cc = m_data['cc_heard']
            m_slope_data.append(('CC Awareness',
                                m_cc.loc[m_cc['Response']=='Yes','Baseline'].values[0]*100,
                                m_cc.loc[m_cc['Response']=='Yes','Midline'].values[0]*100))
            m_nbs = m_data['nbs_heard']
            m_slope_data.append(('NbS Awareness',
                                m_nbs.loc[m_nbs['Response']=='Yes','Baseline'].values[0]*100,
                                m_nbs.loc[m_nbs['Response']=='Yes','Midline'].values[0]*100))
            m_rj = m_data['roles_should_joint']
            m_slope_data.append(('Joint Roles (Should)',
                                m_rj['Baseline'].mean()*100, m_rj['Midline'].mean()*100))
            m_dj = m_data['decision_should_joint']
            m_slope_data.append(('Joint Decisions (Should)',
                                m_dj['Baseline'].mean()*100, m_dj['Midline'].mean()*100))
            m_sn = m_data['socialnorms_agree']
            m_harmful = m_sn[~m_sn['Norm'].str.contains(
                'Planting Crops|Restoring Ecosystems|Express Emotions', regex=True)]
            m_slope_data.append(('Harmful Norms',
                                m_harmful['Baseline'].mean()*100, m_harmful['Midline'].mean()*100))

            fig_mslope = _make_slope_chart(m_slope_data, "Men's Key Indicator Trends")
            st.plotly_chart(fig_mslope, use_container_width=True)

        with me_col2:
            # Men's NbS support comparison across sectors
            sectors = ['mangrove_support', 'seaweed_support', 'forest_support']
            sector_labels = ['Mangrove', 'Seaweed', 'Forest Mgmt']
            sup_bl_vals, sup_ml_vals = [], []
            for key in sectors:
                df = m_data[key]
                yes_df = df[df['Response']=='Yes']
                try:
                    _bl = float(yes_df['Baseline'].values[0])*100 if len(yes_df) else 0.0
                except (ValueError, TypeError):
                    _bl = 0.0
                try:
                    _ml = float(yes_df['Midline'].values[0])*100 if len(yes_df) else 0.0
                except (ValueError, TypeError):
                    _ml = 0.0
                if pd.isna(_bl): _bl = 0.0
                if pd.isna(_ml): _ml = 0.0
                sup_bl_vals.append(_bl)
                sup_ml_vals.append(_ml)

            fig_msup = go.Figure()
            fig_msup.add_trace(go.Bar(x=sector_labels, y=sup_bl_vals, name='Baseline',
                                      marker_color=COLORS['baseline'],
                                      text=[f"{v:.1f}%" for v in sup_bl_vals], textposition='auto'))
            fig_msup.add_trace(go.Bar(x=sector_labels, y=sup_ml_vals, name='Midline',
                                      marker_color=COLORS['midline'],
                                      text=[f"{v:.1f}%" for v in sup_ml_vals], textposition='auto'))
            fig_msup.update_layout(
                title="Men Supporting Women in NbS (% Yes)",
                barmode='group', height=380, yaxis_title='%',
                legend=dict(orientation='h', yanchor='bottom', y=1.02, x=0.5, xanchor='center'),
                font=dict(size=13, color='#333'),
                title_font=dict(size=16, color='#222'),
                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                margin=dict(l=20, r=20, t=60, b=20),
            )
            st.plotly_chart(fig_msup, use_container_width=True)

        for title, body, trend in m_insights:
            _insight_card(title, body, trend)

        st.markdown("---")

    # ====================================================================
    # GJJ KAP WOMEN INSIGHTS
    # ====================================================================
    if gjj_data is not None and gjj_insights:
        _section_header('', 'GJJ KAP Women Insights', 'Baseline/Endline')

        gj_col1, gj_col2 = st.columns([1, 1])
        with gj_col1:
            # GJJ key indicators slope chart
            gj_slope_data = []
            try:
                sa = gjj_data['self_strongly_agree']
                gj_slope_data.append(('Self-Esteem (SA)',
                                     float(sa['Baseline'].mean())*100,
                                     float(sa['Endline'].mean())*100))
                af_rn = gjj_data['rel_af_rn']
                gj_slope_data.append(('Relational (AF)',
                                     float(af_rn['AF_Baseline'].mean())*100,
                                     float(af_rn['AF_Endline'].mean())*100))
                chores = gjj_data['shared_chores_yn']
                ch_yes = chores[chores['Response'].str.strip().str.lower() == 'yes']
                if len(ch_yes):
                    gj_slope_data.append(('Husband Chores',
                                         float(ch_yes['Baseline'].values[0])*100,
                                         float(ch_yes['Endline'].values[0])*100))
                eq = gjj_data['equal_say']
                eq_yes = eq[eq['Response'].str.strip().str.lower() == 'yes']
                if len(eq_yes):
                    gj_slope_data.append(('Equal Say',
                                         float(eq_yes['Baseline'].values[0])*100,
                                         float(eq_yes['Endline'].values[0])*100))
            except Exception:
                pass
            if gj_slope_data:
                fig_gjslope = _make_slope_chart(gj_slope_data, "GJJ KAP Women: Key Indicator Trends")
                st.plotly_chart(fig_gjslope, use_container_width=True)

        with gj_col2:
            # GJJ self-compassion comparison
            try:
                comp = gjj_data['self_compassion']
                fig_comp = go.Figure()
                fig_comp.add_trace(go.Bar(
                    x=comp['Category'], y=comp['Baseline'].apply(lambda x: float(x)*100),
                    name='Baseline', marker_color=COLORS['baseline'],
                    text=comp['Baseline'].apply(lambda x: f"{float(x)*100:.1f}%"),
                    textposition='auto'))
                fig_comp.add_trace(go.Bar(
                    x=comp['Category'], y=comp['Endline'].apply(lambda x: float(x)*100),
                    name='Endline', marker_color=COLORS['midline'],
                    text=comp['Endline'].apply(lambda x: f"{float(x)*100:.1f}%"),
                    textposition='auto'))
                fig_comp.update_layout(
                    title="Self-Compassion Frequency (BL vs EL)",
                    barmode='group', height=380, yaxis_title='%',
                    legend=dict(orientation='h', yanchor='bottom', y=1.02, x=0.5, xanchor='center'),
                    font=dict(size=13, color='#333'),
                    title_font=dict(size=16, color='#222'),
                    plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                    margin=dict(l=20, r=20, t=60, b=20),
                )
                st.plotly_chart(fig_comp, use_container_width=True)
            except Exception:
                pass

        for title, body, trend in gjj_insights:
            _insight_card(title, body, trend)

        st.markdown("---")

    # ====================================================================
    # GJJ KAP MEN INSIGHTS
    # ====================================================================
    if gjj_men_data is not None and gjj_men_insights:
        _section_header('', 'GJJ KAP Men Insights', 'Baseline/Endline')

        gjm_col1, gjm_col2 = st.columns([1, 1])
        with gjm_col1:
            # GJJ Men key indicators slope chart
            gjm_slope_data = []
            try:
                agr = gjj_men_data['self_agreement']
                gjm_slope_data.append(('Self-Esteem (Agree)',
                                       float(agr['Agreement_BL'].mean())*100,
                                       float(agr['Agreement_EL'].mean())*100))
                af_rn = gjj_men_data['rel_af_rn']
                gjm_slope_data.append(('Relational (AF)',
                                       float(af_rn['AF_Baseline'].mean())*100,
                                       float(af_rn['AF_Endline'].mean())*100))
                chores = gjj_men_data['shared_chores_yn']
                ch_yes = chores[chores['Response'].str.strip().str.lower() == 'yes']
                if len(ch_yes):
                    gjm_slope_data.append(('Supports Chores',
                                           float(ch_yes['Baseline'].values[0])*100,
                                           float(ch_yes['Endline'].values[0])*100))
                dc = gjj_men_data['decision_conversations']
                dc_yes = dc[dc['Response'].str.strip().str.lower() == 'yes']
                if len(dc_yes):
                    gjm_slope_data.append(('Decision Conversations',
                                           float(dc_yes['Baseline'].values[0])*100,
                                           float(dc_yes['Endline'].values[0])*100))
            except Exception:
                pass
            if gjm_slope_data:
                fig_gjmslope = _make_slope_chart(gjm_slope_data, "GJJ KAP Men: Key Indicator Trends")
                st.plotly_chart(fig_gjmslope, use_container_width=True)

        with gjm_col2:
            # Leadership + business support comparison
            try:
                lead = gjj_men_data['support_leader']
                biz = gjj_men_data['support_business']
                cats = ['Always (Leader)', 'Definitely (Business)']
                bl_vals, el_vals = [], []
                alw = lead[lead['Category'].str.strip().str.lower() == 'always']
                bl_vals.append(float(alw['Baseline'].values[0])*100 if len(alw) else 0)
                el_vals.append(float(alw['Endline'].values[0])*100 if len(alw) else 0)
                defn = biz[biz['Category'].str.strip().str.lower() == 'definitely']
                bl_vals.append(float(defn['Baseline'].values[0])*100 if len(defn) else 0)
                el_vals.append(float(defn['Endline'].values[0])*100 if len(defn) else 0)
                fig_gjmsup = go.Figure()
                fig_gjmsup.add_trace(go.Bar(x=cats, y=bl_vals, name='Baseline',
                                            marker_color=COLORS['baseline'],
                                            text=[f"{v:.1f}%" for v in bl_vals], textposition='auto'))
                fig_gjmsup.add_trace(go.Bar(x=cats, y=el_vals, name='Endline',
                                            marker_color=COLORS['midline'],
                                            text=[f"{v:.1f}%" for v in el_vals], textposition='auto'))
                fig_gjmsup.update_layout(
                    title="Men's Support: Leadership & Business (BL vs EL)",
                    barmode='group', height=380, yaxis_title='%',
                    legend=dict(orientation='h', yanchor='bottom', y=1.02, x=0.5, xanchor='center'),
                    font=dict(size=13, color='#333'),
                    title_font=dict(size=16, color='#222'),
                    plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                    margin=dict(l=20, r=20, t=60, b=20),
                )
                st.plotly_chart(fig_gjmsup, use_container_width=True)
            except Exception:
                pass

        for title, body, trend in gjj_men_insights:
            _insight_card(title, body, trend)

        st.markdown("---")

    # ====================================================================
    # FOREST TRAINING INSIGHTS
    # ====================================================================
    if ft_data is not None and ft_insights:
        _section_header('', 'Forest Training Insights', 'Pre/Post')

        ft_col1, ft_col2 = st.columns([1, 1])
        with ft_col1:
            # Pass-rate slope chart
            ft_slope_data = []
            try:
                thresholds = ft_data['thresholds']
                for thr_lbl in thresholds['Threshold'].unique():
                    thr_sub = thresholds[thresholds['Threshold'] == thr_lbl]
                    bl_v = float(thr_sub[thr_sub['Timepoint'] == 'Baseline']['Proportion'].values[0]) * 100
                    el_v = float(thr_sub[thr_sub['Timepoint'] == 'Endline']['Proportion'].values[0]) * 100
                    ft_slope_data.append((f'Pass {thr_lbl}', bl_v, el_v))
            except Exception:
                pass
            if ft_slope_data:
                fig_ft_slope = _make_slope_chart(ft_slope_data, "Forest Training: Pass Rate Trends")
                st.plotly_chart(fig_ft_slope, use_container_width=True)

        with ft_col2:
            # Average + min score comparison
            try:
                scores = ft_data['scores']
                cats = ['Average Score', 'Min Score']
                bl_vals = [float(scores[scores['Timepoint']=='Baseline']['AverageScore'].values[0]),
                           float(scores[scores['Timepoint']=='Baseline']['MinScore'].values[0])]
                el_vals = [float(scores[scores['Timepoint']=='Endline']['AverageScore'].values[0]),
                           float(scores[scores['Timepoint']=='Endline']['MinScore'].values[0])]
                fig_ft_bar = go.Figure()
                fig_ft_bar.add_trace(go.Bar(x=cats, y=bl_vals, name='Pre-Training',
                                            marker_color=COLORS['baseline'],
                                            text=[f"{v:.1f}%" for v in bl_vals], textposition='auto'))
                fig_ft_bar.add_trace(go.Bar(x=cats, y=el_vals, name='Post-Training',
                                            marker_color=COLORS['midline'],
                                            text=[f"{v:.1f}%" for v in el_vals], textposition='auto'))
                fig_ft_bar.update_layout(
                    title='Forest Training: Scores (Pre vs Post)',
                    barmode='group', height=380, yaxis_title='Score (%)',
                    legend=dict(orientation='h', yanchor='bottom', y=1.02, x=0.5, xanchor='center'),
                    font=dict(size=13, color='#333'),
                    title_font=dict(size=16, color='#222'),
                    plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                    margin=dict(l=20, r=20, t=60, b=20),
                )
                st.plotly_chart(fig_ft_bar, use_container_width=True)
            except Exception:
                pass

        for title, body, trend in ft_insights:
            _insight_card(title, body, trend)

        st.markdown("---")

    # ====================================================================
    # MANGROVE TRAINING INSIGHTS
    # ====================================================================
    if mg_data is not None and mg_insights:
        _section_header('', 'Mangrove Training Insights', 'Pre/Post')

        # Summary chart — adequate knowledge pre vs post
        try:
            mg_county = mg_data['adequate_county']
            mg_all = mg_county[mg_county['County'] == 'All'].copy()
            mg_all['Pct'] = mg_all['Value'] * 100
            fig_mg_bar = go.Figure()
            for _, row in mg_all.iterrows():
                color = COLORS['baseline'] if row['Timepoint'] == 'Pre-Test' else COLORS['midline']
                fig_mg_bar.add_trace(go.Bar(
                    x=[row['Timepoint']], y=[row['Pct']],
                    name=row['Timepoint'], marker_color=color,
                    text=[f"{row['Pct']:.1f}%"], textposition='auto', width=0.5,
                ))
            fig_mg_bar.update_layout(
                title='Mangrove Training: Adequate Knowledge (≥60%) Pre vs Post',
                height=350, yaxis_title='% of Participants', showlegend=False,
                font=dict(size=13, color='#333'),
                title_font=dict(size=16, color='#222'),
                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                margin=dict(l=20, r=20, t=60, b=20), yaxis=dict(range=[0, 105]),
            )
            st.plotly_chart(fig_mg_bar, use_container_width=True)
        except Exception:
            pass

        for title, body, trend in mg_insights:
            _insight_card(title, body, trend)

        st.markdown("---")

    # ====================================================================
    # SEAWEED PRODUCTION INSIGHTS
    # ====================================================================
    if sw_data is not None and sw_insights:
        _section_header('', 'Seaweed Production Insights', '2025')

        # Summary charts — two columns
        try:
            sw_agg = prepare_seaweed_aggregates(sw_data)
            sw_grp = sw_agg['group_summary'].sort_values('Total_KG', ascending=False).head(7)
            sw_ov = sw_agg['overall']

            # KPI highlights
            ik1, ik2, ik3, ik4 = st.columns(4)
            ik1.markdown(f"""<div class="kpi-card">
                <h3>Total Production</h3>
                <div class="value">{sw_ov['total_kg']:,.0f} kg</div>
            </div>""", unsafe_allow_html=True)
            ik2.markdown(f"""<div class="kpi-card">
                <h3>Avg Achievement</h3>
                <div class="value">{sw_ov['avg_achievement_pct']:.1f}%</div>
            </div>""", unsafe_allow_html=True)
            ik3.markdown(f"""<div class="kpi-card">
                <h3>Rope Gap</h3>
                <div class="value">{sw_ov['gap_total']:,.0f}</div>
            </div>""", unsafe_allow_html=True)
            ik4.markdown(f"""<div class="kpi-card">
                <h3>Multi-Challenge</h3>
                <div class="value">{sw_ov['multi_challenge_pct']:.1f}%</div>
            </div>""", unsafe_allow_html=True)

            ci1, ci2 = st.columns(2)
            with ci1:
                fig_sw_bar = go.Figure(go.Bar(
                    x=sw_grp['Group'], y=sw_grp['Total_KG'],
                    marker_color=COLORS['baseline'],
                    text=sw_grp['Total_KG'].apply(lambda v: f"{v:,.0f} kg"),
                    textposition='auto',
                ))
                fig_sw_bar.update_layout(
                    title='Top Seaweed Groups by Production (kg)',
                    height=350, yaxis_title='Total KG', showlegend=False,
                    font=dict(size=13, color='#333'),
                    title_font=dict(size=16, color='#222'),
                    plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                    margin=dict(l=20, r=20, t=60, b=20), xaxis_tickangle=-25,
                )
                st.plotly_chart(fig_sw_bar, use_container_width=True)
            with ci2:
                # Challenge overview mini chart
                sw_ch = sw_agg['challenge_counts']
                if not sw_ch.empty:
                    sw_ch_sorted = sw_ch.sort_values('Pct', ascending=True)
                    fig_sw_ch = go.Figure(go.Bar(
                        x=sw_ch_sorted['Pct'], y=sw_ch_sorted['Challenge'],
                        orientation='h', marker_color='#FF9800',
                        text=sw_ch_sorted['Pct'].apply(lambda v: f"{v:.1f}%"),
                        textposition='auto',
                    ))
                    fig_sw_ch.update_layout(
                        title='Challenge Prevalence (%)',
                        height=350, xaxis_title='% of Farmers', showlegend=False,
                        font=dict(size=13, color='#333'),
                        title_font=dict(size=15, color='#222'),
                        plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                        margin=dict(l=20, r=20, t=60, b=20),
                    )
                    st.plotly_chart(fig_sw_ch, use_container_width=True)
        except Exception:
            pass

        for title, body, trend in sw_insights:
            _insight_card(title, body, trend)

        st.markdown("---")

    # ====================================================================
    # CROSS-CUTTING INSIGHTS + Performance Quadrant
    # ====================================================================
    _section_header('', 'Cross-Cutting Insights', 'Integrated')

    # Performance Quadrant Scatter — Midline level (x) vs Change (y)
    st.markdown("""<div class="section-narrative">
    <strong>Performance Quadrant:</strong> Each dot is one indicator. The x-axis shows
    the current (Midline) level, the y-axis shows Baseline-to-Midline change. Quadrants
    help identify high performers, rising stars, areas at risk, and stagnating indicators.
    </div>""", unsafe_allow_html=True)

    fig_quad = go.Figure()

    # Separate by dataset for color coding
    for ds, color, symbol in [('Forestry', COLORS['baseline'], 'circle'),
                               ('Women', COLORS['midline'], 'diamond'),
                               ('Men', '#FF9800', 'square'),
                               ('GJJ Women', '#9C27B0', 'star'),
                               ('GJJ Men', '#00BCD4', 'hexagram'),
                               ('Forest Training', '#795548', 'cross'),
                               ('Mangrove Training', '#009688', 'triangle-up')]:
        subset = ind_df[ind_df['Dataset'] == ds]
        fig_quad.add_trace(go.Scatter(
            x=subset['Midline'],
            y=subset['Change'],
            mode='markers+text',
            marker=dict(size=12, color=color, symbol=symbol,
                        line=dict(width=1.5, color='white')),
            text=subset['Label'],
            textposition='top center',
            textfont=dict(size=10),
            name=ds,
            hovertemplate='<b>%{text}</b><br>Midline: %{x:.1f}%<br>Change: %{y:+.1f}pp<extra></extra>'
        ))

    # Quadrant reference lines
    median_ml = ind_df['Midline'].median()
    fig_quad.add_hline(y=0, line_dash='dash', line_color='#999', line_width=1)
    fig_quad.add_vline(x=median_ml, line_dash='dash', line_color='#999', line_width=1)

    # Quadrant labels
    y_range = max(abs(ind_df['Change'].min()), abs(ind_df['Change'].max()))
    x_max = ind_df['Midline'].max()
    x_min = ind_df['Midline'].min()
    annotations = [
        dict(x=x_min + (median_ml - x_min) * 0.5, y=y_range * 0.85,
             text="Needs Attention", showarrow=False,
             font=dict(size=11, color='#C62828'), bgcolor='rgba(255,235,238,0.7)'),
        dict(x=median_ml + (x_max - median_ml) * 0.5, y=y_range * 0.85,
             text="High Performers", showarrow=False,
             font=dict(size=11, color='#2E7D32'), bgcolor='rgba(232,245,233,0.7)'),
        dict(x=x_min + (median_ml - x_min) * 0.5, y=-y_range * 0.85,
             text="At Risk", showarrow=False,
             font=dict(size=11, color='#E65100'), bgcolor='rgba(255,243,224,0.7)'),
        dict(x=median_ml + (x_max - median_ml) * 0.5, y=-y_range * 0.85,
             text="Declining Leaders", showarrow=False,
             font=dict(size=11, color='#37474F'), bgcolor='rgba(236,239,241,0.7)'),
    ]
    fig_quad.update_layout(
        title="Performance Quadrant: Current Level vs Change",
        height=480,
        xaxis_title="Midline Level (%)",
        yaxis_title="Change (pp)",
        legend=dict(orientation='h', yanchor='bottom', y=1.02, x=0.5, xanchor='center'),
        font=dict(size=13, color='#333'),
        title_font=dict(size=16, color='#222'),
        plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
        margin=dict(l=20, r=20, t=60, b=20),
        annotations=annotations,
    )
    st.plotly_chart(fig_quad, use_container_width=True)

    # Insight cards
    for title, body, trend in cc_insights:
        _insight_card(title, body, trend)

    st.markdown("---")

    # ====================================================================
    # CHANGE HEATMAP + Waterfall
    # ====================================================================
    _section_header('', 'Indicator Change Heatmap', 'Overview')
    st.markdown("""<div class="section-narrative">
    A visual summary of percentage-point changes across key indicators from both datasets.
    Green cells indicate positive change; red cells indicate areas that declined.
    </div>""", unsafe_allow_html=True)

    heatmap_df = ind_df.copy()

    # --- Butterfly chart: BL on left, ML on right ---
    hm_sorted = heatmap_df.sort_values('Change')
    hm_col1, hm_col2 = st.columns([1, 1])

    with hm_col1:
        # Diverging bar chart of changes
        colors_bar = [COLORS['good'] if v >= 0 else COLORS['danger'] for v in hm_sorted['Change']]
        fig_hm = go.Figure()
        fig_hm.add_trace(go.Bar(
            y=hm_sorted['Label'],
            x=hm_sorted['Change'],
            orientation='h',
            marker_color=colors_bar,
            text=hm_sorted['Change'].apply(lambda x: f"{x:+.1f} pp"),
            textposition='auto'
        ))
        fig_hm.update_layout(
            title="Baseline-to-Midline Change",
            height=max(400, len(heatmap_df) * 32),
            xaxis_title="Change (pp)",
            yaxis=dict(categoryorder='trace'),
            font=dict(size=13, color='#333'),
            title_font=dict(size=16, color='#222'),
            plot_bgcolor='rgba(0,0,0,0)',
            paper_bgcolor='rgba(0,0,0,0)',
            margin=dict(l=20, r=20, t=60, b=20),
        )
        st.plotly_chart(fig_hm, use_container_width=True)

    with hm_col2:
        # Waterfall chart — cumulative gain/loss
        wf_sorted = heatmap_df.sort_values('Change', ascending=False).reset_index(drop=True)
        wf_colors = [COLORS['good'] if v >= 0 else COLORS['danger'] for v in wf_sorted['Change']]
        fig_wf = go.Figure(go.Waterfall(
            x=wf_sorted['Label'],
            y=wf_sorted['Change'],
            measure=['relative'] * len(wf_sorted),
            text=wf_sorted['Change'].apply(lambda x: f"{x:+.1f}"),
            textposition='outside',
            increasing=dict(marker=dict(color=COLORS['good'])),
            decreasing=dict(marker=dict(color=COLORS['danger'])),
            connector=dict(line=dict(color='#ccc', width=1)),
        ))
        fig_wf.update_layout(
            title="Cumulative Change Waterfall",
            height=max(400, len(heatmap_df) * 32),
            yaxis_title="Change (pp)",
            font=dict(size=13, color='#333'),
            title_font=dict(size=16, color='#222'),
            plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
            margin=dict(l=20, r=20, t=60, b=60),
            xaxis=dict(tickangle=-45),
        )
        st.plotly_chart(fig_wf, use_container_width=True)

    # Data table
    with st.expander("View Full Indicator Change Table"):
        display_df = heatmap_df[['Indicator', 'Dataset', 'Baseline', 'Midline', 'Change']].copy()
        display_df.columns = ['Indicator', 'Dataset', 'Baseline', 'Midline', 'Change (pp)']
        st.dataframe(
            display_df.style.map(
                lambda v: 'color: #2E7D32; font-weight: 700' if isinstance(v, (int, float)) and v > 0
                else ('color: #C62828; font-weight: 700' if isinstance(v, (int, float)) and v < 0 else ''),
                subset=['Change (pp)']
            ),
            use_container_width=True,
            hide_index=True,
        )

    st.markdown("---")

    # ====================================================================
    # KEY RECOMMENDATIONS
    # ====================================================================
    _section_header('', 'Key Recommendations', 'Action Items')

    # Auto-generate recommendations based on the insights
    recs = []
    # Check for negative trends
    all_insights_recs = f_insights + w_insights + m_insights + cc_insights
    neg_areas = [title for title, _, trend in all_insights_recs if trend in ('negative', 'warning')]
    pos_areas = [title for title, _, trend in all_insights_recs if trend == 'positive']

    if any('Functionality' in a for a in neg_areas):
        recs.append(("Strengthen Institutional Capacity",
                     "Targeted support for groups below the functionality threshold — mentorship, "
                     "refresher training on governance, and structured follow-up visits."))
    if any('Leadership' in a or 'Gender' in a for a in neg_areas):
        recs.append(("Accelerate Gender Mainstreaming",
                     "Prioritize women's leadership development within forestry groups, and ensure "
                     "gender considerations are embedded in all planning processes."))
    if any('Savings' in a or 'Income' in a or 'Economic' in a for a in neg_areas):
        recs.append(("Expand Financial Inclusion",
                     "Link VSLA/savings groups to the forest-income value chain. Support women's "
                     "personal savings through financial literacy training and digital payment platforms."))
    if any('Norm' in a for a in neg_areas):
        recs.append(("Address Harmful Social Norms",
                     "Continue community dialogues on gender norms, using evidence from this dashboard "
                     "to show how norms affect women's participation and community outcomes."))
    if any('Care Work' in a or 'Time' in a for a in neg_areas):
        recs.append(("Reduce Women's Time Poverty",
                     "Invest in time-saving technologies (water access, energy-efficient stoves) "
                     "and promote equitable household labour-sharing norms."))
    if any('Prepared' in a or 'Climate' in a or 'NbS' in a for a in neg_areas):
        recs.append(("Scale Climate & DRR Programming",
                     "Expand community-based disaster preparedness planning and integrate NbS training "
                     "into both forestry group activities and women's empowerment programs."))
    if any("Men's" in a and ('Norm' in a or 'Support' in a or 'Role' in a) for a in neg_areas):
        recs.append(("Engage Men as Change Agents",
                     "Strengthen male engagement strategies — community dialogues, role-model campaigns, "
                     "and couple-based interventions to shift attitudes on joint roles and NbS support."))
    if any("Decision" in a and "Practice" in a for a in neg_areas):
        recs.append(("Close the Attitudes-Practice Gap",
                     "While men increasingly say decisions should be joint, actual practice lags behind. "
                     "Facilitate structured household visioning exercises to translate norms into behaviour."))

    # Always add these broad recommendations
    recs.append(("Leverage Positive Trends",
                 f"Build on the {len(pos_areas)} positive trends identified — use these as evidence for "
                 f"donor reporting and community motivation. Document success stories in areas "
                 f"showing clear improvement."))
    recs.append(("Integrated M&E Approach",
                 "Continue tracking community-level (forestry), women's, and men's indicators together. "
                 "The cross-cutting insights show that progress at one level supports the other."))

    for title, body in recs:
        _insight_card(title, body, "neutral")


# ============================================================================
# CROSS-DATASET SYNTHESIS VIEW
# ============================================================================

def render_synthesis_view(f_data, w_data, m_data=None, gjj_data=None, gjj_men_data=None, ft_data=None, mg_data=None, sw_data=None):
    """Combined overview of all datasets — key headline indicators."""
    st.markdown("""<div class="section-narrative">
    <strong> Cross-Dataset Synthesis:</strong> A combined overview comparing headline indicators
    from all programme datasets — Forestry Conservation Groups (community-level), Women's Survey,
    Men's Survey (household-level), GJJ KAP Women, GJJ KAP Men (Baseline/Endline), Forest
    Training, Mangrove Training (Pre/Post), and Seaweed Production. This view highlights key
    programme-wide trends.
    </div>""", unsafe_allow_html=True)

    # ---- Forestry Headline KPIs ----
    st.markdown('<h3 style="margin-top:0.5rem;"> Forestry Conservation Groups — Headlines</h3>',
                unsafe_allow_html=True)
    ft = f_data['functionality_threshold']
    fs = f_data['functionality_scores']
    bl_60 = ft.loc[ft['Timepoint']=='Baseline','Functional_60_pct'].values[0]
    ml_60 = ft.loc[ft['Timepoint']=='Midline','Functional_60_pct'].values[0]
    bl_avg = fs.loc[fs['Timepoint']=='Baseline','Average'].values[0]
    ml_avg = fs.loc[fs['Timepoint']=='Midline','Average'].values[0]
    bl_grp = f_data['num_groups'].loc[f_data['num_groups']['Timepoint']=='Baseline','Groups_Assessed'].values[0]
    ml_grp = f_data['num_groups'].loc[f_data['num_groups']['Timepoint']=='Midline','Groups_Assessed'].values[0]

    fc1, fc2, fc3, fc4 = st.columns(4)
    fc1.markdown(f"""<div class="kpi-card">
        <h3>Groups Assessed</h3>
        <div class="value">{int(ml_grp)}</div>
        <div class="delta-{'positive' if ml_grp>=bl_grp else 'negative'}">{int(ml_grp-bl_grp):+d} from BL</div>
    </div>""", unsafe_allow_html=True)
    fc2.markdown(f"""<div class="kpi-card">
        <h3>Functional ≥60%</h3>
        <div class="value">{ml_60*100:.1f}%</div>
        <div class="delta-{'positive' if ml_60>=bl_60 else 'negative'}">{(ml_60-bl_60)*100:+.1f}pp</div>
    </div>""", unsafe_allow_html=True)
    fc3.markdown(f"""<div class="kpi-card">
        <h3>Overall Score</h3>
        <div class="value">{ml_avg:.1f}/100</div>
        <div class="delta-{'positive' if ml_avg>=bl_avg else 'negative'}">{ml_avg-bl_avg:+.1f} pts</div>
    </div>""", unsafe_allow_html=True)
    # Agroforestry uptake
    af_bl = f_data['agroforestry'].loc[f_data['agroforestry']['Category']=='Yes','Baseline'].values[0]
    af_ml = f_data['agroforestry'].loc[f_data['agroforestry']['Category']=='Yes','Midline'].values[0]
    fc4.markdown(f"""<div class="kpi-card">
        <h3>Agroforestry Yes</h3>
        <div class="value">{af_ml*100:.0f}%</div>
        <div class="delta-{'positive' if af_ml>=af_bl else 'negative'}">{(af_ml-af_bl)*100:+.1f}pp</div>
    </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ---- Women Survey Headline KPIs ----
    st.markdown('<h3> Women\'s Survey — Headlines</h3>', unsafe_allow_html=True)

    # CC heard
    cc_bl = w_data['cc_heard'].loc[w_data['cc_heard']['Response']=='Yes','Baseline'].values[0] if len(w_data['cc_heard'][w_data['cc_heard']['Response']=='Yes']) else 0
    cc_ml = w_data['cc_heard'].loc[w_data['cc_heard']['Response']=='Yes','Midline'].values[0] if len(w_data['cc_heard'][w_data['cc_heard']['Response']=='Yes']) else 0
    # NbS heard
    nbs_bl = w_data['nbs_heard'].loc[w_data['nbs_heard']['Response']=='Yes','Baseline'].values[0] if len(w_data['nbs_heard'][w_data['nbs_heard']['Response']=='Yes']) else 0
    nbs_ml = w_data['nbs_heard'].loc[w_data['nbs_heard']['Response']=='Yes','Midline'].values[0] if len(w_data['nbs_heard'][w_data['nbs_heard']['Response']=='Yes']) else 0
    # Personal saving
    sav_bl = w_data['personal_saving'].loc[w_data['personal_saving']['Response']=='Yes','Baseline'].values[0] if len(w_data['personal_saving'][w_data['personal_saving']['Response']=='Yes']) else 0
    sav_ml = w_data['personal_saving'].loc[w_data['personal_saving']['Response']=='Yes','Midline'].values[0] if len(w_data['personal_saving'][w_data['personal_saving']['Response']=='Yes']) else 0
    # Life skills average
    ls_bl = w_data['lifeskills_agree']['Baseline'].mean()
    ls_ml = w_data['lifeskills_agree']['Midline'].mean()

    wc1, wc2, wc3, wc4 = st.columns(4)
    wc1.markdown(f"""<div class="kpi-card">
        <h3>CC Awareness</h3>
        <div class="value">{cc_ml*100:.0f}%</div>
        <div class="delta-{'positive' if cc_ml>=cc_bl else 'negative'}">{(cc_ml-cc_bl)*100:+.1f}pp</div>
    </div>""", unsafe_allow_html=True)
    wc2.markdown(f"""<div class="kpi-card">
        <h3>NbS Awareness</h3>
        <div class="value">{nbs_ml*100:.0f}%</div>
        <div class="delta-{'positive' if nbs_ml>=nbs_bl else 'negative'}">{(nbs_ml-nbs_bl)*100:+.1f}pp</div>
    </div>""", unsafe_allow_html=True)
    wc3.markdown(f"""<div class="kpi-card">
        <h3>Women Saving</h3>
        <div class="value">{sav_ml*100:.0f}%</div>
        <div class="delta-{'positive' if sav_ml>=sav_bl else 'negative'}">{(sav_ml-sav_bl)*100:+.1f}pp</div>
    </div>""", unsafe_allow_html=True)
    wc4.markdown(f"""<div class="kpi-card">
        <h3>Life Skills (Avg)</h3>
        <div class="value">{ls_ml*100:.0f}%</div>
        <div class="delta-{'positive' if ls_ml>=ls_bl else 'negative'}">{(ls_ml-ls_bl)*100:+.1f}pp</div>
    </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ---- Men Survey Headline KPIs ----
    if m_data is not None:
        st.markdown('<h3> Men\'s Survey — Headlines</h3>', unsafe_allow_html=True)

        # CC heard
        m_cc_bl = m_data['cc_heard'].loc[m_data['cc_heard']['Response']=='Yes','Baseline'].values[0]
        m_cc_ml = m_data['cc_heard'].loc[m_data['cc_heard']['Response']=='Yes','Midline'].values[0]
        # NbS heard
        m_nbs_bl = m_data['nbs_heard'].loc[m_data['nbs_heard']['Response']=='Yes','Baseline'].values[0]
        m_nbs_ml = m_data['nbs_heard'].loc[m_data['nbs_heard']['Response']=='Yes','Midline'].values[0]
        # Roles should joint avg
        m_rj_bl = m_data['roles_should_joint']['Baseline'].mean()
        m_rj_ml = m_data['roles_should_joint']['Midline'].mean()
        # Decision should joint avg
        m_dj_bl = m_data['decision_should_joint']['Baseline'].mean()
        m_dj_ml = m_data['decision_should_joint']['Midline'].mean()

        mc1, mc2, mc3, mc4 = st.columns(4)
        mc1.markdown(f"""<div class="kpi-card">
            <h3>CC Awareness (Men)</h3>
            <div class="value">{m_cc_ml*100:.0f}%</div>
            <div class="delta-{'positive' if m_cc_ml>=m_cc_bl else 'negative'}">{(m_cc_ml-m_cc_bl)*100:+.1f}pp</div>
        </div>""", unsafe_allow_html=True)
        mc2.markdown(f"""<div class="kpi-card">
            <h3>NbS Awareness (Men)</h3>
            <div class="value">{m_nbs_ml*100:.0f}%</div>
            <div class="delta-{'positive' if m_nbs_ml>=m_nbs_bl else 'negative'}">{(m_nbs_ml-m_nbs_bl)*100:+.1f}pp</div>
        </div>""", unsafe_allow_html=True)
        mc3.markdown(f"""<div class="kpi-card">
            <h3>Joint Roles (Should)</h3>
            <div class="value">{m_rj_ml*100:.0f}%</div>
            <div class="delta-{'positive' if m_rj_ml>=m_rj_bl else 'negative'}">{(m_rj_ml-m_rj_bl)*100:+.1f}pp</div>
        </div>""", unsafe_allow_html=True)
        mc4.markdown(f"""<div class="kpi-card">
            <h3>Joint Decisions (Should)</h3>
            <div class="value">{m_dj_ml*100:.0f}%</div>
            <div class="delta-{'positive' if m_dj_ml>=m_dj_bl else 'negative'}">{(m_dj_ml-m_dj_bl)*100:+.1f}pp</div>
        </div>""", unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

    # ---- GJJ KAP Women Headline KPIs ----
    if gjj_data is not None:
        st.markdown('<h3> GJJ KAP Women — Headlines (Baseline/Endline)</h3>', unsafe_allow_html=True)
        try:
            sa = gjj_data['self_strongly_agree']
            sa_bl = float(sa['Baseline'].mean())
            sa_el = float(sa['Endline'].mean())
            af_rn = gjj_data['rel_af_rn']
            af_bl = float(af_rn['AF_Baseline'].mean())
            af_el = float(af_rn['AF_Endline'].mean())
            chores = gjj_data['shared_chores_yn']
            ch_yes = chores[chores['Response'].str.strip().str.lower() == 'yes']
            ch_bl = float(ch_yes['Baseline'].values[0]) if len(ch_yes) else 0.0
            ch_el = float(ch_yes['Endline'].values[0]) if len(ch_yes) else 0.0
            eq = gjj_data['equal_say']
            eq_yes = eq[eq['Response'].str.strip().str.lower() == 'yes']
            eq_bl = float(eq_yes['Baseline'].values[0]) if len(eq_yes) else 0.0
            eq_el = float(eq_yes['Endline'].values[0]) if len(eq_yes) else 0.0

            gc1, gc2, gc3, gc4 = st.columns(4)
            gc1.markdown(f"""<div class="kpi-card">
                <h3>Self-Esteem (SA)</h3>
                <div class="value">{sa_el*100:.0f}%</div>
                <div class="delta-{'positive' if sa_el>=sa_bl else 'negative'}">{(sa_el-sa_bl)*100:+.1f}pp</div>
            </div>""", unsafe_allow_html=True)
            gc2.markdown(f"""<div class="kpi-card">
                <h3>Relational (AF)</h3>
                <div class="value">{af_el*100:.0f}%</div>
                <div class="delta-{'positive' if af_el>=af_bl else 'negative'}">{(af_el-af_bl)*100:+.1f}pp</div>
            </div>""", unsafe_allow_html=True)
            gc3.markdown(f"""<div class="kpi-card">
                <h3>Husband Chores</h3>
                <div class="value">{ch_el*100:.0f}%</div>
                <div class="delta-{'positive' if ch_el>=ch_bl else 'negative'}">{(ch_el-ch_bl)*100:+.1f}pp</div>
            </div>""", unsafe_allow_html=True)
            gc4.markdown(f"""<div class="kpi-card">
                <h3>Equal Say</h3>
                <div class="value">{eq_el*100:.0f}%</div>
                <div class="delta-{'positive' if eq_el>=eq_bl else 'negative'}">{(eq_el-eq_bl)*100:+.1f}pp</div>
            </div>""", unsafe_allow_html=True)
        except Exception:
            st.info("Some GJJ KAP Women headline KPIs could not be computed.")

        st.markdown("<br>", unsafe_allow_html=True)

    # ---- GJJ KAP Men Headline KPIs ----
    if gjj_men_data is not None:
        st.markdown('<h3> GJJ KAP Men — Headlines (Baseline/Endline)</h3>', unsafe_allow_html=True)
        try:
            agr = gjj_men_data['self_agreement']
            agr_bl = float(agr['Agreement_BL'].mean())
            agr_el = float(agr['Agreement_EL'].mean())
            af_rn = gjj_men_data['rel_af_rn']
            af_bl_m = float(af_rn['AF_Baseline'].mean())
            af_el_m = float(af_rn['AF_Endline'].mean())
            ch = gjj_men_data['shared_chores_yn']
            ch_yes = ch[ch['Response'].str.strip().str.lower() == 'yes']
            ch_bl_m = float(ch_yes['Baseline'].values[0]) if len(ch_yes) else 0.0
            ch_el_m = float(ch_yes['Endline'].values[0]) if len(ch_yes) else 0.0
            dc = gjj_men_data['decision_conversations']
            dc_yes = dc[dc['Response'].str.strip().str.lower() == 'yes']
            dc_bl_m = float(dc_yes['Baseline'].values[0]) if len(dc_yes) else 0.0
            dc_el_m = float(dc_yes['Endline'].values[0]) if len(dc_yes) else 0.0

            gmc1, gmc2, gmc3, gmc4 = st.columns(4)
            gmc1.markdown(f"""<div class="kpi-card">
                <h3>Self-Esteem (Men)</h3>
                <div class="value">{agr_el*100:.0f}%</div>
                <div class="delta-{'positive' if agr_el>=agr_bl else 'negative'}">{(agr_el-agr_bl)*100:+.1f}pp</div>
            </div>""", unsafe_allow_html=True)
            gmc2.markdown(f"""<div class="kpi-card">
                <h3>Relational AF (Men)</h3>
                <div class="value">{af_el_m*100:.0f}%</div>
                <div class="delta-{'positive' if af_el_m>=af_bl_m else 'negative'}">{(af_el_m-af_bl_m)*100:+.1f}pp</div>
            </div>""", unsafe_allow_html=True)
            gmc3.markdown(f"""<div class="kpi-card">
                <h3>Supports Chores (Men)</h3>
                <div class="value">{ch_el_m*100:.0f}%</div>
                <div class="delta-{'positive' if ch_el_m>=ch_bl_m else 'negative'}">{(ch_el_m-ch_bl_m)*100:+.1f}pp</div>
            </div>""", unsafe_allow_html=True)
            gmc4.markdown(f"""<div class="kpi-card">
                <h3>Decision Convos (Men)</h3>
                <div class="value">{dc_el_m*100:.0f}%</div>
                <div class="delta-{'positive' if dc_el_m>=dc_bl_m else 'negative'}">{(dc_el_m-dc_bl_m)*100:+.1f}pp</div>
            </div>""", unsafe_allow_html=True)
        except Exception:
            st.info("Some GJJ KAP Men headline KPIs could not be computed.")

        st.markdown("<br>", unsafe_allow_html=True)

    # ---- Forest Training Headline KPIs ----
    if ft_data is not None:
        st.markdown('<h3> Forest Training — Headlines (Pre/Post)</h3>', unsafe_allow_html=True)
        try:
            t_scores = ft_data['scores']
            t_thresh = ft_data['thresholds']
            t_questions = ft_data['questions']

            t_bl_avg = float(t_scores[t_scores['Timepoint']=='Baseline']['AverageScore'].values[0])
            t_el_avg = float(t_scores[t_scores['Timepoint']=='Endline']['AverageScore'].values[0])

            t70 = t_thresh[t_thresh['Threshold'].str.contains('70')]
            t70_bl = float(t70[t70['Timepoint']=='Baseline']['Proportion'].values[0])
            t70_el = float(t70[t70['Timepoint']=='Endline']['Proportion'].values[0])

            t_bl_n = int(t_scores[t_scores['Timepoint']=='Baseline']['Respondents'].values[0])
            t_el_n = int(t_scores[t_scores['Timepoint']=='Endline']['Respondents'].values[0])

            t_avg_correct_bl = float(t_questions['Baseline'].mean()) * 100 if not t_questions.empty else 0
            t_avg_correct_el = float(t_questions['Endline'].mean()) * 100 if not t_questions.empty else 0

            ftc1, ftc2, ftc3, ftc4 = st.columns(4)
            ftc1.markdown(f"""<div class="kpi-card">
                <h3>\u226570% Pass (PMF)</h3>
                <div class="value">{t70_el*100:.1f}%</div>
                <div class="delta-{'positive' if t70_el>=t70_bl else 'negative'}">{(t70_el-t70_bl)*100:+.1f}pp</div>
            </div>""", unsafe_allow_html=True)
            ftc2.markdown(f"""<div class="kpi-card">
                <h3>Average Score</h3>
                <div class="value">{t_el_avg:.1f}%</div>
                <div class="delta-{'positive' if t_el_avg>=t_bl_avg else 'negative'}">{t_el_avg-t_bl_avg:+.1f}pp</div>
            </div>""", unsafe_allow_html=True)
            ftc3.markdown(f"""<div class="kpi-card">
                <h3>Avg % Correct</h3>
                <div class="value">{t_avg_correct_el:.1f}%</div>
                <div class="delta-{'positive' if t_avg_correct_el>=t_avg_correct_bl else 'negative'}">{t_avg_correct_el-t_avg_correct_bl:+.1f}pp</div>
            </div>""", unsafe_allow_html=True)
            ftc4.markdown(f"""<div class="kpi-card">
                <h3>Post-Training N</h3>
                <div class="value">{t_el_n:,}</div>
                <div class="delta-positive">{t_el_n - t_bl_n:+,} vs Pre</div>
            </div>""", unsafe_allow_html=True)
        except Exception:
            st.info("Some Forest Training headline KPIs could not be computed.")

        st.markdown("<br>", unsafe_allow_html=True)

    # ---- Mangrove Training Headline KPIs ----
    if mg_data is not None:
        st.markdown("---")
        st.markdown('<h3>🌊 Mangrove Training (Pre/Post)</h3>', unsafe_allow_html=True)
        try:
            mg_county = mg_data['adequate_county']
            mg_scores = mg_data['scores']
            mg_all = mg_county[mg_county['County'] == 'All']
            mg_pre_val = float(mg_all[mg_all['Timepoint'] == 'Pre-Test']['Value'].values[0]) * 100
            mg_post_val = float(mg_all[mg_all['Timepoint'] == 'Post-Test']['Value'].values[0]) * 100
            mg_change = mg_post_val - mg_pre_val
            mg_all_scores = mg_scores[mg_scores['County'] == 'All']
            mg_avg = float(mg_all_scores['AvgScore'].values[0]) if len(mg_all_scores) else 0
            mg_n = int(mg_all_scores['Respondents'].values[0]) if len(mg_all_scores) else 0

            mk1, mk2, mk3, mk4 = st.columns(4)
            mk1.markdown(f"""<div class="kpi-card">
                <h3>≥60% Pass (Post)</h3>
                <div class="value">{mg_post_val:.1f}%</div>
                <div class="delta-{'positive' if mg_change>0 else 'negative'}">{mg_change:+.1f}pp</div>
            </div>""", unsafe_allow_html=True)
            mk2.markdown(f"""<div class="kpi-card">
                <h3>≥60% Pass (Pre)</h3>
                <div class="value">{mg_pre_val:.1f}%</div>
                <div class="delta-neutral">Baseline</div>
            </div>""", unsafe_allow_html=True)
            mk3.markdown(f"""<div class="kpi-card">
                <h3>Avg Score</h3>
                <div class="value">{mg_avg:.1f}%</div>
                <div class="delta-neutral">Overall</div>
            </div>""", unsafe_allow_html=True)
            mk4.markdown(f"""<div class="kpi-card">
                <h3>Respondents</h3>
                <div class="value">{mg_n:,}</div>
                <div class="delta-neutral">All counties</div>
            </div>""", unsafe_allow_html=True)
        except Exception:
            st.info("Some Mangrove Training headline KPIs could not be computed.")

        st.markdown("<br>", unsafe_allow_html=True)

    # ---- Seaweed Production Headline KPIs ----
    if sw_data is not None:
        st.markdown("---")
        st.markdown('<h3>\U0001f33f Seaweed Production & Challenges (2025)</h3>', unsafe_allow_html=True)
        try:
            sw_agg = prepare_seaweed_aggregates(sw_data)
            sw_ov = sw_agg['overall']
            sw_grp = sw_agg['group_summary']
            sw_ch = sw_agg['challenge_counts']
            sw_avg_ach = sw_data['Ropes_Achievement_pct'].mean()

            # Row 1: 5 KPI cards
            sk1, sk2, sk3, sk4, sk5 = st.columns(5)
            sk1.markdown(f"""<div class="kpi-card">
                <h3>Total Production</h3>
                <div class="value">{sw_ov['total_kg']:,.0f} kg</div>
                <div class="delta-neutral">{sw_ov['n_farmers']:,} farmers</div>
            </div>""", unsafe_allow_html=True)
            sk2.markdown(f"""<div class="kpi-card">
                <h3>Ropes in Ocean</h3>
                <div class="value">{sw_ov['ropes_ocean']:,.0f}</div>
                <div class="delta-neutral">of {sw_ov['ropes_total']:,.0f} total</div>
            </div>""", unsafe_allow_html=True)
            sk3.markdown(f"""<div class="kpi-card">
                <h3>Avg Production/Rope</h3>
                <div class="value">{sw_ov['avg_prod_per_rope']:.2f} kg</div>
                <div class="delta-neutral">{sw_ov['avg_production_per_farmer']:.1f} kg/farmer</div>
            </div>""", unsafe_allow_html=True)
            sk4.markdown(f"""<div class="kpi-card">
                <h3>Target Achievement</h3>
                <div class="value">{sw_avg_ach:.1f}%</div>
                <div class="delta-{'positive' if sw_avg_ach >= 70 else 'neutral'}">
                    {'On track' if sw_avg_ach >= 70 else 'Below target'}</div>
            </div>""", unsafe_allow_html=True)
            sk5.markdown(f"""<div class="kpi-card">
                <h3>Rope Gap</h3>
                <div class="value">{sw_ov['gap_total']:,.0f}</div>
                <div class="delta-negative">ropes needed</div>
            </div>""", unsafe_allow_html=True)

            # Row 2: 3 more KPI cards
            sr1, sr2, sr3 = st.columns(3)
            sr1.markdown(f"""<div class="kpi-card">
                <h3>Dried/Wet Ratio</h3>
                <div class="value">{sw_ov['dried_wet_ratio']:.2f}</div>
                <div class="delta-neutral">{sw_ov['dried_kg']:,.0f} vs {sw_ov['wet_kg']:,.0f}</div>
            </div>""", unsafe_allow_html=True)
            sr2.markdown(f"""<div class="kpi-card">
                <h3>Meet Target</h3>
                <div class="value">{sw_ov['pct_meeting_target']:.0f}%</div>
                <div class="delta-{'positive' if sw_ov['pct_meeting_target'] >= 50 else 'neutral'}">of farmers</div>
            </div>""", unsafe_allow_html=True)
            sr3.markdown(f"""<div class="kpi-card">
                <h3>Multi-Challenge</h3>
                <div class="value">{sw_ov['multi_challenge_pct']:.1f}%</div>
                <div class="delta-{'negative' if sw_ov['multi_challenge_pct'] > 30 else 'neutral'}">face 2+ challenges</div>
            </div>""", unsafe_allow_html=True)

            # Mini charts: production + top challenge
            swc1, swc2 = st.columns(2)
            with swc1:
                fig_sw_mini = go.Figure(go.Bar(
                    x=sw_grp.sort_values('Total_KG', ascending=False)['Group'],
                    y=sw_grp.sort_values('Total_KG', ascending=False)['Total_KG'],
                    marker_color=COLORS['baseline'],
                    text=sw_grp.sort_values('Total_KG', ascending=False)['Total_KG'].apply(lambda v: f"{v:,.0f}"),
                    textposition='auto',
                ))
                fig_sw_mini.update_layout(
                    title='Production by Group (kg)', height=300,
                    yaxis_title='Total KG', showlegend=False,
                    font=dict(size=11, color='#333'), title_font=dict(size=14, color='#222'),
                    plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                    margin=dict(l=10, r=10, t=40, b=10), xaxis_tickangle=-25,
                )
                st.plotly_chart(fig_sw_mini, use_container_width=True)
            with swc2:
                if not sw_ch.empty:
                    fig_sw_ch_mini = go.Figure(go.Bar(
                        x=sw_ch.sort_values('Pct', ascending=True)['Pct'],
                        y=sw_ch.sort_values('Pct', ascending=True)['Challenge'],
                        orientation='h', marker_color='#FF9800',
                        text=sw_ch.sort_values('Pct', ascending=True)['Pct'].apply(lambda v: f"{v:.0f}%"),
                        textposition='auto',
                    ))
                    fig_sw_ch_mini.update_layout(
                        title='Challenges (% farmers)', height=300,
                        xaxis_title='%', showlegend=False,
                        font=dict(size=11, color='#333'), title_font=dict(size=14, color='#222'),
                        plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                        margin=dict(l=10, r=10, t=40, b=10),
                    )
                    st.plotly_chart(fig_sw_ch_mini, use_container_width=True)
        except Exception:
            st.info("Some Seaweed headline KPIs could not be computed.")

        st.markdown("<br>", unsafe_allow_html=True)

    # Side-by-side mini charts
    st.markdown('<h3>Comparative Snapshots</h3>', unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    with c1:
        # Forestry domains radar
        fd = f_data['functionality_domain']
        domain_df = pd.DataFrame({
            'Domain': ['Management','Gender','Effectiveness','Overall'],
            'Baseline': fd[['Management','Gender','Effectiveness','Overall']].iloc[0].values,
            'Midline': fd[['Management','Gender','Effectiveness','Overall']].iloc[1].values,
        })
        fig_r = go.Figure()
        fig_r.add_trace(go.Scatterpolar(
            r=list(domain_df['Baseline'])+[domain_df['Baseline'].iloc[0]],
            theta=list(domain_df['Domain'])+[domain_df['Domain'].iloc[0]],
            fill='toself', name='Baseline',
            fillcolor=COLORS['radar_bl_fill'], line=dict(color=COLORS['baseline'])))
        fig_r.add_trace(go.Scatterpolar(
            r=list(domain_df['Midline'])+[domain_df['Midline'].iloc[0]],
            theta=list(domain_df['Domain'])+[domain_df['Domain'].iloc[0]],
            fill='toself', name='Midline',
            fillcolor=COLORS['radar_ml_fill'], line=dict(color=COLORS['midline'])))
        fig_r.update_layout(title="Forestry Domain Performance",
                            polar=dict(radialaxis=dict(visible=True, range=[0,100])),
                            height=380, legend=dict(orientation='h', yanchor='bottom', y=-0.15),
                            font=dict(size=13, color='#333'), title_font=dict(size=16, color='#222'))
        st.plotly_chart(fig_r, use_container_width=True)
    with c2:
        # Women's life-skills radar
        ls_a = w_data['lifeskills_agree']
        ls_domain = ls_a.groupby('Domain')[['Baseline','Midline']].mean().reset_index()
        domains_l = list(ls_domain['Domain']) + [ls_domain['Domain'].iloc[0]]
        bl_v = list(ls_domain['Baseline']*100) + [ls_domain['Baseline'].iloc[0]*100]
        ml_v = list(ls_domain['Midline']*100) + [ls_domain['Midline'].iloc[0]*100]
        fig_ls = go.Figure()
        fig_ls.add_trace(go.Scatterpolar(
            r=bl_v, theta=domains_l, fill='toself', name='Baseline',
            fillcolor=COLORS['radar_bl_fill'], line=dict(color=COLORS['baseline'])))
        fig_ls.add_trace(go.Scatterpolar(
            r=ml_v, theta=domains_l, fill='toself', name='Midline',
            fillcolor=COLORS['radar_ml_fill'], line=dict(color=COLORS['midline'])))
        fig_ls.update_layout(title="Women's Life Skills Domains",
                             polar=dict(radialaxis=dict(visible=True, range=[0,100])),
                             height=380, legend=dict(orientation='h', yanchor='bottom', y=-0.15),
                             font=dict(size=13, color='#333'), title_font=dict(size=16, color='#222'))
        st.plotly_chart(fig_ls, use_container_width=True)

    # Social norms mini-bar
    c3, c4 = st.columns(2)
    with c3:
        sn = w_data['socialnorms_agree'].copy()
        sn['Change'] = (sn['Midline'] - sn['Baseline'])*100
        sn_sorted = sn.sort_values('Change')
        colors_sn = [COLORS['good'] if v <= 0 else COLORS['danger'] for v in sn_sorted['Change']]
        fig_sn = go.Figure()
        fig_sn.add_trace(go.Bar(y=sn_sorted['Norm'], x=sn_sorted['Change'], orientation='h',
                                marker_color=colors_sn,
                                text=sn_sorted['Change'].apply(lambda x: f"{x:+.1f}pp"), textposition='auto'))
        fig_sn.update_layout(title="Social Norms Change (↓ = positive)", height=400,
                             xaxis_title="Change (pp)",
                             font=dict(size=13, color='#333'), title_font=dict(size=16, color='#222'),
                             plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)')
        st.plotly_chart(fig_sn, use_container_width=True)
    with c4:
        ts = w_data['time_summary'].copy()
        fig_ts = go.Figure()
        fig_ts.add_trace(go.Bar(x=ts['Category'], y=ts['Baseline'], name='Baseline',
                                marker_color=COLORS['baseline'],
                                text=ts['Baseline'].apply(lambda x: f"{x:.1f}h"), textposition='auto'))
        fig_ts.add_trace(go.Bar(x=ts['Category'], y=ts['Midline'], name='Midline',
                                marker_color=COLORS['midline'],
                                text=ts['Midline'].apply(lambda x: f"{x:.1f}h"), textposition='auto'))
        fig_ts.update_layout(title="Women's Time Use (hrs/day)", barmode='group', height=400,
                             yaxis_title='Hours',
                             legend=dict(orientation='h', yanchor='bottom', y=1.02),
                             font=dict(size=13, color='#333'), title_font=dict(size=16, color='#222'),
                             plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)')
        st.plotly_chart(fig_ts, use_container_width=True)

    # ---- Men vs Women comparison charts ----
    if m_data is not None:
        st.markdown('<h3>Men vs Women: Key Comparisons</h3>', unsafe_allow_html=True)
        cmp1, cmp2 = st.columns(2)
        with cmp1:
            # CC awareness comparison
            w_cc_yes_ml = cc_ml  # already computed above
            m_cc_yes_ml = m_data['cc_heard'].loc[m_data['cc_heard']['Response']=='Yes','Midline'].values[0]
            w_nbs_yes_ml = nbs_ml
            m_nbs_yes_ml = m_data['nbs_heard'].loc[m_data['nbs_heard']['Response']=='Yes','Midline'].values[0]
            cmp_df = pd.DataFrame({
                'Indicator': ['CC Awareness', 'NbS Awareness'],
                'Women (Midline)': [w_cc_yes_ml*100, w_nbs_yes_ml*100],
                'Men (Midline)': [m_cc_yes_ml*100, m_nbs_yes_ml*100],
            })
            fig_cmp = go.Figure()
            fig_cmp.add_trace(go.Bar(x=cmp_df['Indicator'], y=cmp_df['Women (Midline)'],
                                     name='Women', marker_color=COLORS['midline'],
                                     text=cmp_df['Women (Midline)'].apply(lambda x: f"{x:.1f}%"),
                                     textposition='auto'))
            fig_cmp.add_trace(go.Bar(x=cmp_df['Indicator'], y=cmp_df['Men (Midline)'],
                                     name='Men', marker_color=COLORS['baseline'],
                                     text=cmp_df['Men (Midline)'].apply(lambda x: f"{x:.1f}%"),
                                     textposition='auto'))
            fig_cmp.update_layout(title="Climate & NbS Awareness (Midline)", barmode='group',
                                  height=380, yaxis_title='%',
                                  legend=dict(orientation='h', yanchor='bottom', y=1.02),
                                  font=dict(size=13, color='#333'),
                                  title_font=dict(size=16, color='#222'),
                                  plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)')
            st.plotly_chart(fig_cmp, use_container_width=True)

        with cmp2:
            # Social norms comparison — harmful norms average
            w_sn = w_data['socialnorms_agree'].copy()
            w_harmful = w_sn[~w_sn['Norm'].str.contains('Planting Crops|Restoring Ecosystems|Express Emotions', regex=True)]
            w_sn_bl = w_harmful['Baseline'].mean() * 100
            w_sn_ml = w_harmful['Midline'].mean() * 100
            m_sn = m_data['socialnorms_agree'].copy()
            m_harmful = m_sn[~m_sn['Norm'].str.contains('Planting Crops|Restoring Ecosystems|Express Emotions', regex=True)]
            m_sn_bl = m_harmful['Baseline'].mean() * 100
            m_sn_ml = m_harmful['Midline'].mean() * 100
            norms_df = pd.DataFrame({
                'Timepoint': ['Baseline', 'Midline'],
                'Women': [w_sn_bl, w_sn_ml],
                'Men': [m_sn_bl, m_sn_ml],
            })
            fig_norms = go.Figure()
            fig_norms.add_trace(go.Bar(x=norms_df['Timepoint'], y=norms_df['Women'],
                                       name='Women', marker_color=COLORS['midline'],
                                       text=norms_df['Women'].apply(lambda x: f"{x:.1f}%"),
                                       textposition='auto'))
            fig_norms.add_trace(go.Bar(x=norms_df['Timepoint'], y=norms_df['Men'],
                                       name='Men', marker_color=COLORS['baseline'],
                                       text=norms_df['Men'].apply(lambda x: f"{x:.1f}%"),
                                       textposition='auto'))
            fig_norms.update_layout(title="Harmful Social Norms (Avg Agree %)", barmode='group',
                                    height=380, yaxis_title='Agree/Strongly Agree (%)',
                                    legend=dict(orientation='h', yanchor='bottom', y=1.02),
                                    font=dict(size=13, color='#333'),
                                    title_font=dict(size=16, color='#222'),
                                    plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)')
            st.plotly_chart(fig_norms, use_container_width=True)

    st.markdown("""<div class="section-narrative" style="margin-top:1rem;">
    <strong>Navigate deeper:</strong> Use the sidebar to select <em>Forestry Groups</em>,
    <em>Women Survey</em>, <em>Men Survey</em>, or <em>GJJ KAP Women</em> for detailed
    breakdowns across all thematic areas.
    </div>""", unsafe_allow_html=True)


# ============================================================================
# MAIN
# ============================================================================

def main():
    st.set_page_config(
        page_title="COSME M&E Dashboard",
        page_icon="bar_chart",
        layout="wide",
        initial_sidebar_state="expanded"
    )

    # ---- THEME ----
    theme_name = st.sidebar.selectbox("Dashboard Theme", list(THEMES.keys()), index=0,
                                       help="Color palette for all charts and UI elements")
    global COLORS
    COLORS = THEMES[theme_name]

    # ---- ENHANCED CSS ----
    st.markdown(f"""
    <style>
    /* ---------- FORCE LIGHT MODE ---------- */
    .stApp, .stApp > header, [data-testid="stAppViewContainer"],
    [data-testid="stHeader"], [data-testid="stToolbar"],
    .main, .main .block-container {{
        background-color: #ffffff !important;
        color: #333333 !important;
    }}
    [data-testid="stAppViewContainer"] > section > div {{
        background-color: #ffffff !important;
    }}

    /* ---------- GLOBAL TYPOGRAPHY ---------- */
    html, body, [class*="css"] {{
        font-family: 'Inter', 'Segoe UI', -apple-system, BlinkMacSystemFont, sans-serif;
        color: #333333;
    }}
    .block-container {{
        max-width: 1200px;
        padding: 1rem 2rem 2rem;
    }}

    /* ---------- HEADER ---------- */
    .main-header {{
        background: {COLORS['header_gradient']};
        color: white; padding: 2rem 2.5rem; border-radius: 14px;
        margin-bottom: 2rem; text-align: center;
        box-shadow: 0 4px 20px rgba(0,0,0,0.18);
        position: relative; overflow: hidden;
    }}
    .main-header::before {{
        content: ''; position: absolute; top: -50%; left: -50%;
        width: 200%; height: 200%; background:
            radial-gradient(circle at 20% 50%, rgba(255,255,255,0.08) 0%, transparent 50%),
            radial-gradient(circle at 80% 20%, rgba(255,255,255,0.05) 0%, transparent 40%);
        pointer-events: none;
    }}
    .main-header h1 {{ margin: 0; font-size: 2.4rem; font-weight: 800; letter-spacing: -0.02em;
        text-shadow: 0 2px 4px rgba(0,0,0,0.2); position: relative; }}
    .main-header p {{ margin: 0.5rem 0 0; opacity: 0.95; font-size: 1.1rem; font-weight: 500;
        position: relative; letter-spacing: 0.01em; }}

    /* ---------- KPI CARDS ---------- */
    .kpi-card {{
        background: white; border-radius: 12px; padding: 1.5rem 1.2rem;
        box-shadow: 0 2px 14px rgba(0,0,0,0.08); text-align: center;
        border-left: 5px solid {COLORS['card_border']};
        transition: transform 0.2s ease, box-shadow 0.2s ease;
        margin-bottom: 1rem;
    }}
    .kpi-card:hover {{ transform: translateY(-3px); box-shadow: 0 6px 24px rgba(0,0,0,0.14); }}
    .kpi-card h3 {{ font-size: 0.85rem; color: #555; margin: 0 0 0.5rem;
        text-transform: uppercase; letter-spacing: 0.05em; font-weight: 700; }}
    .kpi-card .value {{ font-size: 2.2rem; font-weight: 800; color: {COLORS['card_value']};
        line-height: 1.15; }}
    .delta-positive {{ color: {COLORS['good']}; font-weight: 700; font-size: 0.9rem; margin-top: 0.3rem; }}
    .delta-negative {{ color: {COLORS['danger']}; font-weight: 700; font-size: 0.9rem; margin-top: 0.3rem; }}

    /* ---------- SECTION NARRATIVE ---------- */
    .section-narrative {{
        background: {COLORS['narrative_bg']};
        border-left: 5px solid {COLORS['narrative_border']};
        padding: 1.2rem 1.5rem; border-radius: 0 10px 10px 0;
        margin-bottom: 1.4rem; font-size: 1rem;
        color: {COLORS['narrative_text']};
        line-height: 1.65;
        box-shadow: 0 1px 6px rgba(0,0,0,0.05);
    }}
    .section-narrative strong {{ font-size: 1.05rem; }}

    /* ---------- BREADCRUMB / NAV BAR ---------- */
    .nav-breadcrumb {{
        display: flex; align-items: center; gap: 0.5rem;
        padding: 0.7rem 1.2rem; background: rgba(0,0,0,0.035);
        border-radius: 8px; margin-bottom: 1.2rem;
        font-size: 0.92rem; color: #444; font-weight: 500;
        border: 1px solid rgba(0,0,0,0.07);
    }}
    .nav-breadcrumb .sep {{ color: #aaa; font-weight: 400; }}
    .nav-breadcrumb .active {{ font-weight: 700; color: {COLORS['card_value']}; }}

    /* ---------- QUICK NAV PILLS ---------- */
    .quick-nav {{
        display: flex; flex-wrap: wrap; gap: 0.5rem;
        margin-bottom: 1.2rem;
    }}
    .quick-nav-pill {{
        display: inline-block; padding: 0.4rem 1rem;
        background: {COLORS['narrative_bg']}; border: 1px solid {COLORS['narrative_border']};
        border-radius: 20px; font-size: 0.85rem; color: {COLORS['narrative_text']};
        font-weight: 600;
        cursor: pointer; transition: all 0.2s ease; text-decoration: none;
    }}
    .quick-nav-pill:hover {{ background: {COLORS['card_border']}; color: white; }}

    /* ---------- SECTION HEADER ---------- */
    .section-header {{
        display: flex; align-items: center; gap: 0.6rem;
        padding: 0.8rem 0; margin: 1.5rem 0 0.8rem;
        border-bottom: 3px solid {COLORS['card_border']};
    }}
    .section-header h2 {{ margin: 0; font-size: 1.45rem; font-weight: 700; color: {COLORS['card_value']}; }}
    .section-header .badge {{
        background: {COLORS['card_border']}; color: white;
        padding: 0.2rem 0.7rem; border-radius: 10px;
        font-size: 0.75rem; font-weight: 700; letter-spacing: 0.02em;
    }}

    /* ---------- DATA TABLE STYLING ---------- */
    .styled-table {{
        width: 100%; border-collapse: collapse; font-size: 0.92rem;
        border-radius: 8px; overflow: hidden;
        box-shadow: 0 2px 8px rgba(0,0,0,0.07);
    }}
    .styled-table th {{
        background: {COLORS['card_border']}; color: white;
        padding: 0.7rem 1rem; text-align: left; font-weight: 700;
        font-size: 0.9rem; letter-spacing: 0.02em;
    }}
    .styled-table td {{ padding: 0.6rem 1rem; border-bottom: 1px solid #e0e0e0; color: #333; font-weight: 500; }}
    .styled-table tr:hover td {{ background: {COLORS['narrative_bg']}; }}
    .styled-table tr:nth-child(even) td {{ background: rgba(0,0,0,0.015); }}

    /* ---------- DATAFRAME STYLING ---------- */
    div[data-testid="stDataFrame"] {{
        border: 1px solid rgba(0,0,0,0.08);
        border-radius: 8px;
        overflow: hidden;
    }}
    div[data-testid="stDataFrame"] table {{
        font-size: 0.9rem !important;
    }}

    /* ---------- EXPANDER STYLING ---------- */
    div[data-testid="stExpander"] {{
        border: 1px solid rgba(0,0,0,0.1); border-radius: 10px;
        box-shadow: 0 1px 6px rgba(0,0,0,0.05); margin-bottom: 0.8rem;
    }}
    div[data-testid="stExpander"] summary {{
        font-weight: 600 !important; font-size: 0.95rem !important;
        color: #333 !important;
    }}

    /* ---------- METRIC STYLING ---------- */
    div[data-testid="stMetricValue"] {{ font-size: 1.8rem !important; font-weight: 800 !important; color: #222 !important; }}
    div[data-testid="stMetricDelta"] {{ font-size: 0.9rem !important; font-weight: 600 !important; }}
    div[data-testid="stMetricLabel"] {{ font-size: 0.9rem !important; font-weight: 600 !important; color: #555 !important; }}

    /* ---------- TAB STYLING ---------- */
    button[data-baseweb="tab"] {{
        font-size: 0.95rem !important; font-weight: 700 !important;
        padding: 0.7rem 1.2rem !important;
        color: #444 !important;
    }}
    button[data-baseweb="tab"]:hover {{
        background: {COLORS['narrative_bg']} !important;
        color: {COLORS['card_value']} !important;
    }}
    button[data-baseweb="tab"][aria-selected="true"] {{
        color: {COLORS['card_value']} !important;
        border-bottom: 3px solid {COLORS['card_border']} !important;
    }}

    /* ---------- PLOTLY CHART CONTAINER ---------- */
    div[data-testid="stPlotlyChart"] {{
        border: 1px solid rgba(0,0,0,0.06);
        border-radius: 10px;
        padding: 0.5rem;
        margin-bottom: 1rem;
        background: white;
        box-shadow: 0 1px 6px rgba(0,0,0,0.04);
    }}

    /* ---------- SIDEBAR STYLING ---------- */
    section[data-testid="stSidebar"] {{
        background: linear-gradient(180deg, #fafafa 0%, #f0f2f6 100%);
    }}
    section[data-testid="stSidebar"] .block-container {{ padding-top: 1rem; }}
    section[data-testid="stSidebar"] label {{
        font-weight: 600 !important; color: #333 !important; font-size: 0.9rem !important;
    }}
    .sidebar-section {{
        background: white; border-radius: 10px; padding: 1rem 1.2rem;
        margin-bottom: 0.8rem; box-shadow: 0 1px 6px rgba(0,0,0,0.06);
        border: 1px solid rgba(0,0,0,0.07);
    }}
    .sidebar-nav-link {{
        display: block; padding: 0.45rem 0.7rem; margin: 0.2rem 0;
        border-radius: 6px; font-size: 0.88rem; color: #333; font-weight: 500;
        text-decoration: none; transition: all 0.15s ease;
    }}
    .sidebar-nav-link:hover {{ background: {COLORS['narrative_bg']}; color: {COLORS['card_value']}; font-weight: 600; }}

    /* ---------- MARKDOWN TEXT ---------- */
    .stMarkdown p {{ font-size: 0.95rem; color: #333; line-height: 1.65; }}
    .stMarkdown h1 {{ font-size: 1.8rem !important; font-weight: 800 !important; color: #222 !important; }}
    .stMarkdown h2 {{ font-size: 1.4rem !important; font-weight: 700 !important; color: #333 !important; }}
    .stMarkdown h3 {{ font-size: 1.15rem !important; font-weight: 700 !important; color: #444 !important; }}
    .stMarkdown strong {{ color: #222; }}

    /* ---------- HORIZONTAL RULE ---------- */
    hr {{ border: none; border-top: 2px solid rgba(0,0,0,0.08); margin: 1.5rem 0; }}

    /* ---------- FOOTER ---------- */
    .dashboard-footer {{
        text-align: center; color: #777; font-size: 0.88rem;
        padding: 2rem 0 1rem; border-top: 2px solid #e0e0e0;
        margin-top: 2rem;
    }}
    .dashboard-footer strong {{ color: #444; }}
    .dashboard-footer a {{ color: {COLORS['card_border']}; text-decoration: none; font-weight: 600; }}

    /* ---------- BACK TO TOP ---------- */
    .back-to-top {{
        position: fixed; bottom: 2rem; right: 2rem;
        background: {COLORS['card_border']}; color: white;
        width: 44px; height: 44px; border-radius: 50%;
        display: flex; align-items: center; justify-content: center;
        font-size: 1.3rem; font-weight: 700;
        box-shadow: 0 3px 14px rgba(0,0,0,0.25);
        cursor: pointer; z-index: 999; text-decoration: none;
        transition: transform 0.2s ease, opacity 0.2s ease;
        opacity: 0.9;
    }}
    .back-to-top:hover {{ transform: scale(1.1); opacity: 1; }}

    /* ---------- ANIMATIONS ---------- */
    @keyframes fadeInUp {{ from {{ opacity:0; transform:translateY(12px); }} to {{ opacity:1; transform:translateY(0); }} }}
    .stTabs [data-baseweb="tab-panel"] {{ animation: fadeInUp 0.35s ease-out; }}
    </style>

    <a href="#top" class="back-to-top" title="Back to top"></a>
    <div id="top"></div>
    """, unsafe_allow_html=True)

    # ---- SIDEBAR ----
    st.sidebar.markdown("""
    <div style="text-align:center; margin-bottom:0.8rem;">
        <span style="font-size:2.5rem;"></span>
        <h2 style="margin:0.2rem 0 0; font-size:1.15rem; color:#333;">COSME Dashboard</h2>
        <p style="margin:0; font-size:0.78rem; color:#888;">Baseline–Midline M&amp;E</p>
    </div>
    """, unsafe_allow_html=True)
    st.sidebar.markdown("---")

    # Dataset selector
    dataset = st.sidebar.radio(
        "Dataset View",
        ["Combined Overview", "Forestry Groups", "Women Survey", "Men Survey",
         "GJJ KAP \u2013 Women (Baseline/Endline)",
         "GJJ KAP \u2013 Men (Baseline/Endline)",
         "Forest Training (Pre/Post)", "Mangrove Training (Pre/Post)",
         "Seaweed Production & Challenges (2025)",
         "Project Outputs & Activity Indicators",
         "VSLA Functionality (Q1-Q4 2025)",
         "Insights"],
        index=0,
        help="Combined Overview shows headline KPIs from all datasets side by side."
    )

    show_change = st.sidebar.toggle("Show Change (pp) Charts", value=False,
                                     help="Toggle percentage-point change visualisation")

    st.sidebar.markdown("---")

    # ---- QUICK NAVIGATION in sidebar ----
    if dataset == "Forestry Groups":
        st.sidebar.markdown("**Quick Navigate**")
        st.sidebar.markdown("""
        <div class="sidebar-section">
            <span class="sidebar-nav-link">Overview & KPIs</span>
            <span class="sidebar-nav-link">Group Characteristics</span>
            <span class="sidebar-nav-link">Governance & Gender</span>
            <span class="sidebar-nav-link">Training & Assets</span>
            <span class="sidebar-nav-link">Forest Condition</span>
            <span class="sidebar-nav-link">Income & Agroforestry</span>
        </div>
        """, unsafe_allow_html=True)
    elif dataset == "Women Survey":
        st.sidebar.markdown("**Quick Navigate**")
        st.sidebar.markdown("""
        <div class="sidebar-section">
            <span class="sidebar-nav-link">Household Profile</span>
            <span class="sidebar-nav-link">Shocks & Preparedness</span>
            <span class="sidebar-nav-link">Assets & Savings</span>
            <span class="sidebar-nav-link">Roles & Decisions</span>
            <span class="sidebar-nav-link">Climate & NbS</span>
            <span class="sidebar-nav-link">Life Skills & Norms</span>
        </div>
        """, unsafe_allow_html=True)
    elif dataset == "Insights":
        st.sidebar.markdown("**Quick Navigate**")
        st.sidebar.markdown("""
        <div class="sidebar-section">
            <span class="sidebar-nav-link">Forestry Insights</span>
            <span class="sidebar-nav-link">Women Survey Insights</span>
            <span class="sidebar-nav-link">Men Survey Insights</span>
            <span class="sidebar-nav-link">GJJ KAP Women Insights</span>
            <span class="sidebar-nav-link">GJJ KAP Men Insights</span>
            <span class="sidebar-nav-link">Forest Training Insights</span>
            <span class="sidebar-nav-link">Mangrove Training Insights</span>
            <span class="sidebar-nav-link">Seaweed Production Insights</span>
            <span class="sidebar-nav-link">Cross-Cutting Insights</span>
            <span class="sidebar-nav-link">Indicator Change Heatmap</span>
            <span class="sidebar-nav-link">Recommendations</span>
        </div>
        """, unsafe_allow_html=True)
    elif dataset == "Men Survey":
        st.sidebar.markdown("**Quick Navigate**")
        st.sidebar.markdown("""
        <div class="sidebar-section">
            <span class="sidebar-nav-link">Household Profile</span>
            <span class="sidebar-nav-link">Climate & NbS</span>
            <span class="sidebar-nav-link">Support for Women in NbS</span>
            <span class="sidebar-nav-link">Roles & Time Use</span>
            <span class="sidebar-nav-link">Decision-Making</span>
            <span class="sidebar-nav-link">Social Norms</span>
        </div>
        """, unsafe_allow_html=True)
    elif dataset == "GJJ KAP \u2013 Women (Baseline/Endline)":
        st.sidebar.markdown("**Quick Navigate**")
        st.sidebar.markdown("""
        <div class="sidebar-section">
            <span class="sidebar-nav-link">SELF: Confidence & Compassion</span>
            <span class="sidebar-nav-link">Relational Wellbeing</span>
            <span class="sidebar-nav-link">Shared Responsibility</span>
            <span class="sidebar-nav-link">Shared Power & Decisions</span>
            <span class="sidebar-nav-link">Autonomy & Leadership</span>
        </div>
        """, unsafe_allow_html=True)
    elif dataset == "GJJ KAP \u2013 Men (Baseline/Endline)":
        st.sidebar.markdown("**Quick Navigate**")
        st.sidebar.markdown("""
        <div class="sidebar-section">
            <span class="sidebar-nav-link">SELF: Responsibility & Compassion</span>
            <span class="sidebar-nav-link">Relational Wellbeing</span>
            <span class="sidebar-nav-link">Shared Responsibility</span>
            <span class="sidebar-nav-link">Shared Power & Decisions</span>
            <span class="sidebar-nav-link">Leadership & Business Support</span>
        </div>
        """, unsafe_allow_html=True)
    elif dataset == "Forest Training (Pre/Post)":
        st.sidebar.markdown("**Quick Navigate**")
        st.sidebar.markdown("""
        <div class="sidebar-section">
            <span class="sidebar-nav-link">Overview of Training Outcomes</span>
            <span class="sidebar-nav-link">Knowledge by Question</span>
            <span class="sidebar-nav-link">Domain-Level Grouping</span>
        </div>
        """, unsafe_allow_html=True)
    elif dataset == "Mangrove Training (Pre/Post)":
        st.sidebar.markdown("**Quick Navigate**")
        st.sidebar.markdown("""
        <div class="sidebar-section">
            <span class="sidebar-nav-link">Overview & Knowledge Gains</span>
            <span class="sidebar-nav-link">County-Level Performance</span>
            <span class="sidebar-nav-link">Sex Disaggregation</span>
        </div>
        """, unsafe_allow_html=True)
    elif dataset == "Seaweed Production & Challenges (2025)":
        st.sidebar.markdown("**Quick Navigate**")
        st.sidebar.markdown("""
        <div class="sidebar-section">
            <span class="sidebar-nav-link">Overview & KPIs</span>
            <span class="sidebar-nav-link">Group Performance</span>
            <span class="sidebar-nav-link">Production & Yields</span>
            <span class="sidebar-nav-link">Challenges & Constraints</span>
            <span class="sidebar-nav-link">Map View</span>
        </div>
        """, unsafe_allow_html=True)
    elif dataset == "Project Outputs & Activity Indicators":
        st.sidebar.markdown("**Quick Navigate**")
        st.sidebar.markdown("""
        <div class="sidebar-section">
            <span class="sidebar-nav-link">High-Level KPIs</span>
            <span class="sidebar-nav-link">Mangrove Outputs</span>
            <span class="sidebar-nav-link">Seaweed Outputs</span>
            <span class="sidebar-nav-link">Forestry Outputs</span>
            <span class="sidebar-nav-link">GJJ Outputs</span>
        </div>
        """, unsafe_allow_html=True)
    elif dataset == "VSLA Functionality (Q1-Q4 2025)":
        st.sidebar.markdown("**Quick Navigate**")
        st.sidebar.markdown("""
        <div class="sidebar-section">
            <span class="sidebar-nav-link">Overview</span>
            <span class="sidebar-nav-link">Membership & Meetings</span>
            <span class="sidebar-nav-link">Savings & Social Fund</span>
            <span class="sidebar-nav-link">Loans & Repayment</span>
            <span class="sidebar-nav-link">County Comparison</span>
        </div>
        """, unsafe_allow_html=True)
    else:
        st.sidebar.markdown("**Quick Navigate**")
        st.sidebar.markdown("""
        <div class="sidebar-section">
            <span class="sidebar-nav-link">Forestry Headlines</span>
            <span class="sidebar-nav-link">Women Survey Headlines</span>
            <span class="sidebar-nav-link">Men Survey Headlines</span>
            <span class="sidebar-nav-link">GJJ KAP Women Headlines</span>
            <span class="sidebar-nav-link">GJJ KAP Men Headlines</span>
            <span class="sidebar-nav-link">Forest Training Headlines</span>
            <span class="sidebar-nav-link">Mangrove Training Headlines</span>
            <span class="sidebar-nav-link">Seaweed Production Headlines</span>
            <span class="sidebar-nav-link">Comparative Snapshots</span>
            <span class="sidebar-nav-link">Men vs Women Comparisons</span>
        </div>
        """, unsafe_allow_html=True)

    st.sidebar.markdown("---")

    # ---- LOAD DATA ----
    script_dir = os.path.dirname(os.path.abspath(__file__))
    forestry_path = os.path.join(script_dir, FORESTRY_EXCEL)
    women_path = os.path.join(script_dir, WOMEN_EXCEL)
    men_path = os.path.join(script_dir, MEN_EXCEL)
    gjj_kap_path = os.path.join(script_dir, GJJ_KAP_WOMEN_EXCEL)
    gjj_kap_men_path = os.path.join(script_dir, GJJ_KAP_MEN_EXCEL)
    forest_training_path = os.path.join(script_dir, FOREST_TRAINING_EXCEL)
    mangrove_training_path = os.path.join(script_dir, MANGROVE_TRAINING_EXCEL)
    seaweed_path = os.path.join(script_dir, SEAWEED_CSV)
    project_outputs_path = os.path.join(script_dir, PROJECT_OUTPUTS_EXCEL)
    vsla_path = os.path.join(script_dir, VSLA_EXCEL)

    # ---- HEADER ----
    if dataset == "Forestry Groups":
        st.markdown("""<div class="main-header">
            <h1>Community Forest Conservation Dashboard</h1>
            <p>Baseline vs Midline Assessment | Forestry Conservation Groups Functionality</p>
        </div>""", unsafe_allow_html=True)

        # Breadcrumb
        st.markdown('<div class="nav-breadcrumb"><span>COSME</span><span class="sep">›</span>'
                    '<span class="active">Forestry Groups</span></div>', unsafe_allow_html=True)

        data = load_forestry_data(forestry_path)

        # Sidebar KPI summary
        bl_grp = data['num_groups'].loc[data['num_groups']['Timepoint']=='Baseline','Groups_Assessed'].values[0]
        ml_grp = data['num_groups'].loc[data['num_groups']['Timepoint']=='Midline','Groups_Assessed'].values[0]
        st.sidebar.markdown("**Dataset Summary**")
        st.sidebar.markdown(f"""
        <div class="sidebar-section">
            <div style="display:flex; justify-content:space-between; margin-bottom:0.3rem;">
                <span style="font-size:0.82rem; color:#666;">Baseline Groups</span>
                <strong>{int(bl_grp)}</strong>
            </div>
            <div style="display:flex; justify-content:space-between; margin-bottom:0.3rem;">
                <span style="font-size:0.82rem; color:#666;">Midline Groups</span>
                <strong>{int(ml_grp)}</strong>
            </div>
            <div style="display:flex; justify-content:space-between;">
                <span style="font-size:0.82rem; color:#666;">Source File</span>
                <span style="font-size:0.75rem; color:#999;">Forestry Excel</span>
            </div>
        </div>
        """, unsafe_allow_html=True)

        # CSV Downloads
        st.sidebar.markdown("---")
        st.sidebar.markdown("**Export Data**")
        st.sidebar.download_button(
            label="Download Forestry Data (CSV)",
            data=_export_data_csv(data, 'forestry'),
            file_name='forestry_data_export.csv',
            mime='text/csv',
            use_container_width=True,
        )

        render_forestry_tabs(data, show_change)

    elif dataset == "Women Survey":
        st.markdown("""<div class="main-header">
            <h1>Women's Survey Dashboard</h1>
            <p>Baseline vs Midline Assessment | Household-Level Women's Empowerment &amp; Climate Resilience</p>
        </div>""", unsafe_allow_html=True)

        # Breadcrumb
        st.markdown('<div class="nav-breadcrumb"><span>COSME</span><span class="sep">›</span>'
                    '<span class="active">Women Survey</span></div>', unsafe_allow_html=True)

        w = load_women_data(women_path)

        # Sidebar sample size
        st.sidebar.markdown("**Dataset Summary**")
        st.sidebar.markdown("""
        <div class="sidebar-section">
            <div style="display:flex; justify-content:space-between; margin-bottom:0.3rem;">
                <span style="font-size:0.82rem; color:#666;">Baseline N</span>
                <strong>707 women</strong>
            </div>
            <div style="display:flex; justify-content:space-between; margin-bottom:0.3rem;">
                <span style="font-size:0.82rem; color:#666;">Midline N</span>
                <strong>320 women</strong>
            </div>
            <div style="display:flex; justify-content:space-between;">
                <span style="font-size:0.82rem; color:#666;">Source File</span>
                <span style="font-size:0.75rem; color:#999;">Women Survey Excel</span>
            </div>
        </div>
        """, unsafe_allow_html=True)

        # CSV Downloads
        st.sidebar.markdown("---")
        st.sidebar.markdown("**Export Data**")
        st.sidebar.download_button(
            label="Download Women Survey Data (CSV)",
            data=_export_data_csv(w, 'women'),
            file_name='women_survey_data_export.csv',
            mime='text/csv',
            use_container_width=True,
        )

        wt1, wt2, wt3, wt4, wt5, wt6 = st.tabs([
            "Household Profile & Services",
            "Shocks, Coping & Preparedness",
            "Assets, Land, Savings & Loans",
            "Roles, Time Use & Decisions",
            "Climate Change & NbS",
            "Life Skills & Social Norms"
        ])
        with wt1: render_women_tab1(w)
        with wt2: render_women_tab2(w)
        with wt3: render_women_tab3(w)
        with wt4: render_women_tab4(w)
        with wt5: render_women_tab5(w)
        with wt6: render_women_tab6(w)

    elif dataset == "Men Survey":
        st.markdown("""<div class="main-header">
            <h1>Men's Survey Dashboard</h1>
            <p>Baseline vs Midline Assessment | Men's Perspectives on Gender, Climate &amp; NbS</p>
        </div>""", unsafe_allow_html=True)

        # Breadcrumb
        st.markdown('<div class="nav-breadcrumb"><span>COSME</span><span class="sep">›</span>'
                    '<span class="active">Men Survey</span></div>', unsafe_allow_html=True)

        m = load_men_data(men_path)

        # Sidebar sample size
        st.sidebar.markdown("**Dataset Summary**")
        st.sidebar.markdown("""
        <div class="sidebar-section">
            <div style="display:flex; justify-content:space-between; margin-bottom:0.3rem;">
                <span style="font-size:0.82rem; color:#666;">Baseline N</span>
                <strong>661 men</strong>
            </div>
            <div style="display:flex; justify-content:space-between; margin-bottom:0.3rem;">
                <span style="font-size:0.82rem; color:#666;">Midline N</span>
                <strong>176 men</strong>
            </div>
            <div style="display:flex; justify-content:space-between;">
                <span style="font-size:0.82rem; color:#666;">Source File</span>
                <span style="font-size:0.75rem; color:#999;">Men Survey Excel</span>
            </div>
        </div>
        """, unsafe_allow_html=True)

        # CSV Downloads
        st.sidebar.markdown("---")
        st.sidebar.markdown("**Export Data**")
        st.sidebar.download_button(
            label="Download Men Survey Data (CSV)",
            data=_export_data_csv(m, 'men'),
            file_name='men_survey_data_export.csv',
            mime='text/csv',
            use_container_width=True,
        )

        mt1, mt2, mt3, mt4, mt5, mt6 = st.tabs([
            "Household Profile",
            "Climate Change & NbS",
            "Support for Women in NbS",
            "Roles & Time Use",
            "Decision-Making",
            "Social Norms"
        ])
        with mt1: render_men_tab1(m)
        with mt2: render_men_tab2(m)
        with mt3: render_men_tab3(m)
        with mt4: render_men_tab4(m)
        with mt5: render_men_tab5(m)
        with mt6: render_men_tab6(m)

    elif dataset == "GJJ KAP \u2013 Women (Baseline/Endline)":
        st.markdown("""<div class="main-header">
            <h1>GJJ KAP Women Dashboard</h1>
            <p>Baseline vs Endline Assessment | Knowledge, Attitudes &amp; Practices — Gender Justice Journey</p>
        </div>""", unsafe_allow_html=True)

        # Breadcrumb
        st.markdown('<div class="nav-breadcrumb"><span>COSME</span><span class="sep">\u203a</span>'
                    '<span class="active">GJJ KAP \u2013 Women</span></div>', unsafe_allow_html=True)

        gjj = load_gjj_kap_women_data(gjj_kap_path)

        # Sidebar dataset summary
        st.sidebar.markdown("**Dataset Summary**")
        st.sidebar.markdown("""
        <div class="sidebar-section">
            <div style="display:flex; justify-content:space-between; margin-bottom:0.3rem;">
                <span style="font-size:0.82rem; color:#666;">Survey Type</span>
                <strong>KAP Women</strong>
            </div>
            <div style="display:flex; justify-content:space-between; margin-bottom:0.3rem;">
                <span style="font-size:0.82rem; color:#666;">Timepoints</span>
                <strong>Baseline &amp; Endline</strong>
            </div>
            <div style="display:flex; justify-content:space-between;">
                <span style="font-size:0.82rem; color:#666;">Source File</span>
                <span style="font-size:0.75rem; color:#999;">GJJ KAP Women Excel</span>
            </div>
        </div>
        """, unsafe_allow_html=True)

        # CSV Download
        st.sidebar.markdown("---")
        st.sidebar.markdown("**Export Data**")
        st.sidebar.download_button(
            label="Download GJJ KAP Women Data (CSV)",
            data=_export_data_csv(gjj, 'gjj_kap_women'),
            file_name='gjj_kap_women_data_export.csv',
            mime='text/csv',
            use_container_width=True,
        )

        gt1, gt2, gt3, gt4, gt5 = st.tabs([
            "SELF: Confidence & Compassion",
            "Relational Wellbeing",
            "Shared Responsibility",
            "Shared Power & Decisions",
            "Autonomy & Leadership"
        ])
        with gt1: render_gjj_tab1(gjj)
        with gt2: render_gjj_tab2(gjj)
        with gt3: render_gjj_tab3(gjj)
        with gt4: render_gjj_tab4(gjj)
        with gt5: render_gjj_tab5(gjj)

    elif dataset == "GJJ KAP \u2013 Men (Baseline/Endline)":
        st.markdown("""<div class="main-header">
            <h1>GJJ KAP Men Dashboard</h1>
            <p>Baseline vs Endline Assessment | Knowledge, Attitudes &amp; Practices — Gender Justice Journey (Men)</p>
        </div>""", unsafe_allow_html=True)

        # Breadcrumb
        st.markdown('<div class="nav-breadcrumb"><span>COSME</span><span class="sep">\u203a</span>'
                    '<span class="active">GJJ KAP \u2013 Men</span></div>', unsafe_allow_html=True)

        gjj_m = load_gjj_kap_men_data(gjj_kap_men_path)

        # Sidebar dataset summary
        st.sidebar.markdown("**Dataset Summary**")
        st.sidebar.markdown("""
        <div class="sidebar-section">
            <div style="display:flex; justify-content:space-between; margin-bottom:0.3rem;">
                <span style="font-size:0.82rem; color:#666;">Survey Type</span>
                <strong>KAP Men</strong>
            </div>
            <div style="display:flex; justify-content:space-between; margin-bottom:0.3rem;">
                <span style="font-size:0.82rem; color:#666;">Timepoints</span>
                <strong>Baseline &amp; Endline</strong>
            </div>
            <div style="display:flex; justify-content:space-between;">
                <span style="font-size:0.82rem; color:#666;">Source File</span>
                <span style="font-size:0.75rem; color:#999;">GJJ KAP Men Excel</span>
            </div>
        </div>
        """, unsafe_allow_html=True)

        # CSV Download
        st.sidebar.markdown("---")
        st.sidebar.markdown("**Export Data**")
        st.sidebar.download_button(
            label="Download GJJ KAP Men Data (CSV)",
            data=_export_data_csv(gjj_m, 'gjj_kap_men'),
            file_name='gjj_kap_men_data_export.csv',
            mime='text/csv',
            use_container_width=True,
        )

        gmt1, gmt2, gmt3, gmt4, gmt5 = st.tabs([
            "SELF: Responsibility & Compassion",
            "Relational Wellbeing",
            "Shared Responsibility",
            "Shared Power & Decisions",
            "Leadership & Business Support"
        ])
        with gmt1: render_gjj_men_tab1(gjj_m)
        with gmt2: render_gjj_men_tab2(gjj_m)
        with gmt3: render_gjj_men_tab3(gjj_m)
        with gmt4: render_gjj_men_tab4(gjj_m)
        with gmt5: render_gjj_men_tab5(gjj_m)

    elif dataset == "Forest Training (Pre/Post)":
        st.markdown("""<div class="main-header">
            <h1>Forest Training Knowledge Assessment</h1>
            <p>Pre-Training vs Post-Training | Forest Conservation Knowledge Test Results</p>
        </div>""", unsafe_allow_html=True)

        # Breadcrumb
        st.markdown('<div class="nav-breadcrumb"><span>COSME</span><span class="sep">\u203a</span>'
                    '<span class="active">Forest Training (Pre/Post)</span></div>', unsafe_allow_html=True)

        t_data = load_forest_training_data(forest_training_path)

        # Sidebar dataset summary
        bl_n = int(t_data['scores'][t_data['scores']['Timepoint']=='Baseline']['Respondents'].values[0])
        el_n = int(t_data['scores'][t_data['scores']['Timepoint']=='Endline']['Respondents'].values[0])
        st.sidebar.markdown("**Dataset Summary**")
        st.sidebar.markdown(f"""
        <div class="sidebar-section">
            <div style="display:flex; justify-content:space-between; margin-bottom:0.3rem;">
                <span style="font-size:0.82rem; color:#666;">Pre-Training N</span>
                <strong>{bl_n:,}</strong>
            </div>
            <div style="display:flex; justify-content:space-between; margin-bottom:0.3rem;">
                <span style="font-size:0.82rem; color:#666;">Post-Training N</span>
                <strong>{el_n:,}</strong>
            </div>
            <div style="display:flex; justify-content:space-between; margin-bottom:0.3rem;">
                <span style="font-size:0.82rem; color:#666;">Questions</span>
                <strong>{len(t_data['questions'])}</strong>
            </div>
            <div style="display:flex; justify-content:space-between;">
                <span style="font-size:0.82rem; color:#666;">Source File</span>
                <span style="font-size:0.75rem; color:#999;">Forest Training Excel</span>
            </div>
        </div>
        """, unsafe_allow_html=True)

        # Timepoint filter
        st.sidebar.markdown("---")
        tp_filter = st.sidebar.radio(
            "Timepoint View",
            ["Combined", "Baseline", "Endline"],
            index=0,
            help="Show Baseline only, Endline only, or both side by side."
        )

        # CSV Download
        st.sidebar.markdown("---")
        st.sidebar.markdown("**Export Data**")
        st.sidebar.download_button(
            label="Download Forest Training Data (CSV)",
            data=_export_data_csv(t_data, 'forest_training'),
            file_name='forest_training_data_export.csv',
            mime='text/csv',
            use_container_width=True,
        )

        render_forest_training_tabs(t_data, tp_filter)

    elif dataset == "Mangrove Training (Pre/Post)":
        st.markdown("""<div class="main-header">
            <h1>Mangrove Training Knowledge Assessment</h1>
            <p>Pre-Training vs Post-Training | Mangrove Restoration Knowledge Test Results</p>
        </div>""", unsafe_allow_html=True)

        # Breadcrumb
        st.markdown('<div class="nav-breadcrumb"><span>COSME</span><span class="sep">\u203a</span>'
                    '<span class="active">Mangrove Training (Pre/Post)</span></div>', unsafe_allow_html=True)

        mg_data = load_mangrove_training_data(mangrove_training_path)

        # Sidebar dataset summary
        mg_all_scores = mg_data['scores'][mg_data['scores']['County'] == 'All']
        mg_n = int(mg_all_scores['Respondents'].values[0]) if len(mg_all_scores) else 0
        counties_count = len([c for c in mg_data['scores']['County'].unique() if c != 'All'])
        st.sidebar.markdown("**Dataset Summary**")
        st.sidebar.markdown(f"""
        <div class="sidebar-section">
            <div style="display:flex; justify-content:space-between; margin-bottom:0.3rem;">
                <span style="font-size:0.82rem; color:#666;">Total Respondents</span>
                <strong>{mg_n:,}</strong>
            </div>
            <div style="display:flex; justify-content:space-between; margin-bottom:0.3rem;">
                <span style="font-size:0.82rem; color:#666;">Counties</span>
                <strong>{counties_count}</strong>
            </div>
            <div style="display:flex; justify-content:space-between;">
                <span style="font-size:0.82rem; color:#666;">Source File</span>
                <span style="font-size:0.75rem; color:#999;">Mangrove Training Excel</span>
            </div>
        </div>
        """, unsafe_allow_html=True)

        # Timepoint filter
        st.sidebar.markdown("---")
        tp_filter = st.sidebar.radio(
            "Timepoint View",
            ["Combined", "Baseline", "Endline"],
            index=0,
            help="Show Pre-Test only, Post-Test only, or both side by side."
        )

        # CSV Download
        st.sidebar.markdown("---")
        st.sidebar.markdown("**Export Data**")
        st.sidebar.download_button(
            label="Download Mangrove Training Data (CSV)",
            data=_export_data_csv(mg_data, 'mangrove_training'),
            file_name='mangrove_training_data_export.csv',
            mime='text/csv',
            use_container_width=True,
        )

        render_mangrove_training_tabs(mg_data, tp_filter)

    elif dataset == "Seaweed Production & Challenges (2025)":
        st.markdown("""<div class="main-header">
            <h1>Seaweed Production & Challenges Dashboard</h1>
            <p>Seaweed Group Performance, Production Metrics & Constraint Analysis | 2025 Data Collection</p>
        </div>""", unsafe_allow_html=True)

        # Breadcrumb
        st.markdown('<div class="nav-breadcrumb"><span>COSME</span><span class="sep">\u203a</span>'
                    '<span class="active">Seaweed Production & Challenges (2025)</span></div>', unsafe_allow_html=True)

        sw_df = load_seaweed_data(seaweed_path)

        # Sidebar filters
        st.sidebar.markdown("---")
        st.sidebar.markdown("**Seaweed Filters**")
        all_groups = sorted(sw_df['Group'].dropna().unique().tolist())
        sw_group_filter = st.sidebar.multiselect("Filter by Group", all_groups, default=all_groups)
        sw_casual_filter = st.sidebar.radio("Casual Workers", ["All", "Yes", "No"], index=0,
                                              help="Filter by whether farmer uses casual workers.")
        # Challenge flag checkboxes
        challenge_flags = [c for c in sw_df.columns if c.startswith('Challenge_')]
        challenge_labels = [c.replace('Challenge_', '').replace('_', ' ') for c in challenge_flags]
        sw_challenge_filter = st.sidebar.multiselect("Challenge Flags (include if any selected)",
                                                      challenge_labels, default=[])
        # Numeric sliders
        max_kg = int(sw_df['Total_KG'].max()) + 1 if sw_df['Total_KG'].max() > 0 else 100
        sw_min_kg = st.sidebar.slider("Min Total Production (kg)", 0, max_kg, 0)
        sw_min_ach = st.sidebar.slider("Min Ropes Achievement (%)", 0, 100, 0)

        # Sidebar dataset summary
        st.sidebar.markdown("---")
        st.sidebar.markdown("**Dataset Summary**")
        st.sidebar.markdown(f"""
        <div class="sidebar-section">
            <div style="display:flex; justify-content:space-between; margin-bottom:0.3rem;">
                <span style="font-size:0.82rem; color:#666;">Total Farmers</span>
                <strong>{len(sw_df):,}</strong>
            </div>
            <div style="display:flex; justify-content:space-between; margin-bottom:0.3rem;">
                <span style="font-size:0.82rem; color:#666;">Groups</span>
                <strong>{sw_df['Group'].nunique()}</strong>
            </div>
            <div style="display:flex; justify-content:space-between; margin-bottom:0.3rem;">
                <span style="font-size:0.82rem; color:#666;">Total Production</span>
                <strong>{sw_df['Total_KG'].sum():,.0f} kg</strong>
            </div>
            <div style="display:flex; justify-content:space-between;">
                <span style="font-size:0.82rem; color:#666;">Source File</span>
                <span style="font-size:0.75rem; color:#999;">Seaweed CSV (2025)</span>
            </div>
        </div>
        """, unsafe_allow_html=True)

        # CSV Download
        st.sidebar.markdown("---")
        st.sidebar.markdown("**Export Data**")
        st.sidebar.download_button(
            label="Download Seaweed Data (CSV)",
            data=_export_data_csv({'seaweed_production': sw_df}, 'seaweed'),
            file_name='seaweed_production_data_export.csv',
            mime='text/csv',
            use_container_width=True,
        )

        # Map challenge filter labels back to column names
        sw_challenge_cols = [challenge_flags[i] for i, lbl in enumerate(challenge_labels)
                             if lbl in sw_challenge_filter] if sw_challenge_filter else None

        render_seaweed_tabs(sw_df, group_filter=sw_group_filter if sw_group_filter else None,
                            casual_filter=sw_casual_filter,
                            challenge_filter=sw_challenge_cols,
                            min_total_kg=sw_min_kg,
                            min_achievement_pct=sw_min_ach)

    elif dataset == "Project Outputs & Activity Indicators":
        st.markdown("""<div class="main-header">
            <h1>Project Outputs & Activity Indicators</h1>
            <p>COSME Output & Activity Data | Mangrove, Seaweed, Forestry, GJJ | SAR Y3 Reporting Period</p>
        </div>""", unsafe_allow_html=True)

        # Breadcrumb
        st.markdown('<div class="nav-breadcrumb"><span>COSME</span><span class="sep">›</span>'
                    '<span class="active">Project Outputs & Activity Indicators</span></div>', unsafe_allow_html=True)

        po_data = load_project_outputs(project_outputs_path)

        st.sidebar.markdown("**Dataset Summary**")
        st.sidebar.markdown("""
        <div class="sidebar-section">
            <div style="margin-bottom:0.3rem;">Mangrove Outputs</div>
            <div style="margin-bottom:0.3rem;">Seaweed Outputs</div>
            <div style="margin-bottom:0.3rem;">Forestry Outputs</div>
            <div>GJJ Outputs</div>
        </div>
        """, unsafe_allow_html=True)

        render_project_outputs_tabs(po_data)

    elif dataset == "VSLA Functionality (Q1-Q4 2025)":
        st.markdown("""<div class="main-header">
            <h1>VSLA Functionality Dashboard</h1>
            <p>VSLA Savings, Loans, Social Fund & Membership Tracking | Q2–Q4 2025 | Kilifi & Kwale Counties</p>
        </div>""", unsafe_allow_html=True)

        # Breadcrumb
        st.markdown('<div class="nav-breadcrumb"><span>COSME</span><span class="sep">›</span>'
                    '<span class="active">VSLA Functionality (Q1-Q4 2025)</span></div>', unsafe_allow_html=True)

        vsla_data = load_vsla_data(vsla_path)

        # Sidebar filters
        st.sidebar.markdown("---")
        st.sidebar.markdown("**VSLA Filters**")
        vsla_county = st.sidebar.selectbox("County", ['All Counties', 'Kilifi', 'Kwale'], index=0)
        vsla_quarter = st.sidebar.selectbox("Quarter", ['Q4', 'Q3', 'Q2'], index=0)

        # Sidebar summary
        all_vslas_q4 = _vsla_get_cq(vsla_data['vslasAssessed'], 'All Counties', 'Q4')
        all_mem_q4 = _vsla_get_sa(vsla_data['membership']['all'], 'All Counties', 'Q4', 'sum')
        all_sav_q4 = _vsla_get_sa(vsla_data['savings']['value'], 'All Counties', 'Q4', 'sum')
        st.sidebar.markdown("---")
        st.sidebar.markdown("**Dataset Summary**")
        st.sidebar.markdown(f"""
        <div class="sidebar-section">
            <div style="display:flex; justify-content:space-between; margin-bottom:0.3rem;">
                <span style="font-size:0.82rem; color:#666;">VSLAs (Q4)</span>
                <strong>{int(all_vslas_q4):,}</strong>
            </div>
            <div style="display:flex; justify-content:space-between; margin-bottom:0.3rem;">
                <span style="font-size:0.82rem; color:#666;">Total Members (Q4)</span>
                <strong>{int(all_mem_q4):,}</strong>
            </div>
            <div style="display:flex; justify-content:space-between; margin-bottom:0.3rem;">
                <span style="font-size:0.82rem; color:#666;">Total Savings (Q4)</span>
                <strong>{_vsla_fmt_kes(all_sav_q4)}</strong>
            </div>
            <div style="display:flex; justify-content:space-between;">
                <span style="font-size:0.82rem; color:#666;">Source File</span>
                <span style="font-size:0.75rem; color:#999;">VSLA Excel (2025)</span>
            </div>
        </div>
        """, unsafe_allow_html=True)

        render_vsla_tabs(vsla_data, county_filter=vsla_county, quarter_filter=vsla_quarter)

    elif dataset == "Insights":
        st.markdown("""<div class="main-header">
            <h1>COSME Dashboard Insights</h1>
            <p>Automated Data-Driven Insights | Forestry, Women's, Men's, GJJ KAP Women &amp; Men, Forest Training, Mangrove Training, Seaweed Production Analysis</p>
        </div>""", unsafe_allow_html=True)

        # Breadcrumb
        st.markdown('<div class="nav-breadcrumb"><span>COSME</span><span class="sep">›</span>'
                    '<span class="active">Insights</span></div>', unsafe_allow_html=True)

        f_data = load_forestry_data(forestry_path)
        w_data = load_women_data(women_path)
        m_data = load_men_data(men_path)
        gjj_data = load_gjj_kap_women_data(gjj_kap_path)
        gjj_men_data = load_gjj_kap_men_data(gjj_kap_men_path)
        ft_data = load_forest_training_data(forest_training_path)
        mg_data = load_mangrove_training_data(mangrove_training_path)
        sw_data = load_seaweed_data(seaweed_path)

        st.sidebar.markdown("**Datasets Loaded**")
        st.sidebar.markdown("""
        <div class="sidebar-section">
            <div style="margin-bottom:0.3rem;">Forestry Groups (analysis)</div>
            <div style="margin-bottom:0.3rem;">Women Survey (analysis)</div>
            <div style="margin-bottom:0.3rem;">Men Survey (analysis)</div>
            <div style="margin-bottom:0.3rem;">GJJ KAP Women (analysis)</div>
            <div style="margin-bottom:0.3rem;">GJJ KAP Men (analysis)</div>
            <div style="margin-bottom:0.3rem;">Forest Training (analysis)</div>
            <div style="margin-bottom:0.3rem;">Mangrove Training (analysis)</div>
            <div>Seaweed Production (analysis)</div>
        </div>
        """, unsafe_allow_html=True)

        render_insights_tab(f_data, w_data, m_data, gjj_data, gjj_men_data, ft_data, mg_data, sw_data=sw_data)

    else:
        # Combined Overview
        st.markdown("""<div class="main-header">
            <h1>COSME Baseline–Midline Dashboard</h1>
            <p>Integrated M&amp;E Analysis | Forestry, Women's, Men's, GJJ KAP Women &amp; Men, Forest Training, Mangrove Training, Seaweed Production</p>
        </div>""", unsafe_allow_html=True)

        # Breadcrumb
        st.markdown('<div class="nav-breadcrumb"><span class="active">COSME — Combined Overview</span></div>',
                    unsafe_allow_html=True)

        f_data = load_forestry_data(forestry_path)
        w_data = load_women_data(women_path)
        m_data = load_men_data(men_path)
        gjj_data = load_gjj_kap_women_data(gjj_kap_path)
        gjj_men_data = load_gjj_kap_men_data(gjj_kap_men_path)
        ft_data = load_forest_training_data(forest_training_path)
        mg_data = load_mangrove_training_data(mangrove_training_path)
        sw_data = load_seaweed_data(seaweed_path)

        st.sidebar.markdown("**Datasets Loaded**")
        st.sidebar.markdown("""
        <div class="sidebar-section">
            <div style="margin-bottom:0.3rem;">Forestry Groups</div>
            <div style="margin-bottom:0.3rem;">Women Survey</div>
            <div style="margin-bottom:0.3rem;">Men Survey</div>
            <div style="margin-bottom:0.3rem;">GJJ KAP Women</div>
            <div style="margin-bottom:0.3rem;">GJJ KAP Men</div>
            <div style="margin-bottom:0.3rem;">Forest Training</div>
            <div style="margin-bottom:0.3rem;">Mangrove Training</div>
            <div>Seaweed Production</div>
        </div>
        """, unsafe_allow_html=True)

        render_synthesis_view(f_data, w_data, m_data, gjj_data, gjj_men_data, ft_data, mg_data, sw_data=sw_data)

    # ---- FOOTER ----
    st.markdown(f"""
    <div class="dashboard-footer">
        <strong>COSME Baseline–Midline Dashboard</strong><br>
        Community Forest Conservation Groups, Women's Survey, Men's Survey, GJJ KAP Women &amp; Men, Forest Training, Mangrove Training, Seaweed Production | Built with Streamlit + Plotly<br>
        <span style="font-size:0.75rem;">Last updated: February 2026</span>
    </div>""", unsafe_allow_html=True)


if __name__ == "__main__":
    main()

# ============================================================================
# HOW TO RUN:
# streamlit run cosme_dashboard.py
#
# Requirements:
# pip install streamlit pandas numpy plotly openpyxl
#
# Data files (same folder as this script):
# 1. Forest Functionality Basline_midline results.xlsx (sheet: Results)
# 2. Women Survey Basline_midline results.xlsx (sheet: Results Women)
# 3. Men Survey Basline_midline results.xlsx (sheet: Results Men)
# 4. GJJ KAP Women Basline_endline results.xlsx (sheet: Results KAP Women Endline)
# 5. GJJ KAP Men Basline_endline results.xlsx (sheet: Results KAP Men Endline)
# 6. Forest Training Pre_post results.xlsx (sheet: Results)
#
# To adjust data mappings:
# - load_forestry_data(): row/col positions for forestry indicators
# - load_women_data(): row/col positions for women survey indicators
# - load_men_data(): row/col positions for men survey indicators
# - load_gjj_kap_women_data(): row/col positions for GJJ KAP women indicators
# - load_gjj_kap_men_data(): row/col positions for GJJ KAP men indicators
# - Each uses _val(raw, row_0based, col_0based)
#
# To add/remove indicators:
# - Add new DataFrame in the appropriate loader function
# - Add chart call in the corresponding render function
# ============================================================================
